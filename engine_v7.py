"""
engine_v7.py — 이지스 기동전단 통합 방어 시뮬레이터 v7.0
시간 스텝 기반 양방향 교전 엔진

v7.0 핵심 변경:
  · 이벤트 기반 → 시간 스텝(1초 단위) 시뮬레이션
  · 양방향 교전: 아군 공격(해성/하푼) + 적 SAM/CIWS 방어
  · 2D 위치 좌표계 (모든 엔티티 실시간 위치 추적)
  · SimFrame 단위 상태 기록 (애니메이션 지원)

v7.0 패치 이력:
  · 1단계: 시간 스텝 루프 + 위치 모델 + 기본 교전 판정
  · 2단계: 양방향 교전 검증
    - SAM alive=False 조기 설정 버그 수정 (요격 판정 스킵 문제)
    - INTERCEPT_DIST_M 300m → 2000m (1초 스텝 해상도 반영)
    - 대함 탐지 거리 수정 (자함 레이더 45km → 전술 인식 detect_km 병용)
    - 적 함정 시작 위치 수정 (1.8x → 1.0x, 교전 즉시 개시)
    - 적 함정 재발사 허용 (비행 중 미사일 소진 후 재장전)
  · 3단계: ENEMY_DB 전 32종 위협 지원
    - EnemyShipObj → EnemyThreatObj (항공기/함정/잠수함/독립미사일 통합)
    - 항공기: 접근 → 사거리 내 발사 → 이탈 행동 패턴
    - 탄도/순항/HGV/QBM 독립 미사일: _build_enemies()에서 MissileObj로 직접 생성
    - 대잠전: 홍상어/청상어 ASW 운용, 소나 탐지 범위 반영
    - _select_defense_wpn(): 고도/유형 인식 (SM-3 HGV/탄도, SM-6 QBM)
    - _friendly_strike(): 수상함(해성/하푼) + 잠수함(홍상어/청상어) 분리
  · 4단계: 몬테카를로 분석 + 보고서 생성
    - matplotlib 한글 폰트 설정 (Malgun Gothic)
    - monte_carlo_v7(): N회 반복 통계 집계 (요격률/피격/격침/비용)
    - plot_v7(): 6개 서브플롯 PNG 차트 (히스토그램·무기소모·수치요약)
    - save_excel_report_v7(): 4시트 Excel 보고서 (MC요약/무기소모/교전로그/차트)
    - __main__ 인수: python engine_v7.py [시나리오] [MC횟수]
  · 포팅 A: 방어 무기 재고 수동 설정 + 적군 편대 모드
    - _build_friendly(): sm3/sm6/sm2/ram/홍상어/청상어/mk46 수동 재고 지원 (cfg 키로 설정)
    - _build_enemies(): enemy_fleet_mode 4종 지원 (single/preset/custom/random)
    - ENEMY_FLEET_PRESETS / ENEMY_FLEET_RANDOM_CFG / generate_random_enemy_fleet 연동
  · 포팅 B: 전술 기능 — ECM·종말회피·음향기만기·함정회피·적 자체방어
    - ECM_REF_RANGE_M=50km 기준 거리 반비례 Pk 감소, Pk 하한 50% (탄도/HGV 제외)
    - 종말 회피: 20km 이내 terminal_evasion_factor 적용 (ENEMY_DB missile_terminal_evasion)
    - 음향 기만기: DECOY_PK=0.60, 어뢰 전용, decoy_stock 소모
    - 함정 회피 기동: SHIP_EVASION_PK=0.30, 어뢰 전용
    - 적 자체방어: CIWS(enemy_ciws_pk) 요격 → 채프/플레어(self_defense_pk) Pk 감소
  · 포팅 C: 항공 자산 대잠 운용 (FriendlyAircraftObj + _aircraft_asw())
    - AW-159 와일드캣: 함재 헬기, 청상어 2발, 사거리 140km, 폭풍/태풍/농무 불가
    - P-3C 오라이온: 포항기지 출격(+300km), Mk.46 4발, 소노부이 탐지+15km
    - P-8A 포세이돈: 포항기지 출격(+300km), Mk.46 5발, 소노부이 탐지+18km
  · 포팅 D: 분석 기능 — REQ 판정 + 날씨 비교 + A vs B + 저장/불러오기
    - evaluate_req_v7(): REQ-01~08 8항목 시간스텝 기반 판정
    - scenario_comparison_v7(): 날씨 3종(맑음/흐림/폭풍) MC 비교
    - compare_ab_v7(): 두 cfg MC 결과 대비 (Δ요격률·Δ비용)
    - save_scenario_v7() / load_scenario_v7(): JSON 시나리오 저장/불러오기
    - _compile(): remaining_inventory·total_channels·peak_concurrent_threats·t_first_fire 추가
"""

import math, random, os, time as _time, dataclasses
from typing import List, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.font_manager as fm
from matplotlib.figure import Figure as _MplFigure
from matplotlib.backends.backend_agg import FigureCanvasAgg as _FigureCanvasAgg
from matplotlib.ticker import FuncFormatter as _FuncFormatter
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from openpyxl.drawing.image import Image as XLImage
    _CAN_IMG = True
except Exception:
    _CAN_IMG = False

for _fp in ['C:/Windows/Fonts/malgun.ttf', 'C:/Windows/Fonts/malgunbd.ttf']:
    if os.path.exists(_fp):
        fm.fontManager.addfont(_fp)
matplotlib.rcParams['font.family'] = 'Malgun Gothic'
matplotlib.rcParams['axes.unicode_minus'] = False

from engine import (
    ENEMY_DB, FRIENDLY_DB, FRIENDLY_AIRCRAFT_DB, WEATHER_DB,
    SHIP_DB, FLEET_PRESETS,
    ENEMY_FLEET_PRESETS, ENEMY_FLEET_RANDOM_CFG, MIXED_ATTACK_SCENARIOS,
    generate_random_enemy_fleet,
    calculate_detect_range_by_rcs,
    normalize_enemy_db,
)
try:
    from ocean_acoustic_db import get_sonar_depth_factor
    _OCEAN_ACOUSTIC_OK = True
except ImportError:
    _OCEAN_ACOUSTIC_OK = False

try:
    from ocean_environment_db import (
        RADAR_CLUTTER, SONAR_AMBIENT_NOISE, get_current_vector as _get_current_vector
    )
    _RADAR_BF_FACTOR = RADAR_CLUTTER['detection_range_factor']
    _SONAR_BF_FACTOR = SONAR_AMBIENT_NOISE['sonar_range_factor']
    _OCEAN_ENV_OK = True
except ImportError:
    _RADAR_BF_FACTOR = {}
    _SONAR_BF_FACTOR = {}
    _get_current_vector = None
    _OCEAN_ENV_OK = False

try:
    from terrain_db import STRAITS_DB
    _TERRAIN_DB_OK = True
except ImportError:
    STRAITS_DB = {}
    _TERRAIN_DB_OK = False

# ── v9.14: 해협 시나리오 상수 ────────────────────────────────────────────────
# strait_type → (접근 방위 중심각°, 확산각°) 매핑
# korea_west: 위협이 서쪽(일본해→서수도 방향) 270°에서 동쪽으로 접근
# korea_east: 위협이 동쪽(태평양→동수도 방향) 90°에서 서쪽으로 접근
# bilateral: 동서 양방향 협공
STRAIT_BEARING: dict[str, list[float]] = {
    'korea_west': [270.0],
    'korea_east': [90.0],
    'bilateral':  [90.0, 270.0],
}
STRAIT_SPREAD_DEG = 30.0  # 해협 입구 방위 확산각 (±30°)

# 해협 폭 기준 회피 기동 패널티: 폭이 좁을수록 기동 공간 감소
# 기준 200km(개방 해역) 대비 해협 폭 비율로 evade_r 스케일
_STRAIT_OPEN_SEA_KM = 200.0

# ── 시뮬레이션 상수 ──────────────────────────────────────────────────────────
DT               = 1.0    # 시간 스텝 (초)
MAX_SIM_TIME     = 3600   # 최대 시뮬 시간 (초) — 해성 250m/s 기준 250km = 1000초 충분

# ── v10.7: 전술 의사결정 모드 상태 스냅샷 ────────────────────────────────────
@dataclasses.dataclass
class TacticalState:
    t: float
    threats: list       # [{'name', 'type', 'hp', 'dist_km', 'speed_ms'}]
    ships:   list       # [{'name', 'alive', 'hp', 'max_hp', 'radar', 'speed'}]
    intercepted: int    # 현재까지 요격 수
    total_threats: int  # 총 위협 수
    shots_fired: int    # 현재까지 발사 수
INTERCEPT_DIST_M = 200    # BUG-5: SAM 근접 신관 범위 (m). 기존 2000m 과대, 실제 50-200m
ECM_REF_RANGE_M  = 25_000 # MED-9: ECM 재밍 기준 거리 25km (기존 50km 과대)
DECOY_PK         = 0.50   # LOW-7: 0.60→0.50 (AN/SLQ-25 실전 기만 성공률)
SHIP_EVASION_PK  = 0.20   # LOW-8: 0.30→0.20 (회피 기동 성공률 현실화)
MAX_RESPONSE_TIME_S = 120  # 포팅 D: REQ-02 최대 허용 응답시간 (초)

# ── v9.13: 날씨 → Beaufort + 특수 효과 매핑 ─────────────────────────────────
# special_radar: Beaufort 클러터 외 추가 레이더 감쇠 (황사·농무·야간 흡수·산란)
# special_sonar: 모두 1.00 — 황사·야간은 소나 무관, Beaufort 해상 소음으로 충분
WEATHER_BEAUFORT_MAP: dict[str, dict] = {
    '맑음 (주간)':           {'beaufort': 2,  'special_radar': 1.00, 'special_sonar': 1.00},
    '맑음 (야간)':           {'beaufort': 2,  'special_radar': 0.97, 'special_sonar': 1.00},
    '흐림 (박무)':           {'beaufort': 4,  'special_radar': 1.00, 'special_sonar': 1.00},
    '황사 (봄철 황사)':      {'beaufort': 3,  'special_radar': 0.75, 'special_sonar': 1.00},
    '풍랑 (7~8등급)':        {'beaufort': 7,  'special_radar': 1.00, 'special_sonar': 1.00},
    '폭풍 (해상 악화)':      {'beaufort': 9,  'special_radar': 1.00, 'special_sonar': 1.00},
    '태풍 (9~12등급)':       {'beaufort': 11, 'special_radar': 1.00, 'special_sonar': 1.00},
    '농무 (시정 200m 이하)': {'beaufort': 2,  'special_radar': 0.82, 'special_sonar': 1.00},
    '폭풍 (야간)':           {'beaufort': 9,  'special_radar': 0.97, 'special_sonar': 1.00},
    '태풍 (야간)':           {'beaufort': 11, 'special_radar': 0.97, 'special_sonar': 1.00},
    '농무 (야간)':           {'beaufort': 2,  'special_radar': 0.79, 'special_sonar': 1.00},
    '황사 (새벽)':           {'beaufort': 3,  'special_radar': 0.73, 'special_sonar': 1.00},
}

# ── v10.1: ICAO 표준 대기 (ISA) ─────────────────────────────────────────────
def isa_atmosphere(alt_m: float) -> tuple[float, float, float]:
    """ICAO 표준 대기 대류권(0~11000m): 기온(K), 압력(Pa), 밀도(kg/m³)."""
    T0, P0 = 288.15, 101_325.0
    L, g, M, R = 0.0065, 9.80665, 0.028964, 8.31446
    h = max(0.0, min(float(alt_m), 11_000.0))
    T = T0 - L * h
    P = P0 * (T / T0) ** (g * M / (R * L))
    rho = P * M / (R * T)
    return T, P, rho

# ── v9.13: 증발 덕팅 (Evaporation Duct Height) ───────────────────────────────
# (region_key, season) → (edh_m, boost_factor)
# 덕트 내 저고도 표적: 레이더 전파가 해면을 따라 굴절 → 탐지거리 증가
# 출처: 한국 해역 대기 경계층 관측 문헌 평균값
EVAP_DUCT_DB: dict[tuple, tuple] = {
    ('EAST_SEA',     'spring'): (7,  1.15),  # 봄 동해: 기온 상승, 수증기 증가 시작
    ('EAST_SEA',     'summer'): (10, 1.25),  # 여름 동해: 덕트 높이 10m, 탐지 1.25배
    ('EAST_SEA',     'autumn'): (8,  1.18),  # 가을 동해: 여름 근접 수준 유지
    ('EAST_SEA',     'winter'): (6,  1.12),
    ('YELLOW_SEA',   'spring'): (5,  1.10),  # 서해 봄: 황사·박무 시즌, 덕트 약함
    ('YELLOW_SEA',   'summer'): (8,  1.18),  # 서해: 강수·습도로 덕트 발달 약함
    ('YELLOW_SEA',   'autumn'): (6,  1.12),
    ('YELLOW_SEA',   'winter'): (4,  1.08),
    ('KOREA_STRAIT', 'spring'): (6,  1.13),
    ('KOREA_STRAIT', 'summer'): (9,  1.20),
    ('KOREA_STRAIT', 'autumn'): (7,  1.15),
    ('KOREA_STRAIT', 'winter'): (5,  1.10),
}

# ── v9.13: 고층 바람 CEP 배율 ────────────────────────────────────────────────
# (region_key, season) → cep_multiplier
# 순항미사일 Pk = pk_base / cep_factor (CEP 증가 → 명중률 감소)
# 동해 겨울: 편서풍 강함(30~50 m/s), 서해 겨울: 북서풍 (25~40 m/s)
WIND_CEP_FACTOR: dict[tuple, float] = {
    ('EAST_SEA',     'spring'): 1.10,   # 봄: 제트기류 북상 중, 중간 강도
    ('EAST_SEA',     'summer'): 1.05,
    ('EAST_SEA',     'autumn'): 1.12,   # 가을: 제트기류 남하 시작
    ('EAST_SEA',     'winter'): 1.20,
    ('YELLOW_SEA',   'spring'): 1.10,
    ('YELLOW_SEA',   'summer'): 1.08,
    ('YELLOW_SEA',   'autumn'): 1.11,
    ('YELLOW_SEA',   'winter'): 1.15,
    ('KOREA_STRAIT', 'spring'): 1.09,
    ('KOREA_STRAIT', 'summer'): 1.06,
    ('KOREA_STRAIT', 'autumn'): 1.11,
    ('KOREA_STRAIT', 'winter'): 1.18,
}

# ── v10.1: ISA 대기 굴절 + 라디오존데 4계절×5고도층 보정 ──────────────────────
# 키: (region_key, season, alt_layer)
# alt_layer: 1=100~500m / 2=500~1000m / 3=1000~3000m / 4=3000m+
# (alt_layer 0 = <100m 은 _evap_duct_factor 담당)
# 출처: 기상청 고층기상관측 연보 (포항·광주·오산 라디오존데 계절 통계)
# 대기 굴절 지수(N-unit): 여름 N≈330~340 (ISA 315 대비 +5~8%), 겨울 N≈295~310 (+2~3%)
ISA_RADIOSONDE_DB: dict[tuple, float] = {
    # ── EAST_SEA ─────────────────────────────────────────────────────────────
    ('EAST_SEA', 'spring', 1): 1.02, ('EAST_SEA', 'spring', 2): 1.04,
    ('EAST_SEA', 'spring', 3): 1.05, ('EAST_SEA', 'spring', 4): 1.03,
    ('EAST_SEA', 'summer', 1): 1.03, ('EAST_SEA', 'summer', 2): 1.06,  # 쿠로시오 수증기
    ('EAST_SEA', 'summer', 3): 1.07, ('EAST_SEA', 'summer', 4): 1.04,
    ('EAST_SEA', 'autumn', 1): 1.03, ('EAST_SEA', 'autumn', 2): 1.05,
    ('EAST_SEA', 'autumn', 3): 1.06, ('EAST_SEA', 'autumn', 4): 1.03,
    ('EAST_SEA', 'winter', 1): 1.01, ('EAST_SEA', 'winter', 2): 1.03,  # 대륙성 한기
    ('EAST_SEA', 'winter', 3): 1.03, ('EAST_SEA', 'winter', 4): 1.02,
    # ── YELLOW_SEA ────────────────────────────────────────────────────────────
    ('YELLOW_SEA', 'spring', 1): 1.02, ('YELLOW_SEA', 'spring', 2): 1.03,  # 황사 시즌
    ('YELLOW_SEA', 'spring', 3): 1.04, ('YELLOW_SEA', 'spring', 4): 1.02,
    ('YELLOW_SEA', 'summer', 1): 1.02, ('YELLOW_SEA', 'summer', 2): 1.05,  # 장마·고온다습
    ('YELLOW_SEA', 'summer', 3): 1.06, ('YELLOW_SEA', 'summer', 4): 1.04,
    ('YELLOW_SEA', 'autumn', 1): 1.02, ('YELLOW_SEA', 'autumn', 2): 1.04,
    ('YELLOW_SEA', 'autumn', 3): 1.05, ('YELLOW_SEA', 'autumn', 4): 1.03,
    ('YELLOW_SEA', 'winter', 1): 1.00, ('YELLOW_SEA', 'winter', 2): 1.02,  # 북서계절풍
    ('YELLOW_SEA', 'winter', 3): 1.02, ('YELLOW_SEA', 'winter', 4): 1.01,
    # ── KOREA_STRAIT ──────────────────────────────────────────────────────────
    ('KOREA_STRAIT', 'spring', 1): 1.02, ('KOREA_STRAIT', 'spring', 2): 1.04,
    ('KOREA_STRAIT', 'spring', 3): 1.04, ('KOREA_STRAIT', 'spring', 4): 1.03,
    ('KOREA_STRAIT', 'summer', 1): 1.03, ('KOREA_STRAIT', 'summer', 2): 1.05,  # 쓰시마 난류
    ('KOREA_STRAIT', 'summer', 3): 1.06, ('KOREA_STRAIT', 'summer', 4): 1.04,
    ('KOREA_STRAIT', 'autumn', 1): 1.02, ('KOREA_STRAIT', 'autumn', 2): 1.04,
    ('KOREA_STRAIT', 'autumn', 3): 1.05, ('KOREA_STRAIT', 'autumn', 4): 1.03,
    ('KOREA_STRAIT', 'winter', 1): 1.01, ('KOREA_STRAIT', 'winter', 2): 1.03,
    ('KOREA_STRAIT', 'winter', 3): 1.03, ('KOREA_STRAIT', 'winter', 4): 1.02,
}

# ── v10.1: 트로포스캐터 링크 보정 ────────────────────────────────────────────
# 대기 상층 난류 산란 → 수평선 너머 고고도 표적(≥1000m) 탐지거리 추가 증가
# 해양성 기류 안정할수록 산란층 발달 우수. BF6 이상 강풍 시 산란층 붕괴 → 비활성화.
TROPOSCATTER_DB: dict[tuple, float] = {
    ('EAST_SEA',     'spring'): 1.11,
    ('EAST_SEA',     'summer'): 1.16,
    ('EAST_SEA',     'autumn'): 1.13,
    ('EAST_SEA',     'winter'): 1.09,
    ('YELLOW_SEA',   'spring'): 1.09,
    ('YELLOW_SEA',   'summer'): 1.14,
    ('YELLOW_SEA',   'autumn'): 1.11,
    ('YELLOW_SEA',   'winter'): 1.07,
    ('KOREA_STRAIT', 'spring'): 1.10,
    ('KOREA_STRAIT', 'summer'): 1.15,
    ('KOREA_STRAIT', 'autumn'): 1.12,
    ('KOREA_STRAIT', 'winter'): 1.08,
}

# ── v10.2-B: 아군 SAM RCS (Phase C 적 레이더 탐지용) ────────────────────────
# 출처: 각 미사일 동체 크기·형상 기반 추정값 (m²)
_SAM_RCS: dict[str, float] = {
    'SM-3 Block IIA':  0.0008,
    'SM-6':            0.0010,
    'SM-6 Block IB':   0.0010,
    'SM-2 Block IIIB': 0.0015,
    'ESSM Block II':   0.0005,
    '해궁 (K-SAAM)':   0.0005,
    'RIM-116 RAM':     0.0003,
}

# ── v9.12: 해역 매핑 및 지형 레이더 음영 페널티 ─────────────────────────────
# fleet_region UI 문자열 → ocean_acoustic_db 키
REGION_TO_ACOUSTIC_KEY: dict[str, str] = {
    '동해 북부': 'EAST_SEA',
    '동해 중부': 'EAST_SEA',
    '서해':       'YELLOW_SEA',
    '대한해협':   'KOREA_STRAIT',
}

# 해역별 저고도 레이더 음영 배율 (altitude_m → factor)
# 출처: terrain_db.TERRAIN_DB['radar_shadow_reference'] 음영각 기반
# 동해 3.4~8.1° (태백·설악), 서해 0.4° (낭림 원거리), 대한해협 0.9° (소백)
TERRAIN_RADAR_PENALTY: dict[str, list] = {
    'EAST_SEA': [
        (0,    0.78),   # 해면밀착 — 태백·설악 음영 강하게 받음
        (50,   0.88),
        (200,  0.93),
        (1000, 1.00),   # 산능선 위 고고도
    ],
    'YELLOW_SEA': [
        (0,    0.96),   # 낭림산맥 원거리 (0.4°), 영향 미약
        (200,  0.98),
        (1000, 1.00),
    ],
    'KOREA_STRAIT': [
        (0,    0.90),   # 소백산맥 (0.9°)
        (50,   0.94),
        (200,  0.97),
        (1000, 1.00),
    ],
}

# ── 지상 BMD 자산 상수 (이지스 어쇼어 / THAAD) ──────────────────────────────
_ASHORE_SM3_PK     = 0.82     # 지상 고정 설치 → 함정 대비 안정적 추적
_ASHORE_SM3_COST   = 35_000_000
_ASHORE_SM3_SPD_MS = 3_000
_ASHORE_COOLDOWN_S = 5.0
_ASHORE_RANGE_M    = 500_000  # SM-3 최대 교전거리
_ASHORE_ALT_MIN_M  = 40_000   # 중간단계 교전 하한 (40km)

_THAAD_PK          = 0.85     # hit-to-kill 종말 단계
_THAAD_COST        = 3_000_000
_THAAD_SPD_MS      = 2_500
_THAAD_COOLDOWN_S  = 3.0
_THAAD_RANGE_M     = 200_000  # 200km 방어 반경
_THAAD_ALT_MIN_M   = 10_000   # 종말 하한 10km
_THAAD_ALT_MAX_M   = 150_000  # 종말 상한 150km

# 다층 방어 레이어 순서: 가장 먼저 교전하는 함정 유형부터 (BMD 우선 → 방공 우선)
LAYER_ORDER    = ['KDX-III-B2', 'KDX-III-B1', 'KDX-II', 'FFX-III', 'FFX-II', 'FFX-I']
SHIP_LAYER_PRI = {t: i for i, t in enumerate(LAYER_ORDER)}

def _make_physics_wx(weather: str) -> dict:
    """
    v9.13: 날씨 문자열 → Beaufort 물리 기반 wx dict.
    ocean_environment_db 없으면 WEATHER_DB 원본 그대로 반환.
    """
    wx = dict(WEATHER_DB.get(weather, WEATHER_DB['맑음 (주간)']))
    if _OCEAN_ENV_OK:
        bf_info = WEATHER_BEAUFORT_MAP.get(weather)
        if bf_info:
            bf = bf_info['beaufort']
            wx['radar_factor'] = (
                _RADAR_BF_FACTOR.get(bf, 1.0) * bf_info['special_radar']
            )
            wx['sonar_factor'] = (
                _SONAR_BF_FACTOR.get(bf, 1.0) * bf_info['special_sonar']
            )
            wx['beaufort'] = bf
    return wx

# 포팅 C: v7 시뮬 시간 스케일 맞춤 출격 준비 시간 (전시 긴급 출격 기준)
# FRIENDLY_AIRCRAFT_DB의 sortie_time_s(평시)를 v7 700초 시뮬에 맞게 단축
_AIRCRAFT_V7_SORTIE = {
    'AW-159 와일드캣': 300,   # 5분 (평시 동일)
    'P-3C 오라이온':   600,   # 10분 (평시 40분 → 전시 긴급 출격)
    'P-8A 포세이돈':   480,   # 8분 (평시 30분 → 전시 긴급 출격)
}

# ════════════════════════════════════════════════════════════════════════════
#  새 DB: 아군 대함 공격 무기
# ════════════════════════════════════════════════════════════════════════════
FRIENDLY_STRIKE_DB = {
    '해성-I': {
        'speed_ms': 250,   # Mach 0.73 (아음속 순항, SSM-700K 실제 속도)
        'range_km': 180,
        'cost_usd': 800_000,
        'pk_base':  0.80,
        'seeker':   'radar',    # v8.26: 액티브 레이더 탐색기
    },
    '해성-II': {
        'speed_ms': 250,   # Mach 0.73 (해성-I 동일 계열)
        'range_km': 250,
        'cost_usd': 1_200_000,
        'pk_base':  0.82,
        'seeker':   'radar',
    },
    '하푼 Block II': {
        'speed_ms': 245,   # BUG-4: Mach 0.72, 기존 278 m/s 과대
        'range_km': 280,
        'cost_usd': 1_500_000,
        'pk_base':  0.78,
        'seeker':   'radar',    # AN/DSQ-61 액티브 레이더
    },
    # SM-6 Block IB 대함 모드 (OTH 대함 공격, Link-16 유도)
    'SM-6 대함 모드': {
        'speed_ms': 1000,   # Mach 3.5
        'range_km': 370,    # OTH 사거리
        'cost_usd': 4_200_000,
        'pk_base':  0.70,
        'seeker':   'combined', # GPS+능동 레이더 복합
    },
    # 5인치 Mk.45 Mod 4 함포 (근거리 최후 레이어)
    'Mk.45 5인치 함포': {
        'speed_ms': 830,   # 포탄 초속 ~Mach 2.4
        'range_km': 24,    # 유효 사거리 24km
        'cost_usd': 2_000, # 발당 약 $2,000
        'pk_base':  0.40,  # 대함 Pk (표적이 클수록 높음)
        'seeker':   'radar',
    },
    # NEW-P1: Tomahawk Block V 초장거리 대함 타격 (US ships)
    'Tomahawk Block V': {
        'speed_ms': 250,
        'range_km': 1700,
        'cost_usd': 2_000_000,
        'pk_base':  0.80,
        'seeker':   'combined', # DSMAC 영상+GPS 복합
    },
    # KSS-III VLS 탑재 현무-3C 순항미사일 (잠수함 발사)
    '현무-3C': {
        'speed_ms': 250,    # Mach 0.73 (아음속 순항)
        'range_km': 1500,   # 현무-3C 사거리 1,500km
        'cost_usd': 2_000_000,
        'pk_base':  0.80,
        'seeker':   'combined', # GPS+INS+TERCOM 복합
    },
    # NEW-A: 현무-4 지대함 탄도미사일 (한국판 DF-21D, 지상 발사 전용)
    '현무-4 (ASBM)': {
        'speed_ms': 3000,   # 종말 Mach 8~10 (평균 비행 속도 근사)
        'range_km': 800,    # 현무-4 사거리 800km
        'cost_usd': 3_500_000,
        'pk_base':  0.85,   # 고속 종말 — 적 SAM 요격 극히 어려움
        'seeker':   'combined', # 레이더+적외선 복합 종말 탐색
    },
}

# ════════════════════════════════════════════════════════════════════════════
#  새 DB: 적 함정 SAM 시스템
# ════════════════════════════════════════════════════════════════════════════
ENEMY_SAM_DB = {
    'HHQ-9B':    {'range_km': 200, 'speed_ms': 1400, 'pk': 0.82, 'channels': 6},
    'HHQ-16':    {'range_km':  50, 'speed_ms': 1000, 'pk': 0.75, 'channels': 4},
    'HHQ-10':    {'range_km':   9, 'speed_ms':  680, 'pk': 0.70, 'channels': 2},
    '1130-CIWS': {'range_km':   3, 'speed_ms': 1100, 'pk': 0.65, 'channels': 1},
    # NEW-P1: 러시아 함정 SAM
    'S-300F':    {'range_km':  90, 'speed_ms': 1800, 'pk': 0.85, 'channels': 6},  # 슬라바급
    'SA-N-9':    {'range_km':  12, 'speed_ms':  900, 'pk': 0.75, 'channels': 4},  # 우달로이급
}

ENEMY_SHIP_SAM_LOADOUT = {
    '055형 대형 구축함': [
        {'name': 'HHQ-9B',    'stock': 112},
        {'name': 'HHQ-10',    'stock':  24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '052D형 구축함': [
        {'name': 'HHQ-9B',    'stock': 64},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '054A형 호위함': [
        {'name': 'HHQ-16',    'stock': 32},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '056형 초계함': [
        {'name': 'HHQ-10',    'stock':  8},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '022형 미사일 고속정': [
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    # NEW-P1: 신규 적 함정 SAM 탑재 목록
    '052C형 구축함 (HHQ-9)': [
        {'name': 'HHQ-9B',    'stock': 48},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '071형 상륙함': [
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '우달로이급 구축함': [
        {'name': 'SA-N-9',    'stock': 64},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '슬라바급 순양함': [
        {'name': 'S-300F',    'stock': 64},
        {'name': 'HHQ-10',    'stock': 40},   # AK-630 CIWS (HHQ-10 근사)
        {'name': '1130-CIWS', 'stock': 9999},
    ],
}

# 독립 미사일 유형 (EnemyThreatObj 대신 MissileObj로 직접 생성)
_STANDALONE_MISSILE_TYPES = ('탄도미사일', '순항미사일', '극초음속활공체', '저고도기동탄도', '대방사미사일')

# MED-5: 미사일 명칭별 함체 명중 Pk (0.80 하드코드 → 유형 기반 매핑)
# 아음속 대함미사일은 요격 쉬움(Pk 높음), 초음속·어뢰는 요격 어려움(Pk 낮음)
_MISSILE_PK_MAP = {
    # 아음속 순항 대함 (0.70~0.72)
    'YJ-83 대함미사일':          0.72,
    'YJ-83K 주력 대함미사일':    0.70,
    'YJ-8K 대함미사일':          0.72,
    'YJ-91 대함미사일':          0.70,
    'YJ-91 초음속 대함미사일':   0.70,
    'YJ-100 (장거리 순항)':      0.70,
    'CJ-10 (순항미사일)':        0.70,
    # 초음속 순항 대함 (요격 어려움: 0.65~0.68)
    'YJ-12 초음속 대함미사일':   0.65,
    'YJ-12 (초음속 순항)':       0.65,
    'YJ-18 초음속 대함미사일':   0.68,
    'YJ-18B 잠대함미사일':       0.68,
    'Kh-31A 대함미사일':         0.68,
    'Kh-31A (항공기발사 대함)':  0.68,
    'P-800 오닉스 (야혼트)':     0.65,
    # 탄도/HGV/QBM (SM-3/SM-6 전담)
    'DF-11A (단거리 탄도)':      0.75,
    'DF-15 (단거리 탄도)':       0.75,
    'DF-21D (대함 탄도)':        0.70,
    'DF-26 (중장거리 탄도)':     0.65,
    'DF-17 (극초음속 활공)':     0.65,
    'KN-23 (북한 이스칸데르)':   0.72,
    # 어뢰 (수중 유도, 회피 어려움: 0.78)
    'Yu-6 중어뢰':               0.78,
    # NEW-P1: 신규 미사일
    'DF-21D (공중발사)':         0.70,
    'Kh-32 극초음속':            0.65,
    'P-700 그라니트':            0.62,
    'P-1000 (벌칸)':             0.62,
    'Kalibr 3M54 잠대함':        0.68,
    '북극성-1 (SLBM)':           0.70,
    'Kh-31A 대함미사일':         0.68,
    # ARM: 레이더 활성 상태 추적 — 기본 Pk 높음
    'Kh-31P 대방사미사일':       0.88,
    'LD-10 대방사미사일':        0.85,
    'Kh-58U 대방사미사일':       0.90,  # 장거리 고속 — 요격 난이도 높음
    # BUG-6: _MISSILE_PK_MAP 누락 13종 추가 — 기본값 0.72로 잘못 처리되던 항목
    # 극초음속/HGV (요격 매우 어려움: 0.62~0.63)
    'YJ-21 (극초음속 대함)':          0.62,
    '킨잘 (극초음속 탄도)':           0.62,
    '지르콘 (극초음속 순항)':         0.63,
    # ICBM/IRBM (SM-3 전담, 요격 극히 어려움: 0.56~0.65)
    '화성-15 (ICBM급)':               0.58,
    '화성-17 (ICBM 개량)':            0.56,
    '화성-18 (ICBM 고체연료)':        0.58,
    '화성-12 (IRBM)':                 0.65,
    # 기동탄도/QBM (0.68)
    'KN-24 (단거리 기동탄도)':        0.68,
    # 초음속 대함 (0.68)
    'YJ-18 (초음속 대함)':            0.68,
    # 스텔스 순항 (저RCS — 탐지 어려움: 0.68)
    'Kh-101 (스텔스 순항)':           0.68,
    # 아음속 순항 (0.70~0.72)
    'Kalibr (3M14 순항미사일)':       0.70,
    '해성-3 (잠수함발사 순항)':       0.70,
    '북한 순항미사일 (화살-2)':       0.72,
}
_MISSILE_PK_DEFAULT = 0.72  # 미등록 미사일 기본값



# ════════════════════════════════════════════════════════════════════════════
#  Vec2
# ════════════════════════════════════════════════════════════════════════════
#  v10.8: LatLon — 지리 좌표계 (Vec2 완전 대체)
#  x / y 프로퍼티로 기존 시각화 코드 자동 호환
# ════════════════════════════════════════════════════════════════════════════

# 해역별 기준점 (작전 중심 LatLon)
_REGION_REF: dict[str, tuple] = {
    '동해 북부':  (40.5, 130.5),
    '동해 중부':  (37.2, 131.9),
    '서해':       (36.0, 125.0),
    '대한해협':   (34.5, 129.0),
}
_DEFAULT_REF = (37.2, 131.9)  # 기본: 동해 중부

# 시뮬 시작 시 _set_region_ref() 로 갱신
_ref_lat: float = _DEFAULT_REF[0]
_ref_lon: float = _DEFAULT_REF[1]
_ref_m_per_deg_lat: float = 110_540.0
_ref_m_per_deg_lon: float = 111_320.0 * math.cos(math.radians(_DEFAULT_REF[0]))

def _set_region_ref(region: str):
    """시뮬 시작 시 해역에 맞는 기준점 설정."""
    global _ref_lat, _ref_lon, _ref_m_per_deg_lon
    _ref_lat, _ref_lon = _REGION_REF.get(region, _DEFAULT_REF)
    _ref_m_per_deg_lon = 111_320.0 * math.cos(math.radians(_ref_lat))


class LatLon:
    """
    지리 좌표 (위도/경도). x/y 프로퍼티로 기존 Cartesian 코드 자동 호환.
    dist_to(): Haversine 거리(m). bearing_to(): 동=0 북=π/2 평면 방위각(rad).
    """
    __slots__ = ('lat', 'lon')

    def __init__(self, lat: float = 0.0, lon: float = 0.0):
        self.lat = float(lat)
        self.lon = float(lon)

    @classmethod
    def from_xy(cls, x_m: float, y_m: float) -> 'LatLon':
        """미터 좌표(기준점 기준) → LatLon."""
        return cls(
            _ref_lat + y_m / _ref_m_per_deg_lat,
            _ref_lon + x_m / _ref_m_per_deg_lon,
        )

    # ── 시각화 호환 프로퍼티 ──────────────────────────────────────────────────
    @property
    def x(self) -> float:
        """기준점으로부터 동방향(m) — 시각화·거리 계산 호환."""
        return (self.lon - _ref_lon) * _ref_m_per_deg_lon

    @x.setter
    def x(self, v: float):
        self.lon = _ref_lon + v / _ref_m_per_deg_lon

    @property
    def y(self) -> float:
        """기준점으로부터 북방향(m) — 시각화·거리 계산 호환."""
        return (self.lat - _ref_lat) * _ref_m_per_deg_lat

    @y.setter
    def y(self, v: float):
        self.lat = _ref_lat + v / _ref_m_per_deg_lat

    # ── 거리·방위 ─────────────────────────────────────────────────────────────
    def dist_to(self, other: 'LatLon') -> float:
        """Haversine 거리 (m)."""
        R = 6_371_000.0
        φ1 = math.radians(self.lat);  φ2 = math.radians(other.lat)
        Δφ = math.radians(other.lat - self.lat)
        Δλ = math.radians(other.lon - self.lon)
        a = math.sin(Δφ/2)**2 + math.cos(φ1)*math.cos(φ2)*math.sin(Δλ/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(max(0.0, 1.0 - a)))

    def bearing_to(self, other: 'LatLon') -> float:
        """방위각 (rad) — 동=0, 북=π/2 (x/y 평면 관례 유지)."""
        return math.atan2(other.y - self.y, other.x - self.x)

    def move_toward(self, target: 'LatLon', speed_ms: float, dt: float) -> bool:
        """target 방향으로 이동. 도달했으면 True 반환."""
        d = self.dist_to(target)
        step = speed_ms * dt
        if d <= step:
            self.lat, self.lon = target.lat, target.lon
            return True
        angle = self.bearing_to(target)
        self.lat += math.sin(angle) * step / _ref_m_per_deg_lat
        self.lon += math.cos(angle) * step / _ref_m_per_deg_lon
        return False

    def copy(self) -> 'LatLon':
        return LatLon(self.lat, self.lon)

    def __repr__(self) -> str:
        return f"LatLon({self.lat:.4f}°N, {self.lon:.4f}°E)"


# 하위 호환 별칭 — 기존 코드가 Vec2를 참조할 경우 대비
Vec2 = LatLon


# ════════════════════════════════════════════════════════════════════════════
#  MissileObj
# ════════════════════════════════════════════════════════════════════════════
class MissileObj:
    """
    mtype:
      'enemy_strike'   — 적 대함/대지 공격 (아군 함정 또는 독립 미사일 위협)
      'friendly_strike'— 아군 대함/대잠 공격
      'friendly_sam'   — 아군 SAM (적 미사일·항공기 요격)
      'enemy_sam'      — 적 SAM (아군 미사일 요격)
    """
    _id_counter = 0

    def __init__(self, mtype: str, name: str, pos: Vec2,
                 target,
                 speed_ms: float, pk_base: float,
                 owner_id: int, t_spawn: float = 0.0):
        MissileObj._id_counter += 1
        self.uid       = f"M{MissileObj._id_counter:04d}"
        self.mtype     = mtype
        self.name      = name
        self.pos       = pos.copy()
        self.target    = target
        self.speed_ms  = speed_ms
        self.pk_base   = pk_base
        self.owner_id  = owner_id
        self.t_spawn   = t_spawn

        self.alive       = True
        self.intercepted = False
        self.hit         = False
        self.t_intercept: Optional[float] = None
        self.t_hit:       Optional[float] = None

        # 독립 미사일 위협 속성 (탄도/HGV/QBM 등 — _build_enemies에서 설정)
        self.altitude_m:   float = 0.0
        self.is_hgv:       bool  = False
        self.is_qbm:       bool  = False
        self.is_ballistic: bool  = False

        # 포팅 B: 전술 속성
        self.terminal_evasion_factor: float = 1.0     # 종말 회피 계수 (< 20km 적용)
        self.is_torpedo:              bool  = False   # 어뢰 여부 (기만기/회피 판정용)
        self.is_arm:                  bool  = False   # 대방사미사일 여부 (레이더 직격)
        self.seeker:                  str   = 'radar' # v8.26: 탐색기 유형 (채프/플레어 대응)

        # EngagementAnalysis 추적 (A-1)
        self.intercept_weapon: Optional[str] = None  # 격추 무기명
        self.intercept_km:     float         = 0.0   # 격추 거리 (km)
        self.detect_m:         float         = 0.0   # 탐지 시 거리 (m)
        self.enemy_info:       dict          = {}     # ENEMY_DB 원본

        # v10.4: CEC 중계 교전 여부 (자체 탐지 불가 → 아군 네트워크 데이터 의존)
        self.cec_relay: bool = False

    def update(self, dt: float) -> bool:
        """1 tick 이동. alive=False 설정 금지 — 요격/피격 판정은 엔진이 담당."""
        if not self.alive:
            return False
        arrived = self.pos.move_toward(self.target.pos, self.speed_ms, dt)
        if arrived:
            self.hit = True
        return arrived

    @classmethod
    def reset_counter(cls):
        cls._id_counter = 0


# ════════════════════════════════════════════════════════════════════════════
#  EnemyThreatObj — 적 플랫폼 위협 통합 (항공기 / 수상함 / 잠수함)
#  독립 미사일(탄도/순항/HGV/QBM)은 _build_enemies()에서 MissileObj로 생성.
# ════════════════════════════════════════════════════════════════════════════
class EnemyThreatObj:
    _id_counter = 0

    _HP_MAP = {
        '전투기': 1, '폭격기': 1, '전폭기': 1,
        '잠수함': 3,
        '고속정': 2, '초계함': 2,
        '호위함': 3, '구축함': 4,
        '항모': 5,
    }

    def __init__(self, preset_name: str, pos: Vec2):
        EnemyThreatObj._id_counter += 1
        self.uid         = f"ET{EnemyThreatObj._id_counter:03d}"
        self.preset_name = preset_name
        self.name        = preset_name   # 요격 로그 호환
        self.info        = ENEMY_DB[preset_name].copy()
        self.pos         = pos
        self.speed_ms    = self.info['speed_ms']
        self.altitude_m  = self.info.get('altitude_m', 0)

        cat   = self.info.get('category', '대함')
        ttype = self.info.get('type', '')
        self.category    = cat
        self.threat_type = ttype
        self.is_aircraft = ttype in ('전투기', '폭격기', '전폭기')
        self.is_ship     = cat == '대함'
        self.is_sub      = cat == '대잠'

        # ENEMY_DB hp 필드 우선, 없으면 _HP_MAP 기본값
        _db_hp = self.info.get('hp', None)
        self.hp = _db_hp if _db_hp is not None else self._HP_MAP.get(ttype, 2)
        self.high_value_target    = self.info.get('high_value_target', False)
        self.carrier_aircraft     = self.info.get('carrier_aircraft', None)
        self.carrier_wave_interval = self.info.get('carrier_wave_interval', 0)
        self._last_wave_t         = 0.0   # 마지막 함재기 발진 시각

        if self.is_ship:
            loadout = ENEMY_SHIP_SAM_LOADOUT.get(preset_name, [])
            self.sam_inventory = {item['name']: item['stock'] for item in loadout}
            self.sam_max_channels = sum(
                ENEMY_SAM_DB[n]['channels']
                for n in self.sam_inventory if n in ENEMY_SAM_DB
            )
        else:
            self.sam_inventory    = {}
            self.sam_max_channels = 0
        self.sam_channels_used = 0

        self.alive        = True
        self.intercepted  = False
        self.hit_count    = 0
        self.hit_by: list = []
        self.has_fired    = False
        self.t_intercept: Optional[float] = None

        # 항공기 이탈 상태
        self.is_retreating             = False
        self.retreat_pos: Optional[Vec2] = None
        # MED-12: 항공기 재공격 — 이탈 후 재접근 허용
        self.reattack_count = 0
        self.max_reattacks  = 1 if self.is_aircraft else 0

        # v9.6: 잠수함 기습 은닉 상태 — hidden_until 초까지 탐지 불가
        self.hidden_until    = 0.0
        self.ambush_revealed = False

    def take_hit(self, weapon_name: str, t: float):
        self.hit_count += 1
        self.hit_by.append((weapon_name, t))
        self.hp -= 1
        if self.hp <= 0:
            self.alive = False

    def select_sam(self, missile_dist_m: float) -> Optional[str]:
        """수상함용: 사거리 내 SAM 선택 (장거리 우선)"""
        for sam_name in ['HHQ-9B', 'HHQ-16', 'HHQ-10', '1130-CIWS']:
            if self.sam_inventory.get(sam_name, 0) <= 0:
                continue
            sam = ENEMY_SAM_DB.get(sam_name)
            if sam and missile_dist_m <= sam['range_km'] * 1000:
                return sam_name
        return None

    @classmethod
    def reset_counter(cls):
        cls._id_counter = 0


# ════════════════════════════════════════════════════════════════════════════
#  FriendlyShipObj
# ════════════════════════════════════════════════════════════════════════════
class FriendlyShipObj:
    def __init__(self, name: str, ship_type: str, pos: Optional[Vec2] = None):
        self.name        = name
        self.ship_type   = ship_type
        spec             = SHIP_DB[ship_type]
        self.display     = spec['display']
        self.sensor_km   = spec['sensor_km']
        # max_channels는 @property로 계산 (부분 피해 반영)
        self.pos         = pos or LatLon.from_xy(0, 0)

        self.inventory   = spec['default_inventory'].copy()

        self.strike_inventory: dict = {}

        self.is_submarine  = spec.get('is_submarine', False)
        self.nation        = spec.get('nation', 'KOR')  # v8.16: 한미 연합 작전

        # LOW-9: 함정 유형별 HP (함종별 내탄성 차등. 기존 고정값 5 → 실제 격침 내성 반영)
        _hp_map = {
            'KDX-III-B2': 5, 'KDX-III-B1': 5, 'KDX-II': 4,
            'FFX-I': 3, 'FFX-II': 3, 'FFX-III': 3,
            'DDG-51': 5, 'CG-47': 5, 'CVN': 8,
            'LPD': 3, 'SSN': 3, 'LST': 3, 'AO': 2,
            'KSS-I': 2, 'KSS-II': 2, 'KSS-III': 3,
        }
        self.hp            = _hp_map.get(ship_type, 4)
        self._max_hp       = self.hp
        self.alive         = True
        self.hit_count     = 0
        self.hits_taken    = 0  # MC 집계용 피격 횟수 (hit_count와 동일)
        self.total_cost    = 0.0
        self.channels_used = 0
        self.decoy_stock   = 4  # 포팅 B: AN/SLQ-25 음향 기만기 기본 재고
        # 피해 연동: 서브시스템별 성능 배율 (0.0~1.0)
        self.radar_factor   = 1.0  # 레이더 계통 (탐지거리 반영)
        self.channel_factor = 1.0  # 무장 채널 계통 (SAM 동시 교전 수 반영)
        self.speed_factor   = 1.0  # 추진 계통 (회피 기동 효율 반영)
        self.disabled_weapons: set = set()  # 무장 피탄으로 사용 불가 무기 목록
        self._vls_depleted = False  # 탄약 완전 소진 플래그
        self.radar_off_until: float = 0.0  # ARM 회피 전술: 이 시각까지 레이더 OFF
        self.eccm_factor: float = spec.get('eccm_factor', 0.40)  # v8.26: 재밍 상쇄 능력

    @property
    def max_channels(self):
        # 채널 계산 시 부분 피해 반영
        spec = SHIP_DB[self.ship_type]
        return max(1, int(spec['max_channels'] * self.channel_factor))

    @property
    def operational(self) -> bool:
        return self.alive

    def available(self, wpn: str) -> int:
        """무기 가용 재고 (무장 피탄 비활성화 반영)."""
        if wpn in self.disabled_weapons:
            return 0
        return self.inventory.get(wpn, 0)

    def take_hit(self, weapon_name: str, t: float, subsystem: str | None = None):
        """
        subsystem: 'radar' | 'propulsion' | 'weapons' | None
          None이면 서브시스템 피해 없이 HP만 감소 (하위호환).
        """
        self.hit_count  += 1
        self.hits_taken += 1
        self.hp -= 1

        if subsystem == 'radar':
            self.radar_factor = max(0.20, self.radar_factor * 0.50)
        elif subsystem == 'propulsion':
            self.speed_factor = max(0.30, self.speed_factor * 0.60)
        elif subsystem == 'weapons':
            self.channel_factor = max(0.40, self.channel_factor * 0.70)
            # VLS 탄약 25% 직접 손실 (수직발사관 피탄)
            _vls_wpns = ['SM-3 Block IIA', 'SM-6', 'SM-6 Block IB',
                         'SM-2 Block IIIB', 'ESSM Block II']
            for _w in _vls_wpns:
                if self.inventory.get(_w, 0) > 0:
                    self.inventory[_w] = max(0, self.inventory[_w] - max(1, self.inventory[_w] // 4))
            # 무기 비활성화 (SAM·CIWS 전체 대상)
            _candidates = [w for w in [
                'SM-3 Block IIA', 'SM-6', 'SM-6 Block IB',
                'SM-2 Block IIIB', 'ESSM Block II',
                'RIM-116 RAM', '해궁 (K-SAAM)', 'CIWS-II (Phalanx)',
            ] if self.inventory.get(w, 0) > 0 and w not in self.disabled_weapons]
            if _candidates:
                self.disabled_weapons.add(random.choice(_candidates))
            # VLS 고갈 플래그 갱신
            if not self._vls_depleted and all(self.inventory.get(w, 0) == 0 for w in _vls_wpns):
                self._vls_depleted = True

        if self.hp <= 0:
            self.alive = False

    def take_arm_hit(self, t: float):
        """대방사미사일(ARM) 레이더 직격 — 레이더 계통 심각 손상."""
        self.hit_count  += 1
        self.hits_taken += 1
        self.hp -= 1
        # ARM: 레이더 전파를 역추적해 레이더 안테나 직격 → 탐지거리 대폭 감소
        self.radar_factor = max(0.10, self.radar_factor * 0.30)
        if self.hp <= 0:
            self.alive = False


# ════════════════════════════════════════════════════════════════════════════
#  레이더 역할 세분화: 수색 → 추적 → 유도 3단계 파이프라인
# ════════════════════════════════════════════════════════════════════════════
class SearchRadar:
    """편대 수색 레이더 — 신규 위협 초기 탐지 지연 (SPY-1D 빔 회전 주기 6초)."""
    scan_interval: float = 6.0

    def __init__(self):
        self._detect_t: dict[str, float] = {}

    def try_detect(self, uid: str, t: float, radar_off: bool = False) -> bool:
        """처음 호출 시 0~scan_interval 랜덤 딜레이 후 탐지 확정.
        radar_off=True 이면 미탐지 위협은 새로 탐지 불가."""
        if uid in self._detect_t:
            return t >= self._detect_t[uid]
        if radar_off:
            return False
        self._detect_t[uid] = t + random.uniform(0.0, self.scan_interval)
        return False


class TrackRadar:
    """편대 추적 레이더 — 동시 추적 채널 및 획득 지연 관리 (SPY-1D: 18채널)."""
    max_ch:   int   = 18
    acq_time: float = 3.0  # 신규 위협 추적 획득 시간 (초)

    def __init__(self):
        self._track_t: dict[str, float] = {}
        self._ch_used: int = 0

    def try_track(self, uid: str, t: float) -> bool:
        """uid 추적 가능 여부. 처음이면 채널 할당 후 acq_time 딜레이."""
        if uid in self._track_t:
            return t >= self._track_t[uid]
        if self._ch_used >= self.max_ch:
            return False   # 채널 포화 — 추적 거부
        self._ch_used += 1
        self._track_t[uid] = t + self.acq_time
        return False       # 획득 중

    def release(self, uid: str):
        """위협 소멸 시 추적 채널 반환."""
        if uid in self._track_t:
            del self._track_t[uid]
            self._ch_used = max(0, self._ch_used - 1)


# ════════════════════════════════════════════════════════════════════════════
#  포팅 C: FriendlyAircraftObj — 항공 자산 (헬기 / 해상초계기)
# ════════════════════════════════════════════════════════════════════════════
class FriendlyAircraftObj:
    """
    함재 헬기(base_type='ship') / 육상초계기(base_type='land') 통합 클래스.
    매 tick _aircraft_asw()에서 잠수함 표적 확인 후 어뢰 투하.

    v9.8 탐지 상태 머신:
      헬기  — dipping  : idle → hovering(dip_hover_s) → detect_check → [fire | retry | abandon]
      초계기 — sonobuoy : idle → detect_check → [fire | retry(retry_s) | abandon]
    """
    def __init__(self, name: str, home_pos: 'Vec2'):
        self.name              = name
        self.info              = FRIENDLY_AIRCRAFT_DB[name]
        self.home_pos          = home_pos.copy()
        # BUG-7: v7 시뮬(700초)에 맞는 전시 긴급 출격 시간 적용 (engine.py 평시값 무시)
        self.t_available       = float(_AIRCRAFT_V7_SORTIE.get(name, self.info['sortie_time_s']))
        self.payload_remaining = self.info['payload_cnt']
        # v10.6: 공대함 strike payload (KF-21 해성-II 등)
        self.strike_payload_remaining = self.info.get('cap_strike_payload_cnt', 0)
        self.sorties           = 0
        self.total_cost        = 0.0
        # v9.8: 탐지 상태 머신
        self._asw_phase:    str   = 'idle'  # idle | hovering | cooldown
        self._dip_until:    float = 0.0     # 호버링 종료 시각 (dipping 전용)
        self._next_attempt: float = 0.0     # 다음 탐지 시도 가능 시각
        self._detect_fails: int   = 0       # 현재 표적 누적 탐지 실패 수
        self._search_target = None          # 현재 수색 중인 표적 (EnemyThreatObj)
        self._strike_cooldown_until: float = 0.0  # v10.6: AAS 교전 쿨다운


# ════════════════════════════════════════════════════════════════════════════
#  SimFrame
# ════════════════════════════════════════════════════════════════════════════
class SimFrame:
    __slots__ = ('t', 'friendly_ships', 'enemy_ships', 'missiles', 'events',
                 'ship_channels')

    def __init__(self, t: float):
        self.t              = t
        self.friendly_ships = []  # [(name, x, y, alive, hp)]
        self.enemy_ships    = []  # [(uid, preset, x, y, alive, hp, alt_m)]
        self.missiles       = []  # [(uid, x, y, mtype, name, alt_m)]
        self.events         = []  # [str]
        self.ship_channels  = []  # [(name, channels_used, max_channels)]


# ════════════════════════════════════════════════════════════════════════════
#  TimeStepEngine
# ════════════════════════════════════════════════════════════════════════════
class TimeStepEngine:
    """
    매 DT(1초)마다 실행 순서:
      1. 위치 갱신 (적 위협 접근/이탈, 미사일 비행)
      2. 적 발사 조건 확인 (함정/항공기/잠수함)
      3. 아군 TEWA — 방어 (적 미사일·항공기 요격)
      4. 아군 TEWA — 공격 (수상함: 해성/하푼, 잠수함: 홍상어/청상어)
      5. 적 TEWA   — SAM 방어 (수상함만, 아군 미사일 요격)
      6. 교전 결과 판정
      7. 프레임 기록
    """

    def __init__(self, cfg: dict, step_cb=None):
        self.cfg     = cfg
        self.t       = 0.0
        self._step_cb       = step_cb   # (t, t_max, alive, vls, last_log) — 단일 시뮬 진행 콜백
        self._phase_times: dict = {}    # 단계별 누적 실행 시간 (perf_counter 기반)
        # v10.7: 전술 의사결정 모드
        self._tactical_pause_cb = None  # SimWorker가 주입하는 콜백 (state) → choice dict
        self._tactical_interval  = cfg.get('tactical_interval', 30)  # 일시정지 간격(초)

        # BUG-2 fix: LatLon.from_xy() 기준점을 _build_*() 호출 전에 설정
        _set_region_ref(cfg.get('fleet_region', '동해 북부'))
        self._log_entries: list = []
        self._tick_events:  list = []

        # sim_seed 적용 (재현 보장) — python random + numpy 동시 고정
        seed = cfg.get('sim_seed', None)
        if seed:
            random.seed(int(seed))
            np.random.seed(int(seed))

        MissileObj.reset_counter()
        EnemyThreatObj.reset_counter()

        # 레이더 3단계 파이프라인
        self.search_radar = SearchRadar()  # 1단계: 수색 (탐지 지연)
        self.track_radar  = TrackRadar()   # 2단계: 추적 (채널 한계 + 획득 지연)
        # 3단계: 유도 채널은 FriendlyShipObj.channels_used / max_channels 로 관리

        # C&D 딜레이: {target_id → t_fire_allowed}
        # 탐지 시각 + cd_time_s + confirm_time_s + uniform(2,10)s 이후 발사 허용
        self._cd_fire_time: dict = {}
        # VLS 연속 발사 간격: {ship_id → t_last_vls}
        self._vls_last_fire: dict = {}
        # v8.26: 생존 적 항공 전력 중 최대 ECM 재밍 강도 (에어리어 재밍)
        self._active_ecm: float = 0.0

        # stats / wx 먼저 초기화 (build 함수에서 참조 가능)
        self.stats = {
            'total_threats':           0,
            'intercepted_threats':     0,
            'friendly_hits':           0,
            'enemy_hits':              0,
            'friendly_ships_lost':     0,
            'enemy_ships_destroyed':   0,
            'total_cost':              0.0,
            'aircraft_sorties':        0,
            # 포팅 D: REQ 판정용
            'peak_concurrent_threats': 0,
            't_first_fire':            -1.0,
            'total_missiles_fired':    0,
            # v8.16: 한미 연합 기여도
            'kor_shots':               0,
            'usa_shots':               0,
            'kor_cost':                0.0,
            'usa_cost':                0.0,
            # v9.11: 지상 BMD 자산 발사 횟수
            'ashore_sm3_fired':        0,
            'thaad_fired':             0,
        }
        weather = cfg.get('weather', '맑음 (주간)')
        self.wx = _make_physics_wx(weather)  # v9.13: Beaufort 물리값 override

        # v9.3: 아군 공격 임무 격침 기록
        self.strike_log: list = []

        # v9.4: VLS 탄약 고갈 시각 기록 {ship_name: t_depletion}
        self.vls_depletion_t: dict = {}
        # v9.4: 지상 발사 자산 재고
        self.ground_inv: dict = {
            '현무-4 (ASBM)':   cfg.get('hyunmoo4_stock',    0),
            'SM-3 (어쇼어)':   cfg.get('ashore_sm3_stock',  0),
            'THAAD 요격탄':    cfg.get('thaad_stock',        0),
        }
        self._ground_last_fire: float  = -999.0  # 현무-4 마지막 발사
        self._ashore_last_fire: float  = -999.0  # 어쇼어 SM-3 마지막 발사
        self._thaad_last_fire:  float  = -999.0  # THAAD 마지막 발사
        self._ground_cost: float = 0.0

        # NEW-A: 혼합 시나리오 파도 지연 스폰 큐 [(spawn_t, spec_dict), ...]
        self._pending_threats: list = []

        self.friendly_ships: List[FriendlyShipObj]    = self._build_friendly()
        self.missiles:       List[MissileObj]         = []
        self.enemy_threats:  List[EnemyThreatObj]     = self._build_enemies()
        self.aircraft:       List[FriendlyAircraftObj] = self._build_aircraft()
        self.frames:         List[SimFrame]            = []

    # ── 편대 구성 ─────────────────────────────────────────────────────────────

    def _build_friendly(self) -> List[FriendlyShipObj]:
        preset_name = self.cfg.get('fleet_preset', '단독 작전')
        preset = FLEET_PRESETS.get(preset_name, FLEET_PRESETS['단독 작전'])
        ships = []
        for spec in preset:
            s = FriendlyShipObj(spec['name'], spec['type'])
            if s.is_submarine:
                # 잠수함: SHIP_DB default_strike_inventory 사용 (해성/하푼 설정값 무시)
                s.strike_inventory = SHIP_DB[spec['type']].get('default_strike_inventory', {}).copy()
            else:
                s.strike_inventory = {
                    '해성-II':       self.cfg.get('haesong2_stock', 8),
                    '해성-I':        self.cfg.get('haesong1_stock', 0),
                    '하푼 Block II': self.cfg.get('harpoon_stock',  4),
                }
            # 포팅 A: 방어 무기 재고 수동 설정 (설정 없으면 SHIP_DB 기본값 유지)
            _def_map = [
                ('SM-3 Block IIA',  'sm3_stock'),
                ('SM-6',            'sm6_stock'),
                ('SM-2 Block IIIB', 'sm2_stock'),
                ('RIM-116 RAM',     'ram_stock'),
                ('홍상어 (대잠)',    'hongsango_stock'),
                ('청상어 (경어뢰)', 'cheongsango_stock'),
                ('Mk.46 경어뢰',    'mk46_stock'),
            ]
            for wpn, cfg_key in _def_map:
                if cfg_key in self.cfg and wpn in s.inventory:
                    s.inventory[wpn] = self.cfg[cfg_key]
            # 포팅 B: 기만기 재고 수동 설정
            if 'decoy_stock' in self.cfg:
                s.decoy_stock = self.cfg['decoy_stock']
            # NEW-AW: 함정별 위치 분산 (KDX-III 중심, KDX-II 3km, FFX 5km 기준 반경)
            radius = self._FORMATION_RADIUS.get(spec['type'], 3_000)
            if radius > 0:
                angle = math.radians(len(ships) * (360.0 / max(len(preset), 1)))
                s.pos = LatLon.from_xy(math.cos(angle) * radius, math.sin(angle) * radius)
            # NEW-XX: 랜덤 배치 옵션 — 각 함정에 임의 오프셋 추가
            if self.cfg.get('enable_random_placement', False):
                spread_m = self.cfg.get('random_spread_km', 5.0) * 1000.0
                rnd_angle = random.uniform(0, 2 * math.pi)
                rnd_r     = random.uniform(0, spread_m)
                s.pos = LatLon.from_xy(
                    s.pos.x + math.cos(rnd_angle) * rnd_r,
                    s.pos.y + math.sin(rnd_angle) * rnd_r,
                )
            ships.append(s)
        return ships

    def _build_enemies(self) -> List[EnemyThreatObj]:
        """
        플랫폼(항공기/수상함/잠수함) → EnemyThreatObj
        독립 미사일(탄도/순항/HGV/QBM) → MissileObj (self.missiles에 직접 추가)
        """
        # 포팅 A: 적군 편대 모드 (단일/프리셋/커스텀/랜덤/혼합)
        mode = self.cfg.get('enemy_fleet_mode', 'custom')
        if mode == 'preset':
            fleet_cfg = ENEMY_FLEET_PRESETS.get(
                self.cfg.get('enemy_fleet_preset', ''), [])
        elif mode == 'random':
            fleet_cfg = generate_random_enemy_fleet(
                difficulty=self.cfg.get('enemy_fleet_difficulty', '보통'),
                seed=self.cfg.get('enemy_fleet_seed', None))
        elif mode == 'mixed':
            # NEW-A: 혼합 시나리오 — 1파(delay=0)만 즉시 spawn, 나머지는 _pending_threats
            scenario_name = self.cfg.get('mixed_scenario', '')
            scenario = MIXED_ATTACK_SCENARIOS.get(scenario_name, {})
            fleet_cfg = []
            for wave in scenario.get('waves', []):
                delay = wave.get('delay_s', 0)
                if delay == 0:
                    fleet_cfg.extend(wave.get('threats', []))
                else:
                    for spec in wave.get('threats', []):
                        self._pending_threats.append((float(delay), dict(spec)))
        else:
            fleet_cfg = self.cfg.get('enemy_fleet', [])

        detect_km      = self.cfg.get('detect_km', 200)
        surface_det_km = self.cfg.get('surface_detect_km', self.cfg.get('detect_km', 45))
        sub_det_km     = self.cfg.get('sub_detect_km', 50)
        primary    = self._primary()  # 독립 미사일 표적

        # ── 적 전술 AI 전처리 ────────────────────────────────────────────────
        _ai_tactic = self.cfg.get('ai_tactic', None)
        _FAST_TYPES = {'탄도미사일', '극초음속활공체'}

        if _ai_tactic == 'saturation':
            # 채널 포화: 아군 총 교전 채널 ×1.5 가 될 때까지 위협 수 증폭
            total_ch = max(sum(s.max_channels for s in self.friendly_ships if s.alive), 1)
            target   = int(total_ch * 1.5)
            current  = max(sum(s.get('count', 1) for s in fleet_cfg), 1)
            if target > current:
                scale    = target / current
                fleet_cfg = [{'preset': s['preset'],
                               'count':  max(1, round(s.get('count', 1) * scale))}
                              for s in fleet_cfg]

        elif _ai_tactic == 'stagger':
            # 시차 공격: 고속(탄도·HGV≥1500 m/s) 선발 → 중속+30초 → 저속+60초
            new_fleet: list = []
            for spec in fleet_cfg:
                info  = ENEMY_DB.get(spec.get('preset', ''), {})
                speed = info.get('speed_ms', 300)
                ttype = info.get('type', '')
                if speed >= 1500 or ttype in _FAST_TYPES:
                    new_fleet.append(spec)                             # 즉시
                elif speed >= 600:
                    self._pending_threats.append((30.0, dict(spec)))   # +30초
                else:
                    self._pending_threats.append((60.0, dict(spec)))   # +60초
            # 즉시 위협이 아예 없으면 첫 항목은 즉시 등장
            fleet_cfg = new_fleet if new_fleet else fleet_cfg[:1]

        elif _ai_tactic == 'exploit_weakness':
            # 약점 공략: 단일 방향 집중 (다방위 억제)
            self.cfg['enable_multibearing'] = False

        threats: List[EnemyThreatObj] = []
        total = sum(s.get('count', 1) for s in fleet_cfg)
        idx = 0

        # v9.14: 해협 시나리오 — fleet_region='대한해협' 시 자동 활성화
        # 위협 접근 방위를 동/서수도 방향으로 제한 (multibearing 설정 무시)
        _strait_active = (self.cfg.get('fleet_region', '') == '대한해협')
        _strait_type   = self.cfg.get('strait_type', 'korea_west')
        if _strait_active:
            _strait_bases = [
                math.radians(b) for b in STRAIT_BEARING.get(_strait_type, [270.0])
            ]
            _n_strait = len(_strait_bases)

        # enable_multibearing: ON → 2~4개 방위 섹터로 분산 접근
        #                       OFF → 모두 단일 방향(기본 0°)
        # (해협 시나리오 활성 시 아래 분기 대신 _strait_bases 사용)
        _multibearing = self.cfg.get('enable_multibearing', False)
        if not _strait_active:
            if _multibearing:
                _n_sectors = min(4, max(2, total))
                _sector_bases = [
                    math.radians(i * (360 / _n_sectors))
                    for i in range(_n_sectors)
                ]
            else:
                _single_bearing = math.radians(random.uniform(0, 360))

        # 적 편대 전술 기동 — 초기 배치 오프셋
        # 'v_formation': V자 대형 (선두 1기 + 양익)
        # 'encirclement': 포위 기동 (원형 배치)
        # None / 기타: 기본 (bearing 분산)
        _tactics = self.cfg.get('enemy_tactics', None)

        for spec in fleet_cfg:
            name  = spec['preset']
            count = spec.get('count', 1)
            if name not in ENEMY_DB:
                continue
            info  = ENEMY_DB[name]
            ttype = info.get('type', '')

            for _ in range(count):
                if _strait_active:
                    # v9.14: 해협 방위 ± STRAIT_SPREAD_DEG 내 랜덤 배치
                    base = _strait_bases[idx % _n_strait]
                    bearing_rad = base + math.radians(
                        random.uniform(-STRAIT_SPREAD_DEG, STRAIT_SPREAD_DEG))
                elif _multibearing:
                    sector = idx % _n_sectors
                    bearing_rad = _sector_bases[sector] + math.radians(
                        random.uniform(-15, 15))
                else:
                    bearing_rad = _single_bearing

                # BUG-3 연계: 수상함은 대함 레이더 탐지거리(45km)에서 시작
                # 항공·독립미사일은 대공 탐지거리, 잠수함은 소나 탐지거리 유지
                # v9.6: 기습 잠수함은 ambush_start_km (기본 20km) 에서 시작 — 이미 탐지권 내
                if info.get('is_ambush') and info.get('category') == '대잠':
                    start_m = info.get('ambush_start_km', 20) * 1000
                elif info.get('category') == '대잠':
                    start_m = sub_det_km * 1000
                elif info.get('category') == '대함':
                    start_m = surface_det_km * 1000
                else:
                    start_m = detect_km * 1000

                pos = LatLon.from_xy(
                    math.cos(bearing_rad) * start_m,
                    math.sin(bearing_rad) * start_m,
                )

                if ttype in _STANDALONE_MISSILE_TYPES:
                    # 독립 미사일 위협: MissileObj로 직접 생성
                    _tgt = self._pick_target(is_torpedo=False)
                    m = MissileObj(
                        mtype    = 'enemy_strike',
                        name     = name,
                        pos      = pos,
                        target   = _tgt,
                        speed_ms = info['speed_ms'] * self.cfg.get('threat_spd_scale', 1.0),
                        pk_base  = _MISSILE_PK_MAP.get(name, _MISSILE_PK_DEFAULT),  # MED-5
                        owner_id = -1,
                        t_spawn  = 0.0,
                    )
                    m.altitude_m             = float(info.get('altitude_m', 0))
                    m.is_hgv                 = bool(info.get('is_hgv', False))
                    m.is_qbm                 = bool(info.get('is_qbm', False))
                    m.is_ballistic           = (ttype == '탄도미사일')
                    m.is_arm                 = bool(info.get('is_arm', False))
                    m.terminal_evasion_factor = info.get('missile_terminal_evasion', 1.0)
                    m.is_torpedo             = False
                    # 3D 시각화용: 포물선 궤도 계산에 사용할 초기 거리·정점 고도 저장
                    m._init_dist  = m.pos.dist_to(_tgt.pos)
                    m._peak_alt_m = m.altitude_m  # DB 고도 = 정점 고도
                    m.detect_m    = m._init_dist   # A-1: 스폰 거리 = 탐지 거리 근사
                    m.enemy_info  = info.copy()    # A-1: ENEMY_DB 원본
                    self.missiles.append(m)
                    self.stats['total_threats'] += 1
                else:
                    et = EnemyThreatObj(name, pos)
                    et.speed_ms *= self.cfg.get('threat_spd_scale', 1.0)
                    # 전술 기동 대형: V자 or 포위 초기 배치 오프셋
                    if _tactics == 'v_formation':
                        # 선두(idx=0)는 앞쪽, 나머지는 V자 양익
                        if idx == 0:
                            et.pos.x += math.cos(bearing_rad) * (-5_000)
                            et.pos.y += math.sin(bearing_rad) * (-5_000)
                        else:
                            wing_side = 1 if (idx % 2 == 0) else -1
                            perp = bearing_rad + math.pi / 2
                            wing_dist = (idx // 2 + 1) * 3_000
                            et.pos.x += math.cos(perp) * wing_dist * wing_side
                            et.pos.y += math.sin(perp) * wing_dist * wing_side
                            et.pos.x += math.cos(bearing_rad) * 3_000
                            et.pos.y += math.sin(bearing_rad) * 3_000
                    elif _tactics == 'encirclement' and not _strait_active:
                        # 포위: 전체가 원형으로 배치 (다방향 동시 접근 강화)
                        # v9.14: 해협 시나리오 활성 시 포위 기동 무시 (방위 제한 우선)
                        enc_bearing = math.radians((idx / max(total, 1)) * 360)
                        et.pos.x = math.cos(enc_bearing) * (detect_km * 1000)
                        et.pos.y = math.sin(enc_bearing) * (detect_km * 1000)
                    # v9.6: 기습 잠수함 — hidden_until 설정 (은닉 시간 동안 탐지·교전 불가)
                    if info.get('is_ambush') and et.is_sub:
                        et.hidden_until = float(info.get('ambush_hidden_s', 120))
                    # v9.14: 해협 시나리오 — 잠수함 잠항 수심을 해협 임계수심(sill_m)으로 cap
                    if _strait_active and et.is_sub:
                        strait_key = {'korea_west': 'korea_west',
                                      'korea_east': 'korea_east',
                                      'bilateral':  'korea_west'}.get(_strait_type, 'korea_west')
                        sill = STRAITS_DB.get(strait_key, {}).get('sill_m', 0)
                        if sill and et.altitude_m < -sill:
                            et.altitude_m = float(-sill)
                    threats.append(et)

                idx += 1

        return threats

    def _build_aircraft(self) -> List[FriendlyAircraftObj]:
        """포팅 C: enable_helo / enable_p3c / enable_p8a + v10.5 CAP 항공기."""
        aircraft = []
        primary_pos = self._primary().pos
        for en_key, preset_key, default in [
            ('enable_helo',  'helo_preset',  'AW-159 와일드캣'),
            ('enable_p3c',   'p3c_preset',   'P-3C 오라이온'),
            ('enable_p8a',   'p8a_preset',   'P-8A 포세이돈'),
            # v10.5: 한국 공군 CAP
            ('enable_f35a',  'f35a_preset',  'F-35A 라이트닝 II'),
            ('enable_kf21',  'kf21_preset',  'KF-21 보라매'),
            ('enable_fa50',  'fa50_preset',  'FA-50 골든이글'),
        ]:
            if not self.cfg.get(en_key, False):
                continue
            name = self.cfg.get(preset_key, default)
            if name not in FRIENDLY_AIRCRAFT_DB:
                continue
            aircraft.append(FriendlyAircraftObj(name, primary_pos))
        return aircraft

    # ── 헬퍼 ─────────────────────────────────────────────────────────────────

    def _primary(self) -> FriendlyShipObj:
        for t in ('KDX-III-B2', 'KDX-III-B1'):
            for s in self.friendly_ships:
                if s.alive and s.ship_type == t:
                    return s
        if not self.friendly_ships:
            raise RuntimeError("friendly_ships 비어있음 — 편대 프리셋을 확인하세요")
        return next((s for s in self.friendly_ships if s.alive), self.friendly_ships[0])

    def _pick_target(self, is_torpedo: bool = False) -> FriendlyShipObj:
        """적 미사일 타겟 선택 — 전단 내 생존 함정 분산 공격.
        어뢰: 이지스 기함(primary) 우선. 대함미사일: max_hp 가중 랜덤."""
        alive = [s for s in self.friendly_ships if s.alive]
        if len(alive) <= 1:
            return alive[0] if alive else (self.friendly_ships[0] if self.friendly_ships else self._primary())
        if is_torpedo:
            return self._primary()
        weights = [s._max_hp for s in alive]
        return random.choices(alive, weights=weights, k=1)[0]

    def _spawn_pending_threat(self, spec: dict):
        """NEW-A: 혼합 시나리오 파도 지연 스폰 — spec={'preset':name,'count':n}"""
        name  = spec.get('preset', '')
        count = spec.get('count', 1)
        if name not in ENEMY_DB:
            return
        info  = ENEMY_DB[name]
        ttype = info.get('type', '')
        detect_km      = self.cfg.get('detect_km', 200)
        surface_det_km = self.cfg.get('surface_detect_km', detect_km)
        sub_det_km     = self.cfg.get('sub_detect_km', 50)
        primary = self._primary()

        for _ in range(count):
            bearing_rad = math.radians(random.uniform(0, 360))
            if info.get('category') == '대잠':
                start_m = sub_det_km * 1000
            elif info.get('category') == '대함':
                start_m = surface_det_km * 1000
            else:
                start_m = detect_km * 1000
            pos = LatLon.from_xy(math.cos(bearing_rad) * start_m, math.sin(bearing_rad) * start_m)

            if ttype in _STANDALONE_MISSILE_TYPES:
                _tgt = self._pick_target(is_torpedo=False)
                m = MissileObj(
                    mtype='enemy_strike', name=name, pos=pos,
                    target=_tgt,
                    speed_ms=info['speed_ms'] * self.cfg.get('threat_spd_scale', 1.0),
                    pk_base=_MISSILE_PK_MAP.get(name, _MISSILE_PK_DEFAULT),
                    owner_id=-1, t_spawn=self.t,
                )
                m.altitude_m              = float(info.get('altitude_m', 0))
                m.is_hgv                  = bool(info.get('is_hgv', False))
                m.is_qbm                  = bool(info.get('is_qbm', False))
                m.is_ballistic            = (ttype == '탄도미사일')
                m.is_arm                  = bool(info.get('is_arm', False))
                m.terminal_evasion_factor = info.get('missile_terminal_evasion', 1.0)
                m.is_torpedo              = False
                m._init_dist              = m.pos.dist_to(_tgt.pos)
                m._peak_alt_m             = m.altitude_m
                m.detect_m                = m._init_dist
                m.enemy_info              = info.copy()
                self.missiles.append(m)
            else:
                et = EnemyThreatObj(name, pos)
                et.speed_ms *= self.cfg.get('threat_spd_scale', 1.0)
                self.enemy_threats.append(et)
            self.stats['total_threats'] += 1
            self._log(f"[{name}] {self.t:.0f}s 파도 스폰")

    def _log(self, msg: str):
        self._log_entries.append((self.t, msg))
        self._tick_events.append(msg)

    def _detect_range_m(self, ship: FriendlyShipObj, category: str,
                        alt_m: Optional[float] = None) -> float:
        if category == '대잠':
            base_km = ship.sensor_km.get('대잠', 50)
            factor  = self.wx.get('sonar_factor', self.wx.get('detect_range_factor', 1.0))
        else:
            # 대공: 이지스 데이터링크(Link-16) — 편대 최고 성능 레이더 공유
            # 대함: 수상 레이더 한계 — 수평선 넘어 탐지 불가, surface_detect_km 사용
            if category == '대함':
                base_km = max(ship.sensor_km.get('대함', 45),
                              self.cfg.get('surface_detect_km', 45))
            else:
                base_km = ship.sensor_km.get(category, self.cfg.get('detect_km', 200))
            factor = self.wx.get('radar_factor', self.wx.get('detect_range_factor', 1.0))
        # 함정 부분 피해: 레이더 성능 저하 반영
        detect_m = base_km * 1000 * factor * ship.radar_factor * self.cfg.get('detect_scale', 1.0)
        # v8.26: 적 ECM 에어리어 재밍 → 레이더 탐지거리 감소 (소나 무효, enable_ecm 플래그 연동)
        if (category != '대잠' and self._active_ecm > 0
                and self.cfg.get('enable_ecm', True)):
            jam = self._active_ecm * (1.0 - ship.eccm_factor)
            detect_m *= max(0.40, 1.0 - jam)
        # v9.12: 지형 음영 / v9.13: 증발 덕팅 — 레이더 탐지 한정
        if category != '대잠' and alt_m is not None:
            detect_m *= self._terrain_penalty(alt_m)
            if alt_m >= 0:   # 수중 표적은 덕팅·대기 보정 없음
                detect_m *= self._evap_duct_factor(alt_m)
                detect_m *= self._isa_refraction_factor(alt_m)
                detect_m *= self._troposcatter_factor(alt_m)
        return detect_m

    def _thermocline_factor(self, et: 'EnemyThreatObj') -> float:
        """
        수온약층(thermocline) 소나 탐지 보정.
        altitude_m < 0 = 잠항 수심 (음수).

        v9.12: ocean_acoustic_db WOA18 실측값 사용.
        해역(fleet_region)·계절(season) 조합으로 수심별 배율 조회.
        """
        if not et.is_sub:
            return 1.0
        depth = abs(et.altitude_m)
        if _OCEAN_ACOUSTIC_OK:
            region_key = REGION_TO_ACOUSTIC_KEY.get(
                self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
            )
            season = self.cfg.get('season', 'summer')
            return get_sonar_depth_factor(region_key, season, depth)
        # fallback: 기존 하드코딩 (ocean_acoustic_db 없을 때)
        if depth < 100:
            return 1.0
        elif depth < 300:
            return 1.0 - 0.55 * (depth - 100) / 200
        elif depth < 500:
            return 0.45 + 0.20 * (depth - 300) / 200
        else:
            return 0.65

    def _terrain_penalty(self, alt_m: float) -> float:
        """
        지형 레이더 음영 보정 — 해역별 산맥 차폐로 저고도 위협 탐지거리 감소.
        enable_terrain=False 또는 잠수함(소나)이면 1.0 반환.
        """
        if not self.cfg.get('enable_terrain', False):
            return 1.0
        if alt_m < 0:   # 잠수함 — 소나 탐지, 지형 페널티 없음
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        table = TERRAIN_RADAR_PENALTY.get(region_key, [])
        if not table:
            return 1.0
        prev_alt, prev_f = table[0]
        for a, f in table[1:]:
            if alt_m <= a:
                ratio = (alt_m - prev_alt) / max(a - prev_alt, 1)
                return prev_f + ratio * (f - prev_f)
            prev_alt, prev_f = a, f
        return prev_f

    def _evap_duct_factor(self, alt_m: float) -> float:
        """
        증발 덕팅(EDH) 보정 — 대기 하층 수증기 농도 역전으로 레이더 전파 해면 굴절.
        저고도 표적(alt_m ≤ EDH)의 탐지거리 증가. BF7 이상 강풍 시 덕트 파괴.
        enable_evap_duct=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_evap_duct', False):
            return 1.0
        if self.wx.get('beaufort', 2) >= 7:   # 강풍 이상 → 덕트 파괴
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        edh_m, boost = EVAP_DUCT_DB.get((region_key, season), (0, 1.0))
        if edh_m == 0:
            return 1.0
        if alt_m <= edh_m:
            return boost
        if alt_m <= edh_m * 2:
            ratio = (edh_m * 2 - alt_m) / max(edh_m, 1)
            return 1.0 + (boost - 1.0) * ratio
        return 1.0

    def _isa_refraction_factor(self, alt_m: float) -> float:
        """
        ISA 대기 굴절 보정 (라디오존데 4계절×5고도층 실측값 기반).
        100m 미만은 _evap_duct_factor 담당. enable_isa=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_isa', False):
            return 1.0
        if alt_m < 100:
            return 1.0
        if   alt_m < 500:  layer = 1
        elif alt_m < 1000: layer = 2
        elif alt_m < 3000: layer = 3
        else:              layer = 4
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return ISA_RADIOSONDE_DB.get((region_key, season, layer), 1.0)

    def _troposcatter_factor(self, alt_m: float) -> float:
        """
        트로포스캐터 링크 보정 — 대기 상층 난류 산란으로 수평선 너머 탐지거리 증가.
        고고도(≥1000m) 표적에만 적용. BF6 이상 강풍 시 산란층 파괴 → 1.0 반환.
        enable_isa=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_isa', False):
            return 1.0
        if alt_m < 1000:
            return 1.0
        if self.wx.get('beaufort', 2) >= 6:
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return TROPOSCATTER_DB.get((region_key, season), 1.0)

    def _wind_cep_factor(self) -> float:
        """
        고층 바람 CEP 배율 — 순항미사일 탄착 오차 증가로 명중률 저하.
        Pk_eff = pk_base / cep_factor (반환값 > 1 → Pk 감소)
        """
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return WIND_CEP_FACTOR.get((region_key, season), 1.0)

    # ── 1단계: 위치 갱신 ──────────────────────────────────────────────────────

    def _update_positions(self):
        # 아군 함정 회피 기동: 적 미사일이 15km 이내 접근 시 지그재그 위치 보정
        if self.cfg.get('enable_ship_evasion', False):
            # v9.14: 해협 시나리오 — STRAITS_DB 폭 기반 기동 공간 제한
            _evade_r = 15_000
            if self.cfg.get('fleet_region', '') == '대한해협':
                _st = self.cfg.get('strait_type', 'korea_west')
                _sk = 'korea_east' if _st == 'korea_east' else 'korea_west'
                _w  = STRAITS_DB.get(_sk, {}).get('width_km', _STRAIT_OPEN_SEA_KM)
                _evade_r = int(_evade_r * min(1.0, _w / _STRAIT_OPEN_SEA_KM))
            for ship in self.friendly_ships:
                if not ship.alive:
                    continue
                close = any(
                    m.alive and m.mtype == 'enemy_strike'
                    and m.pos.dist_to(ship.pos) < _evade_r
                    for m in self.missiles
                )
                if close:
                    angle = random.uniform(0, 2 * math.pi)
                    # 추진 피탄 시 speed_factor만큼 회피 거리 감소
                    evade_m = random.uniform(300, 800) * ship.speed_factor
                    ship.pos.x += math.cos(angle) * evade_m
                    ship.pos.y += math.sin(angle) * evade_m

        primary_pos = self._primary().pos
        for et in self.enemy_threats:
            if not et.alive:
                continue
            if et.is_retreating and et.retreat_pos:
                arrived = et.pos.move_toward(et.retreat_pos, et.speed_ms, DT)
                if arrived:
                    # MED-12: 재공격 가능 시 재접근, 아니면 전장 이탈
                    if et.reattack_count < et.max_reattacks:
                        et.reattack_count += 1
                        et.is_retreating = False
                        et.retreat_pos   = None
                        self._log(f"[재공격] {et.preset_name} 재접근 개시 ({et.reattack_count}/{et.max_reattacks})")
                    else:
                        et.alive = False
                        self._log(f"[이탈] {et.preset_name} 전장 이탈 완료")
            else:
                et.pos.move_toward(primary_pos, et.speed_ms, DT)

        # NEW-B: 드론/스웜 자폭 — 200m 이내 도달 시 피격 처리
        primary = self._primary()
        for et in self.enemy_threats:
            if not et.alive or not et.is_aircraft or et.is_retreating:
                continue
            if et.info.get('can_fire_missile', True):
                continue  # 일반 전투기는 미사일 발사 후 이탈 — 자폭 없음
            if et.pos.dist_to(primary.pos) > 200:
                continue
            primary.take_hit(et.preset_name, self.t)
            self.stats['friendly_hits'] += 1
            self._log(f"[피격!] {et.preset_name} 자폭")
            et.alive = False

        for m in self.missiles:
            m.update(DT)

        # v10.8: 해류 연동 — 수상함·잠수함 위치에 해류 벡터 누적
        if self.cfg.get('enable_current', False) and _get_current_vector:
            _region_map = {
                '동해 북부': 'east_sea', '동해 중부': 'east_sea',
                '서해': 'west_sea', '대한해협': 'korea_strait',
            }
            _cur_region = _region_map.get(self.cfg.get('fleet_region', '동해 북부'), 'east_sea')
            import datetime
            _month = datetime.date.today().month
            _cv = _get_current_vector(_cur_region, _month)
            _spd_ms = _cv['speed_cms'] / 100.0      # cm/s → m/s
            _dir_rad = math.radians(90 - _cv['direction_deg'])  # 진북 기준 → x/y 평면
            _dx = math.cos(_dir_rad) * _spd_ms * DT
            _dy = math.sin(_dir_rad) * _spd_ms * DT
            for et in self.enemy_threats:
                if et.alive and (et.is_ship or et.is_sub):
                    et.pos.x += _dx
                    et.pos.y += _dy
            for ship in self.friendly_ships:
                if ship.alive and not ship.is_submarine:
                    ship.pos.x += _dx
                    ship.pos.y += _dy

    # ── 2단계: 적 발사 ────────────────────────────────────────────────────────

    def _enemy_fire(self):
        primary = self._primary()
        for et in self.enemy_threats:
            if not et.alive or et.is_retreating:
                continue
            if not et.info.get('can_fire_missile'):
                continue
            # v9.6: 기습 잠수함 은닉 중 — 발사 불가
            if et.is_sub and self.t < et.hidden_until:
                continue

            in_flight = sum(
                1 for m in self.missiles
                if m.alive and m.owner_id == id(et) and m.mtype == 'enemy_strike'
            )
            if in_flight > 0:
                continue

            dist_m       = et.pos.dist_to(primary.pos)
            fire_range_m = et.info.get('missile_range_km', 0) * 1000 * 0.85
            if dist_m > fire_range_m:
                continue

            salvo   = random.randint(
                et.info.get('missile_salvo_min', 1),
                et.info.get('missile_salvo_max', 2),
            )
            m_speed = et.info.get('missile_speed_ms') or 300
            m_name  = et.info.get('missile_name') or '대함미사일'

            for _ in range(salvo):
                _is_torp = '어뢰' in m_name
                offset = LatLon.from_xy(
                    et.pos.x + random.uniform(-500, 500),
                    et.pos.y + random.uniform(-500, 500),
                )
                _m = MissileObj(
                    mtype    = 'enemy_strike',
                    name     = m_name,
                    pos      = offset,
                    target   = self._pick_target(is_torpedo=_is_torp),
                    speed_ms = m_speed,
                    pk_base  = _MISSILE_PK_MAP.get(m_name, _MISSILE_PK_DEFAULT),  # MED-5
                    owner_id = id(et),
                    t_spawn  = self.t,
                )
                # 포팅 B: 전술 속성 설정
                _m.terminal_evasion_factor = et.info.get('missile_terminal_evasion', 1.0)
                _m.is_torpedo = _is_torp
                self.missiles.append(_m)

            et.has_fired = True
            self.stats['total_threats'] += salvo

            if et.is_aircraft:
                et.is_retreating = True
                # 발사 후 200km 후퇴 이탈 (MED-12 재공격 패턴 도입으로 기존 500km에서 단축)
                angle = et.pos.bearing_to(primary.pos) + math.pi
                et.retreat_pos = LatLon.from_xy(
                    et.pos.x + math.cos(angle) * 200_000,
                    et.pos.y + math.sin(angle) * 200_000,
                )
                self._log(
                    f"[적 발사+이탈] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km), 이탈 개시"
                )
            elif et.is_sub:
                # v9.6: dual_weapon — 어뢰 발사 직후 해성-3 순항미사일 동시 발사
                if et.info.get('dual_weapon'):
                    d_name  = et.info.get('dual_missile_name', '')
                    d_spd   = et.info.get('dual_missile_speed_ms', 250)
                    d_salvo = random.randint(
                        et.info.get('dual_salvo_min', 1),
                        et.info.get('dual_salvo_max', 2),
                    )
                    for _ in range(d_salvo):
                        _dm = MissileObj(
                            mtype    = 'enemy_strike',
                            name     = d_name,
                            pos      = LatLon.from_xy(et.pos.x + random.uniform(-300, 300),
                                                      et.pos.y + random.uniform(-300, 300)),
                            target   = self._pick_target(is_torpedo=False),
                            speed_ms = d_spd,
                            pk_base  = _MISSILE_PK_MAP.get(d_name, _MISSILE_PK_DEFAULT),
                            owner_id = id(et),
                            t_spawn  = self.t,
                        )
                        _dm.terminal_evasion_factor = 0.87  # 해성-3: 스텔스 저고도 순항
                        _dm.is_torpedo = False
                        self.missiles.append(_dm)
                    self.stats['total_threats'] += d_salvo
                    self._log(
                        f"[기습 동시발사] {et.preset_name} -> {d_name} {d_salvo}발 "
                        f"(어뢰+순항 동시)"
                    )

                # LOW-18: 잠수함 발사 후 회피 기동 (발사 후 반대 방향 50km 이탈)
                et.is_retreating = True
                angle = et.pos.bearing_to(primary.pos) + math.pi
                et.retreat_pos = LatLon.from_xy(
                    et.pos.x + math.cos(angle) * 50_000,
                    et.pos.y + math.sin(angle) * 50_000,
                )
                self._log(
                    f"[적 발사+잠항회피] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km), 반대 방향 잠항"
                )
            else:
                self._log(
                    f"[적 발사] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km)"
                )

    # ── 3단계: 아군 방어 TEWA ─────────────────────────────────────────────────

    def _cd_allowed(self, target_key: int) -> bool:
        """
        C&D 딜레이 판정.
        첫 탐지 시: 레이더 빔 드웰(1~3s) + cd_time_s*날씨계수 + confirm_time_s + uniform(2,10)s
        야간/악천후: cd_time_factor로 딜레이 자동 증가.
        이후 시각 도달하면 True.
        """
        if target_key not in self._cd_fire_time:
            cd_factor = self.wx.get('cd_time_factor', 1.0)
            cd     = self.cfg.get('cd_time_s', 10) * cd_factor * self.cfg.get('cd_scale', 1.0)
            conf   = self.cfg.get('confirm_time_s', 3)
            dwell  = random.uniform(1, 3)   # 레이더 빔 드웰 타임
            jitter = random.uniform(2, 10)  # 위협 분류 랜덤 편차
            self._cd_fire_time[target_key] = self.t + dwell + cd + conf + jitter
            return False
        return self.t >= self._cd_fire_time[target_key]

    def _vls_interval_ok(self, ship: 'FriendlyShipObj') -> bool:
        """VLS 연속 발사 간격 2.5s 체크."""
        last = self._vls_last_fire.get(id(ship), -999.0)
        return (self.t - last) >= 2.5

    def _arm_radar_off_check(self):
        """ARM 탐지 시 표적 함정 레이더 일시 차단 (ARM 회피 전술, 8초).
        각 ARM 미사일당 레이더 OFF는 한 번만 트리거.
        레이더 OFF 중 ARM은 유도 신호 소실로 빗나감 — _check_hits()에서 처리."""
        if not self.cfg.get('enable_radar_off', True):
            return
        off_dur  = 8.0
        warn_m   = 120_000.0   # ARM이 120km 이내 접근 시 레이더 OFF 결심
        for m in self.missiles:
            if not (m.alive and m.is_arm):
                continue
            if getattr(m, '_radar_off_triggered', False):
                continue   # 이 ARM은 이미 레이더 OFF를 트리거했음
            tgt = m.target
            if not isinstance(tgt, FriendlyShipObj) or not tgt.alive:
                continue
            dist_m = m.pos.dist_to(tgt.pos)
            if dist_m < warn_m:
                m._radar_off_triggered = True
                tgt.radar_off_until = self.t + off_dur
                self._log(
                    f"[레이더 OFF] {tgt.name} ARM {dist_m/1000:.0f}km 접근 탐지 — "
                    f"레이더 {off_dur:.0f}초 차단 ({m.name})"
                )

    def _fire_ground_sam(self, wpn_key: str, target, dist_m: float,
                         name: str, pk: float, speed_ms: float,
                         cost: float, label: str, owner_id: int):
        """지상 BMD 자산 SAM 발사 (어쇼어/THAAD 공용). owner_id: -1=어쇼어, -2=THAAD."""
        self.ground_inv[wpn_key] -= 1
        self._ground_cost += cost
        self.stats['total_missiles_fired'] += 1
        if self.stats['t_first_fire'] < 0:
            self.stats['t_first_fire'] = self.t
        stat_key = 'ashore_sm3_fired' if owner_id == -1 else 'thaad_fired'
        self.stats[stat_key] += 1
        sam = MissileObj(
            mtype    = 'friendly_sam',
            name     = name,
            pos      = self._primary().pos,
            target   = target,
            speed_ms = speed_ms,
            pk_base  = pk,
            owner_id = owner_id,
            t_spawn  = self.t,
        )
        sam.rcs_m2 = _SAM_RCS.get(name, 0.001)  # Phase C: 적 레이더 탐지용
        self.missiles.append(sam)
        tgt_name = target.name if hasattr(target, 'name') else getattr(target, 'preset_name', '?')
        self._log(f"[지상 BMD] {label} → {name} 발사 → {tgt_name} (거리 {dist_m/1000:.0f}km)")

    def _ashore_defense(self, sorted_missiles: list):
        """
        이지스 어쇼어(SM-3) + THAAD 지상 BMD 교전 — 탄도/HGV 전담.

        레이어 순서:
          1. 이지스 어쇼어 SM-3 — 중간단계 (고도 ≥ 40km, 사거리 500km)
          2. THAAD — 종말고고도 (고도 10~150km, 사거리 200km)
        함정 SM-3는 이 두 자산이 소진되거나 실패한 경우의 최후 백업.
        """
        enable_ashore = self.cfg.get('enable_ashore', False)
        enable_thaad  = self.cfg.get('enable_thaad',  False)
        primary_pos   = self._primary().pos

        for m in sorted_missiles:
            if not m.alive or getattr(m, 'intercepted', False):
                continue
            if not (m.is_ballistic or m.is_hgv):
                continue

            dist_m = m.pos.dist_to(primary_pos)
            alt    = m.altitude_m
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is m and s.mtype == 'friendly_sam'
            )
            max_sams = 3 if m.is_hgv else 2

            # ── 1차: 이지스 어쇼어 SM-3 (중간단계) ──────────────────────────
            if (enable_ashore
                    and self.ground_inv.get('SM-3 (어쇼어)', 0) > 0
                    and (self.t - self._ashore_last_fire) >= _ASHORE_COOLDOWN_S
                    and alt >= _ASHORE_ALT_MIN_M
                    and dist_m <= _ASHORE_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'SM-3 (어쇼어)',
                    target   = m, dist_m = dist_m,
                    name     = 'SM-3 (어쇼어)',
                    pk       = _ASHORE_SM3_PK,
                    speed_ms = _ASHORE_SM3_SPD_MS,
                    cost     = _ASHORE_SM3_COST,
                    label    = '어쇼어',
                    owner_id = -1,
                )
                self._ashore_last_fire = self.t
                sams_on += 1

            # ── 2차: THAAD (종말고고도) ───────────────────────────────────────
            if (enable_thaad
                    and self.ground_inv.get('THAAD 요격탄', 0) > 0
                    and (self.t - self._thaad_last_fire) >= _THAAD_COOLDOWN_S
                    and _THAAD_ALT_MIN_M <= alt <= _THAAD_ALT_MAX_M
                    and dist_m <= _THAAD_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'THAAD 요격탄',
                    target   = m, dist_m = dist_m,
                    name     = 'THAAD 요격탄',
                    pk       = _THAAD_PK,
                    speed_ms = _THAAD_SPD_MS,
                    cost     = _THAAD_COST,
                    label    = 'THAAD',
                    owner_id = -2,
                )
                self._thaad_last_fire = self.t
                sams_on += 1

    def _friendly_defense(self):
        """
        다층 방어 (enable_layered_defense=True, 기본 ON):
          KDX-III-B2 → KDX-III-B1 → KDX-II → FFX-III → FFX-II → FFX-I 순서로 위협당 1발씩 배정.

        CEC 사전 동시 배정 (enable_cec_preassign=True, 기본 OFF):
          탐지 즉시 1차+2차 함정 동시 발사. 위협당 최대 2발 허용.

        NEW-AW: 위협 긴급도 정렬 — 속도/잔여거리 내림차순 (빠르고 가까운 위협 먼저)
        NEW-AW: Shoot-Look-Shoot — 탄도/초음속/HGV/QBM 위협은 CEC 없이도 2발 배정
        """
        # CEC 두절 시나리오: enable_cec_jammed=True 이면 CEC 강제 해제 + 독립 교전
        cec_jammed = self.cfg.get('enable_cec_jammed', False)
        # v10.4: enable_cec — 탐지 커버리지 통합 + SAM 사전 동시 배정
        # enable_cec_preassign 하위 호환 유지 (구버전 cfg 지원)
        cec = self.cfg.get('enable_cec', self.cfg.get('enable_cec_preassign', True)) and not cec_jammed

        # CEC 두절 시 각 함정이 독립적으로 교전 (다층 방어 비활성화, 1함정=1교전)
        layered = self.cfg.get('enable_layered_defense', True) and not cec_jammed

        primary_pos = self._primary().pos

        def _urgency(obj):
            # 임박도 = speed / 교전창 잔여 거리 (교전창 하한 근접 시 임박도 급증)
            # floor: 해당 위협 유형의 최소 유효 교전 거리 — 이 거리 이하엔 주 요격 무기 사용 불가
            # 기존 speed/dist 공식은 탄도탄이 SM-3 교전창 끝에 가까울수록 순항미사일 대비
            # urgency 비율이 오히려 하락하는 역전 현상 발생 → floor 도입으로 수정
            is_bal = getattr(obj, 'is_ballistic', False)
            is_hgv = getattr(obj, 'is_hgv',       False)
            is_qbm = getattr(obj, 'is_qbm',       False)
            if is_bal or is_hgv:
                floor = 150_000.0   # SM-3 중간단계 교전창 하한 ~150km
            elif is_qbm:
                floor = 20_000.0    # SM-6 최소 교전거리 ~20km
            else:
                floor = 5_000.0     # CIWS/RAM 최소 교전거리 ~5km
            d_eff = max(obj.pos.dist_to(primary_pos) - floor, 200.0)
            return getattr(obj, 'speed_ms', 300.0) / d_eff

        # 다층 방어 우선순위 정렬 (B2=0, B1=1, KDX-II=2, FFX-III=3, FFX-II=4, FFX-I=5, 나머지=99)
        if layered:
            sorted_ships = sorted(
                [s for s in self.friendly_ships if s.alive],
                key=lambda s: SHIP_LAYER_PRI.get(s.ship_type, 99)
            )
        else:
            # CEC 두절: 함정별로 독립 교전 — 처리 순서는 무작위
            import random as _rnd
            _ships = [s for s in self.friendly_ships if s.alive]
            _rnd.shuffle(_ships)
            sorted_ships = _ships

        # 주요 함정 레이더 OFF 여부 (ARM 회피 전술)
        primary_ship = self._primary()
        is_radar_off = self.t < primary_ship.radar_off_until

        # (A) 적 대함 미사일 요격 — 긴급도 순 정렬
        sorted_missiles = sorted(
            [m for m in self.missiles if m.alive and m.mtype == 'enemy_strike'],
            key=_urgency, reverse=True
        )

        # 지상 BMD 선제 교전 (어쇼어 SM-3 → THAAD, 함정 SM-3보다 우선)
        if self.cfg.get('enable_ashore', False) or self.cfg.get('enable_thaad', False):
            self._ashore_defense(sorted_missiles)

        for m in sorted_missiles:
            # 1단계: 수색 레이더 탐지 확정 (SPY-1D 빔 회전 0~6초 지연)
            if not self.search_radar.try_detect(m.uid, self.t, radar_off=is_radar_off):
                continue
            # 2단계: 추적 채널 획득 (18채널 한계 + 3초 획득 지연)
            if not self.track_radar.try_track(m.uid, self.t):
                continue
            # 3단계(C&D): 분류·교전 결심 딜레이
            if not self._cd_allowed(id(m)):
                continue
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is m and s.mtype == 'friendly_sam'
            )
            # 위협 유형별 살보 수 — CEC 두절 시 1발 고정
            # HGV(극초음속 활공체): SM-3 교전창 극히 좁음 → 3발
            # 탄도탄·QBM·초음속(≥1000m/s): Shoot-Look-Shoot → 2발
            # 기타: 1발 / CEC 추가 협동 배정: +1발
            if m.is_hgv:
                _base = 3
            elif m.is_ballistic or m.is_qbm or m.speed_ms >= 1000:
                _base = 2
            else:
                _base = 1
            max_sams = 1 if cec_jammed else (_base + (1 if cec else 0))
            # BUG-1 fix: 전술 의사결정 모드 살보 수 우선 적용
            _tac_max = self.cfg.get('_tactical_max_salvo')
            if _tac_max is not None:
                max_sams = int(_tac_max)
            if sams_on >= max_sams:
                continue

            shots = 0
            for ship in sorted_ships:
                if sams_on + shots >= max_sams:
                    break
                if not self._vls_interval_ok(ship):
                    continue
                dist_m = ship.pos.dist_to(m.pos)
                # v10.4: CEC 탐지 커버리지 — 함정 자체 탐지거리 초과 여부
                own_detect_m = self._detect_range_m(ship, '대공', m.altitude_m)
                if dist_m > own_detect_m:
                    if not cec:   # CEC 미활성: 자체 탐지 불가 함정 교전 불가
                        continue
                    relay = True  # CEC 중계: 아군 데이터링크로 교전, Pk 패널티 적용
                else:
                    relay = False
                wpn = self._select_defense_wpn(ship, m, dist_m)
                if not wpn or ship.channels_used >= ship.max_channels:
                    continue
                if ship.inventory.get(wpn, 0) <= 0:
                    continue
                self._launch_friendly_sam(ship, wpn, m, dist_m, is_aa=False, cec_relay=relay)
                self._vls_last_fire[id(ship)] = self.t
                shots += 1

        # (B) 적 항공기 직접 요격 — 긴급도 순 정렬
        sorted_ac = sorted(
            [et for et in self.enemy_threats
             if et.alive and et.is_aircraft and not et.is_retreating],
            key=_urgency, reverse=True
        )
        for et in sorted_ac:
            # 1단계: 수색 레이더 탐지 확정
            if not self.search_radar.try_detect(et.uid, self.t, radar_off=is_radar_off):
                continue
            # 2단계: 추적 채널 획득
            if not self.track_radar.try_track(et.uid, self.t):
                continue
            # 3단계(C&D): 분류·교전 결심 딜레이
            if not self._cd_allowed(id(et)):
                continue
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is et and s.mtype == 'friendly_sam'
            )
            # 항공기 위협 유형별 살보 수
            # HGV 항공기(킨잘·지르콘 등) 또는 극초음속(≥1500m/s): 3발
            # 초음속(≥600m/s): 2발 / 기타: 1발 / CEC 추가: +1발
            if getattr(et, 'is_hgv', False) or et.speed_ms >= 1500:
                _base_ac = 3
            elif et.speed_ms >= 600:
                _base_ac = 2
            else:
                _base_ac = 1
            max_sams = 1 if cec_jammed else (_base_ac + (1 if cec else 0))
            # BUG-1 fix: 전술 의사결정 모드 살보 수 우선 적용
            _tac_max = self.cfg.get('_tactical_max_salvo')
            if _tac_max is not None:
                max_sams = int(_tac_max)
            if sams_on >= max_sams:
                continue

            shots = 0
            for ship in sorted_ships:
                if sams_on + shots >= max_sams:
                    break
                if not self._vls_interval_ok(ship):
                    continue
                dist_m = ship.pos.dist_to(et.pos)
                # v10.4: CEC 탐지 커버리지
                own_detect_m = self._detect_range_m(ship, '대공', et.altitude_m)
                if dist_m > own_detect_m:
                    if not cec:
                        continue
                    relay = True
                else:
                    relay = False
                wpn = self._select_aa_wpn(ship, et, dist_m)
                if not wpn or ship.channels_used >= ship.max_channels:
                    continue
                if ship.inventory.get(wpn, 0) <= 0:
                    continue
                self._launch_friendly_sam(ship, wpn, et, dist_m, is_aa=True, cec_relay=relay)
                self._vls_last_fire[id(ship)] = self.t
                shots += 1

    def _launch_friendly_sam(self, ship: FriendlyShipObj, wpn: str, target,
                              dist_m: float, is_aa: bool, cec_relay: bool = False):
        wpn_info = FRIENDLY_DB[wpn]
        ship.inventory[wpn]   -= 1
        ship.channels_used    += 1
        ship.total_cost       += wpn_info['cost_usd']
        # 포팅 D: 발사 통계
        self.stats['total_missiles_fired'] += 1
        if self.stats['t_first_fire'] < 0:
            self.stats['t_first_fire'] = self.t
        # v8.16: 한미 연합 기여도 카운트
        if ship.nation == 'USA':
            self.stats['usa_shots'] += 1
            self.stats['usa_cost']  += wpn_info['cost_usd']
        else:
            self.stats['kor_shots'] += 1
            self.stats['kor_cost']  += wpn_info['cost_usd']
        sam = MissileObj(
            mtype    = 'friendly_sam',
            name     = wpn,
            pos      = ship.pos,
            target   = target,
            speed_ms = wpn_info['speed_ms'],
            pk_base  = wpn_info['pk_dist']['mean'],
            owner_id = id(ship),
            t_spawn  = self.t,
        )
        sam.rcs_m2   = _SAM_RCS.get(wpn, 0.001)  # Phase C: 적 레이더 탐지용
        sam.cec_relay = cec_relay                  # v10.4: CEC 중계 교전 여부
        self.missiles.append(sam)
        prefix = '[대공 방어]' if is_aa else '[방어]'
        tgt_name = target.name if hasattr(target, 'name') else target.preset_name
        self._log(f"{prefix} {ship.name} -> {wpn} 발사 -> {tgt_name} (거리 {dist_m/1000:.1f}km)")
        # 탄약 재보급 한계: VLS 주요 무기 완전 소진 시 경고 + 시각 기록
        vls_wpns = ['SM-3 Block IIA', 'SM-6', 'SM-2 Block IIIB', 'RIM-116 RAM']
        if not ship._vls_depleted:
            if all(ship.inventory.get(w, 0) == 0 for w in vls_wpns):
                ship._vls_depleted = True
                self._log(f"[경고] {ship.name} VLS 탄약 완전 소진 — 방어 불능")
                # v9.4: 고갈 발생 시각 기록 (최초 1회만)
                self.vls_depletion_t.setdefault(ship.name, self.t)

    # LOW-11: 조명기(SPG-62) 가용 채널 (SM-2는 반능동 유도 → 조명기 필요)
    _ILLUMINATOR_MAX = {
        'KDX-III-B2': 3, 'KDX-III-B1': 3, 'KDX-II': 2,
        'FFX-III': 2, 'FFX-II': 2, 'FFX-I': 1,
    }
    # NEW-AW: 함대 포진 기준 반경 (Batch II 중심, B1 1km, KDX-II 3km, FFX 5km)
    _FORMATION_RADIUS = {
        'KDX-III-B2': 0, 'KDX-III-B1': 1_000, 'KDX-II': 3_000,
        'FFX-III': 5_000, 'FFX-II': 5_000, 'FFX-I': 5_000,
    }

    def _sm2_illuminator_ok(self, ship: FriendlyShipObj) -> bool:
        """SM-2 추가 발사 가능 여부: 현재 비행 중 SM-2 수 < 조명기 최대 채널."""
        max_ill = self._ILLUMINATOR_MAX.get(ship.ship_type, 1)
        in_flight = sum(1 for s in self.missiles
                        if s.alive and s.name == 'SM-2 Block IIIB' and s.owner_id == id(ship))
        return in_flight < max_ill

    def _select_defense_wpn(self, ship: FriendlyShipObj, m: MissileObj,
                            dist_m: float) -> Optional[str]:
        """미사일 위협 요격 무기 선택. 고도·유형 인식. 무장 피탄 비활성화 반영.
        v10.7: _tactical_wpn_priority cfg 키가 있으면 해당 무기 우선 반환."""
        # 전술 우선순위 반영: 지정 무기가 재고 있으면 사거리 내에서 먼저 반환
        _prio = self.cfg.get('_tactical_wpn_priority')
        if _prio and ship.available(_prio) > 0:
            wpn_info = FRIENDLY_DB.get(_prio)
            if wpn_info:
                # BUG-5 fix: range_km 없는 무기도 사거리 내 교전 가능하도록 500km 기본값
                max_r = wpn_info.get('range_km', 500) * 1000
                if dist_m <= max_r:
                    return _prio

        alt          = m.altitude_m
        is_hgv       = m.is_hgv
        is_qbm       = m.is_qbm
        is_ballistic = m.is_ballistic

        def ok(wpn):
            return ship.available(wpn) > 0

        # HGV / 고고도 탄도 중간단계 → SM-3
        # 어쇼어 활성·잔여 있으면 함정 SM-3 생략 (지상 자산이 우선 교전)
        if (is_hgv or (is_ballistic and alt >= 40_000)) and dist_m <= 500_000:
            ashore_covers = (
                self.cfg.get('enable_ashore', False)
                and self.ground_inv.get('SM-3 (어쇼어)', 0) > 0
            )
            if not ashore_covers and ok('SM-3 Block IIA'):
                return 'SM-3 Block IIA'

        # QBM (저고도 기동탄도) → SM-6 우선 (SM-3 무효)
        if is_qbm and dist_m <= 240_000:
            if ok('SM-6'): return 'SM-6'

        # 근거리→원거리 표준 다층 (SM-2는 조명기 가용 시에만)
        if dist_m <= 2_000   and ok('CIWS-II (Phalanx)'): return 'CIWS-II (Phalanx)'
        if dist_m <= 9_000   and ok('RIM-116 RAM'):        return 'RIM-116 RAM'
        if dist_m <= 50_000  and ok('ESSM Block II'):      return 'ESSM Block II'
        if dist_m <= 50_000  and ok('해궁 (K-SAAM)'):      return '해궁 (K-SAAM)'
        # LOW-11: SM-2 조명기 채널 확인
        if dist_m <= 170_000 and ok('SM-2 Block IIIB') and self._sm2_illuminator_ok(ship):
            return 'SM-2 Block IIIB'
        if dist_m <= 240_000 and ok('SM-6'):          return 'SM-6'
        if dist_m <= 240_000 and ok('SM-6 Block IB'): return 'SM-6 Block IB'
        if dist_m <= 500_000 and ok('SM-3 Block IIA'): return 'SM-3 Block IIA'  # BUG-2
        return None

    def _select_aa_wpn(self, ship: FriendlyShipObj, et: EnemyThreatObj,
                       dist_m: float) -> Optional[str]:
        """
        항공기 목표 대공 무기 선택 (고도 3단 구분). 무장 피탄 비활성화 반영.

        SM-3는 대기권 외 BMD 전용 → 항공기 요격 불가.
          ≥ 10,000m (고고도): SM-2 → SM-6 (RAM 불필요, 사거리 초과 시 교전 불가)
          3,000–10,000m (중고도): SM-2 → SM-6 → RAM (근접 시)
          < 3,000m (저고도 침투): RAM 우선 → SM-2 → SM-6
        """
        alt    = et.altitude_m
        sm2_ok = self._sm2_illuminator_ok(ship)

        def ok(wpn):
            return ship.available(wpn) > 0

        if alt >= 10_000:
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            return None

        elif alt >= 3_000:
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            if dist_m <= 50_000  and ok('ESSM Block II'):               return 'ESSM Block II'
            if dist_m <= 9_000   and ok('RIM-116 RAM'):                 return 'RIM-116 RAM'
            return None

        else:
            if dist_m <= 9_000   and ok('RIM-116 RAM'):                 return 'RIM-116 RAM'
            if dist_m <= 50_000  and ok('ESSM Block II'):               return 'ESSM Block II'
            if dist_m <= 50_000  and ok('해궁 (K-SAAM)'):               return '해궁 (K-SAAM)'
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            return None

    # ── 4단계: 아군 공격 TEWA ─────────────────────────────────────────────────

    def _friendly_strike(self):
        """
        수상함 → 해성/하푼 (strike_inventory)
        잠수함 → 홍상어/청상어 (inventory)
        """
        if not self.cfg.get('enable_strike', True):
            return
        for ship in self.friendly_ships:
            if not ship.alive:
                continue

            # v10.6: 항모(high_value_target) 우선 정렬 — 화력 집중
            sorted_threats = sorted(
                self.enemy_threats,
                key=lambda e: (0 if e.high_value_target else 1, e.hp),
            )
            for et in sorted_threats:
                if not et.alive:
                    continue
                if not (et.is_ship or et.is_sub):
                    continue
                # v9.6: 기습 잠수함 은닉 중 — 아군 탐지·교전 불가
                if et.is_sub and self.t < et.hidden_until:
                    continue

                dist_m   = ship.pos.dist_to(et.pos)
                category = et.category
                detect_m = self._detect_range_m(ship, category, alt_m=et.altitude_m)
                # 잠수함: 수온약층 소나 보정 추가 적용
                if et.is_sub:
                    detect_m *= self._thermocline_factor(et)
                    # NEW-AW: 이탈 잠수함 — 고속 이탈로 소나 접촉 급감 (탐지 70% 감소)
                    if et.is_retreating:
                        detect_m *= 0.30
                if dist_m > detect_m:
                    continue

                if et.is_ship:
                    en_route = sum(
                        1 for m in self.missiles
                        if m.alive and m.target is et and m.mtype == 'friendly_strike'
                    )
                    # v10.6: 항모(high_value_target) 살보 6발, 일반 수상함 4발
                    max_salvo = 6 if et.high_value_target else 4
                    if en_route >= max_salvo:
                        continue

                    if ship.is_submarine:
                        # 아군 잠수함 → 적 수상함 공격 (현무-3C/하푼/어뢰)
                        wpn = self._select_sub_strike_wpn(ship, dist_m)
                        if not wpn:
                            continue
                        if wpn in FRIENDLY_STRIKE_DB:
                            wpn_info = FRIENDLY_STRIKE_DB[wpn]
                            ship.strike_inventory[wpn] = ship.strike_inventory.get(wpn, 0) - 1
                            pk_b = wpn_info['pk_base']
                            spd  = wpn_info['speed_ms']
                            cost = wpn_info['cost_usd']
                        else:
                            wpn_info = FRIENDLY_DB[wpn]
                            ship.inventory[wpn] -= 1
                            pk_b = wpn_info['pk_dist']['mean']
                            spd  = wpn_info['speed_ms']
                            cost = wpn_info['cost_usd']
                        ship.total_cost += cost
                        _ms = MissileObj(
                            mtype    = 'friendly_strike',
                            name     = wpn,
                            pos      = ship.pos,
                            target   = et,
                            speed_ms = spd,
                            pk_base  = pk_b,
                            owner_id = id(ship),
                            t_spawn  = self.t,
                        )
                        _ms.seeker = (wpn_info.get('seeker', 'radar')
                                      if wpn in FRIENDLY_STRIKE_DB else 'radar')
                        self.missiles.append(_ms)
                        self._log(
                            f"[공격] {ship.name} -> {wpn} -> {et.preset_name} "
                            f"(거리 {dist_m/1000:.0f}km)"
                        )
                    else:
                        wpn = self._select_strike_wpn(ship, dist_m)
                        if not wpn:
                            continue
                        wpn_info = FRIENDLY_STRIKE_DB[wpn]
                        # SM-6 대함 모드: VLS inventory에서 소모
                        if wpn == 'SM-6 대함 모드':
                            ship.inventory['SM-6'] -= 1
                        elif wpn == 'Tomahawk Block V':
                            ship.inventory['Tomahawk Block V'] = ship.inventory.get('Tomahawk Block V', 0) - 1
                        elif wpn == 'Mk.45 5인치 함포':
                            pass  # 함포는 재고 무한 (수백 발 탑재)
                        else:
                            ship.strike_inventory[wpn] = ship.strike_inventory.get(wpn, 0) - 1
                        ship.total_cost += wpn_info['cost_usd']
                        _m = MissileObj(
                            mtype    = 'friendly_strike',
                            name     = wpn,
                            pos      = ship.pos,
                            target   = et,
                            speed_ms = wpn_info['speed_ms'],
                            pk_base  = wpn_info['pk_base'],
                            owner_id = id(ship),
                            t_spawn  = self.t,
                        )
                        _m.seeker = wpn_info.get('seeker', 'radar')
                        self.missiles.append(_m)
                        self._log(
                            f"[공격] {ship.name} -> {wpn} -> {et.preset_name} "
                            f"(거리 {dist_m/1000:.0f}km)"
                        )

                elif et.is_sub:
                    en_route = sum(
                        1 for m in self.missiles
                        if m.alive and m.target is et and m.mtype == 'friendly_strike'
                    )
                    if en_route >= 1:
                        continue
                    wpn = self._select_asw_wpn(ship, dist_m)
                    if not wpn:
                        continue
                    wpn_info = FRIENDLY_DB[wpn]
                    ship.inventory[wpn] -= 1
                    ship.total_cost += wpn_info['cost_usd']
                    self.missiles.append(MissileObj(
                        mtype    = 'friendly_strike',
                        name     = wpn,
                        pos      = ship.pos,
                        target   = et,
                        speed_ms = wpn_info['speed_ms'],
                        pk_base  = wpn_info['pk_dist']['mean'],
                        owner_id = id(ship),
                        t_spawn  = self.t,
                    ))
                    self._log(
                        f"[대잠 공격] {ship.name} -> {wpn} -> {et.preset_name} "
                        f"(거리 {dist_m/1000:.1f}km)"
                    )

        # v9.4: 지상 발사 현무-4 ASBM — 60초 쿨다운, 함정당 최대 2발 비행 중
        _h4_wpn = '현무-4 (ASBM)'
        if (self.cfg.get('enable_strike', True)
                and self.ground_inv.get(_h4_wpn, 0) > 0
                and self.t - self._ground_last_fire >= 60.0):
            wpn_info_h4 = FRIENDLY_STRIKE_DB[_h4_wpn]
            primary_pos  = self._primary().pos
            for et in self.enemy_threats:
                if not et.alive or not et.is_ship:
                    continue
                dist_m = primary_pos.dist_to(et.pos)
                if dist_m > wpn_info_h4['range_km'] * 1000:
                    continue
                en_route = sum(
                    1 for m in self.missiles
                    if m.alive and m.target is et and m.name == _h4_wpn
                )
                if en_route >= 2:
                    continue
                self.ground_inv[_h4_wpn] -= 1
                self._ground_last_fire = self.t
                self._ground_cost += wpn_info_h4['cost_usd']
                self.missiles.append(MissileObj(
                    mtype    = 'friendly_strike',
                    name     = _h4_wpn,
                    pos      = primary_pos,
                    target   = et,
                    speed_ms = wpn_info_h4['speed_ms'],
                    pk_base  = wpn_info_h4['pk_base'],
                    owner_id = 0,
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[지상공격] {_h4_wpn} -> {et.preset_name} "
                    f"(거리 {dist_m/1000:.0f}km)"
                )
                if self.ground_inv.get(_h4_wpn, 0) <= 0:
                    break

    # ── 4.5단계: 항공 자산 대잠 (포팅 C) ─────────────────────────────────────

    def _aircraft_asw(self):
        """
        v9.8: 항공 대잠 탐지 상태 머신 — 헬기(디핑소나)·초계기(소노부이) 방식 분리.

        헬기 흐름:  idle → (범위 진입) → hovering(dip_hover_s) → 탐지 확률 판정
                    탐지 성공 → 어뢰 투하 / 탐지 실패 → cooldown(retry_s) → 재시도
                    max_attempts 초과 → 포기(abandon)

        초계기 흐름: idle → (범위 진입) → 소노부이 투하 → 탐지 확률 판정
                    탐지 성공 → 어뢰 투하 / 탐지 실패 → cooldown(retry_s) → 재투하
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')

        for ac in self.aircraft:
            if ac.payload_remaining <= 0:
                continue
            if self.t < ac.t_available:
                continue
            wx_limits = ac.info.get('weather_limits', {})
            if not wx_limits.get(weather, True):
                continue

            asw_mode = ac.info.get('asw_mode', 'sonobuoy')

            # ── 호버링 단계 처리 (dipping 전용) ─────────────────────────────
            if ac._asw_phase == 'hovering':
                if self.t < ac._dip_until:
                    continue   # 아직 호버링 중
                # 호버링 완료 → 탐지 확률 판정
                et = ac._search_target
                if et is None or not et.alive:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                self._asw_detect_check(ac, et, primary)
                continue

            # ── 쿨다운 단계 (재탐색 대기) ────────────────────────────────────
            if ac._asw_phase == 'cooldown':
                if self.t < ac._next_attempt:
                    continue
                et = ac._search_target
                if et is None or not et.alive:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                # 재시도 — 사거리 재확인 후 탐지
                dist_to_sub = primary.pos.dist_to(et.pos)
                total_dist  = dist_to_sub + (
                    ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)
                if total_dist > ac.info['range_km'] * 1000:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                detect_m = self._asw_detect_range(ac, primary, et)
                bonus_m  = ac.info.get('sonobuoy_detect_bonus_km', 0) * 1000
                if dist_to_sub > detect_m + bonus_m:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                if asw_mode == 'dipping':
                    # 디핑소나 재전개
                    ac._asw_phase = 'hovering'
                    ac._dip_until = self.t + ac.info.get('dip_hover_s', 60)
                    self._log(f"[대잠 헬기] {ac.name} → {et.preset_name} 재탐색 "
                              f"(시도 {ac._detect_fails+1}/{ac.info.get('max_attempts',3)}) "
                              f"디핑소나 재전개…")
                else:
                    self._asw_detect_check(ac, et, primary)
                continue

            # ── idle 단계 — 새 표적 탐색 ─────────────────────────────────────
            for et in self.enemy_threats:
                if not et.alive or not et.is_sub:
                    continue
                # v9.6: 기습 잠수함 은닉 중 탐지 불가
                if self.t < et.hidden_until:
                    continue
                # 이미 어뢰가 향하고 있으면 패스
                if any(m.alive and m.target is et and m.mtype == 'friendly_strike'
                       for m in self.missiles):
                    continue

                # 사거리 체크
                dist_to_sub = primary.pos.dist_to(et.pos)
                total_dist  = dist_to_sub + (
                    ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)
                if total_dist > ac.info['range_km'] * 1000:
                    continue

                # 탐지 범위 체크
                detect_m = self._asw_detect_range(ac, primary, et)
                bonus_m  = ac.info.get('sonobuoy_detect_bonus_km', 0) * 1000
                if dist_to_sub > detect_m + bonus_m:
                    continue

                # 표적 포착 → 탐지 단계 진입
                ac._search_target = et
                ac._detect_fails  = 0
                craft_type = '초계기' if ac.info.get('base_type') == 'land' else '헬기'

                if asw_mode == 'dipping':
                    ac._asw_phase = 'hovering'
                    ac._dip_until = self.t + ac.info.get('dip_hover_s', 60)
                    self._log(
                        f"[대잠 헬기] {ac.name} → {et.preset_name} 수색 구역 도착 "
                        f"(거리 {dist_to_sub/1000:.0f}km) 디핑소나 전개 중 "
                        f"({ac.info.get('dip_hover_s',60):.0f}초 호버링)…")
                else:
                    self._log(
                        f"[대잠 초계] {ac.name}({craft_type}) → {et.preset_name} "
                        f"(거리 {dist_to_sub/1000:.0f}km) 소노부이 투하…")
                    self._asw_detect_check(ac, et, primary)
                break   # 한 tick당 한 표적

    def _asw_detect_range(self, ac: 'FriendlyAircraftObj',
                          primary: 'FriendlyShipObj',
                          et: 'EnemyThreatObj') -> float:
        """항공 대잠 탐지 거리 — 함정 소나 기준 + 수온층 + 이탈 패널티."""
        detect_m = self._detect_range_m(primary, '대잠')
        detect_m *= self._thermocline_factor(et)
        if et.is_retreating:
            detect_m *= 0.30   # 고속 이탈 시 소나 접촉 급감
        return detect_m

    def _asw_weather_factor(self, weather: str) -> float:
        """날씨별 탐지 확률 보정 계수."""
        return {'맑음 (주간)': 1.00, '맑음 (야간)': 0.95,
                '박무 / 흐림': 0.90, '강우 / 강설': 0.80,
                '폭풍':        0.65, '태풍':        0.50}.get(weather, 1.00)

    def _asw_detect_check(self, ac: 'FriendlyAircraftObj',
                          et: 'EnemyThreatObj',
                          primary: 'FriendlyShipObj'):
        """탐지 확률 판정 → 성공 시 어뢰 투하, 실패 시 재탐색 또는 포기."""
        weather   = self.cfg.get('weather', '맑음 (주간)')
        thermo    = self._thermocline_factor(et)
        wx_factor = self._asw_weather_factor(weather)
        base_prob = ac.info.get('detect_base_prob', 0.65)
        prob      = min(base_prob * thermo * wx_factor, 0.97)

        if random.random() < prob:
            # ── 탐지 성공 → 어뢰 투하 ────────────────────────────────────────
            wpn_name = ac.info['payload_wpn']
            wpn_info = FRIENDLY_DB[wpn_name]
            pk       = max(0.0, min(wpn_info['pk_dist']['mean'] + ac.info.get('pk_bonus', 0.0), 0.98))

            ac.payload_remaining -= 1
            ac.sorties           += 1
            ac.total_cost        += ac.info['cost_usd']

            dist_to_sub = primary.pos.dist_to(et.pos)
            total_dist  = dist_to_sub + (
                ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)

            drop_pos = LatLon.from_xy(
                et.pos.x + random.uniform(-300, 300),
                et.pos.y + random.uniform(-300, 300),
            )
            m = MissileObj(
                mtype    = 'friendly_strike',
                name     = f"{wpn_name}({ac.name})",
                pos      = drop_pos,
                target   = et,
                speed_ms = wpn_info['speed_ms'],
                pk_base  = pk,
                owner_id = id(ac),
                t_spawn  = self.t,
            )
            m.is_torpedo = True
            self.missiles.append(m)

            fly_s          = total_dist / max(ac.info['speed_ms'], 1)
            ac.t_available = self.t + ac.info['sortie_time_s'] + fly_s
            ac._asw_phase    = 'idle'
            ac._search_target = None
            ac._detect_fails  = 0

            mode_tag = '디핑소나' if ac.info.get('asw_mode') == 'dipping' else '소노부이'
            self._log(
                f"[대잠 탐지 성공] {ac.name} → {et.preset_name} "
                f"({mode_tag} 확률 {prob:.0%}, 수온층 ×{thermo:.2f}) | "
                f"{wpn_name} Pk={pk:.2f} 투하 | 잔여 {ac.payload_remaining}발"
            )
        else:
            # ── 탐지 실패 ─────────────────────────────────────────────────────
            ac._detect_fails += 1
            max_att = ac.info.get('max_attempts', 3)
            retry_s = ac.info.get('retry_s', 90)
            mode_tag = '디핑소나' if ac.info.get('asw_mode') == 'dipping' else '소노부이'

            if ac._detect_fails >= max_att:
                ac._asw_phase     = 'idle'
                ac._search_target = None
                ac._detect_fails  = 0
                self._log(
                    f"[대잠 탐지 실패] {ac.name} → {et.preset_name} "
                    f"{max_att}회 시도 모두 실패 (확률 {prob:.0%}) — 표적 포기"
                )
            else:
                ac._asw_phase     = 'cooldown'
                ac._next_attempt  = self.t + retry_s
                self._log(
                    f"[대잠 탐지 실패] {ac.name} → {et.preset_name} "
                    f"{mode_tag} 탐지 실패 (확률 {prob:.0%}, 수온층 ×{thermo:.2f}) "
                    f"— {retry_s}초 후 재시도 ({ac._detect_fails}/{max_att})"
                )

    def _aircraft_cap(self):
        """
        v10.5: 한국 공군 CAP — 적 항공기 BVR 요격.
        idle → (적 항공기 패트롤 반경 진입) → AAM 교전(즉시 Pk 판정) → 60s cooldown → idle
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')
        for ac in self.aircraft:
            if ac.info.get('aircraft_role', 'asw') != 'cap':
                continue
            if self.t < ac.t_available:
                continue
            if ac.payload_remaining <= 0:
                continue
            if not ac.info.get('weather_limits', {}).get(weather, True):
                continue
            if self.t < ac._next_attempt:
                continue

            patrol_m   = ac.info.get('cap_patrol_radius_km', 500) * 1000
            aam_range_m= ac.info.get('cap_aam_range_km', 100) * 1000
            aam_pk     = ac.info.get('cap_aam_pk', 0.55)
            base_dist_m= ac.info.get('base_dist_km', 300) * 1000
            wpn_name   = ac.info.get('payload_wpn', 'AAM')

            for et in self.enemy_threats:
                if not et.alive or not et.is_aircraft or et.is_retreating:
                    continue
                dist_threat = primary.pos.dist_to(et.pos)
                # 패트롤 반경 내 + 기지→함대→표적 총 사거리 체크
                if dist_threat > patrol_m:
                    continue
                if dist_threat + base_dist_m > ac.info['range_km'] * 1000:
                    continue
                # AAM 유효 사거리 내 진입 시 교전
                if dist_threat > aam_range_m:
                    continue

                ac.payload_remaining -= 1
                ac.total_cost += ac.info['cost_usd'] / max(ac.info['payload_cnt'], 1)

                if random.random() < aam_pk:
                    et.alive = False
                    et.intercepted = True
                    self.stats['intercepted_threats'] += 1
                    self._log(
                        f"[CAP] {ac.name} → {et.preset_name} "
                        f"{wpn_name} 격추 (거리 {dist_threat/1000:.0f}km, {self.t:.0f}s)"
                    )
                else:
                    self._log(
                        f"[CAP] {ac.name} → {et.preset_name} "
                        f"{wpn_name} 교전 실패 (거리 {dist_threat/1000:.0f}km, Pk {aam_pk:.0%})"
                    )

                # 교전 후 선회·재장전 cooldown
                ac._next_attempt = self.t + 60.0
                break  # 한 tick당 한 표적

    def _aircraft_aas(self):
        """
        v10.6: 항공 대함 공격 (Air-to-Surface Anti-Ship) — KF-21 해성-II 등.
        cap_strike_wpn 속성이 있는 CAP 항공기가 수상 표적(항모 우선) 공격.
        idle → (수상함 탐지) → 해성-II 발사(즉시 Pk) → 90s cooldown → idle
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')
        for ac in self.aircraft:
            if not ac.info.get('cap_strike_wpn'):
                continue
            if self.t < ac.t_available:
                continue
            if ac.strike_payload_remaining <= 0:
                continue
            if not ac.info.get('weather_limits', {}).get(weather, True):
                continue
            if self.t < ac._strike_cooldown_until:
                continue

            strike_range_m = ac.info.get('cap_strike_range_km', 200) * 1000
            strike_pk      = ac.info.get('cap_strike_pk', 0.55)
            base_dist_m    = ac.info.get('base_dist_km', 300) * 1000
            wpn_name       = ac.info['cap_strike_wpn']
            wpn_cost       = ac.info.get('cap_strike_cost_usd', 3_000_000)

            # v10.6: 항모 우선 → 그 외 수상함
            candidates = sorted(
                [et for et in self.enemy_threats if et.alive and et.is_ship],
                key=lambda e: (0 if e.high_value_target else 1),
            )
            for et in candidates:
                dist_threat = primary.pos.dist_to(et.pos)
                if dist_threat + base_dist_m > ac.info['range_km'] * 1000:
                    continue
                if dist_threat > strike_range_m:
                    continue
                # 이미 해당 표적으로 비행 중인 아군 미사일이 4발 이상이면 건너뜀
                en_route = sum(
                    1 for m in self.missiles
                    if m.alive and m.target is et and m.mtype == 'friendly_strike'
                )
                if en_route >= (6 if et.high_value_target else 4):
                    continue

                ac.strike_payload_remaining -= 1
                ac.total_cost += wpn_cost
                _m = MissileObj(
                    mtype    = 'friendly_strike',
                    name     = f"{wpn_name}({ac.name})",
                    pos      = primary.pos,
                    target   = et,
                    speed_ms = 280,          # 해성-II 순항속도 ~280m/s
                    pk_base  = strike_pk,
                    owner_id = id(ac),
                    t_spawn  = self.t,
                )
                self.missiles.append(_m)
                self._log(
                    f"[AAS] {ac.name} → {et.preset_name} "
                    f"{wpn_name} 발사 (거리 {dist_threat/1000:.0f}km, "
                    f"잔여 {ac.strike_payload_remaining}발)"
                )
                ac._strike_cooldown_until = self.t + 90.0
                break  # 한 tick당 한 발사

    def _select_strike_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        # 우선순위: Tomahawk(초장거리) → 해성-II → 해성-I → 하푼 → SM-6 대함(OTH) → Mk.45(근거리)
        # Tomahawk Block V: US 함정 전용 초장거리 대함 타격
        if (ship.inventory.get('Tomahawk Block V', 0) > 0
                and dist_m <= FRIENDLY_STRIKE_DB['Tomahawk Block V']['range_km'] * 1000):
            return 'Tomahawk Block V'
        for wpn in ['해성-II', '해성-I', '하푼 Block II']:
            if ship.strike_inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_STRIKE_DB[wpn]['range_km'] * 1000:
                return wpn
        # SM-6 대함 모드: 해성/하푼 소진 후 OTH 사거리 내 수상함 공격
        if (ship.inventory.get('SM-6', 0) > 0  # SM-6 대함 모드 항상 활성 (cfg 키 제거)
                and dist_m <= FRIENDLY_STRIKE_DB['SM-6 대함 모드']['range_km'] * 1000):
            return 'SM-6 대함 모드'
        # Mk.45 함포: 근거리 최후 수단
        if dist_m <= FRIENDLY_STRIKE_DB['Mk.45 5인치 함포']['range_km'] * 1000:
            return 'Mk.45 5인치 함포'
        return None

    def _select_asw_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        for wpn in ['홍상어 (대잠)', '청상어 (경어뢰)', 'Mk.46 경어뢰']:
            if ship.inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_DB[wpn]['range_km'] * 1000:
                return wpn
        return None

    def _select_sub_strike_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        """아군 잠수함 → 적 수상함 공격 무기 선택 (현무-3C / 하푼 / 청상어)"""
        for wpn in ['현무-3C', '하푼 Block II']:
            if ship.strike_inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_STRIKE_DB[wpn]['range_km'] * 1000:
                return wpn
        for wpn in ['청상어 (경어뢰)', 'Mk.46 경어뢰']:
            if ship.inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_DB[wpn]['range_km'] * 1000:
                return wpn
        return None

    # ── 5단계: 적 SAM 방어 (수상함 전용) ─────────────────────────────────────

    def _enemy_defense(self):
        for et in self.enemy_threats:
            if not et.alive or not et.is_ship or not et.sam_inventory:
                continue

            for m in self.missiles:
                if not m.alive or m.mtype != 'friendly_strike':
                    continue
                if m.target is not et:
                    continue

                already = any(
                    s.alive and s.target is m and s.mtype == 'enemy_sam'
                    for s in self.missiles
                )
                if already:
                    continue
                if et.sam_channels_used >= et.sam_max_channels:
                    continue

                dist_m   = et.pos.dist_to(m.pos)
                sam_name = et.select_sam(dist_m)
                if not sam_name:
                    continue

                sam_info = ENEMY_SAM_DB[sam_name]
                et.sam_inventory[sam_name]  -= 1
                et.sam_channels_used        += 1

                self.missiles.append(MissileObj(
                    mtype    = 'enemy_sam',
                    name     = sam_name,
                    pos      = et.pos,
                    target   = m,
                    speed_ms = sam_info['speed_ms'],
                    pk_base  = sam_info['pk'],
                    owner_id = id(et),
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[적 방어] {et.preset_name} -> {sam_name} 발사 "
                    f"(거리 {dist_m/1000:.1f}km)"
                )

    def _enemy_anti_sam(self):
        """적 함정의 Anti-SAM 방어 — 아군 SAM이 접근 시 CIWS·SAM으로 요격.
        enable_anti_sam=False(기본)이면 즉시 반환 (기존 결과 완전 호환).
        """
        if not self.cfg.get('enable_anti_sam', False):
            return

        for et in self.enemy_threats:
            if not et.alive or not et.is_ship or not et.sam_inventory:
                continue

            for m in list(self.missiles):
                if not m.alive or m.mtype != 'friendly_sam':
                    continue
                if m.target is not et:   # 이 함정을 향하는 SAM만
                    continue

                dist_m = et.pos.dist_to(m.pos)

                # ── CIWS 즉시 판정 (2km 이내) ────────────────────────────────
                ciws = next((n for n in et.sam_inventory if 'CIWS' in n), None)
                if dist_m <= 2000 and ciws:
                    if random.random() < 0.30:
                        m.alive       = False
                        m.intercepted = True
                        # 발사 함정 채널 해제
                        for ship in self.friendly_ships:
                            if id(ship) == m.owner_id:
                                ship.channels_used = max(0, ship.channels_used - 1)
                        self._log(
                            f"[적 Anti-SAM CIWS] {et.preset_name} → {m.name} "
                            f"근접 요격 (Pk 0.30)"
                        )
                    continue   # CIWS 범위는 SAM 발사 없음

                # ── SAM 요격: 탐지거리·채널 확인 후 발사 ─────────────────────
                already = any(s.alive and s.target is m and s.mtype == 'enemy_sam'
                              for s in self.missiles)
                if already:
                    continue
                if et.sam_channels_used >= et.sam_max_channels:
                    continue

                rcs    = getattr(m, 'rcs_m2', 0.001)
                det_km = min(500.0 * ((rcs / 3.0) ** 0.25), 50.0)  # 최대 50km 캡
                if dist_m > det_km * 1000:
                    continue

                sam_name = et.select_sam(dist_m)
                if not sam_name:
                    continue

                sam_info    = ENEMY_SAM_DB[sam_name]
                anti_sam_pk = sam_info['pk'] * 0.35   # SAM vs SAM — 소형 고속 표적

                et.sam_inventory[sam_name] -= 1
                et.sam_channels_used       += 1
                self.missiles.append(MissileObj(
                    mtype    = 'enemy_sam',
                    name     = sam_name,
                    pos      = et.pos,
                    target   = m,
                    speed_ms = sam_info['speed_ms'],
                    pk_base  = anti_sam_pk,
                    owner_id = id(et),
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[적 Anti-SAM] {et.preset_name} → {sam_name} 발사 → {m.name} "
                    f"(거리 {dist_m/1000:.1f}km, Pk {anti_sam_pk:.2f})"
                )

    # ── 6단계: 교전 결과 판정 ─────────────────────────────────────────────────

    def _check_intercepts(self):
        for sam in list(self.missiles):
            if not sam.alive:
                continue
            if sam.mtype not in ('friendly_sam', 'enemy_sam'):
                continue

            tgt = sam.target
            if not tgt.alive:
                # 타겟이 이미 격추됐어도 채널 해제 (누수 방지)
                sam.alive = False
                if sam.mtype == 'friendly_sam':
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)
                continue

            in_range = sam.hit or (sam.pos.dist_to(tgt.pos) <= INTERCEPT_DIST_M)
            if not in_range:
                continue

            sam.alive = False
            tgt_name  = tgt.name if hasattr(tgt, 'name') else str(tgt)

            # 종말 회피 Pk 보정 (아군 SAM vs 적 미사일)
            # BUG-6: ECM은 적 미사일 타격 Pk를 낮추는 것 → _check_hits로 이동
            eff_pk = sam.pk_base
            if sam.mtype == 'friendly_sam' and isinstance(tgt, MissileObj):
                remaining_m = sam.pos.dist_to(tgt.pos)
                if self.cfg.get('enable_evasion', True) and remaining_m < 10_000:  # BUG 수정: 20km→10km
                    eff_pk *= tgt.terminal_evasion_factor

            # pk_scale: LHS/Sobol 분석용 불확실 파라미터 반영
            if sam.mtype == 'friendly_sam':
                eff_pk = min(1.0, eff_pk * self.cfg.get('pk_scale', 1.0))
                # v10.4: CEC 중계 교전 — 자체 탐지 불가, 아군 데이터링크 의존 → Pk 10% 저하
                if getattr(sam, 'cec_relay', False):
                    eff_pk *= 0.90
            if random.random() < eff_pk:
                tgt.alive       = False
                tgt.intercepted = True
                tgt.t_intercept = self.t

                if sam.mtype == 'friendly_sam':
                    self._log(f"[요격 성공] {sam.name} -> {tgt_name} 격추 ({self.t:.0f}s)")
                    # MissileObj만 intercepted_threats에 집계 (BUG 수정: 항공기 플랫폼 격추는 enemy_ships_destroyed로)
                    if isinstance(tgt, MissileObj):
                        self.stats['intercepted_threats'] += 1
                        # A-1: EngagementAnalysis 추적
                        tgt.intercept_weapon = sam.name
                        tgt.intercept_km     = sam.pos.dist_to(tgt.pos) / 1000
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    self._log(f"[적 요격 성공] {sam.name} -> {tgt_name} 격추 ({self.t:.0f}s)")
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)
            else:
                if sam.mtype == 'friendly_sam':
                    self._log(f"[요격 실패] {sam.name} -> {tgt_name} 통과")
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    self._log(f"[적 요격 실패] {sam.name} -> {tgt_name} 통과")
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)

    def _check_hits(self):
        for m in self.missiles:
            # hit=True: 목표 위치 도달. alive=False 또는 intercepted=True면 이미 처리된 미사일.
            if not m.hit or not m.alive or m.intercepted:
                continue

            m.alive = False  # 도달 미사일 소모 (결과 무관)

            if m.mtype == 'enemy_strike':
                tgt = m.target
                if isinstance(tgt, FriendlyShipObj) and tgt.alive:
                    # ARM: ECM 무효 (레이더 전파 역추적 — 재밍이 오히려 표적이 됨)
                    if m.is_arm:
                        # 레이더 OFF 시 유도 신호 소실 → ARM 빗나감
                        if self.t < tgt.radar_off_until:
                            self._log(
                                f"[ARM 회피] {tgt.name} 레이더 OFF — "
                                f"{m.name} 유도 실패 빗나감"
                            )
                            continue
                        if random.random() < m.pk_base:
                            tgt.take_arm_hit(self.t)
                            self.stats['friendly_hits'] += 1
                            self._log(
                                f"[ARM 피격] {tgt.name} 레이더 직격! "
                                f"(레이더 {tgt.radar_factor:.0%}, HP {tgt.hp})")
                        else:
                            self._log(f"[ARM 실패] {m.name} -> {tgt.name} 불발")
                        continue

                    # v9.13: 고층 바람 CEP — 순항·대함미사일 탄착 오차 증가 (탄도/HGV/어뢰 제외)
                    if not m.is_ballistic and not m.is_hgv and not m.is_qbm and not m.is_torpedo:
                        cep_f = self._wind_cep_factor()
                        if cep_f > 1.0 and random.random() > (1.0 / cep_f):
                            self._log(
                                f"[바람 CEP] {m.name} 강풍 탄착 오차 빗나감 "
                                f"(×{cep_f:.2f}, 계절:{self.cfg.get('season','?')})"
                            )
                            continue

                    # BUG-6: 아군 ECM(AN/SLQ-32) — 적 미사일 유도부 교란, Pk 30% 감소
                    # 탄도/HGV는 레이더 유도가 아니므로 ECM 무효
                    if self.cfg.get('enable_ecm', True) and not m.is_ballistic and not m.is_hgv:
                        ecm_red = 0.30 * self.cfg.get('ecm_scale', 1.0)
                        m.pk_base = max(0.0, m.pk_base * (1.0 - ecm_red))
                    # 포팅 B: 음향 기만기 AN/SLQ-25 — 어뢰 전용
                    if m.is_torpedo and self.cfg.get('enable_decoy', True):
                        if tgt.decoy_stock > 0:
                            tgt.decoy_stock -= 1
                            if random.random() < DECOY_PK:
                                self._log(
                                    f"[기만기] {tgt.name} 기만기 성공 — {m.name} 회피 "
                                    f"(잔여 {tgt.decoy_stock}발)")
                                continue
                    # 포팅 B: 함정 회피 기동 — 어뢰 전용
                    if m.is_torpedo and self.cfg.get('enable_evasion', True):
                        # 추진 피탄 시 speed_factor만큼 회피 기동 성공률 저하
                        if random.random() < SHIP_EVASION_PK * tgt.speed_factor:
                            self._log(f"[회피] {tgt.name} 회피 기동 성공 — {m.name}")
                            continue
                    # 서브시스템 피해 롤 (enable_subsystem_damage=True 시)
                    subsystem = None
                    if self.cfg.get('enable_subsystem_damage', True):
                        r = random.random()
                        if getattr(m, 'is_torpedo', False):
                            # 어뢰: 수중 폭발 → 추진·선체 위주, 레이더 거의 없음
                            if   r < 0.05: subsystem = 'radar'
                            elif r < 0.45: subsystem = 'propulsion'
                            elif r < 0.55: subsystem = 'weapons'
                            # else: None (선체 직격, 45%)
                        else:
                            # 대함·탄도·순항미사일
                            if   r < 0.15: subsystem = 'radar'
                            elif r < 0.35: subsystem = 'propulsion'
                            elif r < 0.60: subsystem = 'weapons'
                            # else: None (선체 직격, 40%)
                    if random.random() < m.pk_base:
                        tgt.take_hit(m.name, self.t, subsystem)
                        self.stats['friendly_hits'] += 1
                        _dmg = {
                            'radar':      f'레이더 피탄 (탐지 {tgt.radar_factor:.0%})',
                            'propulsion': f'추진 피탄 (속도 {tgt.speed_factor:.0%})',
                            'weapons':    f'VLS 피탄 (채널 {tgt.channel_factor:.0%}, 탄약 손실)',
                            None:         '선체 직격 (HP 감소)',
                        }
                        _detail = f' — {_dmg[subsystem]}'
                        self._log(f"[피격] {tgt.name} <- {m.name} 명중! HP {tgt.hp}{_detail}")
                    else:
                        self._log(f"[피격 실패] {m.name} -> {tgt.name} 근접 불발")

            elif m.mtype == 'friendly_strike':
                tgt = m.target
                if isinstance(tgt, EnemyThreatObj) and tgt.alive:
                    # 포팅 B: 적 자체방어 — CIWS 요격 → 채프/플레어/DRFM
                    eff_pk = m.pk_base
                    if self.cfg.get('enable_selfdefense', True):
                        ciws_pk = tgt.info.get('enemy_ciws_pk', 0.0)
                        if ciws_pk > 0 and random.random() < ciws_pk:
                            self._log(f"[적 CIWS] {tgt.preset_name} CIWS 요격 — {m.name}")
                            continue
                        # v8.26: seeker 유형별 기만체계 차등 적용
                        seeker = getattr(m, 'seeker', 'radar')
                        if seeker == 'radar':
                            sdpk = tgt.info.get('chaff_pk',
                                   tgt.info.get('self_defense_pk', 0.0))
                        elif seeker == 'ir':
                            sdpk = tgt.info.get('flare_pk',
                                   tgt.info.get('self_defense_pk', 0.0))
                        elif seeker == 'combined':
                            sdpk = tgt.info.get('drfm_pk',
                                   tgt.info.get('self_defense_pk', 0.0) * 0.50)
                        else:  # acoustic 등 — 채프/플레어 무효
                            sdpk = 0.0
                        eff_pk = m.pk_base * (1.0 - sdpk)
                    if random.random() < eff_pk:
                        tgt.take_hit(m.name, self.t)
                        self.stats['enemy_hits'] += 1
                        sunk = not tgt.alive
                        status = '격침' if sunk else f'손상 (HP {tgt.hp})'
                        self._log(f"[적 피격] {tgt.preset_name} <- {m.name} 명중! {status}")
                        # v9.3: 격침 기록
                        self.strike_log.append({
                            'target': tgt.preset_name,
                            'weapon': m.name,
                            't': round(self.t, 1),
                            'sunk': sunk,
                            'hp_remaining': tgt.hp,
                        })
                    else:
                        self._log(f"[적 피격 실패] {m.name} -> {tgt.preset_name} 회피")

    # ── 7단계: 프레임 기록 ────────────────────────────────────────────────────

    def _record_frame(self):
        frame = SimFrame(self.t)
        for s in self.friendly_ships:
            # [5]=radar_factor [6]=speed_factor [7]=disabled_weapons 수 (애니메이션 피해 표시용)
            frame.friendly_ships.append((
                s.name, s.pos.x, s.pos.y, s.alive, s.hp,
                s.radar_factor, s.speed_factor, len(s.disabled_weapons)
            ))
            frame.ship_channels.append((s.name, s.channels_used, s.max_channels))
        for et in self.enemy_threats:
            frame.enemy_ships.append(
                (et.uid, et.preset_name, et.pos.x, et.pos.y, et.alive, et.hp,
                 et.altitude_m))
        for m in self.missiles:
            if m.alive:
                frame.missiles.append(
                    (m.uid, m.pos.x, m.pos.y, m.mtype, m.name,
                     self._missile_disp_alt(m)))
        frame.events = list(self._tick_events)
        self.frames.append(frame)
        self._tick_events.clear()

    def _missile_disp_alt(self, m) -> float:
        """3D 시각화용 미사일 고도(m). 탄도/HGV는 포물선 궤도 추정."""
        if (m.is_ballistic or m.is_hgv) and getattr(m, '_init_dist', 0) > 0:
            target = m.target
            if target and hasattr(target, 'pos'):
                rem = m.pos.dist_to(target.pos)
                progress = max(0.0, min(1.0, 1.0 - rem / m._init_dist))
                peak = getattr(m, '_peak_alt_m', 50_000.0)
                return max(0.0, 4.0 * peak * progress * (1.0 - progress))
        return m.altitude_m

    # ── 종료 조건 ─────────────────────────────────────────────────────────────

    def _is_over(self) -> bool:
        active_threats = [m for m in self.missiles if m.alive and m.mtype == 'enemy_strike']
        # 이탈 중인 항공기: 재공격 가능하면 활성 위협으로 유지, 아니면 종료로 간주
        enemy_active   = [et for et in self.enemy_threats
                          if et.alive and not (et.is_aircraft and et.is_retreating
                                               and et.reattack_count >= et.max_reattacks)]

        if not active_threats and not enemy_active:
            self._log("[종료] 교전 종료 - 모든 위협 소진/격침/이탈")
            return True
        if all(not s.alive for s in self.friendly_ships):
            self._log("[종료] 아군 전멸")
            return True
        return False

    # ── v10.7: 전술 의사결정 헬퍼 ───────────────────────────────────────────────

    def _make_tactical_state(self) -> 'TacticalState':
        """현재 시뮬 상태 스냅샷 → TacticalState."""
        primary = self._primary()
        threats = [
            {
                'name':     et.preset_name,
                'type':     et.info.get('type', '?'),
                'hp':       et.hp,
                'dist_km':  round(primary.pos.dist_to(et.pos) / 1000, 1),
                'speed_ms': et.speed_ms,
            }
            for et in self.enemy_threats if et.alive
        ]
        ships = [
            {
                'name':    s.name,
                'alive':   s.alive,
                'hp':      s.hp,
                'max_hp':  s._max_hp,
                'radar':   round(s.radar_factor, 2),
                'speed':   round(s.speed_factor, 2),
            }
            for s in self.friendly_ships
        ]
        return TacticalState(
            t             = self.t,
            threats       = threats,
            ships         = ships,
            intercepted   = self.stats['intercepted_threats'],
            total_threats = self.stats['total_threats'],
            shots_fired   = self.stats['total_missiles_fired'],
        )

    def _apply_tactical_choice(self, choice: dict):
        """사용자 전술 선택 반영 — cfg 일시 재설정."""
        if not choice:
            return
        # 무기 우선순위: 다음 교전 구간 _select_defense_wpn 재정의 대신 cfg 플래그 사용
        wpn_priority = choice.get('weapon_priority', 'auto')
        if wpn_priority != 'auto':
            self.cfg['_tactical_wpn_priority'] = wpn_priority
        else:
            self.cfg.pop('_tactical_wpn_priority', None)
        # 살보 수 조정
        max_salvo = choice.get('max_salvo', None)
        if max_salvo is not None:
            self.cfg['_tactical_max_salvo'] = int(max_salvo)
        else:
            self.cfg.pop('_tactical_max_salvo', None)

    # ── 메인 루프 ─────────────────────────────────────────────────────────────

    def run(self) -> dict:
        pt = {'위치갱신': 0.0, '적발사': 0.0, '대공방어': 0.0,
              '아군공격': 0.0, '대잠': 0.0, '적방어': 0.0,
              '적Anti-SAM': 0.0, '교전판정': 0.0}
        _pc = _time.perf_counter
        _step_interval = 20   # step_cb 호출 간격(초) — 단일 시뮬 UI 갱신용

        while self.t <= MAX_SIM_TIME:
            # NEW-A: 혼합 시나리오 파도 지연 스폰
            if self._pending_threats:
                due = [s for (spawn_t, s) in self._pending_threats if spawn_t <= self.t]
                self._pending_threats = [(spawn_t, s) for (spawn_t, s) in self._pending_threats
                                         if spawn_t > self.t]
                for spec in due:
                    self._spawn_pending_threat(spec)

            # v8.15: 항모 함재기 파도 스폰
            for et in self.enemy_threats:
                if (et.alive and et.carrier_aircraft
                        and et.carrier_wave_interval > 0
                        and self.t > 0
                        and self.t - et._last_wave_t >= et.carrier_wave_interval):
                    et._last_wave_t = self.t
                    self._pending_threats.append(
                        (self.t + 10.0,          # 10초 후 출격
                         {'preset': et.carrier_aircraft, 'count': 2})
                    )

            # v9.6: 기습 잠수함 은닉 해제 탐지 이벤트
            for et in self.enemy_threats:
                if (et.is_sub and not et.ambush_revealed
                        and et.hidden_until > 0 and self.t >= et.hidden_until):
                    et.ambush_revealed = True
                    dist_km = self._primary().pos.dist_to(et.pos) / 1000
                    self._log(
                        f"[⚠ 기습 탐지!] {et.preset_name} 잠수함 은닉 해제 "
                        f"(거리 {dist_km:.0f}km) — 즉시 교전 개시!"
                    )

            # v8.26: 생존 적 항공기 중 최대 ECM 강도 캐싱 (에어리어 재밍)
            self._active_ecm = max(
                (et.ecm_power for et in self.enemy_threats
                 if et.alive and et.is_aircraft),
                default=0.0
            )
            _t0 = _pc(); self._update_positions(); pt['위치갱신'] += _pc() - _t0
            _t0 = _pc(); self._enemy_fire();       pt['적발사']   += _pc() - _t0
            self._arm_radar_off_check()
            _t0 = _pc(); self._friendly_defense(); pt['대공방어'] += _pc() - _t0
            _t0 = _pc(); self._friendly_strike();  pt['아군공격'] += _pc() - _t0
            _t0 = _pc(); self._aircraft_asw(); self._aircraft_cap(); self._aircraft_aas(); pt['대잠'] += _pc() - _t0
            _t0 = _pc(); self._enemy_defense();    pt['적방어']   += _pc() - _t0
            _t0 = _pc(); self._enemy_anti_sam();   pt['적Anti-SAM'] += _pc() - _t0
            _t0 = _pc()
            self._check_intercepts()
            self._check_hits()
            pt['교전판정'] += _pc() - _t0

            # 소멸·요격 미사일의 추적 채널 반환
            for m in self.missiles:
                if not m.alive or m.intercepted:
                    self.track_radar.release(m.uid)
            self.missiles = [m for m in self.missiles
                             if m.alive and not m.intercepted]

            self._record_frame()

            # 포팅 D: 동시 위협 수 peak 추적
            alive_count = sum(
                1 for et in self.enemy_threats
                if et.alive and not (et.is_aircraft and et.is_retreating)
            )
            if alive_count > self.stats['peak_concurrent_threats']:
                self.stats['peak_concurrent_threats'] = alive_count

            # v10.7: 전술 의사결정 모드 — 구간마다 일시정지 후 사용자 선택 반영
            if (self._tactical_pause_cb and self.t > 0
                    and int(self.t) % self._tactical_interval == 0):
                state = self._make_tactical_state()
                choice = self._tactical_pause_cb(state)   # blocks in worker thread
                self._apply_tactical_choice(choice)

            # 단일 시뮬 진행 콜백 (MC 배치 워커에서는 None)
            if self._step_cb and int(self.t) % _step_interval == 0:
                vls_rem = sum(
                    getattr(s, 'vls_remaining', 0) for s in self.friendly_ships)
                last_log = self._log_entries[-1] if self._log_entries else ""
                self._step_cb(self.t, MAX_SIM_TIME, alive_count, vls_rem, last_log)

            if self._is_over():
                break

            self.t += DT

        self._phase_times = pt
        return self._compile()

    def _compile(self) -> dict:
        self.stats['friendly_ships_lost']   = sum(1 for s in self.friendly_ships if not s.alive)
        # 이탈 항공기(alive=False, is_retreating=True, intercepted=False)는 "격침" 아님
        self.stats['enemy_ships_destroyed'] = sum(
            1 for et in self.enemy_threats
            if not et.alive and (et.intercepted or not et.is_aircraft)
        )
        # 포팅 C: 항공 자산 출격 횟수 + 비용 합산
        self.stats['aircraft_sorties'] = sum(ac.sorties for ac in self.aircraft)
        self.stats['total_cost']       = (
            sum(s.total_cost for s in self.friendly_ships)
            + sum(ac.total_cost for ac in self.aircraft)
            + self._ground_cost  # v9.4: 지상 발사 자산 비용
        )

        intercept_rate = (
            self.stats['intercepted_threats'] / self.stats['total_threats']
            if self.stats['total_threats'] > 0 else 1.0
        )

        # 포팅 D: 잔여 재고 합산 (REQ-07), 총 채널 수 (REQ-08)
        remaining_inv: dict = {}
        for s in self.friendly_ships:
            for wpn, cnt in s.inventory.items():
                remaining_inv[wpn] = remaining_inv.get(wpn, 0) + cnt
        total_channels = sum(s.max_channels for s in self.friendly_ships)

        ship_subsystem_damage = {
            s.name: {
                'radar_factor':    round(s.radar_factor, 3),
                'speed_factor':    round(s.speed_factor, 3),
                'channel_factor':  round(s.channel_factor, 3),
                'disabled_weapons': sorted(s.disabled_weapons),
                'alive':           s.alive,
                'hp':              s.hp,
                'max_hp':          s._max_hp,
                'hits_taken':      s.hits_taken,
            }
            for s in self.friendly_ships
        }

        # v10.6: 항모 격침/전투불능 판정
        carrier_status = {
            et.preset_name: {
                'alive': et.alive,
                'hp': et.hp,
                'max_hp': et._HP_MAP.get('항모', 5),
                'status': '격침' if not et.alive else ('전투불능' if et.hp <= 2 else '정상'),
            }
            for et in self.enemy_threats
            if et.info.get('type') == '항모'
        }

        return {
            **self.stats,
            'intercept_rate':    intercept_rate,
            'sim_time':          self.t,
            'frames':            self.frames,
            'log':               self._log_entries,
            'friendly_ships':    self.friendly_ships,
            'enemy_ships':       self.enemy_threats,   # 하위 호환 키 유지
            'remaining_inventory': remaining_inv,
            'total_channels':      total_channels,
            'used_seed':           self.cfg.get('sim_seed', None),
            'ship_subsystem_damage': ship_subsystem_damage,
            'active_events':     self._build_active_events(),  # A-1
            'strike_log':        self.strike_log,              # v9.3
            'vls_depletion_t':   dict(self.vls_depletion_t),  # v9.4
            'ground_remaining':   dict(self.ground_inv),         # v9.4 / v9.11
            'ashore_sm3_fired':  self.stats['ashore_sm3_fired'],
            'thaad_fired':       self.stats['thaad_fired'],
            'phase_times':       dict(self._phase_times),      # v8.26: 단계별 소요시간
            'carrier_status':    carrier_status,               # v10.6
        }

    def _build_active_events(self) -> list:
        """A-1: enemy_strike MissileObj → EngagementAnalysisTab 어댑터 리스트."""
        class _EvAdapter:
            __slots__ = ('label','is_active','intercepted','intercept_weapon',
                         'intercept_km','t_intercepted','gantt_bars','detect_m','enemy_info')
        evs = []
        for m in self.missiles:
            if m.mtype != 'enemy_strike':
                continue
            ev = _EvAdapter()
            ev.label            = m.name
            ev.is_active        = True
            ev.intercepted      = m.intercepted
            ev.intercept_weapon = m.intercept_weapon
            ev.intercept_km     = m.intercept_km if m.intercept_km else None
            ev.t_intercepted    = m.t_intercept
            ev.detect_m         = m.detect_m
            ev.enemy_info       = m.enemy_info
            # gantt_bars: 위협 비행 전 구간 단일 바
            t_end   = m.t_intercept or self.t
            color   = '#2ecc71' if m.intercepted else '#e74c3c'
            ev.gantt_bars = [(m.name, m.t_spawn, t_end, color)]
            evs.append(ev)
        return evs


# ════════════════════════════════════════════════════════════════════════════
#  탐지거리 자동 계산 (함대 편성 + 날씨 + 데이터링크)
# ════════════════════════════════════════════════════════════════════════════

def calculate_fleet_detect_ranges(fleet_preset_name: str, weather: str) -> dict:
    """
    함대 편성과 날씨를 기반으로 탐지거리를 자동 계산한다.

    데이터링크 원칙:
      - 한국 해군 Link-16/Link-11 적용 — 편대 내 최고 성능 센서 기준 공유
      - 대공·대함 : 편대 내 max(sensor_km['대공'/'대함']) × radar_factor
      - 대잠       : 편대 내 max(sensor_km['대잠']) × sonar_factor
        (황사는 소나에 영향 없음, 풍랑·폭풍은 해상 소음으로 급감)

    반환 예시:
      {'대공': 1140, '대함': 41, '대잠': 30,
       'leading_ship': 'KDX-III', 'radar_factor': 0.95, 'sonar_factor': 0.60}
    """
    preset = FLEET_PRESETS.get(fleet_preset_name, [])
    w = _make_physics_wx(weather)   # v9.13: Beaufort 물리값 적용
    rf = w.get('radar_factor', w.get('detect_range_factor', 1.0))
    sf = w.get('sonar_factor', w.get('detect_range_factor', 1.0))

    max_air = 0; max_surface = 0; max_sub = 0
    leading = '(없음)'
    for ship in preset:
        spec = SHIP_DB.get(ship['type'], {})
        s = spec.get('sensor_km', {})
        air = s.get('대공', 0)
        if air > max_air:
            max_air = air
            leading = ship.get('name', ship['type'])
        max_surface = max(max_surface, s.get('대함', 0))
        max_sub     = max(max_sub,     s.get('대잠', 0))

    return {
        '대공':         max(1, round(max_air     * rf)),
        '대함':         max(1, round(max_surface * rf)),
        '대잠':         max(1, round(max_sub     * sf)),
        'leading_ship': leading,
        'radar_factor': rf,
        'sonar_factor': sf,
    }


# ════════════════════════════════════════════════════════════════════════════
#  외부 API
# ════════════════════════════════════════════════════════════════════════════

def run_v7_simulation(cfg: dict, step_cb=None, tactical_cb=None) -> dict:
    # 탐지거리 자동 계산 (함대 + 날씨 기반, 수동 override 없을 때)
    if not cfg.get('detect_km_manual', False):
        ranges = calculate_fleet_detect_ranges(
            cfg.get('fleet_preset', '단독 작전'),
            cfg.get('weather', '맑음 (주간)'))
        cfg = dict(cfg)
        cfg['detect_km']         = ranges['대공']
        cfg['surface_detect_km'] = ranges['대함']
        cfg['sub_detect_km']     = ranges['대잠']
    sim = TimeStepEngine(cfg, step_cb=step_cb)
    if tactical_cb:
        sim._tactical_pause_cb = tactical_cb  # v10.7: 전술 모드 훅 주입
    return sim.run()


# ════════════════════════════════════════════════════════════════════════════
#  몬테카를로 분석
# ════════════════════════════════════════════════════════════════════════════

def monte_carlo_v7(cfg: dict, n: int = 200, desc: str = '',
                   progress_cb=None) -> dict:
    """
    run_v7_simulation을 n회 반복해 통계를 집계한다.

    반환 dict 키:
      intercept_rates   : list[float]   — 회차별 요격률
      friendly_hits     : list[int]
      enemy_destroyed   : list[int]
      friendly_lost     : list[int]
      total_costs       : list[float]
      weapon_usage      : dict[str, list[int]]  — 무기별 회차별 소모량
      ship_hits         : dict[str, list[int]]  — 함정별 회차별 피격 횟수
      mean_intercept    : float
      std_intercept     : float
      full_pass_rate    : float          — 요격률 1.0 비율
    """
    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}   # {무기명: [회차별 소모량]}
    ship_hits_mc: dict = {}   # {함정명: [회차별 피격]}
    weapon_zero:  dict = {}   # {무기명: 소진(잔여=0) 횟수}
    # v9.4: VLS 고갈 통계
    vls_dep_count: int = 0
    vls_dep_times: list = []

    step = max(1, n // 5)
    if desc:
        print(f'  [{desc}] {n}회 MC 시작... ', end='', flush=True)

    base_seed = cfg.get('sim_seed', None)
    for i in range(n):
        # 회차마다 다른 시드 (기반 시드 + 회차번호)
        run_cfg = dict(cfg)
        if base_seed:
            run_cfg['sim_seed'] = int(base_seed) + i
        r = run_v7_simulation(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])

        # 무기별 소모량 (초기 재고 - 잔여 재고) + 소진 횟수 집계
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            if wpn not in weapon_usage:
                weapon_usage[wpn] = []
            weapon_usage[wpn].append(remaining)
            if remaining == 0:
                weapon_zero[wpn] = weapon_zero.get(wpn, 0) + 1

        # 함정별 피격 횟수
        for ship in r.get('friendly_ships', []):
            sname = ship.name
            hits = getattr(ship, 'hits_taken', 0)
            if sname not in ship_hits_mc:
                ship_hits_mc[sname] = []
            ship_hits_mc[sname].append(hits)

        # v9.4: VLS 고갈 시각 수집
        dep_t = r.get('vls_depletion_t', {})
        if dep_t:
            vls_dep_count += 1
            vls_dep_times.extend(dep_t.values())

        if desc and (i + 1) % step == 0:
            print(f'{(i + 1) * 100 // n}%', end=' ', flush=True)
        if progress_cb:
            progress_cb(i + 1, n)

    if desc:
        print('완료')

    arr = np.array(rates)
    # 무기별 평균 잔여 재고 (소모량 = 초기 - 평균 잔여)
    weapon_avg_remaining = {k: float(np.mean(v)) for k, v in weapon_usage.items()}
    ship_avg_hits = {k: float(np.mean(v)) for k, v in ship_hits_mc.items()}
    dest_arr = np.array(e_dest, dtype=float)
    return {
        'intercept_rates':         rates,
        'friendly_hits':           f_hits,
        'enemy_destroyed':         e_dest,
        'friendly_lost':           f_lost,
        'total_costs':             costs,
        'weapon_avg_remaining':    weapon_avg_remaining,
        'weapon_exhaustion_rates': {k: v / n for k, v in weapon_zero.items()},
        'ship_avg_hits':           ship_avg_hits,
        'mean_intercept':          float(arr.mean()),
        'std_intercept':           float(arr.std()),
        'full_pass_rate':          float((arr == 1.0).mean()),
        'n':                       n,
        # v9.3: 공격 임무 격침 통계
        'mean_enemy_destroyed':    float(dest_arr.mean()),
        'max_enemy_destroyed':     int(dest_arr.max()) if len(dest_arr) else 0,
        # v9.4: VLS 고갈 통계
        'vls_depletion_rate':      vls_dep_count / n,
        'vls_depletion_t_mean':    float(np.mean(vls_dep_times)) if vls_dep_times else None,
    }


def _mc_batch_worker(args: tuple) -> tuple:
    """ProcessPoolExecutor 배치 워커 — PyQt6 의존성 없는 순수 엔진 함수."""
    cfg, n, seed_offset = args
    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}
    weapon_zero:  dict = {}
    ship_hits_mc: dict = {}
    phase_times_acc: dict = {}   # v8.26: 배치 내 단계별 시간 누적
    base_seed = cfg.get('sim_seed', None)
    for i in range(n):
        run_cfg = dict(cfg)
        if base_seed:
            run_cfg['sim_seed'] = int(base_seed) + seed_offset + i
        r = run_v7_simulation(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            weapon_usage.setdefault(wpn, []).append(remaining)
            if remaining == 0:
                weapon_zero[wpn] = weapon_zero.get(wpn, 0) + 1
        for ship in r.get('friendly_ships', []):
            ship_hits_mc.setdefault(ship.name, []).append(
                getattr(ship, 'hits_taken', 0))
        for k, v in r.get('phase_times', {}).items():
            phase_times_acc[k] = phase_times_acc.get(k, 0.0) + v
    # 배치 평균으로 변환
    if n > 0:
        phase_times_avg = {k: v / n for k, v in phase_times_acc.items()}
    else:
        phase_times_avg = {}
    return rates, f_hits, e_dest, f_lost, costs, weapon_usage, ship_hits_mc, weapon_zero, phase_times_avg


# ════════════════════════════════════════════════════════════════════════════
#  분석 고도화: LHS / CVaR / Stress Test / Sobol 민감도
# ════════════════════════════════════════════════════════════════════════════

# LHS 샘플링 대상 불확실 파라미터: (cfg_key, 하한, 상한, 표시명)
_LHS_PARAM_DEFS = [
    ('pk_scale',         0.70, 1.30, 'SAM Pk 배율'),
    ('detect_scale',     0.70, 1.30, '탐지거리 배율'),
    ('cd_scale',         0.80, 1.50, 'C&D 시간 배율'),
    ('ecm_scale',        0.50, 1.50, 'ECM 효과 배율'),
    ('threat_spd_scale', 0.80, 1.30, '위협 속도 배율'),
    ('decoy_stock',      0.0,  4.0,  '기만기 재고'),
]

# 스트레스 테스트 2D 그리드 정의
STRESS_DIMS = {
    'channel_degrade': {
        'label':  '유도 채널 감소 (%)',
        'values': [0, 25, 50, 75],
    },
    'radar_degrade': {
        'label':  '레이더 성능 감소 (%)',
        'values': [0, 25, 50],
    },
}


def compute_cvar(rates: list, alpha: float = 0.05) -> float:
    """하위 alpha% 요격률의 평균 (Conditional Value at Risk — 최악 시나리오 평균)."""
    if not rates:
        return 0.0
    sorted_r = sorted(rates)
    n_tail   = max(1, int(len(sorted_r) * alpha))
    return float(np.mean(sorted_r[:n_tail]))


def monte_carlo_lhs(cfg: dict, n: int = 10_000,
                    progress_cb=None) -> dict:
    """
    Latin Hypercube Sampling 기반 MC.
    불확실 파라미터 6종을 LHS로 공간 균등 샘플링하여 순수 MC 대비 3~5× 빠른 수렴.

    반환: monte_carlo_v7 형식 dict + 'cvar', 'method' 키
    """
    try:
        from scipy.stats.qmc import LatinHypercube
        d = len(_LHS_PARAM_DEFS)
        seed_val = cfg.get('sim_seed', None)
        sampler  = LatinHypercube(d=d, seed=int(seed_val) if seed_val else None)
        samples  = sampler.random(n=n)   # (n, d) in [0,1]
    except ImportError:
        # scipy 미설치 시 균등 랜덤으로 폴백
        samples = np.random.rand(n, len(_LHS_PARAM_DEFS))

    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}
    ship_hits_mc: dict = {}

    for i, sample in enumerate(samples):
        run_cfg = dict(cfg)
        for j, (key, lo, hi, _) in enumerate(_LHS_PARAM_DEFS):
            run_cfg[key] = float(lo + sample[j] * (hi - lo))
        r = run_v7_simulation(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            weapon_usage.setdefault(wpn, []).append(remaining)
        for ship in r.get('friendly_ships', []):
            ship_hits_mc.setdefault(ship.name, []).append(
                getattr(ship, 'hits_taken', 0))
        if progress_cb:
            progress_cb(i + 1, n)

    arr = np.array(rates)
    return {
        'intercept_rates':      rates,
        'friendly_hits':        f_hits,
        'enemy_destroyed':      e_dest,
        'friendly_lost':        f_lost,
        'total_costs':          costs,
        'weapon_avg_remaining': {k: float(np.mean(v)) for k, v in weapon_usage.items()},
        'ship_avg_hits':        {k: float(np.mean(v)) for k, v in ship_hits_mc.items()},
        'mean_intercept':       float(arr.mean()),
        'std_intercept':        float(arr.std()),
        'full_pass_rate':       float((arr == 1.0).mean()),
        'cvar':                 compute_cvar(rates),
        'n':                    n,
        'method':               'LHS',
        # v9.3
        'mean_enemy_destroyed': float(np.mean(e_dest)) if e_dest else 0.0,
        'max_enemy_destroyed':  int(max(e_dest)) if e_dest else 0,
    }


def stress_test_grid(cfg: dict, n_per_cell: int = 500,
                     progress_cb=None) -> dict:
    """
    2D 스트레스 테스트: 채널 감소(%) × 레이더 성능 감소(%) 그리드 요격률 매트릭스.

    n_per_cell: 셀당 시뮬 횟수 (빠름=300, 표준=500, 정밀=3000)
    """
    ch_vals  = STRESS_DIMS['channel_degrade']['values']   # [0, 25, 50, 75]
    rad_vals = STRESS_DIMS['radar_degrade']['values']     # [0, 25, 50]
    grid       = np.zeros((len(ch_vals), len(rad_vals)))
    cvar_grid  = np.zeros_like(grid)
    total_cells = len(ch_vals) * len(rad_vals)
    done = 0

    for i, ch in enumerate(ch_vals):
        for j, rad in enumerate(rad_vals):
            cell_cfg = dict(cfg)
            # 레이더 성능 감소 → detect_scale 감소
            cell_cfg['detect_scale'] = 1.0 - rad / 100.0
            # 채널 감소 → Pk 비례 감소로 근사 (75% 채널 감소 → Pk 약 37.5% 감소)
            cell_cfg['pk_scale'] = max(0.1, 1.0 - ch / 200.0)

            cell_rates = []
            for k in range(n_per_cell):
                run_cfg = dict(cell_cfg)
                base_seed = cfg.get('sim_seed', None)
                if base_seed:
                    run_cfg['sim_seed'] = int(base_seed) + done * n_per_cell + k
                r = run_v7_simulation(run_cfg)
                cell_rates.append(r['intercept_rate'])

            grid[i, j]      = float(np.mean(cell_rates))
            cvar_grid[i, j] = compute_cvar(cell_rates)
            done += 1
            if progress_cb:
                progress_cb(done, total_cells)

    return {
        'grid':      grid.tolist(),
        'cvar_grid': cvar_grid.tolist(),
        'ch_vals':   ch_vals,
        'rad_vals':  rad_vals,
        'ch_label':  STRESS_DIMS['channel_degrade']['label'],
        'rad_label': STRESS_DIMS['radar_degrade']['label'],
        'n_per_cell': n_per_cell,
    }


def sobol_analysis(cfg: dict, n_sobol: int = 4096, n_per_point: int = 1,
                   progress_cb=None) -> dict:
    """
    Sobol 1차/전체 민감도 지수 — 정밀 모드 전용.

    n_per_point: 각 파라미터 샘플 포인트당 시뮬레이션 반복 수.
      - n_per_point=1 (기본): 총 N×(D+2) ≈ 32,768회. 빠르지만 확률 노이즈 있음.
      - n_per_point=3: 총 ≈ 98,304회. 표준편차 √3 ≈ 1.7× 감소.
      - n_per_point=5: 총 ≈ 163,840회. 표준편차 √5 ≈ 2.2× 감소.
    확률적 시뮬레이션에서 n_per_point≥3 권장.
    """
    try:
        from SALib.sample import saltelli
        from SALib.analyze import sobol as sobol_analyze
    except ImportError:
        return {'error': 'SALib 미설치 — pip install SALib'}

    param_names = [p[0] for p in _LHS_PARAM_DEFS]
    problem = {
        'num_vars': len(_LHS_PARAM_DEFS),
        'names':    param_names,
        'bounds':   [[p[1], p[2]] for p in _LHS_PARAM_DEFS],
    }

    param_values  = saltelli.sample(problem, n_sobol, calc_second_order=False)
    n_sobol_pts   = len(param_values)
    total_runs    = n_sobol_pts * n_per_point
    Y = np.zeros(n_sobol_pts)

    for i, pv in enumerate(param_values):
        run_cfg = dict(cfg)
        for j, key in enumerate(param_names):
            run_cfg[key] = float(pv[j])
        if n_per_point > 1:
            # K회 평균으로 확률 노이즈 √K배 감소
            point_rates = []
            for k in range(n_per_point):
                rc = dict(run_cfg)
                base_seed = cfg.get('sim_seed', None)
                if base_seed:
                    rc['sim_seed'] = int(base_seed) + i * n_per_point + k
                point_rates.append(run_v7_simulation(rc)['intercept_rate'])
            Y[i] = float(np.mean(point_rates))
        else:
            Y[i] = run_v7_simulation(run_cfg)['intercept_rate']
        if progress_cb:
            progress_cb(i + 1, n_sobol_pts)

    Si = sobol_analyze.analyze(
        problem, Y, calc_second_order=False, print_to_console=False)
    return {
        'S1':          Si['S1'].tolist(),
        'ST':          Si['ST'].tolist(),
        'S1_conf':     Si['S1_conf'].tolist(),
        'ST_conf':     Si['ST_conf'].tolist(),
        'names':       [p[3] for p in _LHS_PARAM_DEFS],
        'n_runs':      total_runs,
        'n_per_point': n_per_point,
    }


# ════════════════════════════════════════════════════════════════════════════
#  최적 무기 조합 추천 — 그리드 서치 + 정밀 검증
# ════════════════════════════════════════════════════════════════════════════

def optimize_weapon_loadout_v7(cfg: dict,
                                budget:     int = 64,
                                step:       int = 8,
                                max_per:    int = 32,
                                coarse_n:   int = 20,
                                fine_n:     int = 200,
                                top_k:      int = 5,
                                progress_cb = None) -> list:
    """
    VLS 예산 안에서 최적 무기 조합 탐색 (그리드 서치 + 정밀 검증).

    1단계: 모든 조합을 coarse_n 회 MC로 빠르게 평가
    2단계: 상위 top_k 조합을 fine_n 회 MC로 정밀 검증
    반환: [{'combo': dict, 'rate': float, 'std': float, 'total': int}, ...]
    """
    import itertools as _it

    WEAPONS = [
        ('SM-3 Block IIA',  'sm3_stock'),
        ('SM-6',            'sm6_stock'),
        ('SM-2 Block IIIB', 'sm2_stock'),
        ('RIM-116 RAM',     'ram_stock'),
    ]

    vals   = list(range(0, max_per + 1, step))          # [0, 8, 16, 24, 32]
    combos = [
        c for c in _it.product(vals, repeat=len(WEAPONS))
        if sum(c) <= budget and sum(c) > 0              # 0발 조합 제외
    ]

    # 1단계: 조악한 MC로 빠른 탐색
    coarse_results: list = []
    total = len(combos)
    for i, combo in enumerate(combos):
        if progress_cb:
            progress_cb(i, total, 'coarse')
        run_cfg = {**cfg}
        for (_, key), val in zip(WEAPONS, combo):
            run_cfg[key] = val
        mc = monte_carlo_v7(run_cfg, n=coarse_n)
        coarse_results.append((mc['mean_intercept'], combo))

    coarse_results.sort(reverse=True)

    # 2단계: 상위 top_k 정밀 검증
    final: list = []
    for rank, (_, combo) in enumerate(coarse_results[:top_k]):
        if progress_cb:
            progress_cb(total + rank, total + top_k, 'fine')
        run_cfg = {**cfg}
        for (_, key), val in zip(WEAPONS, combo):
            run_cfg[key] = val
        mc = monte_carlo_v7(run_cfg, n=fine_n)
        combo_dict = {wpn: val for (wpn, _), val in zip(WEAPONS, combo)}
        combo_cost = sum(cnt * FRIENDLY_DB.get(wpn, {}).get('cost_usd', 0)
                         for wpn, cnt in combo_dict.items())
        mean_e_dest = float(sum(mc.get('enemy_destroyed', [0])) /
                            max(len(mc.get('enemy_destroyed', [1])), 1))
        final.append({
            'combo':      combo_dict,
            'rate':       mc['mean_intercept'],
            'std':        mc['std_intercept'],
            'total':      sum(combo),
            'combo_cost': combo_cost,
            'mean_cost':  mc.get('mean_cost', 0),
            'cost_per_kill': (mc.get('mean_cost', 0) / mean_e_dest
                              if mean_e_dest > 0 else float('inf')),
        })

    return sorted(final, key=lambda x: -x['rate'])


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: REQ 요구조건 판정
# ════════════════════════════════════════════════════════════════════════════

REQ_ITEMS_V7 = [
    {'id': 'REQ-01', 'name': 'MC 평균 요격률 ≥ 95%',  'desc': 'MC 평균 요격률 ≥ 95% (운 배제, 실력 기준)'},
    {'id': 'REQ-02', 'name': '응답시간 충족',          'desc': f'첫 SAM 발사 ≤ {MAX_RESPONSE_TIME_S}s'},
    {'id': 'REQ-04', 'name': '완전 요격 달성률 ≥ 90%', 'desc': 'MC 완전 요격(100%) 달성 비율 ≥ 90%'},
    {'id': 'REQ-05', 'name': '아군 무피격 (MC)',        'desc': 'MC 시뮬의 85% 이상에서 아군 피격 0회'},
    {'id': 'REQ-06', 'name': '다층 방어 확인',          'desc': '발사 미사일 수 ≥ 위협 수 (재교전 여력)'},
    {'id': 'REQ-07', 'name': '주요 SAM 잔여 ≥ 20%',    'desc': 'SM-3·SM-6·SM-2 합산 잔여 ≥ 초기 재고 20%'},
    {'id': 'REQ-08', 'name': '채널 한계 미초과',        'desc': '최대 동시 위협 ≤ 편대 총 채널'},
]


_SAM_STOCK_KEYS = {
    'SM-3 Block IIA':  'sm3_stock',
    'SM-6':            'sm6_stock',
    'SM-2 Block IIIB': 'sm2_stock',
}


def evaluate_req_v7(result: dict, mc: dict, cfg: dict = None) -> tuple:
    """REQ_ITEMS_V7 7항목 판정. (verdicts: list[bool], details: list[str]) 반환."""
    tfirst   = result.get('t_first_fire', -1.0)
    fired    = result.get('total_missiles_fired', 0)
    threats  = result['total_threats']
    peak_et  = result.get('peak_concurrent_threats', 0)
    tot_ch   = result.get('total_channels', 16)
    rem_inv  = result.get('remaining_inventory', {})

    # REQ-01: MC 평균 요격률 ≥ 95% (단일 시뮬 운 배제)
    req1 = mc['mean_intercept'] >= 0.95

    req2 = 0 <= tfirst <= MAX_RESPONSE_TIME_S

    req4 = mc['full_pass_rate'] >= 0.90

    # REQ-05: MC 시뮬의 85% 이상에서 아군 피격 0회
    hits_list = mc.get('friendly_hits', [])
    zero_hit_rate = (sum(1 for h in hits_list if h == 0) / len(hits_list)) if hits_list else 1.0
    req5 = zero_hit_rate >= 0.85

    req6 = (fired >= threats) if threats > 0 else True

    # REQ-07: 주요 SAM(SM-3·SM-6·SM-2) 잔여 ≥ 초기 재고 20%
    if cfg:
        init_sam = sum(cfg.get(v, 0) for v in _SAM_STOCK_KEYS.values())
        rem_sam  = sum(rem_inv.get(k, 0) for k in _SAM_STOCK_KEYS)
        req7 = (rem_sam / init_sam >= 0.20) if init_sam > 0 else True
        req7_detail = f"주요 SAM 잔여 {rem_sam}발 / 초기 {init_sam}발 ({rem_sam/init_sam:.0%})" if init_sam > 0 else "재고 없음"
    else:
        req7 = any(v > 0 for v in rem_inv.values())
        req7_detail = f"잔여 {'확보됨' if req7 else '전량 소진!'} ({sum(rem_inv.values())}발)"

    req8 = peak_et <= tot_ch

    verdicts = [req1, req2, req4, req5, req6, req7, req8]
    details  = [
        f"MC 평균 요격률 {mc['mean_intercept']:.1%} {'≥' if req1 else '<'} 95%",
        f"첫 발사 {tfirst:.0f}s ≤ {MAX_RESPONSE_TIME_S}s" if tfirst >= 0 else "발사 없음",
        f"MC 완전 성공률 {mc['full_pass_rate']:.1%} {'≥' if req4 else '<'} 90%",
        f"MC 무피격 비율 {zero_hit_rate:.1%} {'≥' if req5 else '<'} 85%",
        f"발사 {fired}발 / 위협 {threats}개",
        req7_detail,
        f"최대 동시 위협 {peak_et} ≤ 채널 {tot_ch}",
    ]
    return verdicts, details


# ════════════════════════════════════════════════════════════════════════════
#  REQ 달성 최소 재고 역산
# ════════════════════════════════════════════════════════════════════════════

# cfg 키 ↔ 무기명 매핑 (포팅 A _def_map과 동일 구조)
_STOCK_CFG_KEY: dict = {
    'SM-3 Block IIA':   'sm3_stock',
    'SM-6':             'sm6_stock',
    'SM-2 Block IIIB':  'sm2_stock',
    'RIM-116 RAM':      'ram_stock',
    '홍상어 (대잠)':    'hongsango_stock',
    '청상어 (경어뢰)':  'cheongsango_stock',
}


def find_min_stock_v7(
    cfg: dict,
    weapon_name: str,
    target_rate: float = 0.90,
    mc_n: int = 40,
) -> int:
    """
    이진 탐색으로 REQ-04(MC 완전 요격 성공률 ≥ target_rate) 달성에 필요한
    weapon_name 의 최소 함정당 재고를 반환.
      ≥ 0 : 달성 가능한 최소 재고
      -1  : 최대값(SHIP_DB 기본)에서도 달성 불가
    """
    stock_key = _STOCK_CFG_KEY.get(weapon_name)
    if stock_key is None:
        return -1

    max_val = FRIENDLY_DB.get(weapon_name, {}).get('stock', 48)

    # 상한에서도 미달성이면 불가
    if monte_carlo_v7({**cfg, stock_key: max_val}, mc_n)['full_pass_rate'] < target_rate:
        return -1

    lo, hi = 0, max_val
    while lo < hi:
        mid = (lo + hi) // 2
        rate = monte_carlo_v7({**cfg, stock_key: mid}, mc_n)['full_pass_rate']
        if rate >= target_rate:
            hi = mid
        else:
            lo = mid + 1
    return lo


def find_all_min_stocks_v7(
    cfg: dict,
    target_rate: float = 0.90,
    mc_n: int = 40,
    progress_cb=None,
) -> dict:
    """
    주요 무기 6종의 최소 함정당 재고를 순서대로 탐색.
    반환: {weapon_name: {'min_stock': int, 'current_stock': int, 'achievable': bool}}
    """
    weapons = list(_STOCK_CFG_KEY.keys())
    results = {}
    for i, wpn in enumerate(weapons):
        if progress_cb:
            progress_cb(i, len(weapons), wpn)
        key     = _STOCK_CFG_KEY[wpn]
        current = cfg.get(key, FRIENDLY_DB.get(wpn, {}).get('stock', 0))
        min_s   = find_min_stock_v7(cfg, wpn, target_rate, mc_n)
        results[wpn] = {
            'min_stock':     min_s,
            'current_stock': current,
            'achievable':    min_s >= 0,
        }
    if progress_cb:
        progress_cb(len(weapons), len(weapons), '완료')
    return results


# ════════════════════════════════════════════════════════════════════════════
#  자동 취약점 진단
# ════════════════════════════════════════════════════════════════════════════

def diagnose_vulnerabilities_v7(result: dict, mc: dict, cfg: dict) -> list:
    """
    MC 결과를 자동 분석하여 취약점 진단 카드 목록을 반환.
    각 카드: {'severity': 'HIGH'|'MED'|'LOW'|'OK', 'title', 'detail', 'suggestion'}
    """
    cards = []

    mean_ir   = mc['mean_intercept']
    full_pass = mc['full_pass_rate']
    std_ir    = mc['std_intercept']
    mean_hits = float(np.mean(mc['friendly_hits'])) if mc['friendly_hits'] else 0.0
    peak_et   = result.get('peak_concurrent_threats', 0)
    tot_ch    = result.get('total_channels', 16)
    rem_inv   = result.get('remaining_inventory', {})
    t_first   = result.get('t_first_fire', -1.0)
    w_avg_rem = mc.get('weapon_avg_remaining', {})

    # ── 1. 완전 요격 성공률 미달 ──────────────────────────────────────────
    if full_pass < 0.90:
        sev = 'HIGH' if full_pass < 0.70 else 'MED'
        # 가장 많이 소진된 무기 파악 → 구체적 개선 제안
        most_depleted = min(w_avg_rem, key=lambda k: w_avg_rem[k], default=None)
        if most_depleted and most_depleted in _STOCK_CFG_KEY:
            key     = _STOCK_CFG_KEY[most_depleted]
            cur_stk = cfg.get(key, FRIENDLY_DB.get(most_depleted, {}).get('stock', 0))
            sugg    = f'• {most_depleted} 재고를 {cur_stk}→{cur_stk + 12}발로 증가 검토\n• MC 횟수 증가로 정밀도 향상'
        else:
            sugg = '• 주요 SAM 재고 증가\n• CEC 활성화 또는 함정 증원 검토'
        cards.append({
            'severity':   sev,
            'title':      f'완전 요격 성공률 미달  ({full_pass:.0%} < REQ 90%)',
            'detail':     (f'MC {mc["n"]}회 중 {full_pass:.0%}만 모든 위협 요격. '
                           f'평균 요격률 {mean_ir:.1%} (편차 ±{std_ir:.1%}).'),
            'suggestion': sugg,
        })

    # ── 2. 아군 피격 빈발 ─────────────────────────────────────────────────
    if mean_hits > 0.3:
        sev = 'HIGH' if mean_hits >= 1.5 else 'MED'
        most_hit_ship = ''
        ship_avg = mc.get('ship_avg_hits', {})
        if ship_avg:
            sh = max(ship_avg, key=ship_avg.get)
            most_hit_ship = f'  가장 많이 피격: {sh} (평균 {ship_avg[sh]:.1f}회)'
        cards.append({
            'severity':   sev,
            'title':      f'아군 함정 피격 빈발  (MC 평균 {mean_hits:.1f}회)',
            'detail':     f'종말 방어 단계(RAM/CIWS) 취약 또는 ECM·회피 기동 효과 부족.{most_hit_ship}',
            'suggestion': '• RIM-116 RAM 재고 증가\n• 함정 회피 기동 활성화\n• ECM(AN/SLQ-32) 옵션 활성화',
        })

    # ── 3. 채널 포화 ──────────────────────────────────────────────────────
    if peak_et > 0 and tot_ch > 0:
        ratio = peak_et / tot_ch
        if ratio >= 1.0:
            cards.append({
                'severity':   'HIGH',
                'title':      f'채널 포화 발생  (동시 위협 {peak_et} > 채널 {tot_ch})',
                'detail':     f'최대 {peak_et}개 위협이 동시 접근 — 교전 채널 {tot_ch}개 초과. 일부 위협 무대응.',
                'suggestion': '• CEC(협동교전능력) 활성화로 채널 공유\n• 추가 함정 편입\n• 발사 간격(launch_interval_s) 단축',
            })
        elif ratio >= 0.80:
            cards.append({
                'severity':   'MED',
                'title':      f'채널 포화 근접  ({peak_et}/{tot_ch} = {ratio:.0%})',
                'detail':     f'채널 사용률 {ratio:.0%} — 위협 추가 시 포화 임박.',
                'suggestion': '• CEC 활성화 또는 함정 증원 검토',
            })

    # ── 4. 주요 무기 소진 ─────────────────────────────────────────────────
    key_weapons = list(_STOCK_CFG_KEY.keys())
    for wpn in key_weapons:
        avg_rem = w_avg_rem.get(wpn, -1.0)
        if avg_rem < 0:
            continue  # 해당 무기 미사용 시나리오
        if avg_rem < 2.0:
            sev     = 'HIGH' if avg_rem < 0.5 else 'MED'
            cfg_key = _STOCK_CFG_KEY[wpn]
            cur_stk = cfg.get(cfg_key, FRIENDLY_DB.get(wpn, {}).get('stock', 0))
            cards.append({
                'severity':   sev,
                'title':      f'{wpn} 재고 고갈 위험  (MC 평균 잔여 {avg_rem:.1f}발)',
                'detail':     f'평균적으로 {wpn}이 거의 소진됨 (현재 재고: 함정당 {cur_stk}발).',
                'suggestion': f'• {wpn} 재고를 {cur_stk}→{cur_stk + 12}발로 증가 검토',
            })

    # ── 5. 응답시간 초과 ──────────────────────────────────────────────────
    if 0 <= t_first > MAX_RESPONSE_TIME_S:
        cards.append({
            'severity':   'MED',
            'title':      f'응답시간 초과  (첫 SAM 발사 {t_first:.0f}s > REQ {MAX_RESPONSE_TIME_S}s)',
            'detail':     'C&D + 확인 절차 후 첫 발사까지 지나치게 오래 소요. REQ-02 불충족.',
            'suggestion': '• C&D 시간(cd_time_s) 단축\n• 탐지거리 확대로 사전 추적 가능',
        })

    # ── 6. 높은 변동성 ────────────────────────────────────────────────────
    if std_ir > 0.15 and mean_ir < 0.98:
        cards.append({
            'severity':   'LOW',
            'title':      f'요격률 불안정  (표준편차 {std_ir:.1%})',
            'detail':     f'MC 결과 편차 큼 — {mean_ir:.0%}±{std_ir:.0%}. 특정 조건에서 방어 붕괴 가능.',
            'suggestion': '• MC 횟수 증가(200회 이상)로 신뢰도 향상\n• CEC 활성화로 일관성 확보',
        })

    # ── 7. 지상 BMD 자산 소진 ────────────────────────────────────────────
    if cfg:
        ground_rem = result.get('ground_remaining', {})
        for asset, cfg_key, label in [
            ('SM-3 (어쇼어)', 'ashore_sm3_stock', '이지스 어쇼어'),
            ('THAAD 요격탄',  'thaad_stock',       'THAAD'),
        ]:
            init_stock = cfg.get(cfg_key, 0)
            if init_stock > 0:
                rem   = ground_rem.get(asset, 0)
                ratio = rem / init_stock
                if ratio < 0.20:
                    sev = 'HIGH' if ratio < 0.05 else 'MED'
                    cards.append({
                        'severity':   sev,
                        'title':      f'{label} 탄약 고갈  (잔여 {rem}발 / 초기 {init_stock}발, {ratio:.0%})',
                        'detail':     '지상 BMD 1차 방어망 약화. 탄도·HGV 위협이 함정까지 도달 가능.',
                        'suggestion': f'• {label} 탄약 재고 {init_stock + 12}발 이상으로 증가\n• 함정 SM-3 재고 보충으로 백업 강화',
                    })

    # ── 이상 없음 ─────────────────────────────────────────────────────────
    if not cards:
        cards.append({
            'severity':   'OK',
            'title':      '취약점 없음 — 방어 태세 양호',
            'detail':     (f'완전 요격 성공률 {full_pass:.0%} · '
                           f'MC 평균 요격률 {mean_ir:.1%} · '
                           f'아군 평균 피격 {mean_hits:.1f}회'),
            'suggestion': '더 어려운 시나리오(전방위 포화·혼합 공격)로 한계 탐색 권장.',
        })

    return cards


# ════════════════════════════════════════════════════════════════════════════
#  교전 후 브리핑 자동 생성
# ════════════════════════════════════════════════════════════════════════════

def generate_briefing(result: dict, mc: dict, cfg: dict) -> str:
    """시뮬 결과를 군사 보고서 형식의 서술형 텍스트로 자동 생성."""
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    mean_ir   = mc['mean_intercept']
    full_pass = mc['full_pass_rate']
    cvar      = mc.get('cvar')
    mean_hits = float(np.mean(mc['friendly_hits'])) if mc['friendly_hits'] else 0.0
    n         = mc['n']

    fleet_preset = cfg.get('fleet_preset', '—')
    weather      = cfg.get('weather', '—')
    enemy_label  = (cfg.get('mixed_scenario')
                    or cfg.get('enemy_fleet_preset')
                    or cfg.get('enemy_fleet_mode', '—'))

    total_threats = result.get('total_threats', 0)
    enemy_dest    = result.get('enemy_ships_destroyed', 0)
    friendly_lost = result.get('friendly_ships_lost', 0)
    mean_cost     = mc.get('mean_cost', result.get('total_cost', 0))
    aircraft_sort = result.get('aircraft_sorties', 0)

    mission_ok  = mean_ir >= 0.95 and full_pass >= 0.90
    mission_str = '✅ 임무 성공' if mission_ok else '❌ 임무 실패'

    verdicts, details = evaluate_req_v7(result, mc, cfg)
    cards = diagnose_vulnerabilities_v7(result, mc, cfg)

    SEP  = '━' * 54
    SEP2 = '─' * 54
    _SEV = {'HIGH': '🔴 위험', 'MED': '🟡 경고', 'LOW': '🔵 주의', 'OK': '🟢 양호'}

    lines = [
        SEP,
        '  이지스 기동전단 교전 후 브리핑 (자동 생성)',
        f'  작성 일시: {now}',
        SEP, '',
        '【1. 작전 개요】',
        f'  편대 구성  : {fleet_preset}',
        f'  위협 환경  : {weather} / {enemy_label}',
        f'  총 위협 수 : {total_threats}개',
        f'  MC 시뮬    : {n}회',
        '',
        '【2. 교전 결과 요약】',
        f'  요격률 (MC 평균)     : {mean_ir:.1%}  →  {mission_str}',
        f'  완전 요격 달성률     : {full_pass:.1%}',
    ]
    if cvar is not None:
        lines.append(f'  최악 CVaR (하위 5%)  : {cvar:.1%}')
    lines += [
        f'  아군 피격 (MC 평균)  : {mean_hits:.1f}회',
        f'  아군 함정 손실       : {friendly_lost}척',
        f'  적 함정 격침         : {enemy_dest}척',
        f'  총 교전 비용 (평균)  : ${mean_cost:,.0f}',
    ]
    if aircraft_sort > 0:
        lines.append(f'  항공 출격 횟수       : {aircraft_sort}회')
    lines += ['', '【3. REQ 판정 결과】']
    for req, v, d in zip(REQ_ITEMS_V7, verdicts, details):
        mark = '✅' if v else '❌'
        lines.append(f'  {mark} {req["id"]}  {req["name"]}  →  {"PASS" if v else "FAIL"}')
        lines.append(f'       {d}')

    lines += ['', '【4. 주요 취약점】']
    for card in cards:
        lines.append(f'  {_SEV.get(card["severity"], "")}  {card["title"]}')
        lines.append(f'       {card["detail"]}')

    lines += ['', '【5. 권고사항】']
    suggestions = []
    for card in cards:
        if card['severity'] in ('HIGH', 'MED') and card.get('suggestion'):
            for s in card['suggestion'].split('\n'):
                s = s.strip().lstrip('•').strip()
                if s and s not in suggestions:
                    suggestions.append(s)
    if suggestions:
        for i, s in enumerate(suggestions, 1):
            lines.append(f'  {i}. {s}')
    else:
        lines.append('  현재 주요 권고사항 없음. 방어 태세 양호.')

    lines += [
        '',
        SEP,
        '  본 브리핑은 시뮬레이션 데이터 기반 자동 생성 문서입니다.',
        '  Pk 수치는 공개 자료 기반 추정값 (±15~20%) — 실측 데이터 아님.',
        SEP,
    ]
    return '\n'.join(lines)


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: 날씨별 시나리오 비교
# ════════════════════════════════════════════════════════════════════════════

_SCENARIO_WEATHERS = [
    ('최선 (맑음)',  '맑음 (주간)'),
    ('평균 (흐림)',  '흐림 (박무)'),
    ('최악 (폭풍)',  '폭풍 (해상 악화)'),
]


def scenario_comparison_v7(cfg: dict, n: int = 200) -> dict:
    """날씨 3종 MC 비교. {label: mc_dict + mean_cost} 반환."""
    results = {}
    for label, weather in _SCENARIO_WEATHERS:
        c = dict(cfg)
        c['weather'] = weather
        mc = monte_carlo_v7(c, n=n, desc=f'시나리오: {label}')
        results[label] = {
            **mc,
            'mean_cost': float(np.mean(mc['total_costs'])),
        }
    return results


# ════════════════════════════════════════════════════════════════════════════
#  CEC 효과 비교
# ════════════════════════════════════════════════════════════════════════════

def cec_comparison_v7(cfg: dict, n: int = 200) -> dict:
    """
    CEC ON/OFF/두절 3종 MC 비교.
    반환: {'CEC ON': mc_dict, 'CEC OFF (독립교전)': mc_dict, 'CEC 두절 (재밍)': mc_dict}
    """
    results = {}
    scenarios = [
        ('CEC ON',            {'enable_cec_preassign': True,  'enable_cec_jammed': False}),
        ('CEC OFF (독립교전)', {'enable_cec_preassign': False, 'enable_cec_jammed': False}),
        ('CEC 두절 (재밍)',   {'enable_cec_preassign': False, 'enable_cec_jammed': True}),
    ]
    for label, overrides in scenarios:
        c = {**cfg, **overrides}
        mc = monte_carlo_v7(c, n=n, desc=f'CEC: {label}')
        results[label] = {
            **mc,
            'mean_cost': float(np.mean(mc['total_costs'])) if mc.get('total_costs') else 0.0,
        }
    return results


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: A vs B 시나리오 비교
# ════════════════════════════════════════════════════════════════════════════

def compare_ab_v7(cfg_a: dict, cfg_b: dict, n: int = 200) -> dict:
    """
    두 cfg로 MC를 각각 실행해 비교 dict를 반환.
    반환: {'a': mc_dict, 'b': mc_dict, 'delta_intercept': float, 'delta_cost': float}
    """
    mc_a = monte_carlo_v7(cfg_a, n=n, desc='A 시나리오')
    mc_b = monte_carlo_v7(cfg_b, n=n, desc='B 시나리오')
    return {
        'a':               mc_a,
        'b':               mc_b,
        'delta_intercept': mc_b['mean_intercept'] - mc_a['mean_intercept'],
        'delta_cost':      float(np.mean(mc_b['total_costs'])) - float(np.mean(mc_a['total_costs'])),
    }


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: 시나리오 저장 / 불러오기
# ════════════════════════════════════════════════════════════════════════════

import json as _json


def save_scenario_v7(cfg: dict, path: str):
    """cfg를 JSON으로 저장. 직렬화 불가능한 값은 제외."""
    serializable = {}
    for k, v in cfg.items():
        try:
            _json.dumps(v)
            serializable[k] = v
        except (TypeError, ValueError):
            pass
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(serializable, f, ensure_ascii=False, indent=2)


def load_scenario_v7(path: str) -> dict:
    """JSON 파일에서 cfg를 불러온다."""
    with open(path, 'r', encoding='utf-8') as f:
        return _json.load(f)


# ════════════════════════════════════════════════════════════════════════════
#  PNG 차트 생성
# ════════════════════════════════════════════════════════════════════════════

_BG   = '#0a0e1a'
_GRID = '#1e2a3a'
_ACC  = '#3498db'

def _ax_style(ax, title: str):
    ax.set_facecolor(_BG)
    ax.tick_params(colors='#aab', labelsize=11)
    for sp in ax.spines.values():
        sp.set_color(_GRID)
    ax.set_title(title, color='#dde', fontsize=13, fontweight='bold', pad=6)


def plot_v7(result: dict, mc: dict, cfg: dict,
            img_path: str = '이지스_기동전단_v7_분석.png') -> str:
    """
    단일 시뮬 결과(result) + MC 통계(mc)를 6개 서브플롯으로 시각화.
    img_path에 저장 후 경로 반환.
    """
    fig = _MplFigure(figsize=(16, 10), facecolor=_BG)
    _FigureCanvasAgg(fig)
    fig.suptitle(
        f"이지스 기동전단 통합 방어 시뮬레이터 v7.0\n"
        f"시나리오: {cfg.get('fleet_preset','?')} | "
        f"날씨: {cfg.get('weather','?')} | "
        f"MC {mc['n']}회",
        color='white', fontsize=16, fontweight='bold', y=0.98,
    )

    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.45, wspace=0.35,
                           left=0.07, right=0.97, top=0.90, bottom=0.08)

    # ── (0,0) 요격률 히스토그램 ──────────────────────────────────────────────
    ax0 = fig.add_subplot(gs[0, 0])
    _ax_style(ax0, '요격률 분포 (MC)')
    ax0.hist(mc['intercept_rates'], bins=20, color=_ACC, edgecolor='#0a0e1a', alpha=0.85)
    ax0.axvline(mc['mean_intercept'], color='#e74c3c', lw=1.5, ls='--',
                label=f"평균 {mc['mean_intercept']:.1%}")
    ax0.set_xlabel('요격률', color='#aab', fontsize=11)
    ax0.set_ylabel('빈도', color='#aab', fontsize=11)
    ax0.legend(fontsize=10, facecolor=_BG, labelcolor='white', edgecolor=_GRID)
    ax0.xaxis.set_major_formatter(_FuncFormatter(lambda v, _: f'{v:.0%}'))

    # ── (0,1) 아군 피격 분포 ─────────────────────────────────────────────────
    ax1 = fig.add_subplot(gs[0, 1])
    _ax_style(ax1, '아군 피격 횟수 분포 (MC)')
    ax1.hist(mc['friendly_hits'], bins=range(0, max(mc['friendly_hits']) + 2),
             color='#e74c3c', edgecolor='#0a0e1a', alpha=0.85, align='left')
    ax1.set_xlabel('피격 횟수', color='#aab', fontsize=11)
    ax1.set_ylabel('빈도', color='#aab', fontsize=11)

    # ── (0,2) 적 격침 분포 ───────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[0, 2])
    _ax_style(ax2, '적 플랫폼 격침 수 분포 (MC)')
    max_dest = max(mc['enemy_destroyed']) if mc['enemy_destroyed'] else 1
    ax2.hist(mc['enemy_destroyed'], bins=range(0, max_dest + 2),
             color='#2ecc71', edgecolor='#0a0e1a', alpha=0.85, align='left')
    ax2.set_xlabel('격침 수', color='#aab', fontsize=11)
    ax2.set_ylabel('빈도', color='#aab', fontsize=11)

    # ── (1,0) 무기 소모 현황 (단일 시뮬) ─────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 0])
    _ax_style(ax3, '무기 소모 현황 (단일 시뮬)')
    ships = result.get('friendly_ships', [])
    wpn_used: dict = {}
    for s in ships:
        spec = SHIP_DB.get(s.ship_type, {})
        default_inv = spec.get('default_inventory', {})
        for wpn, orig in default_inv.items():
            used = orig - s.inventory.get(wpn, orig)
            if used > 0:
                wpn_used[wpn] = wpn_used.get(wpn, 0) + used
    if wpn_used:
        labels = list(wpn_used.keys())
        values = [wpn_used[k] for k in labels]
        colors = [_ACC if i % 2 == 0 else '#5dade2' for i in range(len(labels))]
        bars = ax3.barh(labels, values, color=colors, edgecolor='#0a0e1a')
        ax3.bar_label(bars, padding=3, color='white', fontsize=10)
        ax3.set_xlabel('발사 수', color='#aab', fontsize=11)
        ax3.tick_params(axis='y', labelsize=10)
    else:
        ax3.text(0.5, 0.5, '발사 없음', color='#aab', ha='center', va='center',
                 transform=ax3.transAxes)

    # ── (1,1) 비용 분포 ──────────────────────────────────────────────────────
    ax4 = fig.add_subplot(gs[1, 1])
    _ax_style(ax4, '총 교전 비용 분포 (MC)')
    costs_m = [c / 1_000_000 for c in mc['total_costs']]
    ax4.hist(costs_m, bins=20, color='#f39c12', edgecolor='#0a0e1a', alpha=0.85)
    mean_m = np.mean(costs_m)
    ax4.axvline(mean_m, color='#e74c3c', lw=1.5, ls='--',
                label=f'평균 ${mean_m:.1f}M')
    ax4.set_xlabel('비용 (백만 USD)', color='#aab', fontsize=11)
    ax4.set_ylabel('빈도', color='#aab', fontsize=11)
    ax4.legend(fontsize=10, facecolor=_BG, labelcolor='white', edgecolor=_GRID)

    # ── (1,2) 핵심 수치 요약 ────────────────────────────────────────────────
    ax5 = fig.add_subplot(gs[1, 2])
    ax5.set_facecolor(_BG)
    ax5.axis('off')
    _ax_style(ax5, '핵심 수치 요약')

    enemy_count = sum(s.get('count', 1) for s in cfg.get('enemy_fleet', []))
    summary_lines = [
        ('적 편대 규모',     f"{enemy_count}기/척"),
        ('시뮬 종료 시각',   f"{result['sim_time']:.0f}s"),
        ('총 위협 수',       f"{result['total_threats']}발/기"),
        ('요격 성공 (단일)', f"{result['intercepted_threats']}발/기"),
        ('', ''),
        ('MC 평균 요격률',   f"{mc['mean_intercept']:.1%}"),
        ('MC 표준편차',      f"±{mc['std_intercept']:.1%}"),
        ('완전요격 비율',    f"{mc['full_pass_rate']:.1%}"),
        ('', ''),
        ('아군 피격 (단일)', f"{result['friendly_hits']}회"),
        ('적 격침 (단일)',   f"{result['enemy_ships_destroyed']}기/척"),
        ('아군 손실 (단일)', f"{result['friendly_ships_lost']}척"),
        ('총 비용 (단일)',   f"${result['total_cost']:,.0f}"),
    ]
    y = 0.97
    for label, val in summary_lines:
        if not label:
            y -= 0.04
            continue
        ax5.text(0.04, y, label, color='#7fb3d3', fontsize=11, transform=ax5.transAxes, va='top')
        ax5.text(0.96, y, val,   color='white',   fontsize=11, transform=ax5.transAxes, va='top', ha='right', fontweight='bold')
        y -= 0.075

    fig.savefig(img_path, dpi=150, bbox_inches='tight', facecolor=_BG)
    fig.clf()
    print(f"  그래프 저장: '{img_path}'")
    return img_path


# ════════════════════════════════════════════════════════════════════════════
#  Excel 보고서 생성
# ════════════════════════════════════════════════════════════════════════════

def save_excel_report_v7(result: dict, mc: dict, cfg: dict,
                          img_path: str = '',
                          xlsx_path: str = '이지스_기동전단_v7_보고서.xlsx'):
    """
    Sheet1: MC 통계 요약
    Sheet2: 무기 소모 현황
    Sheet3: 교전 로그
    Sheet4: PNG 차트 삽입 (이미지 있을 때)
    """
    wb = Workbook()
    tb = Border(**{s: Side(style='thin', color='CCCCCC')
                   for s in ('left', 'right', 'top', 'bottom')})

    def cs(ws, r, c, v, bold=False, bg=None, center=True, color='000000'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, size=10, name='Arial', color=color)
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center', wrap_text=True,
        )
        cell.border = tb
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    def title_row(ws, r, text, cols='A:F', bg='1A252F'):
        end_col = cols.split(':')[1]
        ws.merge_cells(f'A{r}:{end_col}{r}')
        cell = ws.cell(row=r, column=1, value=text)
        cell.font      = Font(bold=True, size=13, color='FFFFFF', name='Arial')
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 28

    def hdr(ws, r, headers, bg='2C3E50'):
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.font      = Font(bold=True, size=10, color='FFFFFF', name='Arial')
            cell.fill      = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = tb

    # ── Sheet1: MC 통계 요약 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'MC 통계 요약'
    ws1.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [20, 20, 18, 18, 18, 18]):
        ws1.column_dimensions[col].width = w

    title_row(ws1, 1, f'이지스 기동전단 v7.0 — MC {mc["n"]}회 통계 요약')
    hdr(ws1, 2, ['항목', '단일 시뮬', 'MC 평균', 'MC 표준편차', 'MC 최솟값', 'MC 최댓값'])

    rows = [
        ('요격률',
         f"{result['intercept_rate']:.1%}",
         f"{mc['mean_intercept']:.1%}",
         f"±{mc['std_intercept']:.1%}",
         f"{min(mc['intercept_rates']):.1%}",
         f"{max(mc['intercept_rates']):.1%}"),
        ('아군 피격',
         result['friendly_hits'],
         f"{np.mean(mc['friendly_hits']):.1f}",
         f"±{np.std(mc['friendly_hits']):.1f}",
         min(mc['friendly_hits']),
         max(mc['friendly_hits'])),
        ('적 격침',
         result['enemy_ships_destroyed'],
         f"{np.mean(mc['enemy_destroyed']):.1f}",
         f"±{np.std(mc['enemy_destroyed']):.1f}",
         min(mc['enemy_destroyed']),
         max(mc['enemy_destroyed'])),
        ('아군 함정 손실',
         result['friendly_ships_lost'],
         f"{np.mean(mc['friendly_lost']):.1f}",
         f"±{np.std(mc['friendly_lost']):.1f}",
         min(mc['friendly_lost']),
         max(mc['friendly_lost'])),
        ('총 비용 (USD)',
         f"${result['total_cost']:,.0f}",
         f"${np.mean(mc['total_costs']):,.0f}",
         f"±${np.std(mc['total_costs']):,.0f}",
         f"${min(mc['total_costs']):,.0f}",
         f"${max(mc['total_costs']):,.0f}"),
        ('완전 요격 비율',
         '—',
         f"{mc['full_pass_rate']:.1%}",
         '—', '—', '—'),
    ]
    for i, row in enumerate(rows):
        bg = 'D5F5E3' if i % 2 == 0 else 'EBF5FB'
        for j, val in enumerate(row, 1):
            cs(ws1, i + 3, j, val, bg=bg, center=(j > 1))

    # 시나리오 파라미터 블록
    ws1.cell(row=len(rows) + 5, column=1, value='【시나리오 파라미터】').font = Font(bold=True, size=11)
    params = [
        ('편대 프리셋',   cfg.get('fleet_preset', '?')),
        ('날씨',          cfg.get('weather', '?')),
        ('탐지 거리',     f"{cfg.get('detect_km', '?')} km"),
        ('해성-II 재고',  cfg.get('haesong2_stock', 0)),
        ('해성-I  재고',  cfg.get('haesong1_stock', 0)),
        ('하푼 재고',     cfg.get('harpoon_stock', 0)),
    ]
    for k, (label, val) in enumerate(params):
        cs(ws1, len(rows) + 6 + k, 1, label, center=False)
        cs(ws1, len(rows) + 6 + k, 2, val,   center=False)

    # ── Sheet2: 무기 소모 현황 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('무기 소모 현황')
    ws2.sheet_view.showGridLines = False
    for col, w in zip('ABCDE', [24, 16, 16, 16, 16]):
        ws2.column_dimensions[col].width = w

    title_row(ws2, 1, '함정별 무기 소모 현황 (단일 시뮬)')
    hdr(ws2, 2, ['함정명', '무기', '초기 재고', '소모량', '잔여량'])

    row_idx = 3
    for s in result.get('friendly_ships', []):
        spec        = SHIP_DB.get(s.ship_type, {})
        default_inv = spec.get('default_inventory', {})
        for wpn, orig in default_inv.items():
            remaining = s.inventory.get(wpn, orig)
            used      = orig - remaining
            bg = 'FADBD8' if used > 0 else 'F2F3F4'
            cs(ws2, row_idx, 1, s.name,      bg=bg, center=False)
            cs(ws2, row_idx, 2, wpn,         bg=bg, center=False)
            cs(ws2, row_idx, 3, orig,        bg=bg)
            cs(ws2, row_idx, 4, used,        bg='E74C3C' if used > 0 else bg,
               color='FFFFFF' if used > 0 else '000000', bold=(used > 0))
            cs(ws2, row_idx, 5, remaining,   bg=bg)
            row_idx += 1

    # ── Sheet3: 교전 로그 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('교전 로그')
    ws3.sheet_view.showGridLines = False
    for col, w in zip('AB', [12, 80]):
        ws3.column_dimensions[col].width = w

    title_row(ws3, 1, '교전 로그 (단일 시뮬)', cols='A:B')
    hdr(ws3, 2, ['시각 (s)', '이벤트'])

    for i, (t, msg) in enumerate(result.get('log', [])):
        bg = 'EBF5FB' if i % 2 == 0 else 'FDFEFE'
        cs(ws3, i + 3, 1, f'{t:.0f}', bg=bg)
        cs(ws3, i + 3, 2, msg,        bg=bg, center=False)

    # ── Sheet4: 교전 후 브리핑 ──────────────────────────────────────────────
    ws4 = wb.create_sheet('교전 후 브리핑')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 70
    title_row(ws4, 1, '교전 후 브리핑 (자동 생성)', cols='A:A')
    briefing_text = generate_briefing(result, mc, cfg)
    for i, line in enumerate(briefing_text.split('\n'), 2):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.font      = Font(name='Consolas', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws4.row_dimensions[i].height = 16

    # ── Sheet5: PNG 차트 ─────────────────────────────────────────────────────
    if img_path and _CAN_IMG and os.path.exists(img_path):
        ws5 = wb.create_sheet('분석 차트')
        ws5.sheet_view.showGridLines = False
        title_row(ws5, 1, 'MC 분석 차트', cols='A:L')
        img_obj = XLImage(img_path)
        img_obj.anchor = 'A2'
        ws5.add_image(img_obj)

    wb.save(xlsx_path)
    print(f"  엑셀 보고서 저장: '{xlsx_path}'")
    return xlsx_path


# ════════════════════════════════════════════════════════════════════════════
#  결과 JSON 내보내기
# ════════════════════════════════════════════════════════════════════════════

def save_json_report_v7(result: dict, mc: dict, path: str):
    """
    단일 시뮬 결과 + MC 통계를 JSON으로 저장.
    FriendlyShipObj 등 직렬화 불가 객체는 요약 딕셔너리로 변환.
    """
    def _safe(v):
        if isinstance(v, (int, float, str, bool, type(None))):
            return v
        if isinstance(v, (list, tuple)):
            return [_safe(x) for x in v]
        if isinstance(v, dict):
            return {k2: _safe(v2) for k2, v2 in v.items()}
        # 객체: 기본 속성만 추출
        return str(v)

    summary = {
        'result': {
            'intercept_rate':      result.get('intercept_rate'),
            'total_threats':       result.get('total_threats'),
            'intercepted_threats': result.get('intercepted_threats'),
            'friendly_hits':       result.get('friendly_hits'),
            'enemy_ships_destroyed': result.get('enemy_ships_destroyed'),
            'friendly_ships_lost': result.get('friendly_ships_lost'),
            'total_cost':          result.get('total_cost'),
            'sim_time':            result.get('sim_time'),
            't_first_fire':        result.get('t_first_fire'),
            'total_missiles_fired': result.get('total_missiles_fired'),
            'remaining_inventory': result.get('remaining_inventory', {}),
            'ships': [
                {
                    'name':         s.name,
                    'type':         s.ship_type,
                    'alive':        s.alive,
                    'hits_taken':   getattr(s, 'hits_taken', 0),
                    'total_cost':   s.total_cost,
                    'inventory':    dict(s.inventory),
                }
                for s in result.get('friendly_ships', [])
            ],
        },
        'mc': {
            'n':               mc.get('n'),
            'mean_intercept':  mc.get('mean_intercept'),
            'std_intercept':   mc.get('std_intercept'),
            'full_pass_rate':  mc.get('full_pass_rate'),
            'mean_cost':       float(np.mean(mc['total_costs'])) if mc.get('total_costs') else 0,
            'weapon_avg_remaining': mc.get('weapon_avg_remaining', {}),
            'ship_avg_hits':   mc.get('ship_avg_hits', {}),
        },
    }
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(summary, f, ensure_ascii=False, indent=2, default=_safe)
    print(f"  JSON 보고서 저장: '{path}'")
    return path


# ════════════════════════════════════════════════════════════════════════════
#  단독 실행 테스트 — 32종 혼합 시나리오
# ════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import sys

    scenario = sys.argv[1] if len(sys.argv) > 1 else 'mixed'

    SCENARIOS = {
        # 수상함 대결
        'surface': {
            'fleet_preset': '기동전단 기본',
            'weather':      '맑음 (주간)',
            'detect_km':    200,
            'haesong2_stock': 8,
            'haesong1_stock': 0,
            'harpoon_stock':  4,
            'enemy_fleet': [
                {'preset': '055형 대형 구축함', 'count': 1},
                {'preset': '052D형 구축함',     'count': 2},
            ],
        },
        # 항공 위협
        'air': {
            'fleet_preset': 'BMD 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    300,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': 'J-20 (위룡)',    'count': 2},
                {'preset': 'J-16 (플랭커-D)', 'count': 2},
                {'preset': 'H-6 (폭격기)',    'count': 1},
            ],
        },
        # 탄도/순항 미사일
        'missile': {
            'fleet_preset': 'BMD 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    400,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': 'DF-21D (대함 탄도)',        'count': 2},
                {'preset': 'DF-17 (극초음속 활공)',     'count': 1},
                {'preset': 'KN-23 (북한 이스칸데르)',   'count': 2},
                {'preset': 'YJ-12 (초음속 순항)',       'count': 3},
            ],
        },
        # 잠수함
        'sub': {
            'fleet_preset': '대잠 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    200,
            'sub_detect_km': 50,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': '093형 잠수함 (위안급)', 'count': 2},
                {'preset': '039형 잠수함 (송급)',   'count': 1},
            ],
        },
        # 32종 혼합
        'mixed': {
            'fleet_preset': '최대 편대',
            'weather':      '맑음 (주간)',
            'detect_km':    300,
            'sub_detect_km': 50,
            'haesong2_stock': 12,
            'haesong1_stock': 4,
            'harpoon_stock':  8,
            'enemy_fleet': [
                {'preset': '055형 대형 구축함',         'count': 1},
                {'preset': '052D형 구축함',             'count': 1},
                {'preset': 'J-20 (위룡)',               'count': 1},
                {'preset': 'H-6 (폭격기)',              'count': 1},
                {'preset': 'DF-21D (대함 탄도)',        'count': 1},
                {'preset': 'DF-17 (극초음속 활공)',     'count': 1},
                {'preset': 'KN-23 (북한 이스칸데르)',   'count': 1},
                {'preset': 'YJ-12 (초음속 순항)',       'count': 2},
                {'preset': '039형 잠수함 (송급)', 'count': 1},
            ],
        },
    }

    cfg = SCENARIOS.get(scenario, SCENARIOS['mixed'])

    MC_N = int(sys.argv[2]) if len(sys.argv) > 2 else 200

    print("=" * 66)
    print(f"  이지스 기동전단 통합 방어 시뮬레이터 v7.0  [시나리오: {scenario}]")
    print("=" * 66)

    result = run_v7_simulation(cfg)

    print(f"  시뮬 종료 시각  : {result['sim_time']:.0f}s")
    print(f"  총 위협 수      : {result['total_threats']}발/기")
    print(f"  요격 성공       : {result['intercepted_threats']}발/기")
    print(f"  요격률          : {result['intercept_rate']:.1%}")
    print(f"  아군 피격       : {result['friendly_hits']}회")
    print(f"  적 피격         : {result['enemy_hits']}회")
    print(f"  적 위협 격침    : {result['enemy_ships_destroyed']}기/척")
    print(f"  아군 함정 손실  : {result['friendly_ships_lost']}척")
    print(f"  총 비용         : ${result['total_cost']:,.0f}")
    print("-" * 66)
    print("  교전 로그:")
    for t, msg in result['log']:
        print(f"  [{t:5.0f}s] {msg}")
    print("=" * 66)

    # ── MC 분석 + 보고서 생성 ────────────────────────────────────────────────
    print(f"\n  몬테카를로 분석 ({MC_N}회) 시작...")
    mc = monte_carlo_v7(cfg, n=MC_N, desc=scenario)

    print(f"  MC 평균 요격률 : {mc['mean_intercept']:.1%}  "
          f"(±{mc['std_intercept']:.1%})")
    print(f"  완전 요격 비율 : {mc['full_pass_rate']:.1%}")

    img_path  = f'이지스_기동전단_v7_{scenario}_분석.png'
    xlsx_path = f'이지스_기동전단_v7_{scenario}_보고서.xlsx'

    plot_v7(result, mc, cfg, img_path=img_path)
    save_excel_report_v7(result, mc, cfg,
                         img_path=img_path, xlsx_path=xlsx_path)

    print("\n  ※ 실행 방법:")
    print(f"     python engine_v7.py [시나리오] [MC횟수]")
    print(f"     예) python engine_v7.py mixed 500")
    print("=" * 66)

