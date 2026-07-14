"""
engine_combat.py — 이지스 기동전단 통합 방어 시뮬레이터 v7.0
시간 스텝 기반 양방향 교전 엔진

v7.0 핵심 변경:
  · 이벤트 기반 → 시간 스텝(1초 단위) 시뮬레이션
  · 양방향 교전: 아군 공격(해성/하푼) + 적 SAM/CIWS 방어
  · 2D 위치 좌표계 (모든 엔티티 실시간 위치 추적)
  · SimFrame 단위 상태 기록 (애니메이션 지원)

v7.0 패치 이력:
  · 1단계: 시간 스텝 루프 + 위치 모델 + 기본 교전 판정
  · 2단계: 양방향 교전 검증
    - SAM alive=False 조기 설정 버그 수정 (요격 판정 스킵 문제)
    - INTERCEPT_DIST_M 300m → 2000m (1초 스텝 해상도 반영)
    - 대함 탐지 거리 수정 (자함 레이더 45km → 전술 인식 detect_km 병용)
    - 적 함정 시작 위치 수정 (1.8x → 1.0x, 교전 즉시 개시)
    - 적 함정 재발사 허용 (비행 중 미사일 소진 후 재장전)
  · 3단계: ENEMY_DB 전 32종 위협 지원
    - EnemyShipObj → EnemyThreatObj (항공기/함정/잠수함/독립미사일 통합)
    - 항공기: 접근 → 사거리 내 발사 → 이탈 행동 패턴
    - 탄도/순항/HGV/QBM 독립 미사일: _build_enemies()에서 MissileObj로 직접 생성
    - 대잠전: 홍상어/청상어 ASW 운용, 소나 탐지 범위 반영
    - _select_defense_wpn(): 고도/유형 인식 (SM-3 HGV/탄도, SM-6 QBM)
    - _friendly_strike(): 수상함(해성/하푼) + 잠수함(홍상어/청상어) 분리
  · 4단계: 몬테카를로 분석 + 보고서 생성
    - matplotlib 한글 폰트 설정 (Malgun Gothic)
    - monte_carlo_v7(): N회 반복 통계 집계 (요격률/피격/격침/비용)
    - plot_v7(): 6개 서브플롯 PNG 차트 (히스토그램·무기소모·수치요약)
    - save_excel_report_v7(): 4시트 Excel 보고서 (MC요약/무기소모/교전로그/차트)
    - __main__ 인수: python engine_combat.py [시나리오] [MC횟수]
  · 포팅 A: 방어 무기 재고 수동 설정 + 적군 편대 모드
    - _build_friendly(): sm3/sm6/sm2/ram/홍상어/청상어/mk46 수동 재고 지원 (cfg 키로 설정)
    - _build_enemies(): enemy_fleet_mode 4종 지원 (single/preset/custom/random)
    - ENEMY_FLEET_PRESETS / ENEMY_FLEET_RANDOM_CFG / generate_random_enemy_fleet 연동
  · 포팅 B: 전술 기능 — ECM·종말회피·음향기만기·함정회피·적 자체방어
    - ECM_REF_RANGE_M=50km 기준 거리 반비례 Pk 감소, Pk 하한 50% (탄도/HGV 제외)
    - 종말 회피: 20km 이내 terminal_evasion_factor 적용 (ENEMY_DB missile_terminal_evasion)
    - 음향 기만기: DECOY_PK=0.60, 어뢰 전용, decoy_stock 소모
    - 함정 회피 기동: SHIP_EVASION_PK=0.30, 어뢰 전용
    - 적 자체방어: CIWS(enemy_ciws_pk) 요격 → 채프/플레어(self_defense_pk) Pk 감소
  · 포팅 C: 항공 자산 대잠 운용 (FriendlyAircraftObj + _aircraft_asw())
    - AW-159 와일드캣: 함재 헬기, 청상어 2발, 사거리 140km, 폭풍/태풍/농무 불가
    - P-3C 오라이온: 포항기지 출격(+300km), Mk.46 4발, 소노부이 탐지+15km
    - P-8A 포세이돈: 포항기지 출격(+300km), Mk.46 5발, 소노부이 탐지+18km
  · 포팅 D: 분석 기능 — REQ 판정 + 날씨 비교 + A vs B + 저장/불러오기
    - evaluate_req_v7(): REQ-01~08 8항목 시간스텝 기반 판정
    - scenario_comparison_v7(): 날씨 3종(맑음/흐림/폭풍) MC 비교
    - compare_ab_v7(): 두 cfg MC 결과 대비 (Δ요격률·Δ비용)
    - save_scenario_v7() / load_scenario_v7(): JSON 시나리오 저장/불러오기
    - _compile(): remaining_inventory·total_channels·peak_concurrent_threats·t_first_fire 추가
"""

import math, random, os, time as _time, dataclasses
from typing import List, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.font_manager as fm
from matplotlib.figure import Figure as _MplFigure
from matplotlib.backends.backend_agg import FigureCanvasAgg as _FigureCanvasAgg
from matplotlib.ticker import FuncFormatter as _FuncFormatter
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from openpyxl.drawing.image import Image as XLImage
    _CAN_IMG = True
except Exception:
    _CAN_IMG = False

for _fp in ['C:/Windows/Fonts/malgun.ttf', 'C:/Windows/Fonts/malgunbd.ttf']:
    if os.path.exists(_fp):
        fm.fontManager.addfont(_fp)
matplotlib.rcParams['font.family'] = 'Malgun Gothic'
matplotlib.rcParams['axes.unicode_minus'] = False

from engine_core import (
    ENEMY_DB, FRIENDLY_DB, FRIENDLY_AIRCRAFT_DB, WEATHER_DB,
    SHIP_DB, FLEET_PRESETS, SHIP_PROCUREMENT_USD, ENEMY_PROCUREMENT_USD, SHIP_ENDURANCE,
    ENEMY_MUNITION,
    ENEMY_FLEET_PRESETS, ENEMY_FLEET_RANDOM_CFG, MIXED_ATTACK_SCENARIOS,
    generate_random_enemy_fleet,
    calculate_detect_range_by_rcs,
    normalize_enemy_db,
    SHIP_SURVIVABILITY, FLOOD_WARHEAD_FACTOR, FLOOD_BELOW_WL_PROB, FLOOD_INFLOW_K,
    SHIP_POWER, power_avail_kw,
)
try:
    from db_ocean_acoustic import (
        get_sonar_depth_factor,
        SUBMARINE_ACOUSTIC, SONAR_PLATFORM, WATER_DEPTH_M,
        sonar_detection_range, sonar_detection_prob,
        active_sonar_detection_range,
    )
    _OCEAN_ACOUSTIC_OK = True
except ImportError:
    _OCEAN_ACOUSTIC_OK = False

try:
    from db_ocean_environment import (
        RADAR_CLUTTER, SONAR_AMBIENT_NOISE, get_current_vector as _get_current_vector
    )
    _RADAR_BF_FACTOR = RADAR_CLUTTER['detection_range_factor']
    _SONAR_BF_FACTOR = SONAR_AMBIENT_NOISE['sonar_range_factor']
    _OCEAN_ENV_OK = True
except ImportError:
    _RADAR_BF_FACTOR = {}
    _SONAR_BF_FACTOR = {}
    _get_current_vector = None
    _OCEAN_ENV_OK = False

try:
    from db_terrain import STRAITS_DB
    _TERRAIN_DB_OK = True
except ImportError:
    STRAITS_DB = {}
    _TERRAIN_DB_OK = False

# ── v9.14: 해협 시나리오 상수 ────────────────────────────────────────────────
# strait_type → (접근 방위 중심각°, 확산각°) 매핑
# korea_west: 위협이 서쪽(일본해→서수도 방향) 270°에서 동쪽으로 접근
# korea_east: 위협이 동쪽(태평양→동수도 방향) 90°에서 서쪽으로 접근
# bilateral: 동서 양방향 협공
STRAIT_BEARING: dict[str, list[float]] = {
    'korea_west': [270.0],
    'korea_east': [90.0],
    'bilateral':  [90.0, 270.0],
}
STRAIT_SPREAD_DEG = 30.0  # 해협 입구 방위 확산각 (±30°)

# 해협 폭 기준 회피 기동 패널티: 폭이 좁을수록 기동 공간 감소
# 기준 200km(개방 해역) 대비 해협 폭 비율로 evade_r 스케일
_STRAIT_OPEN_SEA_KM = 200.0

# ── 시뮬레이션 상수 ──────────────────────────────────────────────────────────
DT               = 1.0    # 시간 스텝 (초)
MAX_SIM_TIME     = 3600   # 최대 시뮬 시간 (초) — 해성 250m/s 기준 250km = 1000초 충분

# ── v17.2: 지향성 에너지 무기(레이저·DEW) 상수 ───────────────────────────────
# 회절·대기감쇠로 근거리 한정. ref_km 이내는 정격출력, 그 밖은 도달출력 ∝ 1/거리².
_LASER_RANGE_KM        = 5.0     # 유효 교전거리(HELIOS 실측 최대 6마일≈9.6km의 보수적 유효값)
_LASER_REF_KM          = 2.0     # 정격출력 도달 기준거리(이내=정격, 밖=1/거리² 감쇠)
_LASER_MAX_TARGET_MS   = 340.0   # 표적 상한(아음속). 초음속·탄도·HGV 제외
_LASER_MAX_AIRCRAFT_MS = 120.0   # 드론 상한(유인 전투기 제외 — 저속 무인기만)
_LASER_NOMINAL_SPEED_MS = 10.0   # 방어 순항 근사(전력 마진 산정용). 전장 고속기동 시 저하는 battle 모드에서 발현
# 표적별 격추 소요 에너지(kJ). 드론<자폭정<아음속미사일(내구 순). HELIOS 60kW×5s≈300kJ로 드론 앵커.
_LASER_EKILL_KJ        = {'drone': 300.0, 'boat': 800.0, 'missile': 3000.0}

# ── v10.7: 전술 의사결정 모드 상태 스냅샷 ────────────────────────────────────
@dataclasses.dataclass
class TacticalState:
    t: float
    threats: list       # [{'name', 'type', 'hp', 'dist_km', 'speed_ms'}]
    ships:   list       # [{'name', 'alive', 'hp', 'max_hp', 'radar', 'speed'}]
    intercepted: int    # 현재까지 요격 수
    total_threats: int  # 총 위협 수
    shots_fired: int    # 현재까지 발사 수
    extra: dict = None  # 전장 모드 RL 보조 관측(위협 구성·자원). 부모는 None(단발 무영향)
INTERCEPT_DIST_M = 200    # BUG-5: SAM 근접 신관 범위 (m). 기존 2000m 과대, 실제 50-200m
ECM_REF_RANGE_M  = 25_000 # MED-9: ECM 재밍 기준 거리 25km (기존 50km 과대)
DECOY_PK         = 0.50   # LOW-7: 0.60→0.50 (AN/SLQ-25 실전 기만 성공률)
_DECEPTION_SCALE_M = 150.0  # v16.6 전자 좌표 기만 Pk 감쇠 스케일 (이격 150m마다 e^-1, ARM stale과 동형)
_MINE_SWEEP_FACTOR = 0.4     # v16.7 소해 지원 시 기뢰 접촉 확률 배율 (0.4 = 60% 경감)
# v16.13.02 트랙 C: 기함(지휘 노드) 격침 후 차순위 지휘권 인수까지 CEC relay 저하 지속 시간
_CEC_HANDOVER_DELAY_S = 45.0
SHIP_EVASION_PK  = 0.20   # LOW-8: 0.30→0.20 (회피 기동 성공률 현실화)
# v20.5(B-2) 함정 회피 기동 물리: 종전엔 틱(1초)마다 300~800m를 무작위 방향으로 순간이동해
# 실제 함속의 20~50배로 움직였다(583~1555kts). 구축함·이지스 전속 30kts 기준으로 정정.
_SHIP_EVADE_SPEED_MS = 15.4    # 회피 항주 속도(30kts 전속 — 이지스·구축함 공개 최대속력)
_EVADE_TURN_S        = 30.0    # 좌·우 전환 주기(s)
_EVADE_WEAVE_RAD     = 0.87    # 기준 침로에서 좌·우로 꺾는 각(rad, ≈50°) — weave 진폭
_EVADE_MAX_DRIFT_M   = 5_000.0 # 초기 위치 기준 최대 이탈(m). 전단 대형·중첩 방어 반경 유지
# v16.3 사이버전 효과 계수 (침투 지속 중 곱셈 페널티). 확률·주기·지속은 cfg로 튜닝.
_CYBER_DL_PK   = 0.70     # 데이터링크 변조 시 아군 요격 Pk 배율
_CYBER_CIC_DET = 0.60     # CIC 마비 시 아군 실효 탐지거리 배율
_CYBER_JAM_PK  = 0.75     # 레이더 교란 반격 시 적 발사 미사일 Pk 배율
# v16.2 극초음속 활공 궤적: HGV 비행 단계별 고도 프로파일(활공 완만 하강 → 종말 급강하).
_HGV_GLIDE_END      = 0.65    # 활공 단계 끝(이후 종말 급강하). 비행 진행도 p 기준
_HGV_GLIDE_DESCENT  = 0.35    # 활공 중 고도 하강 비율(peak → (1-0.35)·peak)
_HGV_TERMINAL_ALT_M = 2000.0  # 종말 침투 고도(m)
_HGV_GLIDE_EVADE    = 0.85    # 활공·종말 횡기동 시 아군 요격 Pk 배율
# v20.5(B-2) 표적 난이도 게이트: 요격 Pk가 표적의 속도·RCS를 전혀 보지 않던 공백을 메운다.
# 그전까지 마하 3·RCS 0.02m² ARM이 아음속 대형 표적과 같은 Pk로 맞았고(ARM 24발 중 22.5발
# 요격 = 94%), 그래서 레이더 침묵으로 회피할 이유 자체가 없었다.
#   속도: 접근률이 클수록 종말 유도·근접 신관 여유 감소 → (v_ref+v_sam)/(v_tgt+v_sam)
#   RCS : 레이더 방정식(탐색기 SNR ∝ RCS) → 소형 표적은 종말 획득·신관 신뢰도 저하
# 기준점 앵커링: 아음속 대함미사일(300 m/s·0.1m²)에서 계수 1.0 → 기존 밸런스 보존.
# ⚠ 지수·하한은 모델링 선택이다. 실제 SM-2의 초음속·소형 표적 단발 Pk는 공개되지 않는다
#   (기밀). 다만 '반영률 0'은 명백한 비물리라 유지가 곧 오류이며, 앵커링으로 기존 아음속
#   표적 결과는 그대로 둔다. 하한은 요격이 원천 불가해지지 않도록 건다.
_TDIFF_V_REF    = 300.0   # 기준 표적 속도(아음속 대함미사일급, m/s) → 계수 1.0
_TDIFF_RCS_REF  = 0.1     # 기준 표적 RCS(대함미사일급, m²) → 계수 1.0 (이상은 상한)
_TDIFF_RCS_EXP  = 0.125   # RCS 민감도 지수(작을수록 완만)
_TDIFF_MIN      = 0.35    # 종합 계수 하한(요격 원천 불가 방지)
# v16.1 능동 소나 핑 노출 대가(EMCON 딜레마): 역탐지한 잠수함이 회피 잠항으로 접촉을
# 끊어, ASW 재시도 탐지 보장을 차단한다(능동의 탐지 우위 ↔ 위치 노출 트레이드오프).
_ASW_PING_EVADE_P   = 0.45    # 능동 핑 역탐지 후 잠수함 회피(접촉 단절) 성공 확률
_ASW_CONTACT_LOST_S = 600.0   # 접촉 단절(잠항 도주) 지속 — 재탐지·재발사 불가
# v20.5(B-3) 재탐색 제한: 'max_attempts회 실패 → 표적 포기'가 실제로 포기가 되게 하는 쿨다운.
# 이 상수가 없으면 포기한 항공기가 다음 틱에 같은 잠수함을 처음 보듯 다시 3회 시도해
# **사실상 탐지가 보장**되고, 그러면 EMCON 딜레마('능동을 켜면 들키지만 안 켜면 못 찾는다')가
# 성립할 수 없다 — 잠수함이 접촉을 끊어도 시간만 지나면 반드시 재탐지되기 때문.
# 900초 = 소노부이 패턴을 새로 전개하고 재접촉을 시도하는 데 드는 시간의 보수적 근사
# (접촉 단절 600초보다 길어야 '끊고 도주'가 실제 이득이 된다).
_ASW_GIVEUP_COOLDOWN_S = 900.0
# v20.5(B-3) datum 성장(furthest-on circle) — 표준 대잠 교리 개념.
# 항공 대잠 탐지확률은 '잠수함이 어디 있는지 얼마나 아는가'(datum 오차)에 좌우된다. 접촉이
# 유지되는 동안은 오차가 작지만, 접촉이 끊기면 잠수함이 갈 수 있는 원이 **경과 시간 × 속도**로
# 커져 소노부이 패턴이 그 원을 덮지 못하게 된다. 기존 코드는 datum을 1500m로 **고정**해
# 항공기가 잠수함 위치를 항상 아는 셈이었고, 그래서 탐지확률이 상한 0.97에 붙어
# **탐지가 사실상 보장**됐다 — 그러면 잠수함이 접촉을 끊어도(EMCON 회피) 아무 이득이 없다.
# 임의 계수를 새로 만들지 않는다: 반경 = 잠수함 속도 × 마지막 접촉 이후 경과 시간.
_ASW_DATUM_BASE_M = 1500.0   # 접촉 유지 중 표정 오차(기존 고정값 — 접촉 직후의 값으로 유지)
_ASW_EVADE_JUMP_M   = 6000.0  # 회피 기동 위치 점프(datum 무효화)
MAX_RESPONSE_TIME_S = 120  # 포팅 D: REQ-02 최대 허용 응답시간 (초)

# ── v12.1: 비례항법(PNG) 종말 유도 상수 ──────────────────────────────────────
# enable_png=True일 때만 동작. 아군 SAM↔적 대함미사일 종말 10km 교전을 물리 추격.
PNG_SUBSTEP    = 20       # 종말 구간 substep 분할 → dt_sub = DT/20 = 0.05s
                          #   (폐쇄속도 ~2250m/s에서도 substep당 ~110m로 살상반경 판정 정밀)
PNG_TERMINAL_M = 10_000   # PNG 발동 거리 (종말 10km, _PLANS 권고 최소 적용)
SAM_MAX_G      = 30.0     # 함대공 미사일 기동 한계 (현대 SAM 종말 ~30~50G)
ASM_MAX_G      = 10.0     # 대함미사일 종말 회피 기동 한계 (~10G)
SEEKER_FOV_DEG = 60.0     # 탐색기 반각 시야 — 벗어나면 lock 상실(직진), 무한 기동 방지
PNG_LETHAL_M   = 50.0     # 근접신관 살상 반경 — 블래스트-파편 탄두 + 신관 격발거리(SM-2/6·ESSM)
G_ACCEL        = 9.81     # 중력가속도 (G→m/s² 변환)


def _ang_diff(a: float, b: float) -> float:
    """두 방위각(rad) 차를 (-π, π]로 정규화. PNG LOS rate·FOV 판정용."""
    d = (a - b) % (2 * math.pi)
    if d > math.pi:
        d -= 2 * math.pi
    return d


def _seg_origin_dist(ax: float, ay: float, bx: float, by: float) -> float:
    """선분 (a→b)에서 원점까지 최단거리. PNG substep 구간 내 최근접 통과 거리 계산용
    — 고속 교전에서 substep 끝점만 보면 살상반경 통과를 놓치므로 선분으로 적분."""
    abx, aby = bx - ax, by - ay
    l2 = abx * abx + aby * aby
    if l2 == 0.0:
        return math.hypot(ax, ay)
    t = -(ax * abx + ay * aby) / l2
    t = max(0.0, min(1.0, t))
    cx, cy = ax + t * abx, ay + t * aby
    return math.hypot(cx, cy)

# ── v9.13: 날씨 → Beaufort + 특수 효과 매핑 ─────────────────────────────────
# special_radar: Beaufort 클러터 외 추가 레이더 감쇠 (황사·농무·야간 흡수·산란)
# special_sonar: 모두 1.00 — 황사·야간은 소나 무관, Beaufort 해상 소음으로 충분
WEATHER_BEAUFORT_MAP: dict[str, dict] = {
    '맑음 (주간)':           {'beaufort': 2,  'special_radar': 1.00, 'special_sonar': 1.00},
    '맑음 (야간)':           {'beaufort': 2,  'special_radar': 0.97, 'special_sonar': 1.00},
    '흐림 (박무)':           {'beaufort': 4,  'special_radar': 1.00, 'special_sonar': 1.00},
    '황사 (봄철 황사)':      {'beaufort': 3,  'special_radar': 0.75, 'special_sonar': 1.00},
    '풍랑 (7~8등급)':        {'beaufort': 7,  'special_radar': 1.00, 'special_sonar': 1.00},
    '폭풍 (해상 악화)':      {'beaufort': 9,  'special_radar': 1.00, 'special_sonar': 1.00},
    '태풍 (9~12등급)':       {'beaufort': 11, 'special_radar': 1.00, 'special_sonar': 1.00},
    '농무 (시정 200m 이하)': {'beaufort': 2,  'special_radar': 0.82, 'special_sonar': 1.00},
    '폭풍 (야간)':           {'beaufort': 9,  'special_radar': 0.97, 'special_sonar': 1.00},
    '태풍 (야간)':           {'beaufort': 11, 'special_radar': 0.97, 'special_sonar': 1.00},
    '농무 (야간)':           {'beaufort': 2,  'special_radar': 0.79, 'special_sonar': 1.00},
    '황사 (새벽)':           {'beaufort': 3,  'special_radar': 0.73, 'special_sonar': 1.00},
}

# ── v10.1: ICAO 표준 대기 (ISA) ─────────────────────────────────────────────
def isa_atmosphere(alt_m: float) -> tuple[float, float, float]:
    """ICAO 표준 대기 대류권(0~11000m): 기온(K), 압력(Pa), 밀도(kg/m³)."""
    T0, P0 = 288.15, 101_325.0
    L, g, M, R = 0.0065, 9.80665, 0.028964, 8.31446
    h = max(0.0, min(float(alt_m), 11_000.0))
    T = T0 - L * h
    P = P0 * (T / T0) ** (g * M / (R * L))
    rho = P * M / (R * T)
    return T, P, rho

# ── v9.13: 증발 덕팅 (Evaporation Duct Height) ───────────────────────────────
# (region_key, season) → (edh_m, boost_factor)
# 덕트 내 저고도 표적: 레이더 전파가 해면을 따라 굴절 → 탐지거리 증가
# 출처: 한국 해역 대기 경계층 관측 문헌 평균값
EVAP_DUCT_DB: dict[tuple, tuple] = {
    ('EAST_SEA',     'spring'): (7,  1.15),  # 봄 동해: 기온 상승, 수증기 증가 시작
    ('EAST_SEA',     'summer'): (10, 1.25),  # 여름 동해: 덕트 높이 10m, 탐지 1.25배
    ('EAST_SEA',     'autumn'): (8,  1.18),  # 가을 동해: 여름 근접 수준 유지
    ('EAST_SEA',     'winter'): (6,  1.12),
    ('YELLOW_SEA',   'spring'): (5,  1.10),  # 서해 봄: 황사·박무 시즌, 덕트 약함
    ('YELLOW_SEA',   'summer'): (8,  1.18),  # 서해: 강수·습도로 덕트 발달 약함
    ('YELLOW_SEA',   'autumn'): (6,  1.12),
    ('YELLOW_SEA',   'winter'): (4,  1.08),
    ('KOREA_STRAIT', 'spring'): (6,  1.13),
    ('KOREA_STRAIT', 'summer'): (9,  1.20),
    ('KOREA_STRAIT', 'autumn'): (7,  1.15),
    ('KOREA_STRAIT', 'winter'): (5,  1.10),
}

# ── v9.13: 고층 바람 CEP 배율 ────────────────────────────────────────────────
# (region_key, season) → cep_multiplier
# 순항미사일 Pk = pk_base / cep_factor (CEP 증가 → 명중률 감소)
# 동해 겨울: 편서풍 강함(30~50 m/s), 서해 겨울: 북서풍 (25~40 m/s)
WIND_CEP_FACTOR: dict[tuple, float] = {
    ('EAST_SEA',     'spring'): 1.10,   # 봄: 제트기류 북상 중, 중간 강도
    ('EAST_SEA',     'summer'): 1.05,
    ('EAST_SEA',     'autumn'): 1.12,   # 가을: 제트기류 남하 시작
    ('EAST_SEA',     'winter'): 1.20,
    ('YELLOW_SEA',   'spring'): 1.10,
    ('YELLOW_SEA',   'summer'): 1.08,
    ('YELLOW_SEA',   'autumn'): 1.11,
    ('YELLOW_SEA',   'winter'): 1.15,
    ('KOREA_STRAIT', 'spring'): 1.09,
    ('KOREA_STRAIT', 'summer'): 1.06,
    ('KOREA_STRAIT', 'autumn'): 1.11,
    ('KOREA_STRAIT', 'winter'): 1.18,
}

# ── v12.5: 동적 기상 전이 ─────────────────────────────────────────────────────
# 날씨를 강도 축(주간/야간 계열)으로 정렬한 사다리. 인접 1단계씩만 전이.
WEATHER_INTENSITY_LADDER: dict[str, list] = {
    'day':   ['맑음 (주간)', '흐림 (박무)', '풍랑 (7~8등급)', '폭풍 (해상 악화)', '태풍 (9~12등급)'],
    'night': ['맑음 (야간)', '폭풍 (야간)', '태풍 (야간)'],
    'dust':  ['황사 (봄철 황사)', '흐림 (박무)', '풍랑 (7~8등급)'],  # 황사: 흐림/풍랑으로 합류
    'fog':   ['농무 (시정 200m 이하)', '흐림 (박무)', '풍랑 (7~8등급)'],  # 농무: 흐림/풍랑으로 합류
}

# 전이 간격·악화 확률 기본값 (튜닝 상수 — 작전급 엔진에서 재사용 시 조정)
WEATHER_STEP_INTERVAL_S: float = 300.0   # 5분마다 전이 판정
WEATHER_WORSEN_BASE: float     = 0.15    # 기본 악화 확률
WEATHER_IMPROVE_BASE: float    = 0.10    # 기본 호전 확률

# v12.6: 피아식별 오류 (IFF) 상수
IFF_FAIL_BASE:         float = 0.08   # C&D 통과 후 IFF 실패 기본 확률
IFF_RECHECK_DELAY_S:   float = 15.0  # IFF 실패 시 재확인 대기 (초)
IFF_FRATRICIDE_P:      float = 0.05  # IFF 실패 + 근접 아군 항공기 존재 시 오사 확률
IFF_FRATRICIDE_RANGE_M: float = 80_000.0  # 오사 체크 반경 (80km)

# (region_key, season) → (worsen_delta, improve_delta)
# 기본값에 더해지는 계절·해역 보정값 (기상청 한반도 해역 패턴)
WEATHER_TRANSITION_DB: dict[tuple, tuple] = {
    # 겨울 동해·서해: 북서계절풍 → 악화 확률↑
    ('EAST_SEA',     'winter'): ( 0.15, -0.05),
    ('YELLOW_SEA',   'winter'): ( 0.12, -0.05),
    ('KOREA_STRAIT', 'winter'): ( 0.08, -0.03),
    # 여름 동해·남해: 태풍 시즌 → 악화↑, 태풍 계열 허용
    ('EAST_SEA',     'summer'): ( 0.10,  0.00),
    ('KOREA_STRAIT', 'summer'): ( 0.08,  0.00),
    # 봄 서해: 황사 시즌 → 약간 불안정
    ('YELLOW_SEA',   'spring'): ( 0.05,  0.00),
    # 가을·봄 기타: 비교적 안정
    ('EAST_SEA',     'spring'): ( 0.02,  0.03),
    ('EAST_SEA',     'autumn'): ( 0.03,  0.05),
    ('YELLOW_SEA',   'autumn'): ( 0.02,  0.05),
    ('KOREA_STRAIT', 'spring'): ( 0.02,  0.03),
    ('KOREA_STRAIT', 'autumn'): ( 0.03,  0.05),
    # 여름 서해: 비교적 온화
    ('YELLOW_SEA',   'summer'): ( 0.05,  0.02),
}

def _weather_ladder_key(weather: str) -> str:
    """날씨 문자열 → 소속 사다리 키. 없으면 'day'."""
    for key, ladder in WEATHER_INTENSITY_LADDER.items():
        if weather in ladder:
            return key
    return 'day'

def _weather_ladder_pos(weather: str, ladder_key: str | None = None) -> tuple[list, int]:
    """날씨 문자열 → (사다리 리스트, 현재 인덱스).
    ladder_key를 주면 해당 사다리 안에서만 찾는다 — 공유 항목의 계열 이탈 방지."""
    if ladder_key and ladder_key in WEATHER_INTENSITY_LADDER:
        ladder = WEATHER_INTENSITY_LADDER[ladder_key]
        if weather in ladder:
            return ladder, ladder.index(weather)
    for ladder in WEATHER_INTENSITY_LADDER.values():
        if weather in ladder:
            return ladder, ladder.index(weather)
    return WEATHER_INTENSITY_LADDER['day'], 0

# ── v10.1: ISA 대기 굴절 + 라디오존데 4계절×5고도층 보정 ──────────────────────
# 키: (region_key, season, alt_layer)
# alt_layer: 1=100~500m / 2=500~1000m / 3=1000~3000m / 4=3000m+
# (alt_layer 0 = <100m 은 _evap_duct_factor 담당)
# 출처: 기상청 고층기상관측 연보 (포항·광주·오산 라디오존데 계절 통계)
# 대기 굴절 지수(N-unit): 여름 N≈330~340 (ISA 315 대비 +5~8%), 겨울 N≈295~310 (+2~3%)
ISA_RADIOSONDE_DB: dict[tuple, float] = {
    # ── EAST_SEA ─────────────────────────────────────────────────────────────
    ('EAST_SEA', 'spring', 1): 1.02, ('EAST_SEA', 'spring', 2): 1.04,
    ('EAST_SEA', 'spring', 3): 1.05, ('EAST_SEA', 'spring', 4): 1.03,
    ('EAST_SEA', 'summer', 1): 1.03, ('EAST_SEA', 'summer', 2): 1.06,  # 쿠로시오 수증기
    ('EAST_SEA', 'summer', 3): 1.07, ('EAST_SEA', 'summer', 4): 1.04,
    ('EAST_SEA', 'autumn', 1): 1.03, ('EAST_SEA', 'autumn', 2): 1.05,
    ('EAST_SEA', 'autumn', 3): 1.06, ('EAST_SEA', 'autumn', 4): 1.03,
    ('EAST_SEA', 'winter', 1): 1.01, ('EAST_SEA', 'winter', 2): 1.03,  # 대륙성 한기
    ('EAST_SEA', 'winter', 3): 1.03, ('EAST_SEA', 'winter', 4): 1.02,
    # ── YELLOW_SEA ────────────────────────────────────────────────────────────
    ('YELLOW_SEA', 'spring', 1): 1.02, ('YELLOW_SEA', 'spring', 2): 1.03,  # 황사 시즌
    ('YELLOW_SEA', 'spring', 3): 1.04, ('YELLOW_SEA', 'spring', 4): 1.02,
    ('YELLOW_SEA', 'summer', 1): 1.02, ('YELLOW_SEA', 'summer', 2): 1.05,  # 장마·고온다습
    ('YELLOW_SEA', 'summer', 3): 1.06, ('YELLOW_SEA', 'summer', 4): 1.04,
    ('YELLOW_SEA', 'autumn', 1): 1.02, ('YELLOW_SEA', 'autumn', 2): 1.04,
    ('YELLOW_SEA', 'autumn', 3): 1.05, ('YELLOW_SEA', 'autumn', 4): 1.03,
    ('YELLOW_SEA', 'winter', 1): 1.00, ('YELLOW_SEA', 'winter', 2): 1.02,  # 북서계절풍
    ('YELLOW_SEA', 'winter', 3): 1.02, ('YELLOW_SEA', 'winter', 4): 1.01,
    # ── KOREA_STRAIT ──────────────────────────────────────────────────────────
    ('KOREA_STRAIT', 'spring', 1): 1.02, ('KOREA_STRAIT', 'spring', 2): 1.04,
    ('KOREA_STRAIT', 'spring', 3): 1.04, ('KOREA_STRAIT', 'spring', 4): 1.03,
    ('KOREA_STRAIT', 'summer', 1): 1.03, ('KOREA_STRAIT', 'summer', 2): 1.05,  # 쓰시마 난류
    ('KOREA_STRAIT', 'summer', 3): 1.06, ('KOREA_STRAIT', 'summer', 4): 1.04,
    ('KOREA_STRAIT', 'autumn', 1): 1.02, ('KOREA_STRAIT', 'autumn', 2): 1.04,
    ('KOREA_STRAIT', 'autumn', 3): 1.05, ('KOREA_STRAIT', 'autumn', 4): 1.03,
    ('KOREA_STRAIT', 'winter', 1): 1.01, ('KOREA_STRAIT', 'winter', 2): 1.03,
    ('KOREA_STRAIT', 'winter', 3): 1.03, ('KOREA_STRAIT', 'winter', 4): 1.02,
}

# ── v10.1: 트로포스캐터 링크 보정 ────────────────────────────────────────────
# 대기 상층 난류 산란 → 수평선 너머 고고도 표적(≥1000m) 탐지거리 추가 증가
# 해양성 기류 안정할수록 산란층 발달 우수. BF6 이상 강풍 시 산란층 붕괴 → 비활성화.
TROPOSCATTER_DB: dict[tuple, float] = {
    ('EAST_SEA',     'spring'): 1.11,
    ('EAST_SEA',     'summer'): 1.16,
    ('EAST_SEA',     'autumn'): 1.13,
    ('EAST_SEA',     'winter'): 1.09,
    ('YELLOW_SEA',   'spring'): 1.09,
    ('YELLOW_SEA',   'summer'): 1.14,
    ('YELLOW_SEA',   'autumn'): 1.11,
    ('YELLOW_SEA',   'winter'): 1.07,
    ('KOREA_STRAIT', 'spring'): 1.10,
    ('KOREA_STRAIT', 'summer'): 1.15,
    ('KOREA_STRAIT', 'autumn'): 1.12,
    ('KOREA_STRAIT', 'winter'): 1.08,
}

# ── v10.2-B: 아군 SAM RCS (Phase C 적 레이더 탐지용) ────────────────────────
# 출처: 각 미사일 동체 크기·형상 기반 추정값 (m²)
_SAM_RCS: dict[str, float] = {
    'SM-3 Block IIA':  0.0008,
    'SM-6':            0.0010,
    'SM-6 Block IB':   0.0010,
    'SM-2 Block IIIB': 0.0015,
    'ESSM Block II':   0.0005,
    '해궁 (K-SAAM)':   0.0005,
    'RIM-116 RAM':     0.0003,
    'L-SAM':           0.0010,
    '천궁-II':         0.0006,
    'PAC-3 MSE':       0.0008,
}

# ── v9.12: 해역 매핑 및 지형 레이더 음영 페널티 ─────────────────────────────
# fleet_region UI 문자열 → db_ocean_acoustic 키
REGION_TO_ACOUSTIC_KEY: dict[str, str] = {
    '동해 북부': 'EAST_SEA',
    '동해 중부': 'EAST_SEA',
    '서해':       'YELLOW_SEA',
    '대한해협':   'KOREA_STRAIT',
}

# v14.3: 3D 저고도 탐지 경계 수평선 한계용 안테나·표적 가정 고도(m)
_RADAR_ANT_H_M = 25.0    # 이지스 SPY급 안테나 해발 고도
_SEA_SKIM_H_M  = 10.0    # 해면 밀착 대함미사일 비행 고도

# v14.3 지형 차폐 시각화 — 해역별 (육지 방위 rad: 동=0·북=π/2, 대표 음영각 deg)
# 출처: db_terrain.radar_shadow_reference (동해 태백 3.4°·설악 8.1° → 대표 6°)
_TERRAIN_SHADOW: dict[str, tuple] = {
    'EAST_SEA':     (math.pi,        6.0),   # 한반도 육지 서쪽 — 태백·설악산맥
    'YELLOW_SEA':   (0.0,            0.4),   # 육지 동쪽 — 낭림산맥 원거리(영향 미약)
    'KOREA_STRAIT': (math.pi / 2.0,  0.9),   # 육지 북쪽 — 소백산맥
}

# 해역별 저고도 레이더 음영 배율 (altitude_m → factor)
# 출처: db_terrain.TERRAIN_DB['radar_shadow_reference'] 음영각 기반
# 동해 3.4~8.1° (태백·설악), 서해 0.4° (낭림 원거리), 대한해협 0.9° (소백)
TERRAIN_RADAR_PENALTY: dict[str, list] = {
    'EAST_SEA': [
        (0,    0.78),   # 해면밀착 — 태백·설악 음영 강하게 받음
        (50,   0.88),
        (200,  0.93),
        (1000, 1.00),   # 산능선 위 고고도
    ],
    'YELLOW_SEA': [
        (0,    0.96),   # 낭림산맥 원거리 (0.4°), 영향 미약
        (200,  0.98),
        (1000, 1.00),
    ],
    'KOREA_STRAIT': [
        (0,    0.90),   # 소백산맥 (0.9°)
        (50,   0.94),
        (200,  0.97),
        (1000, 1.00),
    ],
}

# ── 지상 BMD 자산 상수 (이지스 어쇼어 / THAAD) ──────────────────────────────
_ASHORE_SM3_PK     = 0.82     # 지상 고정 설치 → 함정 대비 안정적 추적
_ASHORE_SM3_COST   = 35_000_000
_ASHORE_SM3_SPD_MS = 3_000
_ASHORE_COOLDOWN_S = 5.0
_ASHORE_RANGE_M    = 500_000  # SM-3 최대 교전거리
# SM-3 최소 교전고도 = 100km (v18.04.10 정정, 과거 40km).
# SM-3는 외기권 전용 요격체다(대기권 밖 진공에서만 작동하는 유일한 스탠더드 미사일).
# 공개 자료의 최소 교전고도는 100~120km이고, 대기권 경계(카르만선)가 100km다.
# 과거 40km는 대기권 한복판이라, SM-3가 물리적으로 잡을 수 없는 저고도 표적
# (KN-23 같은 준탄도 SRBM)까지 요격하게 만들었다. 함정 SM-3·지상 어쇼어·최후 폴백
# 세 경로가 모두 이 상수를 참조한다(과거엔 각각 흩어져 하드코딩).
_SM3_ALT_MIN_M     = 100_000
_ASHORE_ALT_MIN_M  = _SM3_ALT_MIN_M   # 지상 어쇼어도 같은 물리 — 별도 값을 두지 않는다

_THAAD_PK          = 0.85     # hit-to-kill 종말 단계
_THAAD_COST        = 3_000_000
_THAAD_SPD_MS      = 2_500
_THAAD_COOLDOWN_S  = 3.0
_THAAD_RANGE_M     = 200_000  # 200km 방어 반경
_THAAD_ALT_MIN_M   = 10_000   # 종말 하한 10km
_THAAD_ALT_MAX_M   = 150_000  # 종말 상한 150km

# ── v20.2a: 탄도미사일 종말 강하 ────────────────────────────────────────────
# 이 모델 이전에는 탄도 고도가 비행 내내 상수라 ICBM이 고도 1200km를 유지한 채 함대에
# 명중했고, 종말 요격층(THAAD·L-SAM·천궁-II)이 교전창에 진입할 수 없었다(SM-3 단독 방어).
# 강하각 일정 근사: 고도 = 잔여 수평거리 × tan(강하각), DB 정점고도를 상한으로 캡.
#   재진입체의 종말 강하각은 대략 40~50° → tan≈1.0(고도 ≈ 잔여거리). 이 기하 덕분에 각
#   요격층이 자기 사거리에서 자기 고도창을 만난다(SM-3 ≥40km·THAAD 10~150km·L-SAM
#   40~70km·천궁-II ≤20km가 거리와 1:1 대응).
#   ⚠ 비행 진행도(p) 기반 곡선은 실패했다 — DB 정점고도(화성-15 1200km, 로프티드 시험궤도
#   값)가 스폰 거리(~877km)보다 커서 강하각이 54°로 과도해지고, 고도가 L-SAM 창에 들어올
#   땐 이미 사거리 밖(164km > 150km)이라 하위 계층이 영구 미발현했다(실측).
# enable_ballistic_descent OFF면 무동작.
_BALLISTIC_DESCENT_TAN = 1.0

# ── v20.2a: 한국형 BMD 계층 (L-SAM 상층 / 천궁-II 하층) ─────────────────────
# L-SAM(장거리 지대공) — 공개 제원: 사거리 150km·요격고도 50~60km·Mach 4~5,
#   3단 hit-to-kill(IIR 탐색기 + DACS). PAC-3와 THAAD 사이 중간 계층.
#   ALT_MIN 40km는 교전 개시 하한이다 — 보도 요격고도(50~60km)는 유효 요격점이고
#   교전을 시작하는 하한은 그보다 낮다. (과거 주석은 이 값을 'SM-3의 40km 문턱과
#   맞물리게 하기 위함'이라 적었으나, SM-3 문턱이 실제 외기권 100km로 정정된 뒤에도
#   L-SAM 하한은 공개 제원에서 나온 값이므로 그대로 둔다 — 40~100km 구간은 L-SAM
#   40~70km와 THAAD 10~150km가 함께 덮어 빈틈이 없다.)
_LSAM_PK           = 0.80     # 국산 신형 hit-to-kill — THAAD(0.85) 소폭 하회로 보수 설정
_LSAM_COST         = 3_000_000   # 단가 미공개. THAAD($3M)급 가정 (PAC-3 $4M·천궁-II $1M 사이)
_LSAM_SPD_MS       = 1_500    # Mach 4~5
_LSAM_COOLDOWN_S   = 4.0
_LSAM_RANGE_M      = 150_000  # 사거리 150km
_LSAM_ALT_MIN_M    = 40_000
_LSAM_ALT_MAX_M    = 70_000   # 요격 상한 (보도 50~60km + 여유)

# 천궁-II(KM-SAM Block-II) — 공개 제원: 대탄도탄 교전 사거리 ~20km·요격고도 ~15km,
#   (대항공기 제원은 사거리 50km·고도 20km이나 여기 대상은 탄도·HGV뿐이라 대탄도탄 값 사용)
#   400kg급 hit-to-kill, 유도탄 단가 약 $1M(PAC-3의 1/4). 최후 종말 점방어 계층.
_CHUNGUNG_PK       = 0.75     # 하층 종말 — 고속 강하 표적이라 상층 자산보다 낮게
_CHUNGUNG_COST     = 1_000_000
_CHUNGUNG_SPD_MS   = 1_400    # Mach 4급
_CHUNGUNG_COOLDOWN_S = 2.0    # 점방어 — 연속 교전 빠름
_CHUNGUNG_RANGE_M  = 20_000   # 대탄도탄 교전 사거리 20km
_CHUNGUNG_ALT_MIN_M = 500     # 종말 점방어 — HGV 종말 침투고도(2km)까지 커버해야 최후 계층 의미
_CHUNGUNG_ALT_MAX_M = 20_000  # 요격고도 15km 내외 + 유도탄 한계 20km

# 패트리엇 PAC-3 MSE(v20.1) — 공개 제원: hit-to-kill, 이중펄스 모터로 사거리·고도 확장,
#   대탄도탄 교전 사거리 ~60km·요격 고도 ~25km, Mach 4.5+, 단가 $4.2M(FY2025 미 육군 예산 기준).
#   천궁-II(하층 점방어)와 L-SAM(상층) 사이를 메우는 중층 — 한·미 연합 종말 방어.
_PATRIOT_PK         = 0.80     # hit-to-kill + Ka대역 능동 탐색기(실전 교전 기록 보유)
_PATRIOT_COST       = 4_200_000
_PATRIOT_SPD_MS     = 1_600    # Mach 4.5+
_PATRIOT_COOLDOWN_S = 3.0
_PATRIOT_RANGE_M    = 60_000   # 대탄도탄 교전 사거리 60km
_PATRIOT_ALT_MIN_M  = 2_000
_PATRIOT_ALT_MAX_M  = 25_000   # 요격 고도 25km

# 한 위협에 지상 BMD가 동시 유도할 수 있는 최대 요격탄 수(계층 각 1발 = 층별 교전 보장)
_GROUND_BMD_MAX_SAMS = 5

# 지상 BMD 자산 owner_id → 발사 통계 키
_GROUND_BMD_STAT_KEY: dict[int, str] = {
    -1: 'ashore_sm3_fired',
    -2: 'thaad_fired',
    -3: 'lsam_fired',
    -4: 'chungung_fired',
    -5: 'patriot_fired',
}

# 다층 방어 레이어 순서: 가장 먼저 교전하는 함정 유형부터 (BMD 우선 → 방공 우선)
LAYER_ORDER    = ['KDX-III-B2', 'KDX-III-B1', 'KDX-II', 'FFX-III', 'FFX-II', 'FFX-I']
SHIP_LAYER_PRI = {t: i for i, t in enumerate(LAYER_ORDER)}

def _make_physics_wx(weather: str) -> dict:
    """
    v9.13: 날씨 문자열 → Beaufort 물리 기반 wx dict.
    db_ocean_environment 없으면 WEATHER_DB 원본 그대로 반환.
    """
    wx = dict(WEATHER_DB.get(weather, WEATHER_DB['맑음 (주간)']))
    if _OCEAN_ENV_OK:
        bf_info = WEATHER_BEAUFORT_MAP.get(weather)
        if bf_info:
            bf = bf_info['beaufort']
            wx['radar_factor'] = (
                _RADAR_BF_FACTOR.get(bf, 1.0) * bf_info['special_radar']
            )
            wx['sonar_factor'] = (
                _SONAR_BF_FACTOR.get(bf, 1.0) * bf_info['special_sonar']
            )
            wx['beaufort'] = bf
    return wx

# 포팅 C: v7 시뮬 시간 스케일 맞춤 출격 준비 시간 (전시 긴급 출격 기준)
# FRIENDLY_AIRCRAFT_DB의 sortie_time_s(평시)를 v7 700초 시뮬에 맞게 단축
_AIRCRAFT_V7_SORTIE = {
    'AW-159 와일드캣': 300,   # 5분 (평시 동일)
    'P-3C 오라이온':   600,   # 10분 (평시 40분 → 전시 긴급 출격)
    'P-8A 포세이돈':   480,   # 8분 (평시 30분 → 전시 긴급 출격)
    # v16.12: 정찰 드론은 교전 개시 전 이미 전개(사전 배치 ISR) — 조기탐지가 존재 이유라
    # 전개 대기시간을 짧게 둬 교전 초반부터 함대 탐지를 확장한다(장기체공 UAV 특성).
    'RQ-101 송골매':   60,    # 근접 전방기지 발진, 1분
    'MQ-9B 시가디언':  0,     # 장기체공 — 교전 개시 시점에 이미 초계 중
    # v18.01.18: CAP 전투기(role=='cap')는 이 목록에 없어도 FriendlyAircraftObj.__init__에서
    # 항모전단 상시 공중초계(24/7 on-station) 교리로 60s 기본값을 받는다. 평시 대기시간을
    # 그대로 두면 적기 종말 교전창(~236s)을 놓쳐 공대공이 죽은 경로가 되므로 role 기준 일반화.
}

# ════════════════════════════════════════════════════════════════════════════
#  새 DB: 아군 대함 공격 무기
# ════════════════════════════════════════════════════════════════════════════
FRIENDLY_STRIKE_DB = {
    '해성-I': {
        'speed_ms': 250,   # Mach 0.73 (아음속 순항, SSM-700K 실제 속도)
        'range_km': 150,   # SSM-700K 해성 공개 사거리 ~150km (기존 180 과대·FRIENDLY_DB와 불일치 → 통일)
        'cost_usd': 800_000,
        'pk_base':  0.80,
        'seeker':   'radar',    # v8.26: 액티브 레이더 탐색기
    },
    '해성-II': {
        'speed_ms': 250,   # Mach 0.73 (해성-I 동일 계열)
        'range_km': 250,
        'cost_usd': 2_200_000,  # PHY-10: $1.2M→$2.2M (실제 단가 반영)
        'pk_base':  0.82,
        'seeker':   'radar',
    },
    '하푼 Block II': {
        'speed_ms': 240,   # Mach 0.71
        'range_km': 140,   # RGM-84 Block II 공개 사거리 ~140km (기존 280 과대)
        'cost_usd': 1_500_000,
        'pk_base':  0.78,
        'seeker':   'radar',    # AN/DSQ-61 액티브 레이더
    },
    # SM-6 Block IB 대함 모드 (OTH 대함 공격, Link-16 유도)
    'SM-6 대함 모드': {
        'speed_ms': 1000,   # Mach 3.5
        'range_km': 370,    # OTH 사거리
        'cost_usd': 4_200_000,
        'pk_base':  0.70,
        'seeker':   'combined', # GPS+능동 레이더 복합
    },
    # 5인치 Mk.45 Mod 4 함포 (근거리 최후 레이어)
    'Mk.45 5인치 함포': {
        'speed_ms': 830,   # 포탄 초속 ~Mach 2.4
        'range_km': 24,    # 유효 사거리 24km
        'cost_usd': 2_000, # 발당 약 $2,000
        'pk_base':  0.40,  # 대함 Pk (표적이 클수록 높음)
        'seeker':   'radar',
    },
    # NEW-P1: Tomahawk Block V 초장거리 대함 타격 (US ships)
    'Tomahawk Block V': {
        'speed_ms': 250,
        'range_km': 1700,
        'cost_usd': 2_000_000,
        'pk_base':  0.80,
        'seeker':   'combined', # DSMAC 영상+GPS 복합
    },
    # KSS-III VLS 탑재 현무-3C 순항미사일 (잠수함 발사)
    '현무-3C': {
        'speed_ms': 250,    # Mach 0.73 (아음속 순항)
        'range_km': 1500,   # 현무-3C 사거리 1,500km
        'cost_usd': 6_000_000,  # PHY-11: $2M→$6M (개발비·단가 반영, 기밀 추정 $5~10M)
        'pk_base':  0.80,
        'seeker':   'combined', # GPS+INS+TERCOM 복합
    },
    # NEW-A: 현무-4 지대함 탄도미사일 (한국판 DF-21D, 지상 발사 전용)
    '현무-4 (ASBM)': {
        'speed_ms': 3000,   # 종말 Mach 8~10 (평균 비행 속도 근사)
        'range_km': 800,    # 현무-4 사거리 800km
        'cost_usd': 8_000_000,  # PHY-12: $3.5M→$8M (고성능 탄도미사일 개발단가)
        'pk_base':  0.85,   # 고속 종말 — 적 SAM 요격 극히 어려움
        'seeker':   'combined', # 레이더+적외선 복합 종말 탐색
    },
}

# ════════════════════════════════════════════════════════════════════════════
#  새 DB: 적 함정 SAM 시스템
# ════════════════════════════════════════════════════════════════════════════
ENEMY_SAM_DB = {
    'HHQ-9B':    {'range_km': 200, 'speed_ms': 1400, 'pk': 0.82, 'channels': 6},
    'HHQ-16':    {'range_km':  50, 'speed_ms': 1000, 'pk': 0.75, 'channels': 4},
    'HHQ-10':    {'range_km':   9, 'speed_ms':  680, 'pk': 0.70, 'channels': 2},
    '1130-CIWS': {'range_km':   3, 'speed_ms': 1100, 'pk': 0.65, 'channels': 1},
    # NEW-P1: 러시아 함정 SAM
    'S-300F':    {'range_km':  90, 'speed_ms': 1800, 'pk': 0.85, 'channels': 6},  # 슬라바급
    'SA-N-9':    {'range_km':  12, 'speed_ms':  900, 'pk': 0.75, 'channels': 4},  # 우달로이급
}

ENEMY_SHIP_SAM_LOADOUT = {
    '055형 대형 구축함': [
        {'name': 'HHQ-9B',    'stock': 112},
        {'name': 'HHQ-10',    'stock':  24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '052D형 구축함': [
        {'name': 'HHQ-9B',    'stock': 64},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '054A형 호위함': [
        {'name': 'HHQ-16',    'stock': 32},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    # v20.1의 054B가 이 표에 없으면 SAM·CIWS를 한 발도 못 쏜다(sam_max=0) — DB·스펙시트가
    #   "054A보다 방공이 강한 발전형"이라 선언한 함정이 실제 교전에선 054A보다 약해지는
    #   정체성 역전이 생긴다(종합 감사 발견). 공개 제원: H/AKJ-16 VLS 32셀 + HHQ-10 + Type 1130.
    '054B형 호위함': [
        {'name': 'HHQ-16',    'stock': 32},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '056형 초계함': [
        {'name': 'HHQ-10',    'stock':  8},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '022형 미사일 고속정': [
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    # NEW-P1: 신규 적 함정 SAM 탑재 목록
    '052C형 구축함 (HHQ-9)': [
        {'name': 'HHQ-9B',    'stock': 48},
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '071형 상륙함': [
        {'name': 'HHQ-10',    'stock': 24},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '우달로이급 구축함': [
        {'name': 'SA-N-9',    'stock': 64},
        {'name': '1130-CIWS', 'stock': 9999},
    ],
    '슬라바급 순양함': [
        {'name': 'S-300F',    'stock': 64},
        {'name': 'HHQ-10',    'stock': 40},   # AK-630 CIWS (HHQ-10 근사)
        {'name': '1130-CIWS', 'stock': 9999},
    ],
}

# 독립 미사일 유형 (EnemyThreatObj 대신 MissileObj로 직접 생성)
_STANDALONE_MISSILE_TYPES = ('탄도미사일', '순항미사일', '극초음속활공체', '저고도기동탄도', '대방사미사일')

# MED-5: 미사일 명칭별 함체 명중 Pk (0.80 하드코드 → 유형 기반 매핑)
# 아음속 대함미사일은 요격 쉬움(Pk 높음), 초음속·어뢰는 요격 어려움(Pk 낮음)
_MISSILE_PK_MAP = {
    # 아음속 순항 대함 (0.70~0.72)
    'YJ-83 대함미사일':          0.72,
    'YJ-62 대함미사일':          0.72,
    'YJ-83K 주력 대함미사일':    0.70,
    'YJ-8K 대함미사일':          0.72,
    'YJ-91 대함미사일':          0.70,
    'YJ-91 초음속 대함미사일':   0.70,
    'YJ-100 (장거리 순항)':      0.70,
    'CJ-10 (순항미사일)':        0.70,
    # 초음속 순항 대함 (요격 어려움: 0.65~0.68)
    'YJ-12 초음속 대함미사일':   0.65,
    'YJ-12 (초음속 순항)':       0.65,
    'YJ-18 초음속 대함미사일':   0.68,
    'YJ-18B 잠대함미사일':       0.68,
    'Kh-31A 대함미사일':         0.68,
    'Kh-31A (항공기발사 대함)':  0.68,
    'P-800 오닉스 (야혼트)':     0.65,
    # 탄도/HGV/QBM (SM-3/SM-6 전담)
    'DF-11A (단거리 탄도)':      0.75,
    'DF-15 (단거리 탄도)':       0.75,
    'DF-21D (대함 탄도)':        0.70,
    'DF-26 (중장거리 탄도)':     0.65,
    'DF-17 (극초음속 활공)':     0.65,
    'KN-23 (북한 이스칸데르)':   0.72,
    # 어뢰 (수중 유도, 회피 어려움: 0.78)
    'Yu-6 중어뢰':               0.78,
    # NEW-P1: 신규 미사일
    'DF-21D (공중발사)':         0.70,
    'Kh-32 극초음속':            0.65,
    'P-700 그라니트':            0.62,
    'P-1000 (벌칸)':             0.62,
    'Kalibr 3M54 잠대함':        0.68,
    '북극성-1 (SLBM)':           0.70,
    'Kh-31A 대함미사일':         0.68,
    # ARM: 레이더 활성 상태 추적 — 기본 Pk 높음
    'Kh-31P 대방사미사일':       0.88,
    'LD-10 대방사미사일':        0.85,
    'Kh-58U 대방사미사일':       0.90,  # 장거리 고속 — 요격 난이도 높음
    # BUG-6: _MISSILE_PK_MAP 누락 13종 추가 — 기본값 0.72로 잘못 처리되던 항목
    # 극초음속/HGV (요격 매우 어려움: 0.62~0.63)
    'YJ-21 (극초음속 대함)':          0.62,
    '킨잘 (극초음속 탄도)':           0.62,
    '지르콘 (극초음속 순항)':         0.63,
    # ICBM/IRBM (SM-3 전담, 요격 극히 어려움: 0.56~0.65)
    '화성-15 (ICBM급)':               0.58,
    '화성-17 (ICBM 개량)':            0.56,
    '화성-18 (ICBM 고체연료)':        0.58,
    '화성-12 (IRBM)':                 0.65,
    # 기동탄도/QBM (0.68)
    'KN-24 (단거리 기동탄도)':        0.68,
    # 초음속 대함 (0.68)
    'YJ-18 (초음속 대함)':            0.68,
    # 스텔스 순항 (저RCS — 탐지 어려움: 0.68)
    'Kh-101 (스텔스 순항)':           0.68,
    # 아음속 순항 (0.70~0.72)
    'Kalibr (3M14 순항미사일)':       0.70,
    '해성-3 (잠수함발사 순항)':       0.70,
    '북한 순항미사일 (화살-2)':       0.72,
}
_MISSILE_PK_DEFAULT = 0.72  # 미등록 미사일 기본값



# ════════════════════════════════════════════════════════════════════════════
#  Vec2
# ════════════════════════════════════════════════════════════════════════════
#  v10.8: LatLon — 지리 좌표계 (Vec2 완전 대체)
#  x / y 프로퍼티로 기존 시각화 코드 자동 호환
# ════════════════════════════════════════════════════════════════════════════

# 해역별 기준점 (작전 중심 LatLon)
_REGION_REF: dict[str, tuple] = {
    '동해 북부':  (40.5, 130.5),
    '동해 중부':  (37.2, 131.9),
    '서해':       (36.0, 125.0),
    '대한해협':   (34.5, 129.0),
}
_DEFAULT_REF = (37.2, 131.9)  # 기본: 동해 중부

# v15.03: 위협 접근 방위 — 적 세력 실제 발진 좌표 (lat, lon)
# 함대 위치에서 이 좌표로 bearing 계산 → 위협이 지리적 방향에서 접근.
_THREAT_ORIGIN: dict[str, tuple] = {
    'china':         (36.1, 120.3),   # 칭다오 (북해함대)
    'china_carrier': (33.5, 129.5),   # 대한해협 남측 — 중국 항모 동해 진입 경유(현실적 경로)
    'nk_west':       (38.7, 125.4),   # 남포 (북한 서해)
    'nk_east':       (39.2, 127.4),   # 원산 (북한 동해)
    'russia':        (43.1, 131.9),   # 블라디보스토크 (태평양함대)
    'japan':         (35.5, 135.4),   # 마이즈루 (동해측 해자대)
}
# 적 편대 프리셋 없음(랜덤·커스텀 모드)일 때 해역별 기본 세력
_REGION_DEFAULT_ORIGIN: dict[str, str] = {
    '동해 북부': 'russia',
    '동해 중부': 'japan',
    '서해':       'china',
    # 대한해협은 STRAIT_BEARING이 처리
}
ORIGIN_SPREAD_DEG = 35.0   # 위협 접근 방위 분산 폭(±)


def _preset_origin(preset_name: str) -> str:
    """적 편대 프리셋 이름 → 세력키(_THREAT_ORIGIN). 'nk'는 함대 해역에서 서/동 결정."""
    n = preset_name or ''
    if '북한' in n:
        return 'nk'
    if '러시아' in n or '쓰시마' in n:   # 쓰시마 봉쇄=Tu-22M3·슬라바·킬로(러 자산)
        return 'russia'
    if '항모전단' in n:                  # 랴오닝·산둥·푸젠 → 대한해협 경유 진입
        return 'china_carrier'
    return 'china'                       # 그 외 중국 계열 기본

# 시뮬 시작 시 _set_region_ref() 로 갱신
_ref_lat: float = _DEFAULT_REF[0]
_ref_lon: float = _DEFAULT_REF[1]
_ref_m_per_deg_lat: float = 110_540.0
_ref_m_per_deg_lon: float = 111_320.0 * math.cos(math.radians(_DEFAULT_REF[0]))

def _set_region_ref(region: str):
    """시뮬 시작 시 해역에 맞는 기준점 설정."""
    global _ref_lat, _ref_lon, _ref_m_per_deg_lon
    _ref_lat, _ref_lon = _REGION_REF.get(region, _DEFAULT_REF)
    _ref_m_per_deg_lon = 111_320.0 * math.cos(math.radians(_ref_lat))


class LatLon:
    """
    지리 좌표 (위도/경도). x/y 프로퍼티로 기존 Cartesian 코드 자동 호환.
    dist_to(): Haversine 거리(m). bearing_to(): 동=0 북=π/2 평면 방위각(rad).
    """
    __slots__ = ('lat', 'lon')

    def __init__(self, lat: float = 0.0, lon: float = 0.0):
        self.lat = float(lat)
        self.lon = float(lon)

    @classmethod
    def from_xy(cls, x_m: float, y_m: float) -> 'LatLon':
        """미터 좌표(기준점 기준) → LatLon."""
        return cls(
            _ref_lat + y_m / _ref_m_per_deg_lat,
            _ref_lon + x_m / _ref_m_per_deg_lon,
        )

    # ── 시각화 호환 프로퍼티 ──────────────────────────────────────────────────
    @property
    def x(self) -> float:
        """기준점으로부터 동방향(m) — 시각화·거리 계산 호환."""
        return (self.lon - _ref_lon) * _ref_m_per_deg_lon

    @x.setter
    def x(self, v: float):
        self.lon = _ref_lon + v / _ref_m_per_deg_lon

    @property
    def y(self) -> float:
        """기준점으로부터 북방향(m) — 시각화·거리 계산 호환."""
        return (self.lat - _ref_lat) * _ref_m_per_deg_lat

    @y.setter
    def y(self, v: float):
        self.lat = _ref_lat + v / _ref_m_per_deg_lat

    # ── 거리·방위 ─────────────────────────────────────────────────────────────
    def dist_to(self, other: 'LatLon') -> float:
        """평면 근사 거리 (m) — 500km 이내 오차 0.1% 미만."""
        dy = (other.lat - self.lat) * _ref_m_per_deg_lat
        dx = (other.lon - self.lon) * _ref_m_per_deg_lon
        return math.sqrt(dx * dx + dy * dy)

    def bearing_to(self, other: 'LatLon') -> float:
        """방위각 (rad) — 동=0, 북=π/2 (x/y 평면 관례 유지)."""
        return math.atan2(other.y - self.y, other.x - self.x)

    def move_toward(self, target: 'LatLon', speed_ms: float, dt: float) -> bool:
        """target 방향으로 이동. 도달했으면 True 반환."""
        d = self.dist_to(target)
        step = speed_ms * dt
        if d <= step:
            self.lat, self.lon = target.lat, target.lon
            return True
        angle = self.bearing_to(target)
        self.lat += math.sin(angle) * step / _ref_m_per_deg_lat
        self.lon += math.cos(angle) * step / _ref_m_per_deg_lon
        return False

    def copy(self) -> 'LatLon':
        return LatLon(self.lat, self.lon)

    def __repr__(self) -> str:
        return f"LatLon({self.lat:.4f}°N, {self.lon:.4f}°E)"


# 하위 호환 별칭 — 기존 코드가 Vec2를 참조할 경우 대비
Vec2 = LatLon


# ════════════════════════════════════════════════════════════════════════════
#  MissileObj
# ════════════════════════════════════════════════════════════════════════════
class MissileObj:
    """
    mtype:
      'enemy_strike'   — 적 대함/대지 공격 (아군 함정 또는 독립 미사일 위협)
      'friendly_strike'— 아군 대함/대잠 공격
      'friendly_sam'   — 아군 SAM (적 미사일·항공기 요격)
      'enemy_sam'      — 적 SAM (아군 미사일 요격)
    """
    _id_counter = 0

    def __init__(self, mtype: str, name: str, pos: Vec2,
                 target,
                 speed_ms: float, pk_base: float,
                 owner_id: int, t_spawn: float = 0.0):
        MissileObj._id_counter += 1
        self.uid       = f"M{MissileObj._id_counter:04d}"
        self.mtype     = mtype
        self.name      = name
        self.pos       = pos.copy()
        self.target    = target
        self.speed_ms  = speed_ms
        self.pk_base   = pk_base
        self.owner_id  = owner_id
        self.t_spawn   = t_spawn

        self.alive       = True
        self.intercepted = False
        self.hit         = False
        self.t_intercept: Optional[float] = None
        self.t_hit:       Optional[float] = None

        # 독립 미사일 위협 속성 (탄도/HGV/QBM 등 — _build_enemies에서 설정)
        self.altitude_m:   float = 0.0
        self.is_hgv:       bool  = False
        self.is_qbm:       bool  = False
        self.is_ballistic: bool  = False

        # 포팅 B: 전술 속성
        self.terminal_evasion_factor: float = 1.0     # 종말 회피 계수 (< 20km 적용)
        self.is_torpedo:              bool  = False   # 어뢰 여부 (기만기/회피 판정용)
        self.is_arm:                  bool  = False   # 대방사미사일 여부 (레이더 직격)
        self.seeker:                  str   = 'radar' # v8.26: 탐색기 유형 (채프/플레어 대응)

        # EngagementAnalysis 추적 (A-1)
        self.intercept_weapon: Optional[str] = None  # 격추 무기명
        self.intercept_km:     float         = 0.0   # 격추 거리 (km)
        self.detect_m:         float         = 0.0   # 탐지 시 거리 (m)
        self.enemy_info:       dict          = {}     # ENEMY_DB 원본

        # v10.4: CEC 중계 교전 여부 (자체 탐지 불가 → 아군 네트워크 데이터 의존)
        self.cec_relay: bool = False

        # v12.1: 비례항법(PNG) 종말 유도 상태 (enable_png=True일 때만 사용)
        # heading_rad: 현재 속도 벡터 방위. spawn 시 표적 방향으로 초기화(정조준 출발).
        self.heading_rad: float = self.pos.bearing_to(target.pos) if target else 0.0
        self.png_active:  bool  = False          # 종말 PNG 진입 여부 (로그·판정 분기)
        self.min_miss_m:  float = float('inf')   # substep 최근접 통과 거리 (miss distance)
        self.jink_phase:  float = 0.0            # 표적 횡기동 위상 (적 대함미사일용)
        self._jink_t:     float = -1.0           # 표적 jink 전진이 적용된 마지막 틱 (다중 SAM 중복 방지)
        # v16.1 ESM→ARM: stale 조준 좌표(레이더 OFF 시 마지막 포착 위치). None이면 실시간 표적 추적.
        self.arm_aim_pos = None
        # v16.6 전자 좌표 기만: 가짜 조준 좌표(아군 ECM 교란). None이면 미기만. 종말권 1회 판정.
        self.decoy_aim_pos = None
        self.decoy_aim_checked = False

    def update(self, dt: float) -> bool:
        """1 tick 이동. alive=False 설정 금지 — 요격/피격 판정은 엔진이 담당."""
        if not self.alive:
            return False
        # v16.1 ESM→ARM: arm_aim_pos(stale 조준 좌표)가 있으면 그쪽으로 유도(레이더 OFF 시
        # 마지막 포착 위치). None이면 기존 실시간 표적 추적(bit-identical).
        # 유도 좌표 우선순위: ARM stale > 전자 좌표 기만 > 실시간 표적
        if self.arm_aim_pos is not None:
            _aim = self.arm_aim_pos
        elif self.decoy_aim_pos is not None:
            _aim = self.decoy_aim_pos
        else:
            _aim = self.target.pos
        arrived = self.pos.move_toward(_aim, self.speed_ms, dt)
        if arrived:
            self.hit = True
        return arrived

    @classmethod
    def reset_counter(cls):
        cls._id_counter = 0


# ════════════════════════════════════════════════════════════════════════════
#  EnemyThreatObj — 적 플랫폼 위협 통합 (항공기 / 수상함 / 잠수함)
#  독립 미사일(탄도/순항/HGV/QBM)은 _build_enemies()에서 MissileObj로 생성.
# ════════════════════════════════════════════════════════════════════════════
_ENEMY_MUNITION_INF = 99999   # 무장 유한화 미대상(1발성 위협 등) 무제한 표현


def _et_col(col: str):
    """EnemyThreatObj proxy 속성 → 엔진 _et_* 컬럼(행 i)으로 위임하는 property.
    v12.7 단계1: 데이터는 엔진 컬럼에 있고 객체는 (엔진, 행 i)만 든다."""
    def _get(self):
        return getattr(self._eng, col)[self._i]
    def _set(self, v):
        getattr(self._eng, col)[self._i] = v
    return property(_get, _set)


class _ThreatPosView:
    """EnemyThreatObj.pos — 엔진 컬럼(_et_lat/_et_lon)의 행 i를 LatLon처럼 노출.
    별도 좌표 저장 없이 컬럼을 직접 읽고 쓴다 (배열이 정본). LatLon과 동일 인터페이스."""
    __slots__ = ('_eng', '_i')

    def __init__(self, eng, i: int):
        self._eng = eng
        self._i   = i

    @property
    def lat(self): return self._eng._et_lat[self._i]
    @lat.setter
    def lat(self, v): self._eng._et_lat[self._i] = v

    @property
    def lon(self): return self._eng._et_lon[self._i]
    @lon.setter
    def lon(self, v): self._eng._et_lon[self._i] = v

    @property
    def x(self): return (self.lon - _ref_lon) * _ref_m_per_deg_lon
    @x.setter
    def x(self, v): self.lon = _ref_lon + v / _ref_m_per_deg_lon

    @property
    def y(self): return (self.lat - _ref_lat) * _ref_m_per_deg_lat
    @y.setter
    def y(self, v): self.lat = _ref_lat + v / _ref_m_per_deg_lat

    def dist_to(self, other) -> float:
        dy = (other.lat - self.lat) * _ref_m_per_deg_lat
        dx = (other.lon - self.lon) * _ref_m_per_deg_lon
        return math.sqrt(dx * dx + dy * dy)

    def bearing_to(self, other) -> float:
        return math.atan2(other.y - self.y, other.x - self.x)

    def move_toward(self, target, speed_ms: float, dt: float) -> bool:
        d = self.dist_to(target)
        step = speed_ms * dt
        if d <= step:
            self.lat, self.lon = target.lat, target.lon
            return True
        angle = self.bearing_to(target)
        self.lat += math.sin(angle) * step / _ref_m_per_deg_lat
        self.lon += math.cos(angle) * step / _ref_m_per_deg_lon
        return False

    def copy(self) -> 'LatLon':
        return LatLon(self.lat, self.lon)


class EnemyThreatObj:
    """v12.7 단계1: SoA 전환 — 실데이터는 엔진 _et_* 컬럼에 저장하고, 이 객체는
    (엔진, 행 인덱스 _i)만 들고 속성 접근을 컬럼으로 위임하는 얇은 뷰(proxy).
    생성은 TimeStepEngine._new_threat()가 담당(컬럼 append + 파생값 계산).
    행은 삭제하지 않으므로(기존도 alive=False만) 인덱스·id(et)가 시뮬 내내 안정."""
    __slots__ = ('_eng', '_i', '_posview')

    _id_counter = 0

    _HP_MAP = {
        '전투기': 1, '폭격기': 1, '전폭기': 1,
        '잠수함': 3,
        '고속정': 2, '초계함': 2,
        '호위함': 3, '구축함': 4,
        '항모': 5,
    }

    def __init__(self, eng, i: int):
        self._eng     = eng
        self._i       = i
        self._posview = _ThreatPosView(eng, i)

    # ── 위치 (컬럼 뷰) ────────────────────────────────────────────────────────
    @property
    def pos(self):
        return self._posview
    @pos.setter
    def pos(self, v):
        self._eng._et_lat[self._i] = v.lat
        self._eng._et_lon[self._i] = v.lon

    # ── 컬럼 위임 속성 (name·preset_name은 동일 컬럼) ──────────────────────────
    name                  = _et_col('_et_preset')
    preset_name           = _et_col('_et_preset')
    info                  = _et_col('_et_info')
    uid                   = _et_col('_et_uid')
    speed_ms              = _et_col('_et_speed')
    altitude_m            = _et_col('_et_alt')
    category              = _et_col('_et_category')
    threat_type           = _et_col('_et_threat_type')
    is_aircraft           = _et_col('_et_is_aircraft')
    is_ship               = _et_col('_et_is_ship')
    is_sub                = _et_col('_et_is_sub')
    _r50_cache            = _et_col('_et_r50_cache')
    hp                    = _et_col('_et_hp')
    high_value_target     = _et_col('_et_high_value')
    carrier_aircraft      = _et_col('_et_carrier_aircraft')
    carrier_wave_interval = _et_col('_et_carrier_wave_interval')
    carrier_wing          = _et_col('_et_carrier_wing')         # 함재 전투기 항공단 규모(0=무제한)
    wing_launched         = _et_col('_et_wing_launched')        # 항모를 떠나 있는 기체 수(공중+손실, 재무장 복귀 시 감소)
    carrier_owner         = _et_col('_et_carrier_owner')        # 함재기의 모항 인덱스(None=함재기 아님)
    _last_wave_t          = _et_col('_et_last_wave_t')
    sam_inventory         = _et_col('_et_sam_inventory')
    munition_remaining    = _et_col('_et_munition_remaining')   # 공격 무장 잔여(무장 유한화)
    sam_max_channels      = _et_col('_et_sam_max_channels')
    sam_channels_used     = _et_col('_et_sam_channels_used')
    alive                 = _et_col('_et_alive')
    intercepted           = _et_col('_et_intercepted')
    hit_count             = _et_col('_et_hit_count')
    hit_by                = _et_col('_et_hit_by')
    has_fired             = _et_col('_et_has_fired')
    t_intercept           = _et_col('_et_t_intercept')
    is_retreating         = _et_col('_et_is_retreating')
    retreat_pos           = _et_col('_et_retreat_pos')
    reattack_count        = _et_col('_et_reattack_count')
    max_reattacks         = _et_col('_et_max_reattacks')
    next_fire_t           = _et_col('_et_next_fire_t')          # Phase 5.2: 이 시각 전엔 발사 금지(재장전 쿨다운)
    ecm_power             = _et_col('_et_ecm_power')
    hidden_until          = _et_col('_et_hidden_until')
    ambush_revealed       = _et_col('_et_ambush_revealed')
    counter_evade_until   = _et_col('_et_counter_evade_until')  # v16.01.03: 능동 핑 역탐지 회피 만기
    contact_lost_until    = _et_col('_et_contact_lost_until')   # v16.1: 능동 핑 노출→회피 잠항 도주(접촉 단절) 만기
    last_contact_t        = _et_col('_et_last_contact_t')       # v20.5(B-3): 마지막 소나 접촉 시각(datum 갱신)

    # ── 메서드 ────────────────────────────────────────────────────────────────
    def take_hit(self, weapon_name: str, t: float):
        eng, i = self._eng, self._i
        eng._et_hit_count[i] += 1
        eng._et_hit_by[i].append((weapon_name, t))
        eng._et_hp[i] -= 1
        if eng._et_hp[i] <= 0:
            eng._et_alive[i] = False

    def select_sam(self, missile_dist_m: float) -> Optional[str]:
        """수상함용: 사거리 내 SAM 선택 (장거리 우선)"""
        inv = self._eng._et_sam_inventory[self._i]
        for sam_name in ['HHQ-9B', 'HHQ-16', 'HHQ-10', '1130-CIWS']:
            if inv.get(sam_name, 0) <= 0:
                continue
            sam = ENEMY_SAM_DB.get(sam_name)
            if sam and missile_dist_m <= sam['range_km'] * 1000:
                return sam_name
        return None

    @classmethod
    def reset_counter(cls):
        cls._id_counter = 0


# ════════════════════════════════════════════════════════════════════════════
#  FriendlyShipObj
# ════════════════════════════════════════════════════════════════════════════
class FriendlyShipObj:
    def __init__(self, name: str, ship_type: str, pos: Optional[Vec2] = None):
        self.name        = name
        self.ship_type   = ship_type
        spec             = SHIP_DB[ship_type]
        self.display     = spec['display']
        self.sensor_km   = spec['sensor_km']
        # max_channels는 @property로 계산 (부분 피해 반영)
        self.pos         = pos or LatLon.from_xy(0, 0)

        self.inventory   = spec['default_inventory'].copy()

        self.strike_inventory: dict = {}

        self.is_submarine  = spec.get('is_submarine', False)
        self.is_unmanned   = spec.get('is_unmanned', False)  # v16.12: 무인정(인명손실 0·저생존)
        self.nation        = spec.get('nation', 'KOR')  # v8.16: 한미 연합 작전

        # LOW-9: 함정 유형별 HP (함종별 내탄성 차등. 기존 고정값 5 → 실제 격침 내성 반영)
        _hp_map = {
            'KDX-III-B2': 5, 'KDX-III-B1': 5, 'KDX-II': 4,
            'FFX-I': 3, 'FFX-II': 3, 'FFX-III': 3,
            'DDG-51': 5, 'CG-47': 5, 'CVN': 8,
            'LPD': 3, 'SSN': 3, 'LST': 3, 'AO': 2,
            'KSS-I': 2, 'KSS-II': 2, 'KSS-III': 3,
            'USV': 2, 'UUV': 1,   # v16.12 무인정: 저생존(소형·경장갑)
        }
        self.hp            = _hp_map.get(ship_type, 4)
        self._max_hp       = self.hp
        self.alive         = True
        self.hit_count     = 0
        self.hits_taken    = 0  # MC 집계용 피격 횟수 (hit_count와 동일)
        self.total_cost    = 0.0
        self.channels_used = 0
        self.decoy_stock   = 4  # 포팅 B: AN/SLQ-25 음향 기만기 기본 재고
        # 피해 연동: 서브시스템별 성능 배율 (0.0~1.0)
        self.radar_factor   = 1.0  # 레이더 계통 (탐지거리 반영)
        self.channel_factor = 1.0  # 무장 채널 계통 (SAM 동시 교전 수 반영)
        self.speed_factor   = 1.0  # 추진 계통 (회피 기동 효율 반영)
        self.disabled_weapons: set = set()  # 무장 피탄으로 사용 불가 무기 목록
        self._vls_depleted = False  # 탄약 완전 소진 플래그
        self.radar_off_until: float = 0.0  # ARM 회피 전술: 이 시각까지 레이더 OFF
        self.eccm_factor: float = spec.get('eccm_factor', 0.40)  # v8.26: 재밍 상쇄 능력

        # v16.5: 해안 고정 방어 포대 — 육상이라 침몰·회피 없음(HP 누적 시 무력화).
        self.is_shore_battery = ship_type in ('CRAM', 'CSAM')
        # v12.4: 동적 침수·복원력 모델 (enable_flooding ON 경로). 잠수함은 제외.
        self._surv = SHIP_SURVIVABILITY.get(ship_type)
        self.is_sub_hull = self.is_submarine or ship_type in ('SSN', 'KSS-I', 'KSS-II', 'KSS-III')
        self.flood       = 0.0    # 현재 침수율 (0~1)
        self.flood_rate  = 0.0    # 침수 진행 속도 (/초)
        self.sunk_by_flood = False  # 침수로 침몰 (HP 즉사와 구분, 집계용)

        # v15.08.01: 연료(자원 지속성 모델, BattleEngine 전용) — 정규화 0~1. 원자력은 무제한.
        self.fuel_max = 1.0
        self.fuel     = 1.0
        _end = SHIP_ENDURANCE.get(ship_type)
        if _end is None:                       # CVN·SSN 원자력 추진 = 사실상 무제한
            self._nuclear = True
            self._fuel_burn_per_s = 0.0
        else:
            self._nuclear = False
            _rng_nm, _cruise_kt = _end
            _t_ship_h = _rng_nm / _cruise_kt   # 함정별 실측 작전가능시간[h]
            _burn_1800 = _FUEL_STD_BURN_1800 * (_FUEL_T_REF_H / _t_ship_h)
            _burn_1800 = max(_FUEL_BURN_MIN, min(_FUEL_BURN_MAX, _burn_1800))
            self._fuel_burn_per_s = _burn_1800 / 1800.0
        self._prev_fuel_pos = None             # 회피 발생 판정용 직전 위치(BattleEngine._fuel_update)
        self._maneuver_fuel_used = 0.0         # 회피 누적 추가 소모(상한 _FUEL_MANEUVER_CAP)
        # v15.09.03: 군수지원함(AOE·AO)은 함대 보급용 화물 연료 보유 — RAS로 전투함에 급유
        self.ras_reserve = _RAS_RESERVE.get(ship_type, 0.0)
        # v17.1: RAS 탄약 재보급(enable_ras_rearm) — 화물 SAM 발수(연료와 별도)·재보급 float 누적기·
        # 초기 방어 재고 고정본(재보급 상한·트리거 기준). _build_friendly의 cfg override 후 재저장됨.
        self.ras_ammo_reserve = _RAS_AMMO_RESERVE.get(ship_type, 0.0)
        self._rearm_acc = 0.0
        self._rearming  = False   # 히스테리시스: TRIGGER서 True, TARGET 도달 시 False
        self._initial_defense_stock = dict(self.inventory)

        # v17.2: 지향성 에너지 무기(레이저) — enable_laser_dew ON 경로에서만 교전.
        # laser_kw=0(미장착)이면 _laser_defense 진입 시 무동작 → 회귀 bit-identical.
        self.laser_kw          = float(SHIP_POWER.get(ship_type, {}).get('laser_kw', 0.0))
        self._laser_target_uid = None   # 현재 조사 중인 표적 uid (dwell lock)
        self._laser_dwell_acc  = 0.0    # 누적 조사 에너지(kJ). E_kill 도달 시 격추

    @property
    def max_channels(self):
        # 채널 계산 시 부분 피해 반영
        spec = SHIP_DB[self.ship_type]
        return max(1, int(spec['max_channels'] * self.channel_factor))

    @property
    def operational(self) -> bool:
        return self.alive

    def available(self, wpn: str) -> int:
        """무기 가용 재고 (무장 피탄 비활성화 반영)."""
        if wpn in self.disabled_weapons:
            return 0
        return self.inventory.get(wpn, 0)

    def take_hit(self, weapon_name: str, t: float, subsystem: str | None = None):
        """
        subsystem: 'radar' | 'propulsion' | 'weapons' | None
          None이면 서브시스템 피해 없이 HP만 감소 (하위호환).
        """
        self.hit_count  += 1
        self.hits_taken += 1
        self.hp -= 1

        if subsystem == 'radar':
            self.radar_factor = max(0.20, self.radar_factor * 0.50)
        elif subsystem == 'propulsion':
            self.speed_factor = max(0.30, self.speed_factor * 0.60)
        elif subsystem == 'weapons':
            self.channel_factor = max(0.40, self.channel_factor * 0.70)
            # VLS 탄약 25% 직접 손실 (수직발사관 피탄)
            _vls_wpns = ['SM-3 Block IIA', 'SM-6', 'SM-6 Block IB',
                         'SM-2 Block IIIB', 'ESSM Block II']
            for _w in _vls_wpns:
                if self.inventory.get(_w, 0) > 0:
                    self.inventory[_w] = max(0, self.inventory[_w] - max(1, self.inventory[_w] // 4))
            # 무기 비활성화 (SAM·CIWS 전체 대상)
            _candidates = [w for w in [
                'SM-3 Block IIA', 'SM-6', 'SM-6 Block IB',
                'SM-2 Block IIIB', 'ESSM Block II',
                'RIM-116 RAM', '해궁 (K-SAAM)', 'CIWS-II (Phalanx)',
            ] if self.inventory.get(w, 0) > 0 and w not in self.disabled_weapons]
            if _candidates:
                self.disabled_weapons.add(random.choice(_candidates))
            # VLS 고갈 플래그 갱신
            if not self._vls_depleted and all(self.inventory.get(w, 0) == 0 for w in _vls_wpns):
                self._vls_depleted = True

        if self.hp <= 0:
            self.alive = False

    def take_arm_hit(self, t: float):
        """대방사미사일(ARM) 레이더 직격 — 레이더 계통 심각 손상."""
        self.hit_count  += 1
        self.hits_taken += 1
        self.hp -= 1
        # ARM: 레이더 전파를 역추적해 레이더 안테나 직격 → 탐지거리 대폭 감소
        self.radar_factor = max(0.10, self.radar_factor * 0.30)
        if self.hp <= 0:
            self.alive = False

    def add_flooding(self, breach: float):
        """v12.4: 수선하 피격 → 격실 침수 누적 (enable_flooding ON 경로).
        breach: 즉시 침수율 증가량(0~1). 잠수함·생존성 데이터 없으면 무시."""
        if self._surv is None or self.is_sub_hull or self.is_shore_battery:
            return
        self.flood      += breach
        self.flood_rate += breach * FLOOD_INFLOW_K   # 파공 유입 속도 가산


# ════════════════════════════════════════════════════════════════════════════
#  레이더 역할 세분화: 수색 → 추적 → 유도 3단계 파이프라인
# ════════════════════════════════════════════════════════════════════════════
class SearchRadar:
    """편대 수색 레이더 — 신규 위협 초기 탐지 지연 (SPY-1D 빔 회전 주기 6초)."""
    scan_interval: float = 6.0

    def __init__(self):
        self._detect_t: dict[str, float] = {}

    def try_detect(self, uid: str, t: float, radar_off: bool = False) -> bool:
        """처음 호출 시 0~scan_interval 랜덤 딜레이 후 탐지 확정.
        radar_off=True 이면 미탐지 위협은 새로 탐지 불가."""
        if uid in self._detect_t:
            return t >= self._detect_t[uid]
        if radar_off:
            return False
        self._detect_t[uid] = t + random.uniform(0.0, self.scan_interval)
        return False


class TrackRadar:
    """편대 추적 레이더 — 동시 추적 채널 및 획득 지연 관리 (SPY-1D: 18채널)."""
    max_ch:   int   = 18
    acq_time: float = 3.0  # 신규 위협 추적 획득 시간 (초)

    def __init__(self):
        self._track_t: dict[str, float] = {}
        self._ch_used: int = 0

    def try_track(self, uid: str, t: float) -> bool:
        """uid 추적 가능 여부. 처음이면 채널 할당 후 acq_time 딜레이."""
        if uid in self._track_t:
            return t >= self._track_t[uid]
        if self._ch_used >= self.max_ch:
            return False   # 채널 포화 — 추적 거부
        self._ch_used += 1
        self._track_t[uid] = t + self.acq_time
        return False       # 획득 중

    def release(self, uid: str):
        """위협 소멸 시 추적 채널 반환."""
        if uid in self._track_t:
            del self._track_t[uid]
            self._ch_used = max(0, self._ch_used - 1)


# ════════════════════════════════════════════════════════════════════════════
#  포팅 C: FriendlyAircraftObj — 항공 자산 (헬기 / 해상초계기)
# ════════════════════════════════════════════════════════════════════════════
class FriendlyAircraftObj:
    """
    함재 헬기(base_type='ship') / 육상초계기(base_type='land') 통합 클래스.
    매 tick _aircraft_asw()에서 잠수함 표적 확인 후 어뢰 투하.

    v9.8 탐지 상태 머신:
      헬기  — dipping  : idle → hovering(dip_hover_s) → detect_check → [fire | retry | abandon]
      초계기 — sonobuoy : idle → detect_check → [fire | retry(retry_s) | abandon]
    """
    def __init__(self, name: str, home_pos: 'Vec2'):
        self.name              = name
        self.info              = FRIENDLY_AIRCRAFT_DB[name]
        self.home_pos          = home_pos.copy()
        # BUG-7: v7 시뮬(700초)에 맞는 전시 긴급 출격 시간 적용 (engine_core.py 평시값 무시)
        # v18.01.18: CAP 전투기는 항모전단 상시 공중초계 교리라 기종별 목록에 없어도 60s 기본값
        # (평시값으로 폴백하면 적기 종말 교전창을 놓쳐 공대공이 죽은 경로가 되는 재발 방지).
        _cap_default = 60.0 if self.info.get('aircraft_role') == 'cap' else self.info['sortie_time_s']
        self.t_available       = float(_AIRCRAFT_V7_SORTIE.get(name, _cap_default))
        self.payload_remaining = self.info['payload_cnt']
        # v10.6: 공대함 strike payload (KF-21 해성-II 등)
        self.strike_payload_remaining = self.info.get('cap_strike_payload_cnt', 0)
        self.sorties           = 0
        self.total_cost        = 0.0
        # v9.8: 탐지 상태 머신
        self._asw_phase:    str   = 'idle'  # idle | transit | hovering | cooldown
        self._dip_until:    float = 0.0     # 호버링 종료 시각 (dipping 전용)
        self._next_attempt: float = 0.0     # 다음 탐지 시도 가능 시각
        self._detect_fails: int   = 0       # 현재 표적 누적 탐지 실패 수
        self._search_target = None          # 현재 수색 중인 표적 (EnemyThreatObj)
        # v20.5(B-3): 포기한 표적별 재접촉 금지 시각 {표적 uid: t_until}.
        # 이게 없으면 '표적 포기'가 포기가 아니다 — 실패 카운터가 0으로 리셋된 채 idle로
        # 돌아가므로 다음 틱에 같은 잠수함을 처음 보듯 다시 탐색해 사실상 무한 재탐색이 된다.
        self._asw_giveup: dict = {}
        self._strike_cooldown_until: float = 0.0  # v10.6: AAS 교전 쿨다운
        # v16.12: 정찰 드론(aircraft_role='recon') 전용 상태
        self._recon_lost:      bool  = False  # 격추되면 True → 탐지 중계 중단
        self._recon_next_roll: float = 0.0    # 다음 격추 판정 가능 시각


# ════════════════════════════════════════════════════════════════════════════
#  SimFrame
# ════════════════════════════════════════════════════════════════════════════
class SimFrame:
    __slots__ = ('t', 'friendly_ships', 'enemy_ships', 'missiles', 'events',
                 'ship_channels')

    def __init__(self, t: float):
        self.t              = t
        self.friendly_ships = []  # [(name, x, y, alive, hp)]
        self.enemy_ships    = []  # [(uid, preset, x, y, alive, hp, alt_m)]
        self.missiles       = []  # [(uid, x, y, mtype, name, alt_m)]
        self.events         = []  # [str]
        self.ship_channels  = []  # [(name, channels_used, max_channels)]


# ════════════════════════════════════════════════════════════════════════════
#  TimeStepEngine
# ════════════════════════════════════════════════════════════════════════════
class TimeStepEngine:
    """
    매 DT(1초)마다 실행 순서:
      1. 위치 갱신 (적 위협 접근/이탈, 미사일 비행)
      2. 적 발사 조건 확인 (함정/항공기/잠수함)
      3. 아군 TEWA — 방어 (적 미사일·항공기 요격)
      4. 아군 TEWA — 공격 (수상함: 해성/하푼, 잠수함: 홍상어/청상어)
      5. 적 TEWA   — SAM 방어 (수상함만, 아군 미사일 요격)
      6. 교전 결과 판정
      7. 프레임 기록
    """

    def __init__(self, cfg: dict, step_cb=None):
        self.cfg     = cfg
        self.t       = 0.0
        # 시뮬 루프 시간 상한(초). 단발 교전은 MAX_SIM_TIME 그대로(동작 보존),
        # 지속 전장(BattleEngine)은 작전급 horizon에 맞춰 상향 오버라이드(Phase 5.1).
        self._sim_time_cap = MAX_SIM_TIME
        self._step_cb       = step_cb   # (t, t_max, alive, vls, last_log) — 단일 시뮬 진행 콜백
        self._phase_times: dict = {}    # 단계별 누적 실행 시간 (perf_counter 기반)
        # v10.7: 전술 의사결정 모드
        self._tactical_pause_cb = None  # SimWorker가 주입하는 콜백 (state) → choice dict
        self._tactical_interval  = cfg.get('tactical_interval', 30)  # 일시정지 간격(초)

        # v15.1: 적응형 전술 AI — ai_tactic='adaptive'일 때만 동작
        self._munition_limit   = bool(cfg.get('enable_munition_limit', True))  # 적 공격 무장 유한화
        # v16.1 ESM→ARM 역탐지: 아군 레이더 방사 중일 때만 적 ESM이 포착 → ARM 실시간 유도.
        # 레이더 OFF면 마지막 포착 위치(stale)로 유도돼 명중 급감. 기본 OFF면 기존 동작 보존.
        self._emcon_arm        = bool(cfg.get('enable_esm_arm', False))
        # v20.5(B-2) 표적 난이도: 요격 Pk에 표적 속도·RCS를 반영. v18.05.07에서 정규 승격·기본 ON
        # — 표적의 속도·크기를 전혀 보지 않는 요격(마하 3 소형 ARM = 아음속 대형 표적)이 비물리라
        # 기본 OFF 자체가 오류였다. 토글은 비교·디버그용으로 유지.
        self._target_difficulty = bool(cfg.get('enable_target_difficulty', True))
        # v16.01.03 능동 소나 핑 역탐지: 능동 소나로 적 잠수함을 탐지하면 잠수함도 핑을 역탐지해
        # 은닉 해제·어뢰 반격 앞당김 + 회피 기동(접촉 급감). 기본 OFF면 기존 동작 보존.
        self._sonar_emcon      = bool(cfg.get('enable_sonar_emcon', False))
        # v20.5(B-3) 대잠 재탐색 제한: '표적 포기' 후 재접촉 쿨다운을 실제로 적용한다.
        # 이것이 없으면 대잠 항공기가 무한 재탐색해 탐지가 사실상 보장되고, 잠수함이 접촉을
        # 끊는 행위(EMCON 회피)가 아무 이득이 없다 → 능동 소나 딜레마가 성립 못 함.
        # 기본 OFF면 기존 동작 보존(실험적).
        self._asw_contact_limit = bool(cfg.get('enable_asw_contact_limit', False))
        # v16.1 ASW 전진 초계: 대잠 항공기가 함대 수동 대기 대신 전개 사거리 안의 잠수함으로
        # 전진(transit) → 탐지권 도착 후 교전. 기본 OFF면 기존(기함 기준 탐지) 보존.
        self._asw_forward      = bool(cfg.get('enable_asw_forward', False))
        # v16.3 사이버전: 적의 데이터링크 변조(아군 요격 Pk↓)·CIC 마비(아군 탐지↓)와
        # 아군 반격의 적 레이더 교란(적 발사 Pk↓). 주기적 침투 굴림으로 성공 시 일정시간
        # 지속(은밀 — 경고 없이 발현·해제). 기본 OFF면 굴림 자체를 건너뜀(회귀 bit-identical).
        self._cyber            = bool(cfg.get('enable_cyber_warfare', False))
        self._cyber_interval   = float(cfg.get('cyber_interval_s', 60.0))   # 침투 굴림 주기
        self._cyber_p          = float(cfg.get('cyber_success_p', 0.30))    # 채널별 침투 성공 확률
        self._cyber_effect_s   = float(cfg.get('cyber_effect_s', 120.0))    # 성공 시 효과 지속
        self._cyber_last_t     = -999.0   # 마지막 침투 굴림 시각
        self._dl_corrupt_until = -1.0     # 데이터링크 변조(아군 요격 Pk↓) 만기
        self._cic_blind_until  = -1.0     # CIC 마비(아군 탐지거리↓) 만기
        self._enemy_jam_until  = -1.0     # 레이더 교란 반격(적 발사 Pk↓) 만기
        # v16.6 전자 좌표 기만: 아군 ECM이 적 레이더 표시 위치를 교란 → 적 대함미사일이 종말권
        # 진입 시 확률적으로 가짜 좌표로 유도(명중 급감). 기본 OFF면 무동작·random 미소비.
        self._coord_deception      = bool(cfg.get('enable_coord_deception', False))
        self._coord_decep_range_m  = float(cfg.get('coord_deception_range_km', 15.0)) * 1000.0
        self._coord_decep_rate     = float(cfg.get('coord_deception_rate', 0.5))     # 종말 기만 성공 확률
        self._coord_decep_offset_m = float(cfg.get('coord_deception_offset_m', 250.0))  # 가짜 좌표 이격
        # v16.7 기뢰전: 작전 해역 진입 시 함정별 확률적 기뢰 접촉(계류·해저·자항 3종 차등).
        # 소해 지원 보유 시 접촉 확률 경감. 기본 OFF면 무동작·random 미소비(회귀 bit-identical).
        self._mine_threat          = bool(cfg.get('enable_mine_threat', False))
        self._mine_density         = float(cfg.get('mine_density', 0.3))      # 기뢰 위협도 0~1
        self._minesweeping         = bool(cfg.get('enable_minesweeping', False))  # 소해 지원
        # v16.2 극초음속 활공 궤적: HGV가 고정 고도가 아니라 활공(완만 하강)→종말 급강하로
        # 고도가 변해, 같은 HGV가 비행 단계별로 다른 요격 층(외기권 SM-3 → 대기권 내
        # SM-6 Block IB)으로 전환된다. 기본 OFF면 altitude_m 고정 → 회귀 bit-identical.
        # v20.5: 둘 다 정규 승격 + 기본 ON. 활공·강하는 '선택적 전술 옵션'이 아니라 위협의
        # 물리적 실제다 — OFF는 극초음속이 60km, 탄도가 1,200km 고도를 유지한 채 함대에
        # 명중하는 비물리 상태이고, 그 상태에선 고도 교전창을 가진 종말 요격층(패트리엇·
        # 천궁-II 등)이 구조적으로 영원히 발사하지 못한다. 토글은 비교·디버그용으로 남긴다.
        self._hgv_glide        = bool(cfg.get('enable_hgv_glide', True))
        self._bal_descent      = bool(cfg.get('enable_ballistic_descent', True))
        # v16.12 정찰 드론(A-1): 무인 ISR 드론이 수평선 너머(OTH) 표적을 함대 데이터링크로
        # 중계 → 생존 중 함대 실효 레이더 탐지거리에 recon_detect_bonus_km 가산.
        # 기본 OFF면 recon 역할 드론이 편성 안 돼 _recon_bonus_km=0 유지 → 회귀 bit-identical.
        self._recon_drone      = bool(cfg.get('enable_recon_drone', False))
        self._recon_bonus_km   = 0.0     # 매 tick _aircraft_recon()이 갱신하는 유효 탐지 확장량
        # v16.12 무인 함정(A-2): USV·UUV를 함대에 편성. UUV=소해(기뢰 접촉 경감)+대잠 피켓,
        # USV=대함 피켓+RAM 점방어. 무인이라 손실 시 인명피해 0(friendly_ships_lost 제외).
        # 기본 OFF면 무인정 미편성 → 피켓 보너스 0·소해 무영향 → 회귀 bit-identical.
        self._unmanned_assets  = bool(cfg.get('enable_unmanned_assets', False))
        self._usv_surf_bonus_km = 0.0    # 매 tick USV 생존 시 대함 탐지 확장량
        self._uuv_asw_bonus_km  = 0.0    # 매 tick UUV 생존 시 대잠 탐지 확장량
        # v16.12 적 무인기 군집(트랙 B): ON 시 적 편대에 자폭 드론 군집을 추가 편성해
        # 함대 SAM·CIWS 채널을 포화시킨다(비대칭 소모). 기본 OFF면 편대 불변 → 회귀 bit-identical.
        self._drone_swarm      = bool(cfg.get('enable_drone_swarm', False))
        self._drone_swarm_size = int(cfg.get('drone_swarm_size', 40))
        self._enforce_wing_cap = False   # 항모 항공단 발진 총량 상한 (전장 모드만 ON — BattleEngine서 설정)
        self._adaptive_ai      = (cfg.get('ai_tactic') == 'adaptive')
        self._adaptive_mode    = 'saturation'   # 초기 전술 (포화)
        self._adaptive_last_t  = -999.0         # 마지막 재평가 시각
        # v15.01.03 슬라이딩 윈도우 — 직전 평가 대비 델타로 최근 요격률 산출
        self._adaptive_prev_fired = 0
        self._adaptive_prev_itc   = 0
        self._adaptive_switches   = 0           # 전술 전환 횟수(결과 표시용)
        # v16.13.02 트랙 C: 함정 자율 교전 + 기함 격침 시 CEC 지휘 저하
        self._cec_degraded_until  = 0.0         # 기함 격침 후 CEC relay 저하 만료 시각
        self._prev_primary_id     = None        # 직전 틱 기함 id — 교체(격침) 감지용
        self._command_handovers   = 0           # 지휘권 인수 횟수(로그·검증용)

        # BUG-2 fix: LatLon.from_xy() 기준점을 _build_*() 호출 전에 설정
        _set_region_ref(cfg.get('fleet_region', '동해 북부'))
        self._log_entries: list = []
        self._tick_events:  list = []
        self._detected_log_set: set = set()   # 첫 탐지 로깅 완료한 위협 uid
        self._thermo_cache: dict = {}          # 수심→수온약층 보정 메모(해역·계절 시뮬 내 불변)

        self._mc_mode: bool = bool(cfg.get('mc_mode', False))

        # sim_seed 적용 (재현 보장) — python random + numpy 동시 고정
        seed = cfg.get('sim_seed', None)
        if seed:
            random.seed(int(seed))
            np.random.seed(int(seed))

        MissileObj.reset_counter()
        EnemyThreatObj.reset_counter()

        # 레이더 3단계 파이프라인
        self.search_radar = SearchRadar()  # 1단계: 수색 (탐지 지연)
        self.track_radar  = TrackRadar()   # 2단계: 추적 (채널 한계 + 획득 지연)
        # 3단계: 유도 채널은 FriendlyShipObj.channels_used / max_channels 로 관리

        # C&D 딜레이: {target_id → t_fire_allowed}
        # 탐지 시각 + cd_time_s + confirm_time_s + uniform(2,10)s 이후 발사 허용
        self._cd_fire_time: dict = {}
        # VLS 연속 발사 간격: {ship_id → t_last_vls}
        self._vls_last_fire: dict = {}
        # v8.26: 생존 적 항공 전력 중 최대 ECM 재밍 강도 (에어리어 재밍)
        self._active_ecm: float = 0.0

        # stats / wx 먼저 초기화 (build 함수에서 참조 가능)
        self.stats = {
            'total_threats':           0,
            'intercepted_threats':     0,
            'friendly_hits':           0,
            'enemy_hits':              0,
            'friendly_ships_lost':     0,
            'enemy_ships_destroyed':   0,
            'laser_kills':             0,   # v17.2: 레이저(DEW) 격추 수(드론·자폭정·아음속미사일)
            'total_cost':              0.0,
            'aircraft_sorties':        0,
            # 포팅 D: REQ 판정용
            'peak_concurrent_threats': 0,
            't_first_fire':            -1.0,
            'total_missiles_fired':    0,
            # v8.16: 한미 연합 기여도
            'kor_shots':               0,
            'usa_shots':               0,
            'kor_cost':                0.0,
            'usa_cost':                0.0,
            # v9.11: 지상 BMD 자산 발사 횟수
            'ashore_sm3_fired':        0,
            'thaad_fired':             0,
            # v20.2a: 한국형 BMD 계층 발사 횟수
            'lsam_fired':              0,
            'chungung_fired':          0,
            'patriot_fired':           0,   # v20.1: 패트리엇 PAC-3 MSE
            # v12.6: IFF 오류
            'iff_failures':            0,
            'iff_fratricide':          0,
            # v16.7: 기뢰전
            'mines_struck':            0,
            'ships_lost_to_mine':      0,
            # v16.12: 정찰 드론 격추 손실
            'recon_losses':            0,
            # v16.12: 무인 함정(USV·UUV) 손실 (인명피해 0 — friendly_ships_lost와 분리)
            'unmanned_lost':           0,
            # v17.1: RAS 탄약 재보급된 주요 SAM 총 발수 (지속 전장 전용 — 단발은 0 유지)
            'ras_missiles_resupplied': 0,
        }
        weather = cfg.get('weather', '맑음 (주간)')
        self.wx = _make_physics_wx(weather)  # v9.13: Beaufort 물리값 override

        # v12.5: 동적 기상 전이 상태
        self._wx_dyn_enabled: bool  = bool(cfg.get('enable_weather_dynamics', False))
        self._wx_next_check:  float = WEATHER_STEP_INTERVAL_S
        self._wx_trend:       str   = cfg.get('weather_trend', '자동')
        self._wx_ladder_key:  str   = _weather_ladder_key(weather)  # 계열 친화성 보존
        # v12.6: IFF 오류
        self._iff_enabled: bool = bool(cfg.get('enable_iff', False))

        # v9.3: 아군 공격 임무 격침 기록
        self.strike_log: list = []

        # v9.4: VLS 탄약 고갈 시각 기록 {ship_name: t_depletion}
        self.vls_depletion_t: dict = {}
        # v9.4: 지상 발사 자산 재고
        self.ground_inv: dict = {
            '현무-4 (ASBM)':   cfg.get('hyunmoo4_stock',    0),
            'SM-3 (어쇼어)':   cfg.get('ashore_sm3_stock',  0),
            'THAAD 요격탄':    cfg.get('thaad_stock',        0),
            'L-SAM':           cfg.get('lsam_stock',         0),
            '천궁-II':         cfg.get('chungung_stock',     0),
            'PAC-3 MSE':       cfg.get('patriot_stock',      0),   # v20.1
        }
        self._ground_last_fire: float  = -999.0  # 현무-4 마지막 발사
        self._ashore_last_fire: float  = -999.0  # 어쇼어 SM-3 마지막 발사
        self._thaad_last_fire:  float  = -999.0  # THAAD 마지막 발사
        self._lsam_last_fire:     float = -999.0  # L-SAM 마지막 발사
        self._chungung_last_fire: float = -999.0  # 천궁-II 마지막 발사
        self._patriot_last_fire:  float = -999.0  # 패트리엇 PAC-3 마지막 발사(v20.1)
        self._ground_cost: float = 0.0

        # NEW-A: 혼합 시나리오 파도 지연 스폰 큐 [(spawn_t, spec_dict), ...]
        self._pending_threats: list = []

        self.friendly_ships: List[FriendlyShipObj]    = self._build_friendly()
        self.missiles:       List[MissileObj]         = []
        # 위협 추적 표/Funnel/타임라인용 — 요격·명중으로 self.missiles에서 빠진
        # 적 공격 미사일을 누적 보관 (객체 참조라 최종 상태가 그대로 반영). 단일 시뮬만.
        self._retired_strikes: List[MissileObj]       = []
        self._retired_uids:    set                    = set()
        self._init_threat_columns()   # v12.7 단계1: SoA 컬럼 (반드시 _build_enemies 전)
        self.enemy_threats:  List[EnemyThreatObj]     = self._build_enemies()
        self.aircraft:       List[FriendlyAircraftObj] = self._build_aircraft()
        self.frames:         List[SimFrame]            = []

    # ── 편대 구성 ─────────────────────────────────────────────────────────────

    def _build_friendly(self) -> List[FriendlyShipObj]:
        # v15.2: 임의 편성(fleet_custom) 지원 — list of {name,type}가 있으면 프리셋 대신 사용.
        # 없으면 기존 프리셋 경로 그대로 → 회귀 bit-identical.
        custom = self.cfg.get('fleet_custom')
        if custom:
            preset = custom
        else:
            preset_name = self.cfg.get('fleet_preset', '단독 작전')
            preset = FLEET_PRESETS.get(preset_name, FLEET_PRESETS['단독 작전'])
        ships = []
        for spec in preset:
            s = FriendlyShipObj(spec['name'], spec['type'])
            if s.is_submarine:
                # 잠수함: SHIP_DB default_strike_inventory 사용 (해성/하푼 설정값 무시)
                s.strike_inventory = SHIP_DB[spec['type']].get('default_strike_inventory', {}).copy()
            else:
                s.strike_inventory = {
                    '해성-II':       self.cfg.get('haesong2_stock', 8),
                    '해성-I':        self.cfg.get('haesong1_stock', 0),
                    '하푼 Block II': self.cfg.get('harpoon_stock',  4),
                }
            # 포팅 A: 방어 무기 재고 수동 설정 (설정 없으면 SHIP_DB 기본값 유지)
            _def_map = [
                ('SM-3 Block IIA',  'sm3_stock'),
                ('SM-6',            'sm6_stock'),
                ('SM-2 Block IIIB', 'sm2_stock'),
                ('RIM-116 RAM',     'ram_stock'),
                ('홍상어 (대잠)',    'hongsango_stock'),
                ('청상어 (경어뢰)', 'cheongsango_stock'),
                ('Mk.46 경어뢰',    'mk46_stock'),
            ]
            for wpn, cfg_key in _def_map:
                if cfg_key in self.cfg and wpn in s.inventory:
                    s.inventory[wpn] = self.cfg[cfg_key]
            # 포팅 B: 기만기 재고 수동 설정
            if 'decoy_stock' in self.cfg:
                s.decoy_stock = self.cfg['decoy_stock']
            # NEW-AW: 함정별 위치 분산 (KDX-III 중심, KDX-II 3km, FFX 5km 기준 반경)
            if self.cfg.get('enable_dmo', False):
                # DMO: 전 함정을 광역 분산 링에 결정적 배치 — 적 집중 포화 회피(이득) +
                # CEC 상호 엄호 약화(대가, 기존 _friendly_defense 거리 판정이 자동 발현).
                # 기함도 중심에서 이탈시켜 표적 분산. 결정론 보존(각도 결정적, random 미소비).
                spread_m = self.cfg.get('dmo_spread_km', 80.0) * 1000.0
                angle = math.radians(len(ships) * (360.0 / max(len(preset), 1)))
                s.pos = LatLon.from_xy(math.cos(angle) * spread_m, math.sin(angle) * spread_m)
            else:
                radius = self._FORMATION_RADIUS.get(spec['type'], 3_000)
                if radius > 0:
                    angle = math.radians(len(ships) * (360.0 / max(len(preset), 1)))
                    s.pos = LatLon.from_xy(math.cos(angle) * radius, math.sin(angle) * radius)
            # NEW-XX: 랜덤 배치 옵션 — 각 함정에 임의 오프셋 추가
            if self.cfg.get('enable_random_placement', False):
                spread_m = self.cfg.get('random_spread_km', 5.0) * 1000.0
                rnd_angle = random.uniform(0, 2 * math.pi)
                rnd_r     = random.uniform(0, spread_m)
                s.pos = LatLon.from_xy(
                    s.pos.x + math.cos(rnd_angle) * rnd_r,
                    s.pos.y + math.sin(rnd_angle) * rnd_r,
                )
            # v17.1: cfg 재고 override 반영된 확정 초기재고 고정(RAS 재보급 상한·트리거 기준)
            s._initial_defense_stock = dict(s.inventory)
            ships.append(s)

        # v16.12: 무인 함정(USV·UUV) 편성 — 토글 ON일 때만 함대 전방 피켓으로 추가.
        # OFF면 이 블록 전체를 건너뜀 → 함대 구성 불변·RNG 미소비(회귀 bit-identical).
        if self.cfg.get('enable_unmanned_assets', False):
            _picket_m = 12_000.0   # 함대 전방 12km 피켓선(결정적 배치 — random 미소비)
            for _uidx, _utype in enumerate(('USV', 'UUV')):
                if _utype not in SHIP_DB:
                    continue
                u = FriendlyShipObj(self.cfg.get(f'{_utype.lower()}_name', _utype), _utype)
                u.strike_inventory = {}
                # 전방(적 접근축 방향 근사 +x)에 좌우로 벌려 배치
                u.pos = LatLon.from_xy(_picket_m, (_uidx - 0.5) * 6_000.0)
                ships.append(u)
        return ships

    def _init_threat_columns(self):
        """v12.7 단계1: 적 위협 SoA 컬럼 초기화 (행=위협 1기, 열=속성). _new_threat가 append.
        numpy 변환은 2단계 hot 경로에서 — 단계1은 파이썬 리스트(동적 증설 단순)."""
        self._et_uid: list                  = []
        self._et_preset: list               = []
        self._et_info: list                 = []
        self._et_lat: list                  = []
        self._et_lon: list                  = []
        self._et_speed: list                = []
        self._et_alt: list                  = []
        self._et_category: list             = []
        self._et_threat_type: list          = []
        self._et_is_aircraft: list          = []
        self._et_is_ship: list              = []
        self._et_is_sub: list               = []
        self._et_r50_cache: list            = []
        self._et_hp: list                   = []
        self._et_high_value: list           = []
        self._et_carrier_aircraft: list     = []
        self._et_carrier_wave_interval: list = []
        self._et_carrier_wing: list         = []
        self._et_wing_launched: list        = []
        self._et_carrier_owner: list        = []
        self._et_last_wave_t: list          = []
        self._et_sam_inventory: list        = []
        self._et_sam_max_channels: list     = []
        self._et_sam_channels_used: list    = []
        self._et_munition_remaining: list   = []   # 공격 무장 잔여(무장 유한화)
        self._et_alive: list                = []
        self._et_intercepted: list          = []
        self._et_hit_count: list            = []
        self._et_hit_by: list               = []
        self._et_has_fired: list            = []
        self._et_t_intercept: list          = []
        self._et_is_retreating: list        = []
        self._et_retreat_pos: list          = []
        self._et_reattack_count: list       = []
        self._et_max_reattacks: list        = []
        self._et_next_fire_t: list          = []   # Phase 5.2: 재장전 쿨다운 만기 시각(기본 0=즉시)
        self._et_ecm_power: list            = []
        self._et_hidden_until: list         = []
        self._et_ambush_revealed: list      = []
        self._et_counter_evade_until: list  = []   # v16.01.03: 능동 핑 역탐지 시 회피 기동 만기 시각
        self._et_contact_lost_until: list   = []   # v16.1: 능동 핑 노출 후 회피 잠항 도주(접촉 단절) 만기
        self._et_last_contact_t: list       = []   # v20.5(B-3): 마지막 소나 접촉 시각(datum 성장 기준)

    def _new_threat(self, preset_name: str, pos: 'LatLon') -> EnemyThreatObj:
        """적 플랫폼 위협 1기 생성 — ENEMY_DB 파생값 계산 후 _et_* 컬럼에 행 추가, proxy 반환.
        (구 EnemyThreatObj.__init__ 로직을 SoA 컬럼 기록으로 이전. RNG 미소비 → 결정론 유지.)"""
        EnemyThreatObj._id_counter += 1
        info    = ENEMY_DB[preset_name].copy()
        cat     = info.get('category', '대함')
        ttype   = info.get('type', '')
        is_ship = (cat == '대함')
        if is_ship:
            loadout = ENEMY_SHIP_SAM_LOADOUT.get(preset_name, [])
            sam_inv = {item['name']: item['stock'] for item in loadout}
            sam_max = sum(ENEMY_SAM_DB[n]['channels']
                          for n in sam_inv if n in ENEMY_SAM_DB)
        else:
            sam_inv, sam_max = {}, 0
        _db_hp = info.get('hp', None)
        _is_air = ttype in ('전투기', '폭격기', '전폭기')

        self._et_uid.append(f"ET{EnemyThreatObj._id_counter:03d}")
        self._et_preset.append(preset_name)
        self._et_info.append(info)
        self._et_lat.append(pos.lat)
        self._et_lon.append(pos.lon)
        self._et_speed.append(info['speed_ms'])
        self._et_alt.append(info.get('altitude_m', 0))
        self._et_category.append(cat)
        self._et_threat_type.append(ttype)
        self._et_is_aircraft.append(_is_air)
        self._et_is_ship.append(is_ship)
        self._et_is_sub.append(cat == '대잠')
        self._et_r50_cache.append({})
        self._et_hp.append(_db_hp if _db_hp is not None else EnemyThreatObj._HP_MAP.get(ttype, 2))
        self._et_high_value.append(info.get('high_value_target', False))
        self._et_carrier_aircraft.append(info.get('carrier_aircraft', None))
        self._et_carrier_wave_interval.append(info.get('carrier_wave_interval', 0))
        self._et_carrier_wing.append(info.get('carrier_air_wing', 0))
        self._et_wing_launched.append(0)
        self._et_carrier_owner.append(None)
        self._et_last_wave_t.append(0.0)
        self._et_sam_inventory.append(sam_inv)
        # 공격 무장 잔여 — 탑재량 목록에 없으면 무제한(1발성 미사일 위협 등). 발사 시 차감.
        self._et_munition_remaining.append(ENEMY_MUNITION.get(preset_name, _ENEMY_MUNITION_INF))
        self._et_sam_max_channels.append(sam_max)
        self._et_sam_channels_used.append(0)
        self._et_alive.append(True)
        self._et_intercepted.append(False)
        self._et_hit_count.append(0)
        self._et_hit_by.append([])
        self._et_has_fired.append(False)
        self._et_t_intercept.append(None)
        self._et_is_retreating.append(False)
        self._et_retreat_pos.append(None)
        self._et_reattack_count.append(0)
        self._et_max_reattacks.append(1 if _is_air else 0)
        self._et_next_fire_t.append(0.0)
        self._et_ecm_power.append(info.get('ecm_power', 0.0))
        self._et_hidden_until.append(0.0)
        self._et_ambush_revealed.append(False)
        self._et_counter_evade_until.append(0.0)
        self._et_contact_lost_until.append(0.0)
        self._et_last_contact_t.append(0.0)

        return EnemyThreatObj(self, len(self._et_uid) - 1)

    def _build_enemies(self) -> List[EnemyThreatObj]:
        """
        플랫폼(항공기/수상함/잠수함) → EnemyThreatObj
        독립 미사일(탄도/순항/HGV/QBM) → MissileObj (self.missiles에 직접 추가)
        """
        # 포팅 A: 적군 편대 모드 (단일/프리셋/커스텀/랜덤/혼합)
        mode = self.cfg.get('enemy_fleet_mode', 'custom')
        if mode == 'preset':
            fleet_cfg = ENEMY_FLEET_PRESETS.get(
                self.cfg.get('enemy_fleet_preset', ''), [])
        elif mode == 'random':
            fleet_cfg = generate_random_enemy_fleet(
                difficulty=self.cfg.get('enemy_fleet_difficulty', '보통'),
                seed=self.cfg.get('enemy_fleet_seed', None))
        elif mode == 'mixed':
            # NEW-A: 혼합 시나리오 — 1파(delay=0)만 즉시 spawn, 나머지는 _pending_threats
            scenario_name = self.cfg.get('mixed_scenario', '')
            scenario = MIXED_ATTACK_SCENARIOS.get(scenario_name, {})
            fleet_cfg = []
            for wave in scenario.get('waves', []):
                delay = wave.get('delay_s', 0)
                if delay == 0:
                    fleet_cfg.extend(wave.get('threats', []))
                else:
                    for spec in wave.get('threats', []):
                        self._pending_threats.append((float(delay), dict(spec)))
        else:
            fleet_cfg = self.cfg.get('enemy_fleet', [])

        detect_km      = self.cfg.get('detect_km', 200)
        surface_det_km = self.cfg.get('surface_detect_km', self.cfg.get('detect_km', 45))
        sub_det_km     = self.cfg.get('sub_detect_km', 50)
        primary    = self._primary()  # 독립 미사일 표적

        # ── 적 전술 AI 전처리 ────────────────────────────────────────────────
        _ai_tactic = self.cfg.get('ai_tactic', None)
        _FAST_TYPES = {'탄도미사일', '극초음속활공체'}

        if _ai_tactic == 'saturation':
            # 채널 포화: 아군 총 교전 채널 ×1.5 가 될 때까지 위협 수 증폭
            total_ch = max(sum(s.max_channels for s in self.friendly_ships if s.alive), 1)
            target   = int(total_ch * 1.5)
            current  = max(sum(s.get('count', 1) for s in fleet_cfg), 1)
            if target > current:
                scale    = target / current
                fleet_cfg = [{'preset': s['preset'],
                               'count':  max(1, round(s.get('count', 1) * scale))}
                              for s in fleet_cfg]

        elif _ai_tactic == 'stagger':
            # 시차 공격: 고속(탄도·HGV≥1500 m/s) 선발 → 중속+30초 → 저속+60초
            new_fleet: list = []
            for spec in fleet_cfg:
                info  = ENEMY_DB.get(spec.get('preset', ''), {})
                speed = info.get('speed_ms', 300)
                ttype = info.get('type', '')
                if speed >= 1500 or ttype in _FAST_TYPES:
                    new_fleet.append(spec)                             # 즉시
                elif speed >= 600:
                    self._pending_threats.append((30.0, dict(spec)))   # +30초
                else:
                    self._pending_threats.append((60.0, dict(spec)))   # +60초
            # 즉시 위협이 아예 없으면 첫 항목은 즉시 등장
            fleet_cfg = new_fleet if new_fleet else fleet_cfg[:1]

        elif _ai_tactic == 'exploit_weakness':
            # 약점 공략: 단일 방향 집중 (다방위 억제)
            self.cfg['enable_multibearing'] = False

        # v16.12 무인기 군집(트랙 B): 자폭 드론 군집을 편대 말미에 추가(기존 위협 스폰 순서·
        # RNG 보존 — 드론은 뒤에 붙어 idx가 이어짐). OFF면 미추가 → 회귀 bit-identical.
        # 다방위(multibearing) 병용 시 드론이 여러 섹터로 분산돼 360° 포화가 된다.
        if self._drone_swarm and self._drone_swarm_size > 0 and '자폭 드론 군집' in ENEMY_DB:
            fleet_cfg = list(fleet_cfg) + [
                {'preset': '자폭 드론 군집', 'count': self._drone_swarm_size}]

        threats: List[EnemyThreatObj] = []
        total = sum(s.get('count', 1) for s in fleet_cfg)
        idx = 0

        # v9.14: 해협 시나리오 — fleet_region='대한해협' 시 자동 활성화
        # 위협 접근 방위를 동/서수도 방향으로 제한 (multibearing 설정 무시)
        _strait_active = (self.cfg.get('fleet_region', '') == '대한해협')
        _strait_type   = self.cfg.get('strait_type', 'korea_west')
        if _strait_active:
            _strait_bases = [
                math.radians(b) for b in STRAIT_BEARING.get(_strait_type, [270.0])
            ]
            _n_strait = len(_strait_bases)

        # enable_multibearing: ON → 2~4개 방위 섹터로 분산 접근
        #                       OFF → 모두 단일 방향(기본 0°)
        # (해협 시나리오 활성 시 아래 분기 대신 _strait_bases 사용)
        _multibearing = self.cfg.get('enable_multibearing', False)
        # v15.03: 적 세력 발진 좌표 기반 접근 방위 (다방위·해협 아닐 때)
        _origin_base = None
        if not _strait_active:
            if _multibearing:
                _n_sectors = min(4, max(2, total))
                _sector_bases = [
                    math.radians(i * (360 / _n_sectors))
                    for i in range(_n_sectors)
                ]
            else:
                _origin_base = self._threat_origin_bearing()
                if _origin_base is None:   # 세력 미해결 → 기존 랜덤 폴백
                    _single_bearing = math.radians(random.uniform(0, 360))

        # 적 편대 전술 기동 — 초기 배치 오프셋
        # 'v_formation': V자 대형 (선두 1기 + 양익)
        # 'encirclement': 포위 기동 (원형 배치)
        # None / 기타: 기본 (bearing 분산)
        _tactics = self.cfg.get('enemy_tactics', None)

        for spec in fleet_cfg:
            name  = spec['preset']
            count = spec.get('count', 1)
            if name not in ENEMY_DB:
                continue
            info  = ENEMY_DB[name]
            ttype = info.get('type', '')

            for _ in range(count):
                if _strait_active:
                    # v9.14: 해협 방위 ± STRAIT_SPREAD_DEG 내 랜덤 배치
                    base = _strait_bases[idx % _n_strait]
                    bearing_rad = base + math.radians(
                        random.uniform(-STRAIT_SPREAD_DEG, STRAIT_SPREAD_DEG))
                elif _multibearing:
                    sector = idx % _n_sectors
                    bearing_rad = _sector_bases[sector] + math.radians(
                        random.uniform(-15, 15))
                elif _origin_base is not None:
                    # v15.03: 적 세력 발진 방위 ± ORIGIN_SPREAD_DEG
                    bearing_rad = _origin_base + math.radians(
                        random.uniform(-ORIGIN_SPREAD_DEG, ORIGIN_SPREAD_DEG))
                else:
                    bearing_rad = _single_bearing

                # BUG-3 연계: 수상함은 대함 레이더 탐지거리(45km)에서 시작
                # 항공·독립미사일은 대공 탐지거리, 잠수함은 소나 탐지거리 유지
                # v9.6: 기습 잠수함은 ambush_start_km (기본 20km) 에서 시작 — 이미 탐지권 내
                if info.get('is_ambush') and info.get('category') == '대잠':
                    start_m = info.get('ambush_start_km', 20) * 1000
                elif info.get('category') == '대잠':
                    start_m = sub_det_km * 1000
                elif info.get('category') == '대함':
                    start_m = surface_det_km * 1000
                elif info.get('is_arm'):
                    # v16.01.02: ARM은 발사 항공기의 SEAD 발사점(사거리 근처)에서 출발.
                    # 대공 탐지거리(~880km) 스폰은 비현실(요격 시간 8배) → 사거리 90%에서 스폰.
                    start_m = info.get('missile_range_km', 110) * 1000 * 0.9
                else:
                    start_m = detect_km * 1000

                pos = LatLon.from_xy(
                    math.cos(bearing_rad) * start_m,
                    math.sin(bearing_rad) * start_m,
                )

                if ttype in _STANDALONE_MISSILE_TYPES:
                    # 독립 미사일 위협: MissileObj로 직접 생성
                    _tgt = self._pick_target(is_torpedo=False, threat_pos=pos)
                    m = MissileObj(
                        mtype    = 'enemy_strike',
                        name     = name,
                        pos      = pos,
                        target   = _tgt,
                        speed_ms = info['speed_ms'] * self.cfg.get('threat_spd_scale', 1.0),
                        pk_base  = _MISSILE_PK_MAP.get(name, _MISSILE_PK_DEFAULT),  # MED-5
                        owner_id = -1,
                        t_spawn  = 0.0,
                    )
                    m.altitude_m             = float(info.get('altitude_m', 0))
                    m.is_hgv                 = bool(info.get('is_hgv', False))
                    m.is_qbm                 = bool(info.get('is_qbm', False))
                    m.is_ballistic           = (ttype == '탄도미사일')
                    m.is_arm                 = bool(info.get('is_arm', False))
                    m.terminal_evasion_factor = info.get('missile_terminal_evasion', 1.0)
                    m.is_torpedo             = False
                    # 3D 시각화용: 포물선 궤도 계산에 사용할 초기 거리·정점 고도 저장
                    m._init_dist  = m.pos.dist_to(_tgt.pos)
                    m._peak_alt_m = m.altitude_m  # DB 고도 = 정점 고도
                    m.detect_m    = m._init_dist   # A-1: 스폰 거리 = 탐지 거리 근사
                    m.enemy_info  = info.copy()    # A-1: ENEMY_DB 원본
                    self.missiles.append(m)
                    self.stats['total_threats'] += 1
                else:
                    et = self._new_threat(name, pos)
                    et.speed_ms *= self.cfg.get('threat_spd_scale', 1.0)
                    # 전술 기동 대형: V자 or 포위 초기 배치 오프셋
                    if _tactics == 'v_formation':
                        # 선두(idx=0)는 앞쪽, 나머지는 V자 양익
                        if idx == 0:
                            et.pos.x += math.cos(bearing_rad) * (-5_000)
                            et.pos.y += math.sin(bearing_rad) * (-5_000)
                        else:
                            wing_side = 1 if (idx % 2 == 0) else -1
                            perp = bearing_rad + math.pi / 2
                            wing_dist = (idx // 2 + 1) * 3_000
                            et.pos.x += math.cos(perp) * wing_dist * wing_side
                            et.pos.y += math.sin(perp) * wing_dist * wing_side
                            et.pos.x += math.cos(bearing_rad) * 3_000
                            et.pos.y += math.sin(bearing_rad) * 3_000
                    elif _tactics == 'encirclement' and not _strait_active:
                        # 포위: 전체가 원형으로 배치 (다방향 동시 접근 강화)
                        # v9.14: 해협 시나리오 활성 시 포위 기동 무시 (방위 제한 우선)
                        enc_bearing = math.radians((idx / max(total, 1)) * 360)
                        et.pos.x = math.cos(enc_bearing) * (detect_km * 1000)
                        et.pos.y = math.sin(enc_bearing) * (detect_km * 1000)
                    # v9.6: 기습 잠수함 — hidden_until 설정 (은닉 시간 동안 탐지·교전 불가)
                    if info.get('is_ambush') and et.is_sub:
                        et.hidden_until = float(info.get('ambush_hidden_s', 120))
                    # v9.14: 해협 시나리오 — 잠수함 잠항 수심을 해협 임계수심(sill_m)으로 cap
                    if _strait_active and et.is_sub:
                        strait_key = {'korea_west': 'korea_west',
                                      'korea_east': 'korea_east',
                                      'bilateral':  'korea_west'}.get(_strait_type, 'korea_west')
                        sill = STRAITS_DB.get(strait_key, {}).get('sill_m', 0)
                        if sill and et.altitude_m < -sill:
                            et.altitude_m = float(-sill)
                    threats.append(et)

                idx += 1

        return threats

    def _build_aircraft(self) -> List[FriendlyAircraftObj]:
        """포팅 C: enable_helo / enable_p3c / enable_p8a + v10.5 CAP 항공기."""
        aircraft = []
        primary_pos = self._primary().pos
        for en_key, preset_key, default in [
            ('enable_helo',  'helo_preset',  'AW-159 와일드캣'),
            ('enable_p3c',   'p3c_preset',   'P-3C 오라이온'),
            ('enable_p8a',   'p8a_preset',   'P-8A 포세이돈'),
            # v10.5: 한국 공군 CAP
            ('enable_f35a',  'f35a_preset',  'F-35A 라이트닝 II'),
            ('enable_kf21',  'kf21_preset',  'KF-21 보라매'),
            ('enable_fa50',  'fa50_preset',  'FA-50 파이팅이글'),
            # v16.12: 아군 무인 정찰 드론 (ISR 전용)
            ('enable_recon_drone', 'recon_preset', 'MQ-9B 시가디언'),
        ]:
            if not self.cfg.get(en_key, False):
                continue
            name = self.cfg.get(preset_key, default)
            if name not in FRIENDLY_AIRCRAFT_DB:
                continue
            aircraft.append(FriendlyAircraftObj(name, primary_pos))
        return aircraft

    # ── 헬퍼 ─────────────────────────────────────────────────────────────────

    def _primary(self) -> FriendlyShipObj:
        for t in ('KDX-III-B2', 'KDX-III-B1'):
            for s in self.friendly_ships:
                if s.alive and s.ship_type == t:
                    return s
        if not self.friendly_ships:
            raise RuntimeError("friendly_ships 비어있음 — 편대 프리셋을 확인하세요")
        return next((s for s in self.friendly_ships if s.alive), self.friendly_ships[0])

    def _pick_target(self, is_torpedo: bool = False,
                     threat_pos: Optional['LatLon'] = None) -> FriendlyShipObj:
        """적 미사일 타겟 선택 — 전단 내 생존 함정 분산 공격.
        어뢰: 이지스 기함(primary) 우선. 대함미사일: max_hp 가중 랜덤.
        DMO ON + threat_pos: 위협 접근축 거리 기반 가중 — 광역 분산 시 원거리
        함정은 표적 확률이 급감(적 집중 포화 회피). OFF면 기존 경로 그대로(회귀 보존)."""
        alive = [s for s in self.friendly_ships if s.alive]
        if len(alive) <= 1:
            return alive[0] if alive else (self.friendly_ships[0] if self.friendly_ships else self._primary())
        if is_torpedo:
            return self._primary()
        # DMO: 위협 위치에서 가까운(접근축 상) 함정에 가중 집중 → 멀리 분산된 함정은 안전
        if self.cfg.get('enable_dmo', False) and threat_pos is not None:
            _D = max(self.cfg.get('dmo_spread_km', 80.0) * 1000.0 * 0.75, 1.0)  # 특성 감쇠거리(dmo_spread_km=0 시 0 나눗셈 방어)
            w = [s._max_hp * math.exp(-threat_pos.dist_to(s.pos) / _D) for s in alive]
            if sum(w) > 0:
                return random.choices(alive, weights=w, k=1)[0]
        weights = [s._max_hp for s in alive]
        return random.choices(alive, weights=weights, k=1)[0]

    def _spawn_pending_threat(self, spec: dict):
        """NEW-A: 혼합 시나리오 파도 지연 스폰 — spec={'preset':name,'count':n}"""
        name  = spec.get('preset', '')
        count = spec.get('count', 1)
        if name not in ENEMY_DB:
            return
        info  = ENEMY_DB[name]
        ttype = info.get('type', '')
        detect_km      = self.cfg.get('detect_km', 200)
        surface_det_km = self.cfg.get('surface_detect_km', detect_km)
        sub_det_km     = self.cfg.get('sub_detect_km', 50)
        primary = self._primary()

        for _ in range(count):
            bearing_rad = math.radians(random.uniform(0, 360))
            if info.get('category') == '대잠':
                start_m = sub_det_km * 1000
            elif info.get('category') == '대함':
                start_m = surface_det_km * 1000
            else:
                start_m = detect_km * 1000
            pos = LatLon.from_xy(math.cos(bearing_rad) * start_m, math.sin(bearing_rad) * start_m)

            if ttype in _STANDALONE_MISSILE_TYPES:
                _tgt = self._pick_target(is_torpedo=False, threat_pos=pos)
                m = MissileObj(
                    mtype='enemy_strike', name=name, pos=pos,
                    target=_tgt,
                    speed_ms=info['speed_ms'] * self.cfg.get('threat_spd_scale', 1.0),
                    pk_base=_MISSILE_PK_MAP.get(name, _MISSILE_PK_DEFAULT),
                    owner_id=-1, t_spawn=self.t,
                )
                m.altitude_m              = float(info.get('altitude_m', 0))
                m.is_hgv                  = bool(info.get('is_hgv', False))
                m.is_qbm                  = bool(info.get('is_qbm', False))
                m.is_ballistic            = (ttype == '탄도미사일')
                m.is_arm                  = bool(info.get('is_arm', False))
                m.terminal_evasion_factor = info.get('missile_terminal_evasion', 1.0)
                m.is_torpedo              = False
                m._init_dist              = m.pos.dist_to(_tgt.pos)
                m._peak_alt_m             = m.altitude_m
                m.detect_m                = m._init_dist
                m.enemy_info              = info.copy()
                self.missiles.append(m)
            else:
                et = self._new_threat(name, pos)
                et.speed_ms *= self.cfg.get('threat_spd_scale', 1.0)
                et.carrier_owner = spec.get('carrier_owner')   # v15.09.02: 함재기면 모항 인덱스(재무장 복귀)
                self.enemy_threats.append(et)
            self.stats['total_threats'] += 1
            self._log(f"[{name}] {self.t:.0f}s 파도 스폰")

    def _log(self, msg: str):
        if self._mc_mode:
            return
        self._log_entries.append((self.t, msg))
        self._tick_events.append(msg)

    def _threat_origin_bearing(self):
        """v15.03: 함대 위치에서 적 세력 발진 좌표로의 접근 방위(rad). 해결 불가 시 None.
        프리셋 모드면 이름으로 세력 분류, 아니면 해역 기본 세력. 북한은 함대 경도로 서/동."""
        region = self.cfg.get('fleet_region', '동해 중부')
        mode   = self.cfg.get('enemy_fleet_mode', 'preset')
        preset = self.cfg.get('enemy_fleet_preset') if mode == 'preset' else None
        origin = _preset_origin(preset) if preset else _REGION_DEFAULT_ORIGIN.get(region)
        if not origin:
            return None
        flat, flon = _REGION_REF.get(region, _DEFAULT_REF)
        if origin == 'nk':   # 함대 경도로 북한 서해(남포)/동해(원산) 발진점 선택
            origin = 'nk_west' if flon < 127.0 else 'nk_east'
        coord = _THREAT_ORIGIN.get(origin)
        if not coord:
            return None
        olat, olon = coord
        dx = (olon - flon) * math.cos(math.radians(flat))   # 동(+)
        dy = (olat - flat)                                   # 북(+)
        if dx == 0.0 and dy == 0.0:
            return None
        return math.atan2(dy, dx)

    def _log_detections(self):
        """위협이 처음 레이더/소나 탐지 범위에 들어온 시점을 1회 로깅 (스탠드오프 가시화).
        함대 공유(Link-16) 실효 탐지거리(detect_km/surface/sub) 기준 — 실제 교전 탐지와 일치.
        순수 거리 판정만 사용 — RNG 미소비라 결정론 보존. MC 모드는 _log가 자동으로 무시."""
        if self._mc_mode:
            return
        ships = [s for s in self.friendly_ships if s.alive]
        if not ships:
            return
        # v16.12: 정찰 드론 OTH 중계 — 대공·대함 탐지거리 확장(교전 판정과 동일 기준)
        air_km  = self.cfg.get('detect_km', 200) + self._recon_bonus_km
        surf_km = self.cfg.get('surface_detect_km', self.cfg.get('detect_km', 200)) + self._recon_bonus_km
        sub_km  = self.cfg.get('sub_detect_km', 50)

        # 함정·항공기·잠수함 위협 — 개체별 최초 탐지
        for et in self.enemy_threats:
            if not et.alive or et.uid in self._detected_log_set:
                continue
            if et.is_sub:
                sensor, rng_m = '소나', sub_km * 1000
            elif et.is_aircraft:
                sensor, rng_m = '레이더', air_km * 1000
            else:
                sensor, rng_m = '레이더', surf_km * 1000
            best = min(ships, key=lambda s: s.pos.dist_to(et.pos))
            dist = best.pos.dist_to(et.pos)
            if dist <= rng_m:
                self._detected_log_set.add(et.uid)
                self._log(f"[탐지] {best.name} {sensor} → {et.preset_name} 포착 "
                          f"(거리 {dist/1000:.0f}km)")

        # 독립 미사일 위협(탄도·순항 등) — 종류별 최초 1회만 (살보 폭증 방지)
        for m in self.missiles:
            if not m.alive or m.mtype != 'enemy_strike':
                continue
            key = f"missile:{m.name}"
            if key in self._detected_log_set:
                continue
            best = min(ships, key=lambda s: s.pos.dist_to(m.pos))
            dist = best.pos.dist_to(m.pos)
            if dist <= air_km * 1000:
                self._detected_log_set.add(key)
                self._log(f"[탐지] {best.name} 레이더 → {m.name} 포착 "
                          f"(거리 {dist/1000:.0f}km)")

    def _detect_range_m(self, ship: FriendlyShipObj, category: str,
                        alt_m: Optional[float] = None) -> float:
        if category == '대잠':
            base_km = ship.sensor_km.get('대잠', 50)
            # v16.12: 무인 잠수정(UUV) 전방 수중 피켓 — 대잠 탐지거리 확장(생존 UUV 있을 때만>0)
            base_km += self._uuv_asw_bonus_km
            factor  = self.wx.get('sonar_factor', self.wx.get('detect_range_factor', 1.0))
        else:
            # 대공: 이지스 데이터링크(Link-16) — 편대 최고 성능 레이더 공유
            # 대함: 수상 레이더 한계 — 수평선 넘어 탐지 불가, surface_detect_km 사용
            if category == '대함':
                base_km = max(ship.sensor_km.get('대함', 45),
                              self.cfg.get('surface_detect_km', 45))
                # v16.12: 무인 수상정(USV) 전방 피켓 — 대함 탐지거리 확장(생존 USV 있을 때만>0)
                base_km += self._usv_surf_bonus_km
            else:
                base_km = ship.sensor_km.get(category, self.cfg.get('detect_km', 200))
            # v16.12: 정찰 드론 OTH 중계 — 대공·대함 레이더 탐지거리 확장(생존 드론 있을 때만>0)
            base_km += self._recon_bonus_km
            factor = self.wx.get('radar_factor', self.wx.get('detect_range_factor', 1.0))
        # 함정 부분 피해: 레이더 성능 저하 반영
        detect_m = base_km * 1000 * factor * ship.radar_factor * self.cfg.get('detect_scale', 1.0)
        # v16.3 사이버전 CIC 마비: 침투 지속 중 전투정보실 처리 지연 → 실효 탐지거리 저하
        if self._cyber and self.t < self._cic_blind_until:
            detect_m *= _CYBER_CIC_DET
        # v8.26: 적 ECM 에어리어 재밍 → 레이더 탐지거리 감소 (소나 무효, enable_ecm 플래그 연동)
        if (category != '대잠' and self._active_ecm > 0
                and self.cfg.get('enable_ecm', True)):
            jam = self._active_ecm * (1.0 - ship.eccm_factor)
            detect_m *= max(0.40, 1.0 - jam)
        # v9.12: 지형 음영 / v9.13: 증발 덕팅 — 레이더 탐지 한정
        if category != '대잠' and alt_m is not None:
            detect_m *= self._terrain_penalty(alt_m)
            if alt_m >= 0:   # 수중 표적은 덕팅·대기 보정 없음
                detect_m *= self._evap_duct_factor(alt_m)
                detect_m *= self._isa_refraction_factor(alt_m)
                detect_m *= self._troposcatter_factor(alt_m)
        return detect_m

    def _thermocline_factor(self, et: 'EnemyThreatObj') -> float:
        """
        수온약층(thermocline) 소나 탐지 보정.
        altitude_m < 0 = 잠항 수심 (음수).

        v9.12: db_ocean_acoustic WOA18 실측값 사용.
        해역(fleet_region)·계절(season) 조합으로 수심별 배율 조회.
        """
        if not et.is_sub:
            return 1.0
        depth = abs(et.altitude_m)
        # 함정 무관(수심+해역·계절 상수)이라 함정마다 재계산은 낭비 → 수심별 메모
        cached = self._thermo_cache.get(depth)
        if cached is not None:
            return cached
        if _OCEAN_ACOUSTIC_OK:
            region_key = REGION_TO_ACOUSTIC_KEY.get(
                self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
            )
            season = self.cfg.get('season', 'summer')
            factor = get_sonar_depth_factor(region_key, season, depth)
        # fallback: 기존 하드코딩 (db_ocean_acoustic 없을 때)
        elif depth < 100:
            factor = 1.0
        elif depth < 300:
            factor = 1.0 - 0.55 * (depth - 100) / 200
        elif depth < 500:
            factor = 0.45 + 0.20 * (depth - 300) / 200
        else:
            factor = 0.65
        self._thermo_cache[depth] = factor
        return factor

    # ── v12.3: dB 소나 방정식 (enable_sonar_equation ON 경로) ─────────────────
    def _sonar_env(self):
        """현재 교전 환경 → (해역키, 계절, 대표수심, 주변소음 NL dB)."""
        region = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA')
        season = self.cfg.get('season', 'summer')
        water_depth = WATER_DEPTH_M.get(region, 1500.0)
        return region, season, water_depth, self._sonar_ambient_dB(region)

    def _sonar_ambient_dB(self, region: str) -> float:
        """Beaufort 기반 1kHz 주변소음 스펙트럼 레벨 + 해역 선박소음 보정."""
        if not _OCEAN_ENV_OK or 'beaufort' not in self.wx:
            return 46.0   # BF4 근사 폴백
        bf = max(0, min(12, int(round(self.wx.get('beaufort', 4)))))
        try:
            n1k = SONAR_AMBIENT_NOISE['by_beaufort_spectral_dB'][bf][2]
        except Exception:
            return 46.0
        shipping = {'EAST_SEA': 0, 'YELLOW_SEA': 5, 'KOREA_STRAIT': 10}.get(region, 0)
        return n1k + shipping

    def _sonar_r50(self, et: 'EnemyThreatObj', sensor_key: str, active: bool = False) -> float:
        """잠수함·센서 조합의 50% 탐지거리(m). 데이터/모듈 없으면 -1.0(레거시 폴백).
        active=True면 능동 소나 R50(왕복 TL+TS+잔향), False면 수동 R50."""
        if not _OCEAN_ACOUSTIC_OK:
            return -1.0   # db_ocean_acoustic 미탑재 — 레거시 경로로 폴백
        region, season, water_depth, ambient = self._sonar_env()
        depth = abs(et.altitude_m)
        key = (sensor_key, active, round(depth / 10.0) * 10, round(ambient), region, season)
        cached = et._r50_cache.get(key)
        if cached is not None:
            return cached
        fn = active_sonar_detection_range if active else sonar_detection_range
        r50 = fn(et.preset_name, sensor_key, region, season, depth, ambient, water_depth)
        et._r50_cache[key] = r50
        return r50

    def _sonar_eq_pd(self, et: 'EnemyThreatObj', dist_m: float, sensor_key: str):
        """수동·능동 소나 통합 탐지확률 Pd. 'both' 센서면 max(수동,능동).
        반환: (pd, max_r50) / 데이터 없으면 (None, None)."""
        region, season, water_depth, _ = self._sonar_env()
        f = SONAR_PLATFORM[sensor_key]['freq_khz']
        depth = abs(et.altitude_m)
        best_pd, best_r50, have = 0.0, 0.0, False
        for active in (False, True):
            r50 = self._sonar_r50(et, sensor_key, active=active)
            if r50 < 0:
                continue   # 해당 모드 미지원(능동 불가 센서 등)
            have = True
            if r50 == 0:
                continue
            # 고속 이탈 또는 v16.01.03 능동 핑 역탐지 회피 기동 시 접촉 급감
            _evading = et.is_retreating or (self._sonar_emcon and self.t < et.counter_evade_until)
            r50_eff = r50 * 0.30 if _evading else r50
            best_r50 = max(best_r50, r50_eff)
            pd = sonar_detection_prob(
                dist_m, r50_eff, water_depth, f, region, season, depth,
                tl_mult=2.0 if active else 1.0)
            best_pd = max(best_pd, pd)
        if not have:
            return None, None
        return best_pd, best_r50

    def _sonar_eq_detect(self, et: 'EnemyThreatObj', dist_m: float, sensor_key: str):
        """
        dB 소나 방정식 확률 탐지 판정 (수동·능동 통합).
        반환: True=탐지 / False=실패 / None=음향 데이터 없음(레거시 폴백).
        """
        pd, max_r50 = self._sonar_eq_pd(et, dist_m, sensor_key)
        if pd is None:
            return None
        if max_r50 <= 0:
            return False
        if dist_m > max_r50 * 3.0:
            return False   # 3·R50 너머는 Pd≈0
        return random.random() < pd

    def _terrain_penalty(self, alt_m: float) -> float:
        """
        지형 레이더 음영 보정 — 해역별 산맥 차폐로 저고도 위협 탐지거리 감소.
        enable_terrain=False 또는 잠수함(소나)이면 1.0 반환.
        """
        if not self.cfg.get('enable_terrain', False):
            return 1.0
        if alt_m < 0:   # 잠수함 — 소나 탐지, 지형 페널티 없음
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        table = TERRAIN_RADAR_PENALTY.get(region_key, [])
        if not table:
            return 1.0
        prev_alt, prev_f = table[0]
        for a, f in table[1:]:
            if alt_m <= a:
                ratio = (alt_m - prev_alt) / max(a - prev_alt, 1)
                return prev_f + ratio * (f - prev_f)
            prev_alt, prev_f = a, f
        return prev_f

    # ── v12.5: 동적 기상 전이 ────────────────────────────────────────────────

    # ── v12.6: 피아식별 오류 (IFF) ─────────────────────────────────────────
    def _iff_check(self, target) -> bool:
        """C&D 통과 직후 IFF 판정. False 반환 시 교전 대기 재진입."""
        if not self._iff_enabled:
            return True
        n_alive = getattr(self, '_n_alive_threats', len(self.enemy_threats))
        p_fail = min(0.30,
                     IFF_FAIL_BASE
                     + self._active_ecm * 0.10
                     + max(0, n_alive - 3) * 0.02)
        if random.random() >= p_fail:
            return True
        self._cd_fire_time[target.uid] = self.t + IFF_RECHECK_DELAY_S
        self.stats['iff_failures'] += 1
        if not self._mc_mode:
            self._log(f"⚠️ IFF 식별 실패 → {IFF_RECHECK_DELAY_S:.0f}s 재확인")
        self._check_iff_fratricide(target)
        return False

    def _check_iff_fratricide(self, target) -> None:
        """IFF 실패 시 근접 아군 항공기가 있으면 확률적으로 오사 발생."""
        nearby = [ac for ac in self.aircraft
                  if ac.payload_remaining > 0
                  and ac.home_pos.dist_to(target.pos) <= IFF_FRATRICIDE_RANGE_M]
        if not nearby or random.random() >= IFF_FRATRICIDE_P:
            return
        victim = random.choice(nearby)
        victim.payload_remaining = 0
        victim.strike_payload_remaining = 0
        self.stats['iff_fratricide'] += 1
        if not self._mc_mode:
            self._log(f"💥 아군 오사: SAM → {victim.name} (IFF 오인식)")

    def _update_weather(self) -> None:
        """매 WEATHER_STEP_INTERVAL_S마다 호출. 확률적으로 날씨를 1단계 전이."""
        region_key = {
            '동해 북부': 'EAST_SEA', '동해 남부': 'EAST_SEA',
            '서해 북부': 'YELLOW_SEA', '서해 남부': 'YELLOW_SEA',
            '남해': 'KOREA_STRAIT', '대한해협': 'KOREA_STRAIT',
        }.get(self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA')
        season = self.cfg.get('season', 'summer')

        delta_w, delta_i = WEATHER_TRANSITION_DB.get((region_key, season), (0.0, 0.0))
        p_worsen  = max(0.0, min(0.6, WEATHER_WORSEN_BASE  + delta_w))
        p_improve = max(0.0, min(0.6, WEATHER_IMPROVE_BASE + delta_i))

        # 추세 강제
        trend = self._wx_trend
        if trend == '악화':
            p_worsen  = min(0.6, p_worsen  + 0.20)
            p_improve = max(0.0, p_improve - 0.08)
        elif trend == '호전':
            p_improve = min(0.6, p_improve + 0.20)
            p_worsen  = max(0.0, p_worsen  - 0.08)
        elif trend == '안정':
            p_worsen  = max(0.0, p_worsen  - 0.10)
            p_improve = max(0.0, p_improve - 0.05)

        roll = random.random()
        if roll < p_worsen:
            direction = 1
        elif roll < p_worsen + p_improve:
            direction = -1
        else:
            return  # 변화 없음

        current = self.cfg.get('weather', '맑음 (주간)')
        ladder, idx = _weather_ladder_pos(current, self._wx_ladder_key)
        new_idx = idx + direction
        if not (0 <= new_idx < len(ladder)):
            return  # 사다리 끝 → 더 이상 이동 불가
        new_weather = ladder[new_idx]
        self._apply_weather_transition(new_weather)

    def _apply_weather_transition(self, new_weather: str) -> None:
        """날씨를 new_weather로 교체하고 wx 계수 갱신."""
        old_weather = self.cfg.get('weather', '맑음 (주간)')
        if old_weather == new_weather:
            return
        self.cfg['weather'] = new_weather
        self.wx = _make_physics_wx(new_weather)
        if not self._mc_mode:
            self._log(f"🌦️ 기상 변화: {old_weather} → {new_weather}")

    # ── 물리 환경 보정 함수들 ─────────────────────────────────────────────────

    def _evap_duct_factor(self, alt_m: float) -> float:
        """
        증발 덕팅(EDH) 보정 — 대기 하층 수증기 농도 역전으로 레이더 전파 해면 굴절.
        저고도 표적(alt_m ≤ EDH)의 탐지거리 증가. BF7 이상 강풍 시 덕트 파괴.
        enable_evap_duct=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_evap_duct', False):
            return 1.0
        if self.wx.get('beaufort', 2) >= 7:   # 강풍 이상 → 덕트 파괴
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        edh_m, boost = EVAP_DUCT_DB.get((region_key, season), (0, 1.0))
        if edh_m == 0:
            return 1.0
        if alt_m <= edh_m:
            return boost
        if alt_m <= edh_m * 2:
            ratio = (edh_m * 2 - alt_m) / max(edh_m, 1)
            return 1.0 + (boost - 1.0) * ratio
        return 1.0

    def _isa_refraction_factor(self, alt_m: float) -> float:
        """
        ISA 대기 굴절 보정 (라디오존데 4계절×5고도층 실측값 기반).
        100m 미만은 _evap_duct_factor 담당. enable_isa=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_isa', False):
            return 1.0
        if alt_m < 100:
            return 1.0
        if   alt_m < 500:  layer = 1
        elif alt_m < 1000: layer = 2
        elif alt_m < 3000: layer = 3
        else:              layer = 4
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return ISA_RADIOSONDE_DB.get((region_key, season, layer), 1.0)

    def _troposcatter_factor(self, alt_m: float) -> float:
        """
        트로포스캐터 링크 보정 — 대기 상층 난류 산란으로 수평선 너머 탐지거리 증가.
        고고도(≥1000m) 표적에만 적용. BF6 이상 강풍 시 산란층 파괴 → 1.0 반환.
        enable_isa=False이면 1.0 반환.
        """
        if not self.cfg.get('enable_isa', False):
            return 1.0
        if alt_m < 1000:
            return 1.0
        if self.wx.get('beaufort', 2) >= 6:
            return 1.0
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return TROPOSCATTER_DB.get((region_key, season), 1.0)

    def _wind_cep_factor(self) -> float:
        """
        고층 바람 CEP 배율 — 순항미사일 탄착 오차 증가로 명중률 저하.
        Pk_eff = pk_base / cep_factor (반환값 > 1 → Pk 감소)
        """
        region_key = REGION_TO_ACOUSTIC_KEY.get(
            self.cfg.get('fleet_region', '동해 북부'), 'EAST_SEA'
        )
        season = self.cfg.get('season', 'summer')
        return WIND_CEP_FACTOR.get((region_key, season), 1.0)

    # ── 1단계: 위치 갱신 ──────────────────────────────────────────────────────

    def _apply_ship_evasion(self, evade_r_base: float = 15_000,
                            evade_speed_ms: float = _SHIP_EVADE_SPEED_MS):
        """아군 함정 회피 기동: 적 미사일이 evade_r 이내면 전속 지그재그 침로로 항주한다.

        v20.5(B-2) — 두 가지를 함께 고친다.

        ① **물리 정정**: 종전에는 틱(1초)마다 **300~800m를 무작위 방향으로 순간이동**했다.
           초속 300~800m = 583~1555kts로, 구축함 전속(30kts≈15m/s)의 20~50배다. 함대가 매
           초 순간이동하니 교전 기하가 망가져 회피를 켤수록 오히려 요격률이 떨어지고 피격이
           늘었다(툴팁은 '피탄율을 낮춥니다'인데 실제로는 순손해였다).
           → 함속(_SHIP_EVADE_SPEED_MS) × DT만큼만 **연속 이동**하고, 침로는 _EVADE_TURN_S
           마다 새로 뽑아 지그재그(weave)를 만든다. 추진 피탄 시 speed_factor만큼 감속.

        ② **기본 ON 승격**: 실제 함대는 항주한다 — 정지한 함대가 비물리다. 게다가 이 기동은
           레이더 침묵(ARM 회피)과 **짝**이다. ESM stale 유도는 '침묵 동안 함정이 움직여 생긴
           조준 좌표 이격'으로 ARM 명중을 깎는데, 함대가 서 있으면 이격이 0이라 레이더를 꺼도
           ARM이 같은 좌표로 날아왔다(실측: 회피 OFF면 enable_radar_off ON/OFF가
           bit-identical = 침묵이 죽은 기능이었다). 둘 다 켜야 실제 교리인 '방사 중단 +
           침로 변경'이 성립한다.
        """
        if not self.cfg.get('enable_ship_evasion', True):
            return
        # v9.14: 해협 시나리오 — STRAITS_DB 폭 기반 기동 공간 제한
        _evade_r = evade_r_base
        if self.cfg.get('fleet_region', '') == '대한해협':
            _st = self.cfg.get('strait_type', 'korea_west')
            _sk = 'korea_east' if _st == 'korea_east' else 'korea_west'
            _w  = STRAITS_DB.get(_sk, {}).get('width_km', _STRAIT_OPEN_SEA_KM)
            _evade_r = int(_evade_r * min(1.0, _w / _STRAIT_OPEN_SEA_KM))
        for ship in self.friendly_ships:
            if not ship.alive or ship.is_shore_battery:
                continue   # 해안 고정 포대는 회피 기동 없음
            close = any(
                m.alive and m.mtype == 'enemy_strike'
                and m.pos.dist_to(ship.pos) < _evade_r
                for m in self.missiles
            )
            if close:
                # 지그재그(weave): **기준 침로를 유지한 채** _EVADE_TURN_S마다 좌·우로 번갈아
                # 꺾는다. 매번 침로를 완전 무작위로 다시 뽑으면 그건 weave가 아니라 랜덤워크라,
                # 순 변위가 √n로만 쌓이고 되돌아오기까지 한다 — ARM 조준 좌표 이격이 시드마다
                # 요동쳐 stale 감쇠가 걸리다 말다 한다. 기준 침로 ±_EVADE_WEAVE_RAD로 흔들면
                # 변위가 시간에 비례해 안정적으로 쌓인다(실제 회피 기동의 모습이기도 하다).
                if not hasattr(ship, '_evade_base'):
                    ship._evade_base = random.uniform(0, 2 * math.pi)   # 함대 항주 침로
                    ship._evade_side = 1
                    ship._evade_turn_t = self.t + _EVADE_TURN_S
                    ship._evade_origin = ship.pos.copy()
                elif self.t >= ship._evade_turn_t:
                    ship._evade_side *= -1                              # 좌 ↔ 우 전환
                    ship._evade_turn_t = self.t + _EVADE_TURN_S

                heading = ship._evade_base + ship._evade_side * _EVADE_WEAVE_RAD
                step_m  = evade_speed_ms * DT * ship.speed_factor   # 추진 피탄 시 감속
                nx = ship.pos.x + math.cos(heading) * step_m
                ny = ship.pos.y + math.sin(heading) * step_m

                # 대형 유지: 초기 위치에서 _EVADE_MAX_DRIFT_M 밖으로는 벗어나지 않는다.
                # (상한이 없으면 장시간 포화 교전에서 함정마다 제 침로로 흩어져, 초기 배치로
                #  성립하던 중첩 방어·CEC 중계 거리가 무너진다. 회피는 전단 대형 안에서 한다.)
                _ox = nx - ship._evade_origin.x
                _oy = ny - ship._evade_origin.y
                if math.hypot(_ox, _oy) <= _EVADE_MAX_DRIFT_M:
                    ship.pos.x, ship.pos.y = nx, ny
                else:
                    ship._evade_base = random.uniform(0, 2 * math.pi)   # 대형 밖 → 침로 재설정
                    ship._evade_turn_t = self.t + _EVADE_TURN_S

    def _threat_move_target(self, et, primary_pos):
        """Phase 5.3.1: 위협이 향하는 표적 좌표. 부모(단발)는 항상 기함 → 동작 보존.
        BattleEngine이 목표지향(수상함=돌파선/그 외=기함)으로 오버라이드."""
        return primary_pos

    def _update_positions(self):
        self._apply_ship_evasion()
        self._arm_esm_update()   # v16.1: ARM 유도를 아군 레이더 방사 상태에 연동(미사일 이동 전)
        self._cyber_update()     # v16.3: 사이버전 침투 굴림(주기적, OFF면 무동작)
        self._coord_deception_update()  # v16.6: 종말권 진입 대함미사일 좌표 기만(미사일 이동 전)

        primary_pos = self._primary().pos
        for et in self.enemy_threats:
            if not et.alive:
                continue
            if et.is_retreating and et.retreat_pos:
                arrived = et.pos.move_toward(et.retreat_pos, et.speed_ms, DT)
                if arrived:
                    self._on_retreat_arrived(et)
            else:
                et.pos.move_toward(self._threat_move_target(et, primary_pos), et.speed_ms, DT)

        # NEW-B: 드론/스웜 자폭 — 200m 이내 도달 시 피격 처리
        # v16.8: 공중 자폭 드론(is_aircraft) + 수상 자폭정(is_suicide) 공통 처리.
        primary = self._primary()
        for et in self.enemy_threats:
            if not et.alive or et.is_retreating:
                continue
            if not (et.is_aircraft or et.info.get('is_suicide')):
                continue
            if et.info.get('can_fire_missile', True):
                continue  # 일반 전투기는 미사일 발사 후 이탈 — 자폭 없음
            if et.pos.dist_to(primary.pos) > 200:
                continue
            primary.take_hit(et.preset_name, self.t)
            self.stats['friendly_hits'] += 1
            self._log(f"[피격!] {et.preset_name} 자폭")
            et.alive = False

        # v12.1: 미사일 이동 — PNG 종말 유도(enable_png) 분기
        # PNG 교전 SAM은 표적(적 대함미사일)을 jink로 직접 전진시키므로, 그 표적은
        # 자신의 move_toward(update)에서 제외한다 — 빼지 않으면 한 틱에 이중 이동.
        png_on = self.cfg.get('enable_png', False)
        png_targets = set()
        if png_on:
            for m in self.missiles:
                if m.alive and self._png_engaged(m):
                    png_targets.add(id(m.target))
        for m in self.missiles:
            if not m.alive:
                continue
            if png_on and self._png_engaged(m):
                self._update_missile_png(m)   # 종말 10km SAM↔대함 — SAM·표적 함께 전진
            elif id(m) in png_targets:
                continue                      # PNG SAM이 이 미사일을 이미 전진시킴 (이중 이동 방지)
            else:
                m.update(DT)                  # 기존 완벽 정조준 이동

        self._hgv_glide_update()   # v16.2: 활공 HGV 교전 고도 갱신(이동 후, OFF면 무동작)
        self._ballistic_descent_update()   # v20.2a: 탄도 종말 강하(OFF면 무동작)

        # v10.8: 해류 연동 — 수상함·잠수함 위치에 해류 벡터 누적
        if self.cfg.get('enable_current', False) and _get_current_vector:
            _region_map = {
                '동해 북부': 'east_sea', '동해 중부': 'east_sea',
                '서해': 'west_sea', '대한해협': 'korea_strait',
            }
            _cur_region = _region_map.get(self.cfg.get('fleet_region', '동해 북부'), 'east_sea')
            import datetime
            _month = datetime.date.today().month
            _cv = _get_current_vector(_cur_region, _month)
            _spd_ms = _cv['speed_cms'] / 100.0      # cm/s → m/s
            _dir_rad = math.radians(90 - _cv['direction_deg'])  # 진북 기준 → x/y 평면
            _dx = math.cos(_dir_rad) * _spd_ms * DT
            _dy = math.sin(_dir_rad) * _spd_ms * DT
            for et in self.enemy_threats:
                if et.alive and (et.is_ship or et.is_sub):
                    et.pos.x += _dx
                    et.pos.y += _dy
            for ship in self.friendly_ships:
                if ship.alive and not ship.is_submarine:
                    ship.pos.x += _dx
                    ship.pos.y += _dy

    # ── v12.1: 비례항법(PNG) 종말 유도 ───────────────────────────────────────
    def _png_engaged(self, m: 'MissileObj') -> bool:
        """PNG 물리 추격 대상 여부: 아군 SAM이 적 대함미사일을 종말 10km 내 추격 중."""
        if m.mtype != 'friendly_sam':
            return False
        tgt = m.target
        if not isinstance(tgt, MissileObj) or not tgt.alive:
            return False
        if tgt.mtype != 'enemy_strike':
            return False
        return m.pos.dist_to(tgt.pos) < PNG_TERMINAL_M

    def _update_missile_png(self, sam: 'MissileObj'):
        """종말 substep 적분: SAM은 비례항법 추격, 표적(적 대함미사일)은 jink 기동.
        둘을 같은 dt_sub로 함께 전진시켜 상대운동을 정확히 반영하고, 매 substep
        최근접 거리(min_miss_m)를 기록한다. 명중/빗나감 판정은 _check_intercepts."""
        tgt = sam.target
        dt_s = DT / PNG_SUBSTEP
        sam_max_turn = (SAM_MAX_G * G_ACCEL) / max(sam.speed_ms, 1.0)  # rad/s 상한
        fov = math.radians(SEEKER_FOV_DEG)
        # PNG 첫 진입: 종말 전까지 move_toward로 비행해 heading_rad가 낡았으므로
        # 예측 요격점(lead) 방향으로 정렬 — 실제 SAM도 발사 시 lead각으로 조준한다.
        # (LOS로 정렬하면 근접 교전에서 lead각까지 G 한계로 다 못 돌려 빗나감)
        if not sam.png_active:
            sam.heading_rad = self._lead_intercept_heading(sam, tgt)
        # 표적 전진은 이번 틱 첫 SAM만 수행 — 같은 표적을 여러 SAM(살보)이 노려도
        # 표적이 한 틱에 한 번만 jink 전진하도록 가드 (중복 이동 방지).
        advance_tgt = (tgt._jink_t != self.t)
        for _ in range(PNG_SUBSTEP):
            if not tgt.alive:
                break
            # (1) 표적 종말 회피 기동 (적 대함미사일 jink) — 틱당 1회만
            if advance_tgt:
                self._advance_target_jink(tgt, dt_s)
            # (2) SAM 유도: 예측 요격점(lead) 방향으로 G 한계 내 선회
            desired = self._lead_intercept_heading(sam, tgt)
            los = sam.pos.bearing_to(tgt.pos)
            # 탐색기 FOV: 표적이 시야 밖이면 lock 상실 → 직진(무한 기동 방지)
            if abs(_ang_diff(los, sam.heading_rad)) <= fov:
                turn = _ang_diff(desired, sam.heading_rad)
                turn = max(-sam_max_turn * dt_s,
                           min(sam_max_turn * dt_s, turn))     # G 한계 clamp
                sam.heading_rad += turn
            # (3) SAM heading 방향으로 전진 (정조준 아님). 이동 전 상대위치 기록
            rel_ax, rel_ay = sam.pos.x - tgt.pos.x, sam.pos.y - tgt.pos.y
            sam.pos.x += math.cos(sam.heading_rad) * sam.speed_ms * dt_s
            sam.pos.y += math.sin(sam.heading_rad) * sam.speed_ms * dt_s
            # (4) substep 구간(선분) 내 최근접 통과 거리 — 고속 교전 건너뜀 방지
            rel_bx, rel_by = sam.pos.x - tgt.pos.x, sam.pos.y - tgt.pos.y
            seg_miss = _seg_origin_dist(rel_ax, rel_ay, rel_bx, rel_by)
            sam.min_miss_m = min(sam.min_miss_m, seg_miss)
        if advance_tgt:
            tgt._jink_t = self.t          # 이번 틱 표적 전진 완료 표시
        sam.png_active = True

    def _lead_intercept_heading(self, sam: 'MissileObj', tgt: 'MissileObj') -> float:
        """예측 요격점(lead collision point) 방향 반환 — 증강 비례항법.
        표적 평균 진로(함정 방향, jink 평균 0)로 등속 외삽한 미래 위치와 SAM이
        만나는 시간 t를 요격 삼각형 2차방정식으로 풀어, 그 점으로 향하는 방위를 준다.
        속도 열세로 해가 없으면(순수 후미추격 불가) 표적 현재 위치 직접 조준으로 fallback."""
        base = tgt.pos.bearing_to(tgt.target.pos)        # 표적 평균 진로 (함정 방향)
        vtx = math.cos(base) * tgt.speed_ms
        vty = math.sin(base) * tgt.speed_ms
        rx = tgt.pos.x - sam.pos.x
        ry = tgt.pos.y - sam.pos.y
        vs = sam.speed_ms
        a = vtx * vtx + vty * vty - vs * vs
        b = 2.0 * (rx * vtx + ry * vty)
        c = rx * rx + ry * ry
        t = None
        if abs(a) < 1e-6:                                # 속도 동일 — 선형
            if abs(b) > 1e-6 and -c / b > 0:
                t = -c / b
        else:
            disc = b * b - 4 * a * c
            if disc >= 0:
                sq = math.sqrt(disc)
                for tt in ((-b - sq) / (2 * a), (-b + sq) / (2 * a)):
                    if tt > 0 and (t is None or tt < t):
                        t = tt
        if t is None:                                    # 요격해 없음 → 직접 추격
            return sam.pos.bearing_to(tgt.pos)
        lead_x = tgt.pos.x + vtx * t
        lead_y = tgt.pos.y + vty * t
        return math.atan2(lead_y - sam.pos.y, lead_x - sam.pos.x)

    def _advance_target_jink(self, tgt: 'MissileObj', dt_s: float):
        """적 대함미사일 종말 회피: 함정 향한 평균 진로 + 횡방향 정현파 weave.
        진폭은 기존 terminal_evasion_factor에서 파생(회피 강할수록 큼) — 새 DB 필드 없음.
        횡가속은 ASM_MAX_G로 제한해 비현실적 기동 방지."""
        base = tgt.pos.bearing_to(tgt.target.pos)         # 함정 방향
        # terminal_evasion_factor 1.0=회피없음 → 진폭 0, 작을수록 강한 회피
        evade = max(0.0, 1.0 - tgt.terminal_evasion_factor)
        jink_omega = 0.4                                  # rad/s (주기 ~16s weave)
        # 횡가속 한계 a=ASM_MAX_G → 최대 횡속도 성분 = a/omega, 진폭 비율로 환산
        max_lat_ms = min(tgt.speed_ms, (ASM_MAX_G * G_ACCEL) / jink_omega)
        tgt.jink_phase += jink_omega * dt_s
        lat = evade * max_lat_ms * math.sin(tgt.jink_phase)
        # 진로 방향 전진 + 수직(좌현) 성분 합성
        fwd = tgt.speed_ms
        tgt.pos.x += (math.cos(base) * fwd - math.sin(base) * lat) * dt_s
        tgt.pos.y += (math.sin(base) * fwd + math.cos(base) * lat) * dt_s

    # ── 2단계: 적 발사 ────────────────────────────────────────────────────────

    _ADAPTIVE_MODES = ('saturation', 'dispersal', 'deception')

    def _adaptive_tactic_update(self):
        """
        v15.1: 적응형 전술 AI — 주기(20초)마다 전장 상태를 평가해 전술 전환.
        규칙기반 3단계(포화→분산→기만). adaptive 모드에서만 호출.
          sat    = 비행 중 위협 / 아군 총 교전 채널 (채널 포화도)
          r_recent = 직전 평가 이후(최근 20초 구간) 요격률 — 슬라이딩 윈도우
        v15.01.03: 누적 요격률 대신 최근 구간으로 판단(관성 제거) + 목표 단계로
        한 칸씩만 이동(포화→분산→기만 점진적으로 밟음, 급점프 방지).
        """
        AI_REEVAL_S = 20.0
        if self.t - self._adaptive_last_t < AI_REEVAL_S:
            return
        self._adaptive_last_t = self.t

        total_ch  = max(sum(s.max_channels for s in self.friendly_ships if s.alive), 1)
        in_flight = sum(1 for m in self.missiles
                        if m.alive and m.mtype == 'enemy_strike')
        sat   = in_flight / total_ch
        fired = self.stats['total_threats']
        itc   = self.stats['intercepted_threats']
        # 슬라이딩 윈도우: 직전 평가 이후 발사/요격 델타로 최근 요격률
        d_fired = fired - self._adaptive_prev_fired
        d_itc   = itc   - self._adaptive_prev_itc
        self._adaptive_prev_fired, self._adaptive_prev_itc = fired, itc
        if d_fired > 0:
            r_recent = d_itc / d_fired
        else:   # 이번 구간 발사 없음 — 누적치로 폴백
            r_recent = (itc / fired) if fired > 0 else 0.0

        # 최근 요격률 → 목표 단계 (0=포화, 1=분산, 2=기만)
        if r_recent > 0.6:
            target = 2
        elif r_recent > 0.35:
            target = 1
        else:
            target = 0
        cur  = self._ADAPTIVE_MODES.index(self._adaptive_mode)
        prev = self._adaptive_mode
        if   target > cur: cur += 1   # 한 칸씩만 상승
        elif target < cur: cur -= 1   # 한 칸씩만 하강
        new = self._ADAPTIVE_MODES[cur]

        if new != prev:
            self._adaptive_mode = new
            self._adaptive_switches += 1
            if not self._mc_mode:
                _lbl = {'saturation': '포화 공격', 'dispersal': '분산 접근',
                        'deception': '기만 침투'}
                self._log(f"⚡[적 전술 전환] {_lbl.get(prev, prev)} → {_lbl.get(new, new)} "
                          f"(최근 요격률 {r_recent*100:.0f}%, 채널포화 {sat*100:.0f}%)")

    def _on_retreat_arrived(self, et):
        """후퇴 도착 위협 처리 — 기본(단발): 재공격 횟수 한도 내 재접근, 아니면 이탈.
        (BattleEngine은 목표지향 압박 유지로 오버라이드 — max_reattacks 캡 제거)"""
        # MED-12: 재공격 가능 시 재접근, 아니면 전장 이탈
        # 무장 소진 시 재접근 불가 — 재무장 위해 복귀(전장 이탈)
        _no_munition = self._munition_limit and et.munition_remaining <= 0
        if et.reattack_count < et.max_reattacks and not _no_munition:
            et.reattack_count += 1
            et.is_retreating = False
            et.retreat_pos   = None
            self._log(f"[재공격] {et.preset_name} 재접근 개시 ({et.reattack_count}/{et.max_reattacks})")
        else:
            et.alive = False
            self._log(f"[이탈] {et.preset_name} 전장 이탈 완료")

    def _reload_delay_s(self, et) -> float:
        """Phase 5.2: 살보 발사 후 다음 발사까지 재장전 지연(초). 부모(단발)는 0 → 동작 보존.
        BattleEngine이 오버라이드해 화력을 시간축으로 분산(파상공격)."""
        return 0.0

    def _enemy_fire(self):
        primary = self._primary()
        for et in self.enemy_threats:
            if not et.alive or et.is_retreating:
                continue
            if not et.info.get('can_fire_missile'):
                continue
            # 무장 유한화: 공격 무장 소진 시 발사 중단 (CIWS 등 방어무장 무관)
            if self._munition_limit and et.munition_remaining <= 0:
                continue
            # v9.6: 기습 잠수함 은닉 중 — 발사 불가
            if et.is_sub and self.t < et.hidden_until:
                continue
            # v16.1: 능동 핑 노출 후 회피 잠항 도주 중 — 발사 불가(도주 우선)
            if et.is_sub and self.t < et.contact_lost_until:
                continue

            in_flight = sum(
                1 for m in self.missiles
                if m.alive and m.owner_id == id(et) and m.mtype == 'enemy_strike'
            )
            if in_flight > 0:
                continue
            # Phase 5.2: 재장전 쿨다운 — 만기 전엔 발사 금지(부모는 next_fire_t=0 → 무영향)
            if self.t < et.next_fire_t:
                continue

            dist_m       = et.pos.dist_to(primary.pos)
            fire_range_m = et.info.get('missile_range_km', 0) * 1000 * 0.85
            if dist_m > fire_range_m:
                continue

            _smin = et.info.get('missile_salvo_min', 1)
            _smax = et.info.get('missile_salvo_max', 2)
            # v15.1 적응형: 포화=최대 / 분산=중간(시차 유도) / 기만=최소(시차 극대)
            if self._adaptive_ai:
                if self._adaptive_mode == 'saturation':
                    salvo = _smax
                elif self._adaptive_mode == 'deception':
                    salvo = _smin
                else:  # dispersal
                    salvo = max(_smin, (_smin + _smax) // 2)
            else:
                salvo = random.randint(_smin, _smax)
            if self._munition_limit:
                salvo = min(salvo, et.munition_remaining)   # 잔여 무장 한도 내로
            m_speed = et.info.get('missile_speed_ms') or 300
            m_name  = et.info.get('missile_name') or '대함미사일'

            for _ in range(salvo):
                _is_torp = '어뢰' in m_name
                offset = LatLon.from_xy(
                    et.pos.x + random.uniform(-500, 500),
                    et.pos.y + random.uniform(-500, 500),
                )
                # v15.1 적응형 포화: 기함(primary) 집중 / 그 외: 기존 분산 선택
                # (random 호출 순서 보존 위해 offset 뒤에서 표적 선택)
                if self._adaptive_ai and self._adaptive_mode == 'saturation' and not _is_torp:
                    _tgt = primary
                else:
                    _tgt = self._pick_target(is_torpedo=_is_torp, threat_pos=et.pos)
                _m = MissileObj(
                    mtype    = 'enemy_strike',
                    name     = m_name,
                    pos      = offset,
                    target   = _tgt,
                    speed_ms = m_speed,
                    pk_base  = _MISSILE_PK_MAP.get(m_name, _MISSILE_PK_DEFAULT),  # MED-5
                    owner_id = id(et),
                    t_spawn  = self.t,
                )
                # 포팅 B: 전술 속성 설정
                _ev = et.info.get('missile_terminal_evasion', 1.0)
                # v15.1 기만: 종말 회피 강화(계수↓ → 요격 더 어려움). 하한 0.2.
                if self._adaptive_ai and self._adaptive_mode == 'deception':
                    _ev = max(0.2, _ev * 0.6)
                _m.terminal_evasion_factor = _ev
                _m.is_torpedo = _is_torp
                self.missiles.append(_m)

            et.has_fired = True
            # Phase 5.2: 다음 발사까지 재장전 지연(부모 _reload_delay_s=0 → 다음 틱 즉시 가능, bit-identical)
            et.next_fire_t = self.t + self._reload_delay_s(et)
            if self._munition_limit:
                et.munition_remaining -= salvo
            self.stats['total_threats'] += salvo
            _munition_out = self._munition_limit and et.munition_remaining <= 0

            if et.is_aircraft:
                et.is_retreating = True
                # 발사 후 200km 후퇴 이탈 (MED-12 재공격 패턴 도입으로 기존 500km에서 단축)
                angle = et.pos.bearing_to(primary.pos) + math.pi
                et.retreat_pos = LatLon.from_xy(
                    et.pos.x + math.cos(angle) * 200_000,
                    et.pos.y + math.sin(angle) * 200_000,
                )
                self._log(
                    f"[적 발사+이탈] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km), 이탈 개시"
                )
            elif et.is_sub:
                # v9.6: dual_weapon — 어뢰 발사 직후 해성-3 순항미사일 동시 발사
                if et.info.get('dual_weapon'):
                    d_name  = et.info.get('dual_missile_name', '')
                    d_spd   = et.info.get('dual_missile_speed_ms', 250)
                    d_salvo = random.randint(
                        et.info.get('dual_salvo_min', 1),
                        et.info.get('dual_salvo_max', 2),
                    )
                    for _ in range(d_salvo):
                        _dm = MissileObj(
                            mtype    = 'enemy_strike',
                            name     = d_name,
                            pos      = LatLon.from_xy(et.pos.x + random.uniform(-300, 300),
                                                      et.pos.y + random.uniform(-300, 300)),
                            target   = self._pick_target(is_torpedo=False, threat_pos=et.pos),
                            speed_ms = d_spd,
                            pk_base  = _MISSILE_PK_MAP.get(d_name, _MISSILE_PK_DEFAULT),
                            owner_id = id(et),
                            t_spawn  = self.t,
                        )
                        _dm.terminal_evasion_factor = 0.87  # 해성-3: 스텔스 저고도 순항
                        _dm.is_torpedo = False
                        self.missiles.append(_dm)
                    self.stats['total_threats'] += d_salvo
                    self._log(
                        f"[기습 동시발사] {et.preset_name} -> {d_name} {d_salvo}발 "
                        f"(어뢰+순항 동시)"
                    )

                # LOW-18: 잠수함 발사 후 회피 기동 (발사 후 반대 방향 50km 이탈)
                et.is_retreating = True
                angle = et.pos.bearing_to(primary.pos) + math.pi
                et.retreat_pos = LatLon.from_xy(
                    et.pos.x + math.cos(angle) * 50_000,
                    et.pos.y + math.sin(angle) * 50_000,
                )
                self._log(
                    f"[적 발사+잠항회피] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km), 반대 방향 잠항"
                )
            else:
                self._log(
                    f"[적 발사] {et.preset_name} -> {m_name} {salvo}발 "
                    f"(거리 {dist_m/1000:.0f}km)"
                )
                # 무장 소진 수상함 — 재무장 위해 전장 이탈(복귀). 도착 시 위협 소멸.
                if _munition_out:
                    et.is_retreating = True
                    _ang = et.pos.bearing_to(primary.pos) + math.pi
                    et.retreat_pos = LatLon.from_xy(
                        et.pos.x + math.cos(_ang) * 300_000,
                        et.pos.y + math.sin(_ang) * 300_000,
                    )
                    self._log(f"[무장 소진·복귀] {et.preset_name} 전장 이탈(재무장)")

    # ── 3단계: 아군 방어 TEWA ─────────────────────────────────────────────────

    def _cd_allowed(self, target_key: str) -> bool:
        """
        C&D 딜레이 판정. target_key는 위협/미사일의 uid(안정 문자열).
        (id() 키는 객체 소멸 후 주소 재사용으로 stale 충돌 위험 → uid로 고정)
        첫 탐지 시: 레이더 빔 드웰(1~3s) + cd_time_s*날씨계수 + confirm_time_s + uniform(2,10)s
        야간/악천후: cd_time_factor로 딜레이 자동 증가.
        이후 시각 도달하면 True.
        """
        if target_key not in self._cd_fire_time:
            cd_factor = self.wx.get('cd_time_factor', 1.0)
            cd     = self.cfg.get('cd_time_s', 10) * cd_factor * self.cfg.get('cd_scale', 1.0)
            conf   = self.cfg.get('confirm_time_s', 3)
            dwell  = random.uniform(1, 3)   # 레이더 빔 드웰 타임
            jitter = random.uniform(2, 10)  # 위협 분류 랜덤 편차
            self._cd_fire_time[target_key] = self.t + dwell + cd + conf + jitter
            return False
        return self.t >= self._cd_fire_time[target_key]

    def _vls_interval_ok(self, ship: 'FriendlyShipObj') -> bool:
        """VLS 연속 발사 간격 2.5s 체크."""
        last = self._vls_last_fire.get(id(ship), -999.0)
        return (self.t - last) >= 2.5

    def _arm_esm_update(self):
        """v16.1 ESM→ARM 역탐지: 아군 레이더 방사 상태로 ARM 조준 좌표를 갱신한다.
        레이더 ON(방사 중)이면 적 ESM이 신호를 실시간 포착 → 조준 좌표를 표적 현재 위치로
        갱신. 레이더 OFF면 ESM 미포착 → 마지막 포착 좌표(stale) 유지 → 표적이 떠난 만큼
        명중 급감(_check_hits에서 판정). enable_esm_arm OFF면 무동작(회귀 bit-identical)."""
        if not self._emcon_arm:
            return
        for m in self.missiles:
            if not (m.alive and m.is_arm):
                continue
            tgt = m.target
            if not isinstance(tgt, FriendlyShipObj):
                continue
            if self.t >= tgt.radar_off_until:
                # 레이더 ON → ESM 실시간 포착, 조준 좌표 갱신
                m.arm_aim_pos = LatLon.from_xy(tgt.pos.x, tgt.pos.y)
            elif m.arm_aim_pos is None:
                # 발사부터 레이더 OFF → 현 위치를 '마지막 포착'으로 1회 고정(이후 stale 유지)
                m.arm_aim_pos = LatLon.from_xy(tgt.pos.x, tgt.pos.y)
            # else: 레이더 OFF 지속 → 기존 stale 좌표 유지(갱신 안 함)

    def _coord_deception_update(self):
        """v16.6 전자 좌표 기만: 아군 ECM이 적 레이더 화면상 함정 표시 위치를 교란 →
        적 대함미사일이 종말권 진입 시 확률적으로 가짜 좌표(decoy_aim_pos)로 유도된다.
        _check_hits에서 실제 표적과의 이격만큼 명중 급감. 레이더 유도 대함만(탄도·HGV·
        ARM·어뢰는 레이더 유도가 아니라 무효 — 기존 ECM 무효 항목과 동일).
        enable_coord_deception OFF면 무동작·random 미소비(회귀 bit-identical)."""
        if not self._coord_deception:
            return
        for m in self.missiles:
            if not (m.alive and m.mtype == 'enemy_strike') or m.decoy_aim_checked:
                continue
            if m.is_ballistic or m.is_hgv or m.is_arm or m.is_torpedo:
                continue
            tgt = m.target
            if not isinstance(tgt, FriendlyShipObj) or not tgt.alive:
                continue
            if m.pos.dist_to(tgt.pos) > self._coord_decep_range_m:
                continue  # 종말권 밖 — 유도 안정, 기만 미적용
            m.decoy_aim_checked = True   # 종말 진입 1회만 판정(매틱 재시도 방지)
            if random.random() < self._coord_decep_rate:
                ang = random.uniform(0, 2 * math.pi)
                off = self._coord_decep_offset_m
                m.decoy_aim_pos = LatLon.from_xy(
                    tgt.pos.x + math.cos(ang) * off,
                    tgt.pos.y + math.sin(ang) * off)

    def _apply_mine_exposure(self):
        """v16.7 기뢰전: 작전 해역 진입 시 함정별 확률적 기뢰 접촉을 1회 판정한다.
        소해 지원(enable_minesweeping) 보유 시 접촉 확률 경감. 계류·해저·자항 3종 차등
        피해(take_hit + 침수). 배수량 큰 함정일수록 감응 기뢰에 취약(소형함은 회피 유리).
        잠수함·해안 포대는 제외. enable_mine_threat OFF면 무동작·random 미소비(회귀 보존)."""
        if not self._mine_threat:
            return
        # v16.12: 무인 소해정(UUV) 편성 시에도 소해 활성화 — 기존 enable_minesweeping과 OR.
        _uuv_sweep = any(s.alive and s.is_unmanned
                         and SHIP_DB[s.ship_type].get('is_minesweeper')
                         for s in self.friendly_ships)
        sweep = _MINE_SWEEP_FACTOR if (self._minesweeping or _uuv_sweep) else 1.0
        for ship in self.friendly_ships:
            # 무인정(USV·UUV)은 소해·정찰 자산이라 기뢰 접촉 판정에서 제외(expendable 피켓)
            if (not ship.alive or ship.is_submarine or ship.is_shore_battery
                    or ship.is_unmanned):
                continue
            disp = (ship._surv or {}).get('displacement_t', 4000)
            suscept = min(1.0, 0.5 + disp / 20000.0)   # 소형~대형 취약도 0.5~1.0
            if random.random() >= self._mine_density * suscept * sweep:
                continue
            # 접촉 — 기뢰 종류 선택(계류 0.5·해저 0.3·자항 0.2)
            r = random.random()
            if r < 0.5:
                mtype, hits, breach = '계류기뢰', 1, 0.35
            elif r < 0.8:
                mtype, hits, breach = '해저감응기뢰', 2, 0.55
            else:
                mtype, hits, breach = '자항기뢰', 2, 0.50
            self.stats['mines_struck'] += 1
            for _ in range(hits):
                if ship.alive:
                    ship.take_hit(mtype, self.t)
            if ship.alive and self.cfg.get('enable_flooding', True):
                ship.add_flooding(breach)
            if not ship.alive:
                self.stats['ships_lost_to_mine'] += 1
            if not self._mc_mode:
                self._log(f"[기뢰] {ship.name} {mtype} 접촉 — HP {ship.hp}"
                          + ("" if ship.alive else " 침몰"))

    def _cyber_update(self):
        """v16.3 사이버전: 주기적으로 3개 채널의 침투를 독립 시도한다. 데이터링크 변조·CIC
        마비는 적→아군(요격 Pk·탐지거리 저하), 레이더 교란은 아군→적(적 발사 Pk 저하).
        각 채널 성공 시 효과가 일정시간 지속되며 경고 로그 없이 발현·해제(은밀성). 전자전이
        탐지거리를 상시 감소시키는 것과 달리 확률적·지속적이다.
        enable_cyber_warfare OFF면 굴림 자체를 건너뜀(random 미소비 → 회귀 bit-identical)."""
        if not self._cyber:
            return
        if self.t < self._cyber_last_t + self._cyber_interval:
            return
        self._cyber_last_t = self.t
        if random.random() < self._cyber_p:
            self._dl_corrupt_until = self.t + self._cyber_effect_s
        if random.random() < self._cyber_p:
            self._cic_blind_until = self.t + self._cyber_effect_s
        if random.random() < self._cyber_p:
            self._enemy_jam_until = self.t + self._cyber_effect_s

    def _hgv_glide_alt(self, p: float, peak: float) -> float:
        """v16.2 HGV 활공 고도 프로파일. p=비행 진행도(0 스폰~1 표적). 교전 구간은
        활공~종말(부스트 상승은 사거리 밖에서 종료) → 스폰점=활공 정점에서 완만히
        하강, 종말에 침투 고도로 급강하."""
        if p < _HGV_GLIDE_END:
            f = p / _HGV_GLIDE_END
            return peak * (1.0 - _HGV_GLIDE_DESCENT * f)
        f = (p - _HGV_GLIDE_END) / (1.0 - _HGV_GLIDE_END)
        start = peak * (1.0 - _HGV_GLIDE_DESCENT)
        # 종말 목표 고도는 활공 끝 고도(start) 이하로 클램프 — 정점이 2km 미만인 초저고도
        # 활공체가 종말에 오히려 상승하는 비물리 역전 방지(감사 발견, 현 DB는 미발현)
        term = min(_HGV_TERMINAL_ALT_M, start)
        return max(term, start + (term - start) * f)

    def _hgv_glide_update(self):
        """v16.2: 활공 HGV의 교전 고도를 비행 진행도에 따라 갱신한다. 이로써
        _select_defense_wpn이 비행 단계별로 다른 요격 층을 고르고, 한 HGV가 고고도
        SM-3 누출 시 저고도 SM-6 Block IB로 재교전된다(미사일이 살아 다음 틱 재판정).
        enable_hgv_glide OFF면 무동작(altitude_m 고정 → 회귀 bit-identical)."""
        if not self._hgv_glide:
            return
        for m in self.missiles:
            if not (m.alive and m.is_hgv):
                continue
            init = getattr(m, '_init_dist', 0)
            tgt  = m.target
            if init <= 0 or not (tgt and hasattr(tgt, 'pos')):
                continue
            p = max(0.0, min(1.0, 1.0 - m.pos.dist_to(tgt.pos) / init))
            m.altitude_m = self._hgv_glide_alt(p, m._peak_alt_m)

    def _ballistic_descent_update(self):
        """v20.2a: 탄도미사일의 교전 고도를 표적까지 잔여거리에 따라 강하시킨다.
        강하각 일정 근사(고도 = 잔여거리 × tan) + DB 정점고도 캡.
        이로써 SM-3(≥40km) → THAAD(10~150km) → L-SAM(40~70km) → 천궁-II(≤20km)가
        접근 거리대별로 순차 교전창을 갖는다.
        enable_ballistic_descent OFF면 무동작(altitude_m 고정 → 회귀 bit-identical)."""
        if not self._bal_descent:
            return
        for m in self.missiles:
            if not (m.alive and m.is_ballistic):
                continue
            tgt = m.target
            if not (tgt and hasattr(tgt, 'pos')):
                continue
            rem  = m.pos.dist_to(tgt.pos)
            peak = getattr(m, '_peak_alt_m', 0.0)
            m.altitude_m = max(0.0, min(peak, rem * _BALLISTIC_DESCENT_TAN))

    def _target_difficulty_factor(self, sam, tgt) -> float:
        """요격 Pk에 곱할 표적 난이도 계수 (속도·RCS 기반). enable_target_difficulty OFF면 1.0.

        고속·소형 표적일수록 요격이 어렵다는 물리를 Pk에 반영한다. 기준 표적(아음속
        대함미사일 300 m/s·0.1m²)에서 정확히 1.0이라 기존 결과는 보존되고, 마하 3급
        소형 ARM 같은 극단 표적에서만 유의하게 내려간다. 상수 근거는 _TDIFF_* 참조."""
        if not self._target_difficulty:
            return 1.0
        v_tgt = getattr(tgt, 'speed_ms', 0.0) or 0.0
        v_sam = getattr(sam, 'speed_ms', 0.0) or 0.0
        if v_tgt <= 0.0 or v_sam <= 0.0:
            return 1.0

        # 속도: 접근률이 클수록 종말 유도 여유가 준다. 기준 표적에서 1.0.
        # ⚠ 단, 탄도·극초음속 표적은 면제한다. pk_base는 이미 '그 무기의 설계 표적'에 대한
        # 값이다 — SM-3·THAAD·L-SAM의 Pk는 애초에 마하 10급 탄도를 상대로 매긴 수치이고,
        # 거기에 속도 페널티를 또 곱하면 이중 계상이 된다(요격 전용 계층이 자기 설계 표적에
        # 벌점을 받는 꼴). 속도 계수는 대공 SAM이 '설계 표적을 벗어난' 초음속 대함미사일·
        # 대방사미사일을 상대할 때 걸리는 것이다. 소형 표적 페널티(RCS)는 탄도에도 그대로 적용.
        if getattr(tgt, 'is_ballistic', False) or getattr(tgt, 'is_hgv', False):
            f_v = 1.0
        else:
            f_v = (_TDIFF_V_REF + v_sam) / (v_tgt + v_sam)

        # RCS: 표적 DB가 있는 위협만(적 함정·항공기가 발사한 대함미사일은 원본 dict가 없어
        # 중립 1.0 — 속도만 반영된다). 기준 RCS 이상은 상한 1.0으로 잘라 큰 표적이
        # 기존보다 더 잘 맞는 일이 없게 한다(하위 호환 방향).
        _src = getattr(tgt, 'enemy_info', None) or getattr(tgt, 'info', None) or {}
        rcs  = _src.get('rcs_m2') if isinstance(_src, dict) else None
        if rcs and rcs > 0.0:
            f_rcs = min(1.0, (rcs / _TDIFF_RCS_REF) ** _TDIFF_RCS_EXP)
        else:
            f_rcs = 1.0

        return max(_TDIFF_MIN, min(1.0, f_v * f_rcs))

    def _arm_radar_off_check(self):
        """ARM 탐지 시 표적 함정 레이더 침묵 (ARM 회피 전술 — ARM이 지나갈 때까지 유지).

        v20.5(B-2) 게이트 정정. 과거에는 ARM당 **한 번만** 8초 끄고 말았다. ARM이 120km
        밖에 있을 때 8초 껐다가 다시 켜면, 초음속 ARM이 남은 100여 초를 날아오는 동안
        레이더가 켜져 있어 **정작 명중 시점엔 회피가 작동하지 않았다** — ARM 명중 판정
        (`_check_hits`)과 ESM 역탐지(`_arm_esm_update`)가 둘 다 '명중 순간 레이더 OFF'를
        조건으로 보는데, 그 순간엔 이미 레이더가 복귀해 있었던 것. 그래서 ESM→ARM 역탐지
        (`enable_esm_arm`)의 효과가 편성을 24발로 늘려도 **완전히 0**이었다(bit-identical).

        실제 교리는 **ARM이 종말 유도 구간에 들어올 때 방사를 끊는 것**이고, 그 대가로
        그 구간 동안 대공 탐지·교전이 저하되는 것이 EMCON 딜레마의 본질이다(대가는
        `is_radar_off`로 이미 구현돼 있다).
        → ARM이 **종말권(20km) 안에 살아 있는 동안** 침묵을 연장한다. ARM이 요격·소멸하면
        갱신이 멈춰 off_dur 뒤 레이더가 자동 복귀한다.

        ⚠ 침묵 구간을 넓게 잡으면(예: 과거 트리거 거리인 120km 내내) 레이더가 사실상 영구
        침묵해 SAM 교전 자체가 불가능해지고 함대가 전멸한다(실측: 요격률 0.52→0.05).
        그건 회피가 아니라 방어 포기다 — 그래서 종말 구간으로 좁힌다.
        """
        if not self.cfg.get('enable_radar_off', True):
            return
        off_dur  = 8.0         # 침묵 연장 단위 — ARM이 종말권에 살아 있는 한 매 틱 갱신된다
        warn_m   = 20_000.0    # ARM 종말 유도 구간(20km) 진입 시 레이더 침묵
        for m in self.missiles:
            if not (m.alive and m.is_arm):
                continue
            tgt = m.target
            if not isinstance(tgt, FriendlyShipObj) or not tgt.alive:
                continue
            dist_m = m.pos.dist_to(tgt.pos)
            if dist_m >= warn_m:
                continue
            if not getattr(m, '_radar_off_triggered', False):
                m._radar_off_triggered = True     # 로그는 최초 1회만
                self._log(
                    f"[레이더 OFF] {tgt.name} ARM {dist_m/1000:.0f}km 접근 탐지 — "
                    f"ARM 통과까지 레이더 침묵 ({m.name})"
                )
            # ARM이 살아 있는 동안 침묵 유지(연장). 여러 ARM이면 가장 늦은 시각까지.
            tgt.radar_off_until = max(tgt.radar_off_until, self.t + off_dur)

    def _fire_ground_sam(self, wpn_key: str, target, dist_m: float,
                         name: str, pk: float, speed_ms: float,
                         cost: float, label: str, owner_id: int):
        """지상 BMD 자산 SAM 발사 (4계층 공용).
        owner_id: -1=어쇼어 SM-3, -2=THAAD, -3=L-SAM, -4=천궁-II."""
        self.ground_inv[wpn_key] -= 1
        self._ground_cost += cost
        self.stats['total_missiles_fired'] += 1
        if self.stats['t_first_fire'] < 0:
            self.stats['t_first_fire'] = self.t
        stat_key = _GROUND_BMD_STAT_KEY.get(owner_id)
        if stat_key is None:   # 새 지상 자산 추가 시 통계 키 등록 누락을 조용히 넘기지 않는다
            raise ValueError(f'지상 BMD 통계 키 미등록: owner_id={owner_id} ({name})')
        self.stats[stat_key] += 1
        sam = MissileObj(
            mtype    = 'friendly_sam',
            name     = name,
            pos      = self._primary().pos,
            target   = target,
            speed_ms = speed_ms,
            pk_base  = pk,
            owner_id = owner_id,
            t_spawn  = self.t,
        )
        sam.rcs_m2 = _SAM_RCS.get(name, 0.001)  # Phase C: 적 레이더 탐지용
        self.missiles.append(sam)
        tgt_name = target.name if hasattr(target, 'name') else getattr(target, 'preset_name', '?')
        self._log(f"[지상 BMD] {label} → {name} 발사 → {tgt_name} (거리 {dist_m/1000:.0f}km)")

    def _ashore_defense(self, sorted_missiles: list):
        """
        지상 BMD 5계층 교전 — 탄도/HGV 전담.

        레이어 순서 (교전거리 내림차순 = 먼 층부터):
          1. 이지스 어쇼어 SM-3 — 중간단계 (고도 ≥ 40km, 사거리 500km)
          2. THAAD — 종말고고도 (고도 10~150km, 사거리 200km)
          3. L-SAM — 종말 상층 (고도 40~70km, 사거리 150km)          [v20.2a]
          4. 패트리엇 PAC-3 MSE — 종말 중층 (고도 2~25km, 사거리 60km) [v20.1]
          5. 천궁-II — 종말 하층 점방어 (고도 0.5~20km, 사거리 20km) [v20.2a]
        함정 SM-3는 이 자산들이 소진되거나 실패한 경우의 최후 백업.

        ⚠ 사거리(dist_m)는 포대 기준(기함 위치로 추상화)이고 고도는 표적함 접근 기하로
        정해지므로, 함대가 넓게 분산(DMO)해 표적이 기함에서 멀면 사거리 짧은 천궁-II는
        고도창을 만족해도 사거리에서 탈락할 수 있다(최후 계층이 구조적으로 희소).
        """
        enable_ashore   = self.cfg.get('enable_ashore',   False)
        enable_thaad    = self.cfg.get('enable_thaad',    False)
        enable_lsam     = self.cfg.get('enable_lsam',     False)
        enable_chungung = self.cfg.get('enable_chungung', False)
        enable_patriot  = self.cfg.get('enable_patriot',  False)   # v20.1
        primary_pos     = self._primary().pos

        for m in sorted_missiles:
            if not m.alive or getattr(m, 'intercepted', False):
                continue
            if not (m.is_ballistic or m.is_hgv):
                continue

            dist_m = m.pos.dist_to(primary_pos)
            alt    = m.altitude_m
            # v20.2a: 지상 BMD는 함정과 별개 사격통제 — 함정 SAM 유도 수와 무관하게
            # 자체 상한을 갖는다(owner_id<0 = 지상 자산). 이전처럼 함정·지상을 합산하면
            # 함정 SM-6가 상한을 선점해 종말 계층(THAAD·L-SAM·천궁-II)이 영구 차단된다.
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is m and s.mtype == 'friendly_sam'
                and s.owner_id < 0
            )
            max_sams = _GROUND_BMD_MAX_SAMS   # 5계층 각 1발 기회

            # ── 1차: 이지스 어쇼어 SM-3 (중간단계) ──────────────────────────
            if (enable_ashore
                    and self.ground_inv.get('SM-3 (어쇼어)', 0) > 0
                    and (self.t - self._ashore_last_fire) >= _ASHORE_COOLDOWN_S
                    and alt >= _ASHORE_ALT_MIN_M
                    and dist_m <= _ASHORE_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'SM-3 (어쇼어)',
                    target   = m, dist_m = dist_m,
                    name     = 'SM-3 (어쇼어)',
                    pk       = _ASHORE_SM3_PK,
                    speed_ms = _ASHORE_SM3_SPD_MS,
                    cost     = _ASHORE_SM3_COST,
                    label    = '어쇼어',
                    owner_id = -1,
                )
                self._ashore_last_fire = self.t
                sams_on += 1

            # ── 2차: THAAD (종말고고도) ───────────────────────────────────────
            if (enable_thaad
                    and self.ground_inv.get('THAAD 요격탄', 0) > 0
                    and (self.t - self._thaad_last_fire) >= _THAAD_COOLDOWN_S
                    and _THAAD_ALT_MIN_M <= alt <= _THAAD_ALT_MAX_M
                    and dist_m <= _THAAD_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'THAAD 요격탄',
                    target   = m, dist_m = dist_m,
                    name     = 'THAAD 요격탄',
                    pk       = _THAAD_PK,
                    speed_ms = _THAAD_SPD_MS,
                    cost     = _THAAD_COST,
                    label    = 'THAAD',
                    owner_id = -2,
                )
                self._thaad_last_fire = self.t
                sams_on += 1

            # ── 3차: L-SAM (종말 상층 — 한국형 BMD) ──────────────────────────
            if (enable_lsam
                    and self.ground_inv.get('L-SAM', 0) > 0
                    and (self.t - self._lsam_last_fire) >= _LSAM_COOLDOWN_S
                    and _LSAM_ALT_MIN_M <= alt <= _LSAM_ALT_MAX_M
                    and dist_m <= _LSAM_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'L-SAM',
                    target   = m, dist_m = dist_m,
                    name     = 'L-SAM',
                    pk       = _LSAM_PK,
                    speed_ms = _LSAM_SPD_MS,
                    cost     = _LSAM_COST,
                    label    = 'L-SAM',
                    owner_id = -3,
                )
                self._lsam_last_fire = self.t
                sams_on += 1

            # ── 4차: 패트리엇 PAC-3 MSE (종말 중층 — 한·미 연합) ─────────────
            if (enable_patriot
                    and self.ground_inv.get('PAC-3 MSE', 0) > 0
                    and (self.t - self._patriot_last_fire) >= _PATRIOT_COOLDOWN_S
                    and _PATRIOT_ALT_MIN_M <= alt <= _PATRIOT_ALT_MAX_M
                    and dist_m <= _PATRIOT_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = 'PAC-3 MSE',
                    target   = m, dist_m = dist_m,
                    name     = 'PAC-3 MSE',
                    pk       = _PATRIOT_PK,
                    speed_ms = _PATRIOT_SPD_MS,
                    cost     = _PATRIOT_COST,
                    label    = '패트리엇',
                    owner_id = -5,
                )
                self._patriot_last_fire = self.t
                sams_on += 1

            # ── 5차: 천궁-II (종말 하층 점방어 — 최후 계층) ──────────────────
            if (enable_chungung
                    and self.ground_inv.get('천궁-II', 0) > 0
                    and (self.t - self._chungung_last_fire) >= _CHUNGUNG_COOLDOWN_S
                    and _CHUNGUNG_ALT_MIN_M <= alt <= _CHUNGUNG_ALT_MAX_M
                    and dist_m <= _CHUNGUNG_RANGE_M
                    and sams_on < max_sams):
                self._fire_ground_sam(
                    wpn_key  = '천궁-II',
                    target   = m, dist_m = dist_m,
                    name     = '천궁-II',
                    pk       = _CHUNGUNG_PK,
                    speed_ms = _CHUNGUNG_SPD_MS,
                    cost     = _CHUNGUNG_COST,
                    label    = '천궁-II',
                    owner_id = -4,
                )
                self._chungung_last_fire = self.t
                sams_on += 1

    def _target_sort_key(self, obj, primary_pos):
        """위협 교전 우선순위 정렬 키 (sorted reverse=True, 클수록 먼저 교전).
        기본 = 임박도(speed / 교전창 잔여거리). BattleEngine이 오버라이드해
        표적 우선순위 전술을 주입한다 — 부모 동작 보존 훅(v15.11.02)."""
        is_bal = getattr(obj, 'is_ballistic', False)
        is_hgv = getattr(obj, 'is_hgv',       False)
        is_qbm = getattr(obj, 'is_qbm',       False)
        if is_bal or is_hgv:
            floor = 150_000.0   # SM-3 중간단계 교전창 하한 ~150km
        elif is_qbm:
            floor = 20_000.0    # SM-6 최소 교전거리 ~20km
        else:
            floor = 5_000.0     # CIWS/RAM 최소 교전거리 ~5km
        d_eff = max(obj.pos.dist_to(primary_pos) - floor, 200.0)
        return getattr(obj, 'speed_ms', 300.0) / d_eff

    def _friendly_defense(self):
        """
        다층 방어 (enable_layered_defense=True, 기본 ON):
          KDX-III-B2 → KDX-III-B1 → KDX-II → FFX-III → FFX-II → FFX-I 순서로 위협당 1발씩 배정.

        CEC 사전 동시 배정 (enable_cec_preassign=True, 기본 OFF):
          탐지 즉시 1차+2차 함정 동시 발사. 위협당 최대 2발 허용.

        NEW-AW: 위협 긴급도 정렬 — 속도/잔여거리 내림차순 (빠르고 가까운 위협 먼저)
        NEW-AW: Shoot-Look-Shoot — 탄도/초음속/HGV/QBM 위협은 CEC 없이도 2발 배정
        """
        # CEC 두절 시나리오: enable_cec_jammed=True 이면 CEC 강제 해제 + 독립 교전
        cec_jammed = self.cfg.get('enable_cec_jammed', False)
        # v16.13.02 트랙 C: 함정 자율 교전 — 중앙 조율 없이 독립(메시 relay 유지, 협동 살보 +1만 제외)
        autonomous = self.cfg.get('enable_autonomous_engagement', False)
        # v10.4: enable_cec — 탐지 커버리지 통합 + SAM 사전 동시 배정
        # enable_cec_preassign 하위 호환 유지 (구버전 cfg 지원)
        cec_base = self.cfg.get('enable_cec', self.cfg.get('enable_cec_preassign', True)) and not cec_jammed

        # v16.13.02 기함(지휘 노드) 격침 감지 → CEC relay 저하 + 차순위 지휘권 인수 지연.
        # 자율 교전은 애초에 지휘 노드에 의존하지 않아 이 저하에 면역(강건성).
        cur_primary = self._primary()
        if self._prev_primary_id is not None and id(cur_primary) != self._prev_primary_id:
            self._command_handovers += 1
            if cec_base and not autonomous:
                self._cec_degraded_until = self.t + _CEC_HANDOVER_DELAY_S
                if not self._mc_mode:
                    self._log(f"⚠[지휘권 인수] 기함 격침 — 차순위 함정 지휘 승계, "
                              f"CEC 지휘 저하 {_CEC_HANDOVER_DELAY_S:.0f}s")
        self._prev_primary_id = id(cur_primary)

        # 실효 CEC relay(협동 탐지 커버리지) + 협동 살보 보너스 결정.
        # 자율 교전 = 메시 네트워크: relay(분산 협동 탐지)는 유지하되 중앙 조율 살보 +1은 없음.
        # 대신 지휘 노드(기함)에 의존하지 않아 기함 격침 시 CEC 저하에 면역(강건성).
        if autonomous:
            cec = cec_base                                       # 메시 relay 유지 — 기함 저하 면역
            salvo_bonus = 0                                      # 중앙 조율 없음 — 협동 살보 +1 제외
        elif cec_base and self.t < self._cec_degraded_until:
            cec = False                                          # 지휘권 인수 중 relay 일시 저하
            salvo_bonus = 0
        else:
            cec = cec_base
            salvo_bonus = 1 if cec_base else 0

        # CEC 두절 시 각 함정이 독립적으로 교전 (다층 방어 비활성화, 1함정=1교전)
        layered = self.cfg.get('enable_layered_defense', True) and not cec_jammed

        primary_pos = cur_primary.pos   # 위에서 구한 기함 재사용(틱당 _primary 중복 조회 제거)
        _urgency = lambda obj: self._target_sort_key(obj, primary_pos)  # noqa: E731

        # 다층 방어 우선순위 정렬 (B2=0, B1=1, KDX-II=2, FFX-III=3, FFX-II=4, FFX-I=5, 나머지=99)
        if layered:
            sorted_ships = sorted(
                [s for s in self.friendly_ships if s.alive],
                key=lambda s: SHIP_LAYER_PRI.get(s.ship_type, 99)
            )
        else:
            # CEC 두절: 함정별로 독립 교전 — 처리 순서는 무작위
            import random as _rnd
            _ships = [s for s in self.friendly_ships if s.alive]
            _rnd.shuffle(_ships)
            sorted_ships = _ships

        # 주요 함정 레이더 OFF 여부 (ARM 회피 전술)
        primary_ship = cur_primary
        is_radar_off = self.t < primary_ship.radar_off_until

        # (A) 적 대함 미사일 요격 — 긴급도 순 정렬
        # ⚠ 어뢰 제외(v20.5 B-3): 어뢰는 **수중** 표적이다. 함대공 미사일(SM-2·SM-6·해궁)과
        # CIWS는 대공 무기라 물속을 달리는 어뢰를 요격할 수 없다. 그런데 과거에는 어뢰가
        # mtype=='enemy_strike'라는 이유로 이 목록에 섞여 들어와, 레이더가 "어뢰를 포착"하고
        # SM-2가 "어뢰를 격추"했다(로그로 확인). 그 결과 **잠수함이 쏜 어뢰가 전부 SAM에
        # 요격돼 아군 피격이 사실상 0**이었고(대잠 시나리오 피격 0.00~0.02), 잠수함 위협 자체가
        # 무력했다 — 그래서 능동 소나 EMCON 딜레마("핑을 켜면 들킨다")도 성립할 수 없었다.
        # 들켜봤자 잠수함이 아무것도 못 하니 숨을 이유가 없었던 것.
        # 어뢰 대응은 대공 요격이 아니라 **기만기(enable_decoy)·회피 기동(enable_evasion)**이
        # 담당한다(_check_hits에 이미 구현돼 있다).
        sorted_missiles = sorted(
            [m for m in self.missiles
             if m.alive and m.mtype == 'enemy_strike' and not m.is_torpedo],
            key=_urgency, reverse=True
        )

        # 지상 BMD 선제 교전 (어쇼어 SM-3 → THAAD, 함정 SM-3보다 우선)
        if (self.cfg.get('enable_ashore',   False)
                or self.cfg.get('enable_thaad',    False)
                or self.cfg.get('enable_lsam',     False)
                or self.cfg.get('enable_chungung', False)
                or self.cfg.get('enable_patriot',  False)):
            self._ashore_defense(sorted_missiles)

        for m in sorted_missiles:
            # 1단계: 수색 레이더 탐지 확정 (SPY-1D 빔 회전 0~6초 지연)
            if not self.search_radar.try_detect(m.uid, self.t, radar_off=is_radar_off):
                continue
            # 2단계: 추적 채널 획득 (18채널 한계 + 3초 획득 지연)
            if not self.track_radar.try_track(m.uid, self.t):
                continue
            # 3단계(C&D): 분류·교전 결심 딜레이
            if not self._cd_allowed(m.uid):
                continue
            # 4단계: 피아식별 (IFF)
            if not self._iff_check(m):
                continue
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is m and s.mtype == 'friendly_sam'
            )
            # 위협 유형별 살보 수 — CEC 두절 시 1발 고정
            # HGV(극초음속 활공체): SM-3 교전창 극히 좁음 → 3발
            # 탄도탄·QBM·초음속(≥1000m/s): Shoot-Look-Shoot → 2발
            # 기타: 1발 / CEC 추가 협동 배정: +1발
            if m.is_hgv:
                _base = 3
            elif m.is_ballistic or m.is_qbm or m.speed_ms >= 1000:
                _base = 2
            else:
                _base = 1
            max_sams = 1 if cec_jammed else (_base + salvo_bonus)
            # BUG-1 fix: 전술 살보 수는 위협별 최솟값(_base) 이하로 내려가지 않음
            _tac_max = self.cfg.get('_tactical_max_salvo')
            if _tac_max is not None:
                max_sams = max(int(_tac_max), _base)
            if sams_on >= max_sams:
                continue

            shots = 0
            for ship in sorted_ships:
                if sams_on + shots >= max_sams:
                    break
                if not self._vls_interval_ok(ship):
                    continue
                dist_m = ship.pos.dist_to(m.pos)
                # v10.4: CEC 탐지 커버리지 — 함정 자체 탐지거리 초과 여부
                own_detect_m = self._detect_range_m(ship, '대공', m.altitude_m)
                if dist_m > own_detect_m:
                    if not cec:   # CEC 미활성: 자체 탐지 불가 함정 교전 불가
                        continue
                    relay = True  # CEC 중계: 아군 데이터링크로 교전, Pk 패널티 적용
                else:
                    relay = False
                wpn = self._select_defense_wpn(ship, m, dist_m)
                if not wpn or ship.channels_used >= ship.max_channels:
                    continue
                if ship.inventory.get(wpn, 0) <= 0:
                    continue
                self._launch_friendly_sam(ship, wpn, m, dist_m, is_aa=False, cec_relay=relay)
                self._vls_last_fire[id(ship)] = self.t
                shots += 1

        # (B) 적 항공기 직접 요격 — 긴급도 순 정렬
        sorted_ac = sorted(
            [et for et in self.enemy_threats
             if et.alive and et.is_aircraft and not et.is_retreating],
            key=_urgency, reverse=True
        )
        for et in sorted_ac:
            # 1단계: 수색 레이더 탐지 확정
            if not self.search_radar.try_detect(et.uid, self.t, radar_off=is_radar_off):
                continue
            # 2단계: 추적 채널 획득
            if not self.track_radar.try_track(et.uid, self.t):
                continue
            # 3단계(C&D): 분류·교전 결심 딜레이
            if not self._cd_allowed(et.uid):
                continue
            # 4단계: 피아식별 (IFF)
            if not self._iff_check(et):
                continue
            sams_on = sum(
                1 for s in self.missiles
                if s.alive and s.target is et and s.mtype == 'friendly_sam'
            )
            # 항공기 위협 유형별 살보 수
            # HGV 항공기(킨잘·지르콘 등) 또는 극초음속(≥1500m/s): 3발
            # 초음속(≥600m/s): 2발 / 기타: 1발 / CEC 추가: +1발
            if getattr(et, 'is_hgv', False) or et.speed_ms >= 1500:
                _base_ac = 3
            elif et.speed_ms >= 600:
                _base_ac = 2
            else:
                _base_ac = 1
            max_sams = 1 if cec_jammed else (_base_ac + salvo_bonus)
            # BUG-1 fix: 전술 살보 수는 위협별 최솟값(_base_ac) 이하로 내려가지 않음
            _tac_max = self.cfg.get('_tactical_max_salvo')
            if _tac_max is not None:
                max_sams = max(int(_tac_max), _base_ac)
            if sams_on >= max_sams:
                continue

            shots = 0
            for ship in sorted_ships:
                if sams_on + shots >= max_sams:
                    break
                if not self._vls_interval_ok(ship):
                    continue
                dist_m = ship.pos.dist_to(et.pos)
                # v10.4: CEC 탐지 커버리지
                own_detect_m = self._detect_range_m(ship, '대공', et.altitude_m)
                if dist_m > own_detect_m:
                    if not cec:
                        continue
                    relay = True
                else:
                    relay = False
                wpn = self._select_aa_wpn(ship, et, dist_m)
                if not wpn or ship.channels_used >= ship.max_channels:
                    continue
                if ship.inventory.get(wpn, 0) <= 0:
                    continue
                self._launch_friendly_sam(ship, wpn, et, dist_m, is_aa=True, cec_relay=relay)
                self._vls_last_fire[id(ship)] = self.t
                shots += 1

    def _laser_defense(self):
        """v17.2: 지향성 에너지 무기(레이저·DEW) 방어. enable_laser_dew ON 경로.

        CIWS 채널과 **독립된 별도 경로**로, 레이저 장착 생존함마다 저속 표적(드론·자폭정·
        아음속 순항미사일)을 1개씩 조사(dwell)해 격추한다. 표적당 조사시간(누적 에너지)이
        필요해 사실상 1채널 — 동시 다표적엔 약하나, CIWS 동시교전 채널 포화와 무관하게
        처치량을 더한다(설계: 채널 포화 우회 vs 1채널 dwell 트레이드오프).

        OFF이거나 레이저 미장착(laser_kw=0)이면 무동작 → 회귀 bit-identical."""
        if not self.cfg.get('enable_laser_dew', False):
            return
        for ship in self.friendly_ships:
            if ship.alive and ship.laser_kw > 0:
                self._laser_engage_ship(ship)

    def _laser_lookup(self, uid):
        """lock된 표적 uid를 현재 생존 미사일·위협에서 조회(없으면 None)."""
        if uid is None:
            return None
        for m in self.missiles:
            if m.uid == uid:
                return m if (m.alive and not m.intercepted) else None
        for et in self.enemy_threats:
            if et.uid == uid:
                return et if (et.alive and not et.intercepted) else None
        return None

    def _laser_eligible_missile(self, m) -> bool:
        # 아음속 순항미사일만(초음속·탄도·HGV·QBM 제외)
        return (m.alive and not m.intercepted and m.mtype == 'enemy_strike'
                and not m.is_ballistic and not m.is_hgv and not m.is_qbm
                and m.speed_ms <= _LASER_MAX_TARGET_MS)

    def _laser_eligible_threat(self, et) -> bool:
        if not et.alive or et.intercepted:
            return False
        # 저속 무인기(자폭 드론) — 유인 전투기 제외
        if et.is_aircraft and not et.is_retreating and et.speed_ms <= _LASER_MAX_AIRCRAFT_MS:
            return True
        # 수상 자폭정(고속정 type, is_suicide)
        if et.is_ship and et.info.get('is_suicide') and et.speed_ms <= _LASER_MAX_TARGET_MS:
            return True
        return False

    def _laser_ekill(self, tgt) -> float:
        if isinstance(tgt, MissileObj):
            return _LASER_EKILL_KJ['missile']
        # 드론은 is_aircraft·is_ship 양쪽 True일 수 있어 항공기(드론)를 먼저 판정
        if getattr(tgt, 'is_aircraft', False):
            return _LASER_EKILL_KJ['drone']
        if tgt.is_ship:
            return _LASER_EKILL_KJ['boat']
        return _LASER_EKILL_KJ['drone']

    def _laser_acquire(self, ship):
        """사거리 내 유효 표적 중 최근접 1개 lock(없으면 None)."""
        rng_m = _LASER_RANGE_KM * 1000
        best, best_d = None, rng_m
        for m in self.missiles:
            if self._laser_eligible_missile(m):
                d = ship.pos.dist_to(m.pos)
                if d <= best_d:
                    best, best_d = m, d
        for et in self.enemy_threats:
            if self._laser_eligible_threat(et):
                d = ship.pos.dist_to(et.pos)
                if d <= best_d:
                    best, best_d = et, d
        return best

    def _laser_engage_ship(self, ship: FriendlyShipObj):
        # 1) 현재 조사 표적 유지(lock) — 유효하지 않으면 신규 획득
        tgt = self._laser_lookup(ship._laser_target_uid)
        if tgt is None:
            tgt = self._laser_acquire(ship)
            ship._laser_dwell_acc = 0.0
            ship._laser_target_uid = tgt.uid if tgt is not None else None
            if tgt is None:
                return
        # 2) 사거리 이탈 시 lock 해제
        dist_m = ship.pos.dist_to(tgt.pos)
        if dist_m > _LASER_RANGE_KM * 1000:
            ship._laser_target_uid = None
            ship._laser_dwell_acc = 0.0
            return
        # 3) 도달출력 P[kW] = 정격 × (ref/거리)² (이내=정격 상한). 잉여 전력 한도 적용.
        ref_m = _LASER_REF_KM * 1000
        reach = ship.laser_kw * min(1.0, (ref_m / max(dist_m, 1.0)) ** 2)
        avail = power_avail_kw(ship.ship_type, _LASER_NOMINAL_SPEED_MS)
        power = min(reach, avail)   # kW = kJ/s
        # 4) 누적 조사 에너지 → E_kill 도달 시 격추
        ship._laser_dwell_acc += power * DT
        if ship._laser_dwell_acc >= self._laser_ekill(tgt):
            self._laser_kill(ship, tgt, dist_m)
            ship._laser_target_uid = None
            ship._laser_dwell_acc = 0.0

    def _laser_kill(self, ship: FriendlyShipObj, tgt, dist_m: float):
        """레이저 격추 처리 — 기존 SAM/CIWS 격추와 동일 상태 규약(집계 일관)."""
        tgt.alive = False
        tgt.intercepted = True
        self.stats['laser_kills'] += 1
        if isinstance(tgt, MissileObj):
            tgt.t_intercept = self.t
            # 아음속 미사일 격추는 SAM과 동일하게 intercepted_threats에 집계(집계 규약 일치)
            self.stats['intercepted_threats'] += 1
            tgt.intercept_weapon = '레이저(DEW)'
            tgt.intercept_km = dist_m / 1000
            tgt_name = tgt.name
        else:
            # 항공기·자폭정 플랫폼은 enemy_ships_destroyed로 집계(SAM 규약과 일치, 4459~ 참조)
            tgt.t_intercept = self.t
            tgt_name = tgt.preset_name
        if not self._mc_mode:
            self._log(f"[레이저 격추] {ship.name} → {tgt_name} "
                      f"(거리 {dist_m/1000:.1f}km, {self.t:.0f}s)")

    def _launch_friendly_sam(self, ship: FriendlyShipObj, wpn: str, target,
                              dist_m: float, is_aa: bool, cec_relay: bool = False):
        wpn_info = FRIENDLY_DB[wpn]
        ship.inventory[wpn]   -= 1
        ship.channels_used    += 1
        ship.total_cost       += wpn_info['cost_usd']
        # 포팅 D: 발사 통계
        self.stats['total_missiles_fired'] += 1
        if self.stats['t_first_fire'] < 0:
            self.stats['t_first_fire'] = self.t
        # v8.16: 한미 연합 기여도 카운트
        if ship.nation == 'USA':
            self.stats['usa_shots'] += 1
            self.stats['usa_cost']  += wpn_info['cost_usd']
        else:
            self.stats['kor_shots'] += 1
            self.stats['kor_cost']  += wpn_info['cost_usd']
        sam = MissileObj(
            mtype    = 'friendly_sam',
            name     = wpn,
            pos      = ship.pos,
            target   = target,
            speed_ms = wpn_info['speed_ms'],
            pk_base  = wpn_info['pk_dist']['mean'],
            owner_id = id(ship),
            t_spawn  = self.t,
        )
        sam.rcs_m2   = _SAM_RCS.get(wpn, 0.001)  # Phase C: 적 레이더 탐지용
        sam.cec_relay = cec_relay                  # v10.4: CEC 중계 교전 여부
        self.missiles.append(sam)
        prefix = '[대공 방어]' if is_aa else '[방어]'
        tgt_name = target.name if hasattr(target, 'name') else target.preset_name
        self._log(f"{prefix} {ship.name} -> {wpn} 발사 -> {tgt_name} (거리 {dist_m/1000:.1f}km)")
        # 탄약 재보급 한계: VLS 주요 무기 완전 소진 시 경고 + 시각 기록
        vls_wpns = ['SM-3 Block IIA', 'SM-6', 'SM-2 Block IIIB', 'RIM-116 RAM']
        if not ship._vls_depleted:
            if all(ship.inventory.get(w, 0) == 0 for w in vls_wpns):
                ship._vls_depleted = True
                self._log(f"[경고] {ship.name} VLS 탄약 완전 소진 — 방어 불능")
                # v9.4: 고갈 발생 시각 기록 (최초 1회만)
                self.vls_depletion_t.setdefault(ship.name, self.t)

    # LOW-11: 조명기(SPG-62) 가용 채널 (SM-2는 반능동 유도 → 조명기 필요)
    _ILLUMINATOR_MAX = {
        'KDX-III-B2': 3, 'KDX-III-B1': 3, 'KDX-II': 2,
        'FFX-III': 2, 'FFX-II': 2, 'FFX-I': 1,
    }
    # NEW-AW: 함대 포진 기준 반경 (Batch II 중심, B1 1km, KDX-II 3km, FFX 5km)
    _FORMATION_RADIUS = {
        'KDX-III-B2': 0, 'KDX-III-B1': 1_000, 'KDX-II': 3_000,
        'FFX-III': 5_000, 'FFX-II': 5_000, 'FFX-I': 5_000,
    }

    def _sm2_illuminator_ok(self, ship: FriendlyShipObj) -> bool:
        """SM-2 추가 발사 가능 여부: 현재 비행 중 SM-2 수 < 조명기 최대 채널."""
        max_ill = self._ILLUMINATOR_MAX.get(ship.ship_type, 1)
        in_flight = sum(1 for s in self.missiles
                        if s.alive and s.name == 'SM-2 Block IIIB' and s.owner_id == id(ship))
        return in_flight < max_ill

    def _select_defense_wpn(self, ship: FriendlyShipObj, m: MissileObj,
                            dist_m: float) -> Optional[str]:
        """미사일 위협 요격 무기 선택. 고도·유형 인식. 무장 피탄 비활성화 반영.
        v10.7: _tactical_wpn_priority cfg 키가 있으면 해당 무기 우선 반환."""
        # 전술 우선순위 반영: 지정 무기가 재고 있으면 사거리 내에서 먼저 반환
        _prio = self.cfg.get('_tactical_wpn_priority')
        if _prio and ship.available(_prio) > 0:
            wpn_info = FRIENDLY_DB.get(_prio)
            if wpn_info:
                # BUG-5 fix: range_km 없는 무기도 사거리 내 교전 가능하도록 500km 기본값
                max_r = wpn_info.get('range_km', 500) * 1000
                if dist_m <= max_r:
                    return _prio

        alt          = m.altitude_m
        is_hgv       = m.is_hgv
        is_qbm       = m.is_qbm
        is_ballistic = m.is_ballistic

        def ok(wpn):
            return ship.available(wpn) > 0

        # HGV / 고고도 탄도 중간단계 → SM-3 (외기권 ≥100km — _SM3_ALT_MIN_M 참조)
        # 어쇼어 활성·잔여 있으면 함정 SM-3 생략 (지상 자산이 우선 교전)
        if ((is_hgv or is_ballistic) and alt >= _SM3_ALT_MIN_M) and dist_m <= 500_000:
            ashore_covers = (
                self.cfg.get('enable_ashore', False)
                and self.ground_inv.get('SM-3 (어쇼어)', 0) > 0
            )
            if not ashore_covers and ok('SM-3 Block IIA'):
                return 'SM-3 Block IIA'

        # v16.2: HGV 저고도(<40km) 글라이드/종말 페이즈 → SM-6 Block IB (대기권 내 기동 요격).
        # SM-3(외기권 탄도 요격)는 저고도 활공 HGV에 부적합 — 킨잘·지르콘 등은 이 층이 주력.
        # 고고도 HGV도 SM-3 소진·누출 시 이 층으로 폴백(370km 이내).
        if is_hgv and dist_m <= 370_000 and ok('SM-6 Block IB'):
            return 'SM-6 Block IB'

        # QBM (저고도 기동탄도) → SM-6 우선 (SM-3 무효)
        if is_qbm and dist_m <= 240_000:
            if ok('SM-6'): return 'SM-6'

        # 근거리→원거리 표준 다층 (SM-2는 조명기 가용 시에만)
        if dist_m <= 2_000   and ok('CIWS-II (Phalanx)'): return 'CIWS-II (Phalanx)'
        if dist_m <= 9_000   and ok('RIM-116 RAM'):        return 'RIM-116 RAM'
        if dist_m <= 50_000  and ok('ESSM Block II'):      return 'ESSM Block II'
        if dist_m <= 50_000  and ok('해궁 (K-SAAM)'):      return '해궁 (K-SAAM)'
        # LOW-11: SM-2 조명기 채널 확인
        if dist_m <= 170_000 and ok('SM-2 Block IIIB') and self._sm2_illuminator_ok(ship):
            return 'SM-2 Block IIIB'
        if dist_m <= 240_000 and ok('SM-6'):          return 'SM-6'
        if dist_m <= 240_000 and ok('SM-6 Block IB'): return 'SM-6 Block IB'
        # 최후 폴백 SM-3 — 외기권 고도에서만. 과거엔 고도 조건이 없어, 대기권 내를 나는
        # 저고도 표적(준탄도 SRBM·대함미사일)까지 SM-3가 요격했다(외기권 전용인데).
        if (dist_m <= 500_000 and alt >= _SM3_ALT_MIN_M
                and ok('SM-3 Block IIA')): return 'SM-3 Block IIA'
        return None

    def _select_aa_wpn(self, ship: FriendlyShipObj, et: EnemyThreatObj,
                       dist_m: float) -> Optional[str]:
        """
        항공기 목표 대공 무기 선택 (고도 3단 구분). 무장 피탄 비활성화 반영.

        SM-3는 대기권 외 BMD 전용 → 항공기 요격 불가.
          ≥ 10,000m (고고도): SM-2 → SM-6 (RAM 불필요, 사거리 초과 시 교전 불가)
          3,000–10,000m (중고도): SM-2 → SM-6 → RAM (근접 시)
          < 3,000m (저고도 침투): RAM 우선 → SM-2 → SM-6
        """
        alt    = et.altitude_m
        sm2_ok = self._sm2_illuminator_ok(ship)

        def ok(wpn):
            return ship.available(wpn) > 0

        if alt >= 10_000:
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            return None

        elif alt >= 3_000:
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            if dist_m <= 50_000  and ok('ESSM Block II'):               return 'ESSM Block II'
            if dist_m <= 9_000   and ok('RIM-116 RAM'):                 return 'RIM-116 RAM'
            return None

        else:
            if dist_m <= 9_000   and ok('RIM-116 RAM'):                 return 'RIM-116 RAM'
            if dist_m <= 50_000  and ok('ESSM Block II'):               return 'ESSM Block II'
            if dist_m <= 50_000  and ok('해궁 (K-SAAM)'):               return '해궁 (K-SAAM)'
            if dist_m <= 170_000 and ok('SM-2 Block IIIB') and sm2_ok: return 'SM-2 Block IIIB'
            if dist_m <= 240_000 and ok('SM-6'):                        return 'SM-6'
            if dist_m <= 240_000 and ok('SM-6 Block IB'):               return 'SM-6 Block IB'
            return None

    # ── 4단계: 아군 공격 TEWA ─────────────────────────────────────────────────

    def _friendly_strike(self):
        """
        수상함 → 해성/하푼 (strike_inventory)
        잠수함 → 홍상어/청상어 (inventory)
        """
        if not self.cfg.get('enable_strike', True):
            return
        # v10.6: 항모(high_value_target) 우선 정렬 — 화력 집중.
        # 정렬 키가 함정과 무관하고 _friendly_strike 내에서 위협 hp가 불변이므로
        # 함정 루프 밖에서 1회만 계산 (대잠 등 함정·위협 多 시나리오 성능 핫스팟 해소).
        sorted_threats = sorted(
            self.enemy_threats,
            key=lambda e: (0 if e.high_value_target else 1, e.hp),
        )
        for ship in self.friendly_ships:
            if not ship.alive:
                continue

            for et in sorted_threats:
                if not et.alive:
                    continue
                if not (et.is_ship or et.is_sub):
                    continue
                # v9.6: 기습 잠수함 은닉 중 — 아군 탐지·교전 불가
                if et.is_sub and self.t < et.hidden_until:
                    continue
                # v20.5(B-3): 능동 핑 노출 후 잠항 도주 중이면 **함정 소나도 접촉을 잃는다**.
                # 이게 없으면 잠수함이 항공기를 따돌려도 함정이 계속 접촉을 유지해 datum이
                # 갱신되고, 결국 항공기가 다시 찾아낸다 — 도주가 아무 이득이 없어진다.
                # (OFF면 기존 동작 그대로: 함정은 도주 중에도 접촉 유지)
                if (self._asw_contact_limit and et.is_sub
                        and self.t < et.contact_lost_until):
                    continue

                dist_m   = ship.pos.dist_to(et.pos)
                category = et.category
                detect_m = self._detect_range_m(ship, category, alt_m=et.altitude_m)
                # 잠수함: 수온약층 소나 보정 추가 적용
                if et.is_sub and self.cfg.get('enable_sonar_equation', False):
                    # v12.3: dB 소나 방정식 — R50 기반 정규 CDF 확률 탐지
                    res = self._sonar_eq_detect(et, dist_m, 'hull')
                    if res is False:
                        continue
                    elif res is None:
                        # 음향 데이터 없음 → 레거시 소나 보정 폴백
                        detect_m *= self._thermocline_factor(et)
                        if et.is_retreating:
                            detect_m *= 0.30
                        if dist_m > detect_m:
                            continue
                    # res True → 탐지 성공, 통과
                    if et.is_sub:
                        et.last_contact_t = self.t   # v20.5(B-3): 접촉 유지 → datum 초기화
                else:
                    if et.is_sub:
                        detect_m *= self._thermocline_factor(et)
                        # NEW-AW: 이탈 잠수함 — 고속 이탈로 소나 접촉 급감 (탐지 70% 감소)
                        if et.is_retreating:
                            detect_m *= 0.30
                    if dist_m > detect_m:
                        continue
                    if et.is_sub:
                        et.last_contact_t = self.t   # v20.5(B-3): 접촉 유지 → datum 초기화

                if et.is_ship:
                    en_route = sum(
                        1 for m in self.missiles
                        if m.alive and m.target is et and m.mtype == 'friendly_strike'
                    )
                    # v10.6: 항모(high_value_target) 살보 6발, 일반 수상함 4발
                    max_salvo = 6 if et.high_value_target else 4
                    if en_route >= max_salvo:
                        continue

                    if ship.is_submarine:
                        # 아군 잠수함 → 적 수상함 공격 (현무-3C/하푼/어뢰)
                        wpn = self._select_sub_strike_wpn(ship, dist_m)
                        if not wpn:
                            continue
                        if wpn in FRIENDLY_STRIKE_DB:
                            wpn_info = FRIENDLY_STRIKE_DB[wpn]
                            ship.strike_inventory[wpn] = ship.strike_inventory.get(wpn, 0) - 1
                            pk_b = wpn_info['pk_base']
                            spd  = wpn_info['speed_ms']
                            cost = wpn_info['cost_usd']
                        else:
                            wpn_info = FRIENDLY_DB[wpn]
                            ship.inventory[wpn] -= 1
                            pk_b = wpn_info['pk_dist']['mean']
                            spd  = wpn_info['speed_ms']
                            cost = wpn_info['cost_usd']
                        ship.total_cost += cost
                        _ms = MissileObj(
                            mtype    = 'friendly_strike',
                            name     = wpn,
                            pos      = ship.pos,
                            target   = et,
                            speed_ms = spd,
                            pk_base  = pk_b,
                            owner_id = id(ship),
                            t_spawn  = self.t,
                        )
                        _ms.seeker = (wpn_info.get('seeker', 'radar')
                                      if wpn in FRIENDLY_STRIKE_DB else 'radar')
                        self.missiles.append(_ms)
                        self._log(
                            f"[공격] {ship.name} -> {wpn} -> {et.preset_name} "
                            f"(거리 {dist_m/1000:.0f}km)"
                        )
                    else:
                        # v16.8: 소형 자폭정(USV)엔 고가 대함미사일 낭비 방지 —
                        # Mk.45 5인치 함포 근접 격퇴만 (재고 무한·근거리 최후 레이어)
                        if et.info.get('is_suicide'):
                            _gun_rng = FRIENDLY_STRIKE_DB['Mk.45 5인치 함포']['range_km'] * 1000
                            wpn = 'Mk.45 5인치 함포' if dist_m <= _gun_rng else None
                        else:
                            wpn = self._select_strike_wpn(ship, dist_m)
                        if not wpn:
                            continue
                        wpn_info = FRIENDLY_STRIKE_DB[wpn]
                        # SM-6 대함 모드: VLS inventory에서 소모
                        if wpn == 'SM-6 대함 모드':
                            ship.inventory['SM-6'] -= 1
                        elif wpn == 'Tomahawk Block V':
                            ship.inventory['Tomahawk Block V'] = ship.inventory.get('Tomahawk Block V', 0) - 1
                        elif wpn == 'Mk.45 5인치 함포':
                            pass  # 함포는 재고 무한 (수백 발 탑재)
                        else:
                            ship.strike_inventory[wpn] = ship.strike_inventory.get(wpn, 0) - 1
                        ship.total_cost += wpn_info['cost_usd']
                        _m = MissileObj(
                            mtype    = 'friendly_strike',
                            name     = wpn,
                            pos      = ship.pos,
                            target   = et,
                            speed_ms = wpn_info['speed_ms'],
                            pk_base  = wpn_info['pk_base'],
                            owner_id = id(ship),
                            t_spawn  = self.t,
                        )
                        _m.seeker = wpn_info.get('seeker', 'radar')
                        self.missiles.append(_m)
                        self._log(
                            f"[공격] {ship.name} -> {wpn} -> {et.preset_name} "
                            f"(거리 {dist_m/1000:.0f}km)"
                        )

                elif et.is_sub:
                    # v16.1: 능동 핑 노출 후 회피 잠항 도주 중 — 함정 대잠 공격 불가(접촉 단절)
                    if self.t < et.contact_lost_until:
                        continue
                    en_route = sum(
                        1 for m in self.missiles
                        if m.alive and m.target is et and m.mtype == 'friendly_strike'
                    )
                    if en_route >= 1:
                        continue
                    wpn = self._select_asw_wpn(ship, dist_m)
                    if not wpn:
                        continue
                    # v16.1 능동 소나 핑 노출 대가(EMCON 딜레마): 대잠 공격(능동 핑)을 시도하면
                    # 잠수함이 핑을 역탐지 → 회피 굴림. 성공 시 datum 무효화·잠항 도주로 접촉
                    # 단절(어뢰 발사 무산, 한동안 재탐지·재발사 불가). OFF면 무동작.
                    if self._sonar_emcon:
                        if random.random() < _ASW_PING_EVADE_P:
                            # 능동 핑으로 노출된 함정에서 멀어지는 방향으로 잠항 이탈(standoff 복귀)
                            ang = et.pos.bearing_to(ship.pos) + math.pi
                            et.pos.x += math.cos(ang) * _ASW_EVADE_JUMP_M
                            et.pos.y += math.sin(ang) * _ASW_EVADE_JUMP_M
                            et.contact_lost_until = self.t + _ASW_CONTACT_LOST_S
                            if not self._mc_mode:
                                self._log(f"[접촉 단절] {et.preset_name} 능동 소나 핑 노출 — "
                                          f"회피 잠항 도주(어뢰 발사 무산)")
                            continue
                    wpn_info = FRIENDLY_DB[wpn]
                    ship.inventory[wpn] -= 1
                    ship.total_cost += wpn_info['cost_usd']
                    self.missiles.append(MissileObj(
                        mtype    = 'friendly_strike',
                        name     = wpn,
                        pos      = ship.pos,
                        target   = et,
                        speed_ms = wpn_info['speed_ms'],
                        pk_base  = wpn_info['pk_dist']['mean'],
                        owner_id = id(ship),
                        t_spawn  = self.t,
                    ))
                    self._log(
                        f"[대잠 공격] {ship.name} -> {wpn} -> {et.preset_name} "
                        f"(거리 {dist_m/1000:.1f}km)"
                    )

        # v9.4: 지상 발사 현무-4 ASBM — 60초 쿨다운, 함정당 최대 2발 비행 중
        _h4_wpn = '현무-4 (ASBM)'
        if (self.cfg.get('enable_strike', True)
                and self.ground_inv.get(_h4_wpn, 0) > 0
                and self.t - self._ground_last_fire >= 60.0):
            wpn_info_h4 = FRIENDLY_STRIKE_DB[_h4_wpn]
            primary_pos  = self._primary().pos
            for et in self.enemy_threats:
                if not et.alive or not et.is_ship:
                    continue
                dist_m = primary_pos.dist_to(et.pos)
                if dist_m > wpn_info_h4['range_km'] * 1000:
                    continue
                en_route = sum(
                    1 for m in self.missiles
                    if m.alive and m.target is et and m.name == _h4_wpn
                )
                if en_route >= 2:
                    continue
                self.ground_inv[_h4_wpn] -= 1
                self._ground_last_fire = self.t
                self._ground_cost += wpn_info_h4['cost_usd']
                self.missiles.append(MissileObj(
                    mtype    = 'friendly_strike',
                    name     = _h4_wpn,
                    pos      = primary_pos,
                    target   = et,
                    speed_ms = wpn_info_h4['speed_ms'],
                    pk_base  = wpn_info_h4['pk_base'],
                    owner_id = 0,
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[지상공격] {_h4_wpn} -> {et.preset_name} "
                    f"(거리 {dist_m/1000:.0f}km)"
                )
                if self.ground_inv.get(_h4_wpn, 0) <= 0:
                    break

    # ── 4.5단계: 항공 자산 대잠 (포팅 C) ─────────────────────────────────────

    def _aircraft_asw(self):
        """
        v9.8: 항공 대잠 탐지 상태 머신 — 헬기(디핑소나)·초계기(소노부이) 방식 분리.

        헬기 흐름:  idle → (범위 진입) → hovering(dip_hover_s) → 탐지 확률 판정
                    탐지 성공 → 어뢰 투하 / 탐지 실패 → cooldown(retry_s) → 재시도
                    max_attempts 초과 → 포기(abandon)

        초계기 흐름: idle → (범위 진입) → 소노부이 투하 → 탐지 확률 판정
                    탐지 성공 → 어뢰 투하 / 탐지 실패 → cooldown(retry_s) → 재투하
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')

        for ac in self.aircraft:
            # CAP 전투기·정찰기는 대잠 무장이 없다 — ASW 순회에서 제외한다. 미제외 시
            # 공대공 payload(KF-21 'IRIS-T SL')를 어뢰로 FRIENDLY_DB 조회 → KeyError 크래시
            # (잠수함 위협이 CAP 탐지권에 든 편성·시드에서만 발현). 대잠은 헬기·초계기만.
            if ac.info.get('aircraft_role') in ('cap', 'recon'):
                continue
            if ac.payload_remaining <= 0:
                continue
            if self.t < ac.t_available:
                continue
            wx_limits = ac.info.get('weather_limits', {})
            if not wx_limits.get(weather, True):
                continue

            asw_mode = ac.info.get('asw_mode', 'sonobuoy')

            # ── 호버링 단계 처리 (dipping 전용) ─────────────────────────────
            if ac._asw_phase == 'hovering':
                if self.t < ac._dip_until:
                    continue   # 아직 호버링 중
                # 호버링 완료 → 탐지 확률 판정
                et = ac._search_target
                if et is None or not et.alive or self.t < et.contact_lost_until:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                self._asw_detect_check(ac, et, primary)
                continue

            # ── 쿨다운 단계 (재탐색 대기) ────────────────────────────────────
            if ac._asw_phase == 'cooldown':
                if self.t < ac._next_attempt:
                    continue
                et = ac._search_target
                if et is None or not et.alive or self.t < et.contact_lost_until:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                # 재시도 — 사거리 재확인 후 탐지
                dist_to_sub = primary.pos.dist_to(et.pos)
                total_dist  = dist_to_sub + (
                    ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)
                if total_dist > ac.info['range_km'] * 1000:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                detect_m = self._asw_detect_range(ac, primary, et)
                bonus_m  = ac.info.get('sonobuoy_detect_bonus_km', 0) * 1000
                if dist_to_sub > detect_m + bonus_m:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                if asw_mode == 'dipping':
                    # 디핑소나 재전개
                    ac._asw_phase = 'hovering'
                    ac._dip_until = self.t + ac.info.get('dip_hover_s', 60)
                    self._log(f"[대잠 헬기] {ac.name} → {et.preset_name} 재탐색 "
                              f"(시도 {ac._detect_fails+1}/{ac.info.get('max_attempts',3)}) "
                              f"디핑소나 재전개…")
                else:
                    self._asw_detect_check(ac, et, primary)
                continue

            # ── 전진(transit) 단계 — 잠수함으로 이동 중 (v16.1 ASW 전진 초계) ──
            if ac._asw_phase == 'transit':
                if self.t < ac._dip_until:
                    continue   # 아직 전진 비행 중
                et = ac._search_target
                if et is None or not et.alive or self.t < et.contact_lost_until:
                    ac._asw_phase = 'idle'; ac._search_target = None; continue
                # 탐지권 가장자리 도착 → 탐지 단계 진입(전진 완료, 재확인 없이 교전)
                if asw_mode == 'dipping':
                    ac._asw_phase = 'hovering'
                    ac._dip_until = self.t + ac.info.get('dip_hover_s', 60)
                else:
                    self._asw_detect_check(ac, et, primary)
                    if ac._asw_phase == 'transit':   # detect_check가 phase 미변경 시 idle 복귀
                        ac._asw_phase = 'idle'
                continue

            # ── idle 단계 — 새 표적 탐색 ─────────────────────────────────────
            for et in self.enemy_threats:
                if not et.alive or not et.is_sub:
                    continue
                # v9.6: 기습 잠수함 은닉 중 탐지 불가
                if self.t < et.hidden_until:
                    continue
                # v16.1: 능동 핑 노출 후 회피 잠항 도주 중 — 접촉 단절, 재탐지 불가
                if self.t < et.contact_lost_until:
                    continue
                # v20.5(B-3): 이 기체가 포기한 표적은 재접촉 쿨다운 동안 다시 붙잡지 않는다.
                # (OFF면 _asw_giveup이 비어 있어 무조건 통과 → 기존 동작 bit-identical)
                if self.t < ac._asw_giveup.get(id(et), 0.0):
                    continue
                # 이미 어뢰가 향하고 있으면 패스
                if any(m.alive and m.target is et and m.mtype == 'friendly_strike'
                       for m in self.missiles):
                    continue

                # 사거리 체크
                dist_to_sub = primary.pos.dist_to(et.pos)
                total_dist  = dist_to_sub + (
                    ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)
                if total_dist > ac.info['range_km'] * 1000:
                    continue

                # 탐지 범위 체크
                detect_m = self._asw_detect_range(ac, primary, et)
                bonus_m  = ac.info.get('sonobuoy_detect_bonus_km', 0) * 1000
                if dist_to_sub > detect_m + bonus_m:
                    # v16.1 ASW 전진 초계: 탐지권 밖이지만 전개 사거리 안 → 잠수함 방위로 전진
                    if self._asw_forward:
                        transit_s = (dist_to_sub - (detect_m + bonus_m)) / max(ac.info.get('speed_ms', 200), 1)
                        ac._search_target = et
                        ac._detect_fails  = 0
                        ac._asw_phase     = 'transit'
                        ac._dip_until     = self.t + transit_s
                        if not self._mc_mode:
                            self._log(f"[대잠 전진] {ac.name} → {et.preset_name} 잠수함 방위로 전진 "
                                      f"({dist_to_sub/1000:.0f}km, {transit_s:.0f}초 비행)…")
                        break
                    continue

                # 표적 포착 → 탐지 단계 진입
                ac._search_target = et
                ac._detect_fails  = 0
                craft_type = '초계기' if ac.info.get('base_type') == 'land' else '헬기'

                if asw_mode == 'dipping':
                    ac._asw_phase = 'hovering'
                    ac._dip_until = self.t + ac.info.get('dip_hover_s', 60)
                    self._log(
                        f"[대잠 헬기] {ac.name} → {et.preset_name} 수색 구역 도착 "
                        f"(거리 {dist_to_sub/1000:.0f}km) 디핑소나 전개 중 "
                        f"({ac.info.get('dip_hover_s',60):.0f}초 호버링)…")
                else:
                    self._log(
                        f"[대잠 초계] {ac.name}({craft_type}) → {et.preset_name} "
                        f"(거리 {dist_to_sub/1000:.0f}km) 소노부이 투하…")
                    self._asw_detect_check(ac, et, primary)
                break   # 한 tick당 한 표적

    def _asw_detect_range(self, ac: 'FriendlyAircraftObj',
                          primary: 'FriendlyShipObj',
                          et: 'EnemyThreatObj') -> float:
        """항공 대잠 탐지 거리 — 함정 소나 기준 + 수온층 + 이탈 패널티."""
        detect_m = self._detect_range_m(primary, '대잠')
        detect_m *= self._thermocline_factor(et)
        if et.is_retreating:
            detect_m *= 0.30   # 고속 이탈 시 소나 접촉 급감
        return detect_m

    def _asw_weather_factor(self, weather: str) -> float:
        """날씨별 탐지 확률 보정 계수."""
        return {'맑음 (주간)': 1.00, '맑음 (야간)': 0.95,
                '박무 / 흐림': 0.90, '강우 / 강설': 0.80,
                '폭풍':        0.65, '태풍':        0.50}.get(weather, 1.00)

    def _asw_datum_m(self, et) -> float:
        """항공 대잠 탐지의 표정 오차(datum) 반경 — 마지막 소나 접촉 이후 성장.

        표준 대잠 교리의 furthest-on circle: 접촉이 끊긴 뒤 잠수함이 도달 가능한 반경은
        (경과 시간 × 잠수함 속도)로 커진다. 접촉을 유지 중이면 기존 고정값과 같다.
        enable_asw_contact_limit OFF면 기존 상수(1500m) 그대로 → bit-identical.
        """
        if not self._asw_contact_limit:
            return _ASW_DATUM_BASE_M
        elapsed = max(0.0, self.t - getattr(et, 'last_contact_t', 0.0))
        v_sub   = max(getattr(et, 'speed_ms', 0.0) or 0.0, 1.0)
        return _ASW_DATUM_BASE_M + v_sub * elapsed

    def _asw_detect_check(self, ac: 'FriendlyAircraftObj',
                          et: 'EnemyThreatObj',
                          primary: 'FriendlyShipObj'):
        """탐지 확률 판정 → 성공 시 어뢰 투하, 실패 시 재탐색 또는 포기."""
        weather   = self.cfg.get('weather', '맑음 (주간)')
        thermo    = self._thermocline_factor(et)
        wx_factor = self._asw_weather_factor(weather)
        base_prob = ac.info.get('detect_base_prob', 0.65)
        if self.cfg.get('enable_sonar_equation', False):
            # v12.3: dB 소나 방정식 — 디핑/소노부이 수동·능동 통합 Pd (날씨만 곱)
            sensor_key = 'dipping' if ac.info.get('asw_mode') == 'dipping' else 'sonobuoy'
            # 항공기는 표적 상공으로 전개 — 함정 거리가 아닌 표정 오차 거리에서 탐지.
            # 정온 잠수함은 능동(디핑·소노부이)이 있어야 탐지.
            # v20.5(B-3): datum을 1500m로 고정하면 항공기가 잠수함 위치를 늘 아는 셈이 되어
            # 탐지확률이 상한(0.97)에 붙고 탐지가 사실상 보장된다 → EMCON 딜레마 성립 불가.
            # 접촉이 끊긴 시간만큼 오차원이 커지게 한다(furthest-on circle).
            pd, _ = self._sonar_eq_pd(et, self._asw_datum_m(et), sensor_key)
            if pd is not None:
                prob = min(pd * wx_factor, 0.97)
            else:
                prob = min(base_prob * thermo * wx_factor, 0.97)
        else:
            prob = min(base_prob * thermo * wx_factor, 0.97)

        if random.random() < prob:
            # v16.01.03 능동 핑 역탐지: 능동 소나(디핑/소노부이)로 탐지된 잠수함은 그 핑을
            # 역포착 → 은닉 해제 + 어뢰 반격 앞당김 + 회피 기동(접촉 급감). 기본 OFF면 무동작.
            if self._sonar_emcon and et.is_sub:
                # 능동 핑 노출 → 잠수함 회피 굴림. 성공 시 datum을 무효화하며 잠항 도주해
                # 접촉을 끊는다(재시도 탐지 보장 차단) → 어뢰 투하 무산, 한동안 재탐지·발사 불가.
                if random.random() < _ASW_PING_EVADE_P:
                    ang = random.uniform(0.0, 2.0 * math.pi)
                    et.pos.x += math.cos(ang) * _ASW_EVADE_JUMP_M
                    et.pos.y += math.sin(ang) * _ASW_EVADE_JUMP_M
                    et.contact_lost_until = self.t + _ASW_CONTACT_LOST_S
                    ac._asw_phase     = 'idle'
                    ac._search_target = None
                    ac._detect_fails  = 0
                    if not self._mc_mode:
                        self._log(f"[접촉 단절] {et.name} 능동 핑 노출 — 회피 잠항 도주, "
                                  f"접촉 상실(어뢰 투하 무산)")
                    return
                # 회피 실패 → 발각 확정: 은닉 해제·어뢰 반격·근거리 회피
                et.hidden_until      = min(et.hidden_until, self.t)   # 은닉 해제(발각됨)
                et.next_fire_t       = min(et.next_fire_t, self.t)    # 즉시 어뢰 반격 가능
                et.counter_evade_until = self.t + 90.0               # 90초 회피 기동(재탐지 급감)
                if not self._mc_mode:
                    self._log(f"[역탐지] {et.name} 능동 핑 포착 — 은닉 해제·어뢰 반격·회피 기동")
            # ── 탐지 성공 → 어뢰 투하 ────────────────────────────────────────
            wpn_name = ac.info['payload_wpn']
            wpn_info = FRIENDLY_DB[wpn_name]
            pk       = max(0.0, min(wpn_info['pk_dist']['mean'] + ac.info.get('pk_bonus', 0.0), 0.98))

            ac.payload_remaining -= 1
            ac.sorties           += 1
            ac.total_cost        += ac.info['cost_usd']

            dist_to_sub = primary.pos.dist_to(et.pos)
            total_dist  = dist_to_sub + (
                ac.info.get('base_dist_km', 0) * 1000 if ac.info.get('base_type') == 'land' else 0)

            drop_pos = LatLon.from_xy(
                et.pos.x + random.uniform(-300, 300),
                et.pos.y + random.uniform(-300, 300),
            )
            m = MissileObj(
                mtype    = 'friendly_strike',
                name     = f"{wpn_name}({ac.name})",
                pos      = drop_pos,
                target   = et,
                speed_ms = wpn_info['speed_ms'],
                pk_base  = pk,
                owner_id = id(ac),
                t_spawn  = self.t,
            )
            m.is_torpedo = True
            self.missiles.append(m)

            fly_s          = total_dist / max(ac.info['speed_ms'], 1)
            ac.t_available = self.t + ac.info['sortie_time_s'] + fly_s
            ac._asw_phase    = 'idle'
            ac._search_target = None
            ac._detect_fails  = 0

            mode_tag = '디핑소나' if ac.info.get('asw_mode') == 'dipping' else '소노부이'
            self._log(
                f"[대잠 탐지 성공] {ac.name} → {et.preset_name} "
                f"({mode_tag} 확률 {prob:.0%}, 수온층 ×{thermo:.2f}) | "
                f"{wpn_name} Pk={pk:.2f} 투하 | 잔여 {ac.payload_remaining}발"
            )
        else:
            # ── 탐지 실패 ─────────────────────────────────────────────────────
            ac._detect_fails += 1
            max_att = ac.info.get('max_attempts', 3)
            retry_s = ac.info.get('retry_s', 90)
            mode_tag = '디핑소나' if ac.info.get('asw_mode') == 'dipping' else '소노부이'

            if ac._detect_fails >= max_att:
                ac._asw_phase     = 'idle'
                ac._search_target = None
                ac._detect_fails  = 0
                # v20.5(B-3): 포기를 실제 포기로 만든다 — 쿨다운 동안 이 표적 재탐색 금지.
                # OFF면 기록만 안 해 기존(무한 재탐색) 동작 그대로 보존.
                _giveup_s = 0.0
                if self._asw_contact_limit:
                    _giveup_s = _ASW_GIVEUP_COOLDOWN_S
                    ac._asw_giveup[id(et)] = self.t + _giveup_s
                self._log(
                    f"[대잠 탐지 실패] {ac.name} → {et.preset_name} "
                    f"{max_att}회 시도 모두 실패 (확률 {prob:.0%}) — 표적 포기"
                    + (f" (재접촉 시도까지 {_giveup_s:.0f}초)" if _giveup_s else "")
                )
            else:
                ac._asw_phase     = 'cooldown'
                ac._next_attempt  = self.t + retry_s
                self._log(
                    f"[대잠 탐지 실패] {ac.name} → {et.preset_name} "
                    f"{mode_tag} 탐지 실패 (확률 {prob:.0%}, 수온층 ×{thermo:.2f}) "
                    f"— {retry_s}초 후 재시도 ({ac._detect_fails}/{max_att})"
                )

    def _aircraft_recon(self):
        """v16.12: 아군 무인 정찰 드론 — OTH 표적 탐지 중계로 함대 실효 탐지거리 확장.
        생존 중인 recon 드론의 recon_detect_bonus_km 중 최댓값을 self._recon_bonus_km에 실어
        _detect_range_m·_log_detections가 대공·대함 레이더 탐지거리에 가산한다(무장 없음).
        적 항공위협 존재 시 recon_roll_s 주기로 survive_prob 격추 판정 — 실패 시 중계 중단.
        recon 역할 드론이 없으면(기본 OFF) _recon_bonus_km=0·RNG 미소비 → 회귀 bit-identical."""
        self._recon_bonus_km = 0.0
        weather = self.cfg.get('weather', '맑음 (주간)')
        # 격추 노출 판정용 — 생존 적 항공위협 존재 여부(퇴각기 제외)
        air_threat = any(et.alive and et.is_aircraft and not et.is_retreating
                         for et in self.enemy_threats)
        for ac in self.aircraft:
            if ac.info.get('aircraft_role') != 'recon':
                continue
            if ac._recon_lost:
                continue
            if self.t < ac.t_available:
                continue
            if not ac.info.get('weather_limits', {}).get(weather, True):
                continue
            # 격추 판정: 적 항공위협에 노출된 동안 주기적으로 생존 굴림
            if air_threat and self.t >= ac._recon_next_roll:
                ac._recon_next_roll = self.t + ac.info.get('recon_roll_s', 300)
                if random.random() >= ac.info.get('survive_prob', 0.9):
                    ac._recon_lost = True
                    self.stats['recon_losses'] += 1
                    self._log(f"[정찰 드론] {ac.name} 피격 격추 — OTH 탐지 중계 중단 "
                              f"({self.t:.0f}s)")
                    continue
            ac.sorties = max(ac.sorties, 1)   # 전개 1회 계상(비용 집계용)
            ac.total_cost = ac.info['cost_usd']
            bonus = ac.info.get('recon_detect_bonus_km', 0)
            if bonus > self._recon_bonus_km:
                self._recon_bonus_km = bonus

    def _unmanned_picket_update(self):
        """v16.12: 무인 함정 전방 피켓 — 생존 USV는 대함, UUV는 대잠 탐지거리를 확장한다.
        _detect_range_m이 category별로 이 보너스를 가산(대함=USV·대잠=UUV). 무인정이 없으면
        (기본 OFF) 보너스 0·상태 무변 → 회귀 bit-identical. 파괴되면 다음 틱부터 보너스 소멸."""
        self._usv_surf_bonus_km = 0.0
        self._uuv_asw_bonus_km  = 0.0
        for s in self.friendly_ships:
            if not s.alive or not s.is_unmanned:
                continue
            spec = SHIP_DB[s.ship_type]
            sb = spec.get('picket_surface_bonus_km', 0)
            ab = spec.get('picket_asw_bonus_km', 0)
            if sb > self._usv_surf_bonus_km:
                self._usv_surf_bonus_km = sb
            if ab > self._uuv_asw_bonus_km:
                self._uuv_asw_bonus_km = ab

    def _cap_posture_factors(self):
        """CAP 전개 자세 → (패트롤 반경 배율, 교전 후 쿨다운 초). 부모는 현행 고정값."""
        return (1.0, 60.0)

    def _aircraft_cap(self):
        """
        v10.5: 한국 공군 CAP — 적 항공기 BVR 요격.
        idle → (적 항공기 패트롤 반경 진입) → AAM 교전(즉시 Pk 판정) → cooldown → idle
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')
        cap_rscale, cap_cooldown = self._cap_posture_factors()
        for ac in self.aircraft:
            if ac.info.get('aircraft_role', 'asw') != 'cap':
                continue
            if self.t < ac.t_available:
                continue
            if ac.payload_remaining <= 0:
                continue
            if not ac.info.get('weather_limits', {}).get(weather, True):
                continue
            if self.t < ac._next_attempt:
                continue

            patrol_m   = ac.info.get('cap_patrol_radius_km', 500) * 1000 * cap_rscale
            aam_range_m= ac.info.get('cap_aam_range_km', 100) * 1000
            aam_pk     = ac.info.get('cap_aam_pk', 0.55)
            base_dist_m= ac.info.get('base_dist_km', 300) * 1000
            wpn_name   = ac.info.get('payload_wpn', 'AAM')

            for et in self.enemy_threats:
                if not et.alive or not et.is_aircraft or et.is_retreating:
                    continue
                dist_threat = primary.pos.dist_to(et.pos)
                # 패트롤 반경 내 + 기지→함대→표적 총 사거리 체크
                if dist_threat > patrol_m:
                    continue
                if dist_threat + base_dist_m > ac.info['range_km'] * 1000:
                    continue
                # AAM 유효 사거리 내 진입 시 교전
                if dist_threat > aam_range_m:
                    continue

                ac.payload_remaining -= 1
                ac.sorties += 1   # v18.01.18: BVR 교전 = 출격 계상(헬기 _asw_detect_check 4121과 대칭)
                ac.total_cost += ac.info['cost_usd'] / max(ac.info['payload_cnt'], 1)

                if random.random() < aam_pk:
                    et.alive = False
                    et.intercepted = True   # 격침 집계용(enemy_ships_destroyed) — 항공기 격추는 요격률(발사 살보 대비)에 미포함, SAM 경로(4594) 규약과 일치
                    self._log(
                        f"[CAP] {ac.name} → {et.preset_name} "
                        f"{wpn_name} 격추 (거리 {dist_threat/1000:.0f}km, {self.t:.0f}s)"
                    )
                else:
                    self._log(
                        f"[CAP] {ac.name} → {et.preset_name} "
                        f"{wpn_name} 교전 실패 (거리 {dist_threat/1000:.0f}km, Pk {aam_pk:.0%})"
                    )

                # 교전 후 선회·재장전 cooldown (자세에 따라 가변)
                ac._next_attempt = self.t + cap_cooldown
                break  # 한 tick당 한 표적

    def _aircraft_aas(self):
        """
        v10.6: 항공 대함 공격 (Air-to-Surface Anti-Ship) — KF-21 해성-II 등.
        cap_strike_wpn 속성이 있는 CAP 항공기가 수상 표적(항모 우선) 공격.
        idle → (수상함 탐지) → 해성-II 발사(즉시 Pk) → 90s cooldown → idle
        """
        primary = self._primary()
        weather = self.cfg.get('weather', '맑음 (주간)')
        for ac in self.aircraft:
            if not ac.info.get('cap_strike_wpn'):
                continue
            if self.t < ac.t_available:
                continue
            if ac.strike_payload_remaining <= 0:
                continue
            if not ac.info.get('weather_limits', {}).get(weather, True):
                continue
            if self.t < ac._strike_cooldown_until:
                continue

            strike_range_m = ac.info.get('cap_strike_range_km', 200) * 1000
            strike_pk      = ac.info.get('cap_strike_pk', 0.55)
            base_dist_m    = ac.info.get('base_dist_km', 300) * 1000
            wpn_name       = ac.info['cap_strike_wpn']
            wpn_cost       = ac.info.get('cap_strike_cost_usd', 3_000_000)

            # v10.6: 항모 우선 → 그 외 수상함
            candidates = sorted(
                [et for et in self.enemy_threats if et.alive and et.is_ship],
                key=lambda e: (0 if e.high_value_target else 1),
            )
            for et in candidates:
                dist_threat = primary.pos.dist_to(et.pos)
                if dist_threat + base_dist_m > ac.info['range_km'] * 1000:
                    continue
                if dist_threat > strike_range_m:
                    continue
                # 이미 해당 표적으로 비행 중인 아군 미사일이 4발 이상이면 건너뜀
                en_route = sum(
                    1 for m in self.missiles
                    if m.alive and m.target is et and m.mtype == 'friendly_strike'
                )
                if en_route >= (6 if et.high_value_target else 4):
                    continue

                ac.strike_payload_remaining -= 1
                ac.total_cost += wpn_cost
                _m = MissileObj(
                    mtype    = 'friendly_strike',
                    name     = f"{wpn_name}({ac.name})",
                    pos      = primary.pos,
                    target   = et,
                    speed_ms = 280,          # 해성-II 순항속도 ~280m/s
                    pk_base  = strike_pk,
                    owner_id = id(ac),
                    t_spawn  = self.t,
                )
                self.missiles.append(_m)
                self._log(
                    f"[AAS] {ac.name} → {et.preset_name} "
                    f"{wpn_name} 발사 (거리 {dist_threat/1000:.0f}km, "
                    f"잔여 {ac.strike_payload_remaining}발)"
                )
                ac._strike_cooldown_until = self.t + 90.0
                break  # 한 tick당 한 발사

    def _select_strike_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        # 우선순위: Tomahawk(초장거리) → 해성-II → 해성-I → 하푼 → SM-6 대함(OTH) → Mk.45(근거리)
        # Tomahawk Block V: US 함정 전용 초장거리 대함 타격
        if (ship.inventory.get('Tomahawk Block V', 0) > 0
                and dist_m <= FRIENDLY_STRIKE_DB['Tomahawk Block V']['range_km'] * 1000):
            return 'Tomahawk Block V'
        for wpn in ['해성-II', '해성-I', '하푼 Block II']:
            if ship.strike_inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_STRIKE_DB[wpn]['range_km'] * 1000:
                return wpn
        # SM-6 대함 모드: 해성/하푼 소진 후 OTH 사거리 내 수상함 공격
        if (ship.inventory.get('SM-6', 0) > 0  # SM-6 대함 모드 항상 활성 (cfg 키 제거)
                and dist_m <= FRIENDLY_STRIKE_DB['SM-6 대함 모드']['range_km'] * 1000):
            return 'SM-6 대함 모드'
        # Mk.45 함포: 근거리 최후 수단
        if dist_m <= FRIENDLY_STRIKE_DB['Mk.45 5인치 함포']['range_km'] * 1000:
            return 'Mk.45 5인치 함포'
        return None

    def _select_asw_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        for wpn in ['홍상어 (대잠)', '청상어 (경어뢰)', 'Mk.46 경어뢰']:
            if ship.inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_DB[wpn]['range_km'] * 1000:
                return wpn
        return None

    def _select_sub_strike_wpn(self, ship: FriendlyShipObj, dist_m: float) -> Optional[str]:
        """아군 잠수함 → 적 수상함 공격 무기 선택 (현무-3C / 하푼 / 청상어)"""
        for wpn in ['현무-3C', '하푼 Block II']:
            if ship.strike_inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_STRIKE_DB[wpn]['range_km'] * 1000:
                return wpn
        for wpn in ['청상어 (경어뢰)', 'Mk.46 경어뢰']:
            if ship.inventory.get(wpn, 0) <= 0:
                continue
            if dist_m <= FRIENDLY_DB[wpn]['range_km'] * 1000:
                return wpn
        return None

    # ── 5단계: 적 SAM 방어 (수상함 전용) ─────────────────────────────────────

    def _enemy_defense(self):
        for et in self.enemy_threats:
            if not et.alive or not et.is_ship or not et.sam_inventory:
                continue

            for m in self.missiles:
                if not m.alive or m.mtype != 'friendly_strike':
                    continue
                if m.target is not et:
                    continue

                already = any(
                    s.alive and s.target is m and s.mtype == 'enemy_sam'
                    for s in self.missiles
                )
                if already:
                    continue
                if et.sam_channels_used >= et.sam_max_channels:
                    continue

                dist_m   = et.pos.dist_to(m.pos)
                sam_name = et.select_sam(dist_m)
                if not sam_name:
                    continue

                sam_info = ENEMY_SAM_DB[sam_name]
                et.sam_inventory[sam_name]  -= 1
                et.sam_channels_used        += 1

                self.missiles.append(MissileObj(
                    mtype    = 'enemy_sam',
                    name     = sam_name,
                    pos      = et.pos,
                    target   = m,
                    speed_ms = sam_info['speed_ms'],
                    pk_base  = sam_info['pk'],
                    owner_id = id(et),
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[적 방어] {et.preset_name} -> {sam_name} 발사 "
                    f"(거리 {dist_m/1000:.1f}km)"
                )

    def _enemy_anti_sam(self):
        """적 함정의 Anti-SAM 방어 — 아군 SAM이 접근 시 CIWS·SAM으로 요격.
        enable_anti_sam=False(기본)이면 즉시 반환 (기존 결과 완전 호환).
        """
        if not self.cfg.get('enable_anti_sam', False):
            return

        for et in self.enemy_threats:
            if not et.alive or not et.is_ship or not et.sam_inventory:
                continue

            for m in list(self.missiles):
                if not m.alive or m.mtype != 'friendly_sam':
                    continue
                if m.target is not et:   # 이 함정을 향하는 SAM만
                    continue

                dist_m = et.pos.dist_to(m.pos)

                # ── CIWS 즉시 판정 (2km 이내) ────────────────────────────────
                ciws = next((n for n in et.sam_inventory if 'CIWS' in n), None)
                if dist_m <= 2000 and ciws:
                    if random.random() < 0.30:
                        m.alive       = False
                        m.intercepted = True
                        # 발사 함정 채널 해제
                        for ship in self.friendly_ships:
                            if id(ship) == m.owner_id:
                                ship.channels_used = max(0, ship.channels_used - 1)
                        self._log(
                            f"[적 Anti-SAM CIWS] {et.preset_name} → {m.name} "
                            f"근접 요격 (Pk 0.30)"
                        )
                    continue   # CIWS 범위는 SAM 발사 없음

                # ── SAM 요격: 탐지거리·채널 확인 후 발사 ─────────────────────
                already = any(s.alive and s.target is m and s.mtype == 'enemy_sam'
                              for s in self.missiles)
                if already:
                    continue
                if et.sam_channels_used >= et.sam_max_channels:
                    continue

                rcs    = getattr(m, 'rcs_m2', 0.001)
                det_km = min(500.0 * ((rcs / 3.0) ** 0.25), 50.0)  # 최대 50km 캡
                if dist_m > det_km * 1000:
                    continue

                sam_name = et.select_sam(dist_m)
                if not sam_name:
                    continue

                sam_info    = ENEMY_SAM_DB[sam_name]
                anti_sam_pk = sam_info['pk'] * 0.35   # SAM vs SAM — 소형 고속 표적

                et.sam_inventory[sam_name] -= 1
                et.sam_channels_used       += 1
                self.missiles.append(MissileObj(
                    mtype    = 'enemy_sam',
                    name     = sam_name,
                    pos      = et.pos,
                    target   = m,
                    speed_ms = sam_info['speed_ms'],
                    pk_base  = anti_sam_pk,
                    owner_id = id(et),
                    t_spawn  = self.t,
                ))
                self._log(
                    f"[적 Anti-SAM] {et.preset_name} → {sam_name} 발사 → {m.name} "
                    f"(거리 {dist_m/1000:.1f}km, Pk {anti_sam_pk:.2f})"
                )

    # ── 6단계: 교전 결과 판정 ─────────────────────────────────────────────────

    def _check_intercepts(self):
        for sam in list(self.missiles):
            if not sam.alive:
                continue
            if sam.mtype not in ('friendly_sam', 'enemy_sam'):
                continue

            tgt = sam.target
            if not tgt.alive:
                # 타겟이 이미 격추됐어도 채널 해제 (누수 방지)
                sam.alive = False
                if sam.mtype == 'friendly_sam':
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)
                continue

            # v12.1: PNG 활성 SAM은 substep 최근접거리(min_miss_m)로 물리 판정
            png = getattr(sam, 'png_active', False)
            if png:
                if sam.min_miss_m <= PNG_LETHAL_M or sam.hit:
                    in_range = True                       # 신관 도달(또는 직접 도달) → 명중 후보
                elif sam.pos.dist_to(tgt.pos) > sam.min_miss_m + 1.0:
                    # 최근접점 통과 후 살상반경 밖 → 빗나감 확정, SAM 소멸(무한 추격 방지)
                    sam.alive = False
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                    self._log(f"[요격 실패] {sam.name} 빗나감 "
                              f"(최근접 {sam.min_miss_m:.0f}m, 종말 회피 기동)")
                    continue
                else:
                    in_range = False                      # 아직 접근 중
            else:
                in_range = sam.hit or (sam.pos.dist_to(tgt.pos) <= INTERCEPT_DIST_M)
            if not in_range:
                continue

            sam.alive = False
            tgt_name  = tgt.name if hasattr(tgt, 'name') else str(tgt)

            # 종말 회피 Pk 보정 (아군 SAM vs 적 미사일)
            # BUG-6: ECM은 적 미사일 타격 Pk를 낮추는 것 → _check_hits로 이동
            # v12.1: PNG 활성 시 회피는 물리 추격으로 이미 반영 → 곱셈 할인 건너뜀
            eff_pk = sam.pk_base
            # v20.5(B-2): 표적 난이도(속도·RCS) — 발사자와 무관한 물리라 양쪽 SAM에 공통 적용.
            eff_pk *= self._target_difficulty_factor(sam, tgt)
            if (not png) and sam.mtype == 'friendly_sam' and isinstance(tgt, MissileObj):
                remaining_m = sam.pos.dist_to(tgt.pos)
                if self.cfg.get('enable_evasion', True) and remaining_m < 10_000:  # BUG 수정: 20km→10km
                    eff_pk *= tgt.terminal_evasion_factor

            # pk_scale: LHS/Sobol 분석용 불확실 파라미터 반영
            if sam.mtype == 'friendly_sam':
                eff_pk = min(1.0, eff_pk * self.cfg.get('pk_scale', 1.0))
                # v10.4: CEC 중계 교전 — 자체 탐지 불가, 아군 데이터링크 의존 → Pk 10% 저하
                if getattr(sam, 'cec_relay', False):
                    eff_pk *= 0.90
                # v16.3 사이버전 데이터링크 변조: 침투 지속 중 표적 데이터 오염 → 요격 Pk 저하
                if self._cyber and self.t < self._dl_corrupt_until:
                    eff_pk *= _CYBER_DL_PK
                # v16.2 활공 HGV 횡기동: 활공·종말 기동 회피로 요격 Pk 소폭 저하
                if self._hgv_glide and getattr(tgt, 'is_hgv', False):
                    eff_pk *= _HGV_GLIDE_EVADE
            if random.random() < eff_pk:
                tgt.alive       = False
                tgt.intercepted = True
                tgt.t_intercept = self.t

                if sam.mtype == 'friendly_sam':
                    self._log(f"[요격 성공] {sam.name} -> {tgt_name} 격추 ({self.t:.0f}s)")
                    # MissileObj만 intercepted_threats에 집계 (BUG 수정: 항공기 플랫폼 격추는 enemy_ships_destroyed로)
                    if isinstance(tgt, MissileObj):
                        self.stats['intercepted_threats'] += 1
                        # A-1: EngagementAnalysis 추적
                        tgt.intercept_weapon = sam.name
                        tgt.intercept_km     = sam.pos.dist_to(tgt.pos) / 1000
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    self._log(f"[적 요격 성공] {sam.name} -> {tgt_name} 격추 ({self.t:.0f}s)")
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)
            else:
                if sam.mtype == 'friendly_sam':
                    self._log(f"[요격 실패] {sam.name} -> {tgt_name} 통과")
                    for ship in self.friendly_ships:
                        if id(ship) == sam.owner_id:
                            ship.channels_used = max(0, ship.channels_used - 1)
                else:
                    self._log(f"[적 요격 실패] {sam.name} -> {tgt_name} 통과")
                    for et in self.enemy_threats:
                        if id(et) == sam.owner_id:
                            et.sam_channels_used = max(0, et.sam_channels_used - 1)

    def _check_hits(self):
        for m in self.missiles:
            # hit=True: 목표 위치 도달. alive=False 또는 intercepted=True면 이미 처리된 미사일.
            if not m.hit or not m.alive or m.intercepted:
                continue

            m.alive = False  # 도달 미사일 소모 (결과 무관)

            if m.mtype == 'enemy_strike':
                tgt = m.target
                if isinstance(tgt, FriendlyShipObj) and tgt.alive:
                    # ARM: ECM 무효 (레이더 전파 역추적 — 재밍이 오히려 표적이 됨)
                    if m.is_arm:
                        # v16.1 EMCON ON: stale 조준 좌표로 유도된 ARM은 표적 이격만큼 명중 급감.
                        if self._emcon_arm and m.arm_aim_pos is not None:
                            _STALE_SCALE_M = 150.0   # 이격 150m마다 명중률 e^-1 감쇠
                            miss_d = m.arm_aim_pos.dist_to(tgt.pos)
                            if self.t < tgt.radar_off_until and miss_d > 200.0:
                                # 레이더 OFF + 큰 이격 → 명백히 빗나감(stale 유도)
                                self._log(
                                    f"[ARM 회피] {tgt.name} 레이더 OFF — {m.name} "
                                    f"{miss_d:.0f}m 빗나감(stale 유도)")
                                continue
                            pk_eff = m.pk_base * math.exp(-miss_d / _STALE_SCALE_M)
                            if random.random() < pk_eff:
                                tgt.take_arm_hit(self.t)
                                self.stats['friendly_hits'] += 1
                                self._log(
                                    f"[ARM 피격] {tgt.name} 레이더 직격! "
                                    f"(이격 {miss_d:.0f}m, Pk {pk_eff:.0%}, HP {tgt.hp})")
                            else:
                                self._log(f"[ARM 실패] {m.name} -> {tgt.name} "
                                          f"불발(이격 {miss_d:.0f}m)")
                            continue
                        # EMCON OFF(기본): 기존 로직 — 명중 순간 레이더 OFF면 빗나감
                        if self.t < tgt.radar_off_until:
                            self._log(
                                f"[ARM 회피] {tgt.name} 레이더 OFF — "
                                f"{m.name} 유도 실패 빗나감"
                            )
                            continue
                        if random.random() < m.pk_base:
                            tgt.take_arm_hit(self.t)
                            self.stats['friendly_hits'] += 1
                            self._log(
                                f"[ARM 피격] {tgt.name} 레이더 직격! "
                                f"(레이더 {tgt.radar_factor:.0%}, HP {tgt.hp})")
                        else:
                            self._log(f"[ARM 실패] {m.name} -> {tgt.name} 불발")
                        continue

                    # v9.13: 고층 바람 CEP — 순항·대함미사일 탄착 오차 증가 (탄도/HGV/어뢰 제외)
                    if not m.is_ballistic and not m.is_hgv and not m.is_qbm and not m.is_torpedo:
                        cep_f = self._wind_cep_factor()
                        if cep_f > 1.0 and random.random() > (1.0 / cep_f):
                            self._log(
                                f"[바람 CEP] {m.name} 강풍 탄착 오차 빗나감 "
                                f"(×{cep_f:.2f}, 계절:{self.cfg.get('season','?')})"
                            )
                            continue

                    # BUG-6: 아군 ECM(AN/SLQ-32) — 적 미사일 유도부 교란, Pk 30% 감소
                    # 탄도/HGV는 레이더 유도가 아니므로 ECM 무효
                    if self.cfg.get('enable_ecm', True) and not m.is_ballistic and not m.is_hgv:
                        ecm_red = 0.30 * self.cfg.get('ecm_scale', 1.0)
                        m.pk_base = max(0.0, m.pk_base * (1.0 - ecm_red))
                    # v16.6 전자 좌표 기만: 가짜 좌표로 유도된 미사일은 실제 표적 이격만큼 명중 급감
                    # (ARM stale 유도와 동형). 기존 ECM Pk 감소와 별개인 위치 기만 레이어.
                    if m.decoy_aim_pos is not None:
                        miss_d = m.decoy_aim_pos.dist_to(tgt.pos)
                        m.pk_base = m.pk_base * math.exp(-miss_d / _DECEPTION_SCALE_M)
                    # v16.3 사이버전 레이더 교란 반격(아군→적): 침투 지속 중 적 사격통제 교란 →
                    # 적 발사 미사일 Pk 저하. 레이더 유도가 아닌 탄도/HGV는 무효(ECM과 동형).
                    if (self._cyber and self.t < self._enemy_jam_until
                            and not m.is_ballistic and not m.is_hgv):
                        m.pk_base = max(0.0, m.pk_base * _CYBER_JAM_PK)
                    # 포팅 B: 음향 기만기 AN/SLQ-25 — 어뢰 전용
                    # v20.5(B-3): 성공 시 어뢰를 **무력화 처리**한다. 과거에는 continue만 해서
                    # ①어뢰가 살아남아 다음 틱에 또 판정 → 기만기를 반복 소모했고 ②막아낸 것이
                    # intercepted_threats에 안 잡혀 **요격률(방어 성공률)에 반영되지 않았다**.
                    # 어뢰는 SAM으로 못 잡으므로(수중), 기만기·회피가 곧 어뢰 방어의 성패다.
                    if m.is_torpedo and self.cfg.get('enable_decoy', True):
                        if tgt.decoy_stock > 0:
                            tgt.decoy_stock -= 1
                            if random.random() < DECOY_PK:
                                m.alive = False
                                m.intercepted = True
                                self.stats['intercepted_threats'] += 1
                                self._log(
                                    f"[기만기] {tgt.name} 기만기 성공 — {m.name} 회피 "
                                    f"(잔여 {tgt.decoy_stock}발)")
                                continue
                    # 포팅 B: 함정 회피 기동 — 어뢰 전용
                    if m.is_torpedo and self.cfg.get('enable_evasion', True):
                        # 추진 피탄 시 speed_factor만큼 회피 기동 성공률 저하
                        if random.random() < SHIP_EVASION_PK * tgt.speed_factor:
                            m.alive = False
                            m.intercepted = True
                            self.stats['intercepted_threats'] += 1
                            self._log(f"[회피] {tgt.name} 회피 기동 성공 — {m.name}")
                            continue
                    # 서브시스템 피해 롤 (enable_subsystem_damage=True 시)
                    subsystem = None
                    if self.cfg.get('enable_subsystem_damage', True):
                        r = random.random()
                        if getattr(m, 'is_torpedo', False):
                            # 어뢰: 수중 폭발 → 추진·선체 위주, 레이더 거의 없음
                            if   r < 0.05: subsystem = 'radar'
                            elif r < 0.45: subsystem = 'propulsion'
                            elif r < 0.55: subsystem = 'weapons'
                            # else: None (선체 직격, 45%)
                        else:
                            # 대함·탄도·순항미사일
                            # PHY-5: 레이더15%→25%, 추진20%→15%, 무장25%→20% (포클랜드전 피탄 통계 반영)
                            # 레이더 안테나는 선체 상부 노출 → 피탄율 높음
                            if   r < 0.25: subsystem = 'radar'
                            elif r < 0.40: subsystem = 'propulsion'
                            elif r < 0.60: subsystem = 'weapons'
                            # else: None (선체 직격, 40%)
                    if random.random() < m.pk_base:
                        tgt.take_hit(m.name, self.t, subsystem)
                        self.stats['friendly_hits'] += 1
                        # v12.4: 수선하 피격 시 격실 침수 (HP와 병행 레이어)
                        self._apply_flood_hit(m, tgt)
                        _dmg = {
                            'radar':      f'레이더 피탄 (탐지 {tgt.radar_factor:.0%})',
                            'propulsion': f'추진 피탄 (속도 {tgt.speed_factor:.0%})',
                            'weapons':    f'VLS 피탄 (채널 {tgt.channel_factor:.0%}, 탄약 손실)',
                            None:         '선체 직격 (HP 감소)',
                        }
                        _detail = f' — {_dmg[subsystem]}'
                        self._log(f"[피격] {tgt.name} <- {m.name} 명중! HP {tgt.hp}{_detail}")
                    else:
                        self._log(f"[피격 실패] {m.name} -> {tgt.name} 근접 불발")

            elif m.mtype == 'friendly_strike':
                tgt = m.target
                if isinstance(tgt, EnemyThreatObj) and tgt.alive:
                    # 포팅 B: 적 자체방어 — CIWS 요격 → 채프/플레어/DRFM
                    eff_pk = m.pk_base
                    if self.cfg.get('enable_selfdefense', True):
                        ciws_pk = tgt.info.get('enemy_ciws_pk', 0.0)
                        if ciws_pk > 0 and random.random() < ciws_pk:
                            self._log(f"[적 CIWS] {tgt.preset_name} CIWS 요격 — {m.name}")
                            continue
                        # v8.26: seeker 유형별 기만체계 차등 적용
                        seeker = getattr(m, 'seeker', 'radar')
                        if seeker == 'radar':
                            sdpk = tgt.info.get('chaff_pk',
                                   tgt.info.get('self_defense_pk', 0.0))
                        elif seeker == 'ir':
                            sdpk = tgt.info.get('flare_pk',
                                   tgt.info.get('self_defense_pk', 0.0))
                        elif seeker == 'combined':
                            sdpk = tgt.info.get('drfm_pk',
                                   tgt.info.get('self_defense_pk', 0.0) * 0.50)
                        else:  # acoustic 등 — 채프/플레어 무효
                            sdpk = 0.0
                        eff_pk = m.pk_base * (1.0 - sdpk)
                    if random.random() < eff_pk:
                        tgt.take_hit(m.name, self.t)
                        self.stats['enemy_hits'] += 1
                        sunk = not tgt.alive
                        status = '격침' if sunk else f'손상 (HP {tgt.hp})'
                        self._log(f"[적 피격] {tgt.preset_name} <- {m.name} 명중! {status}")
                        # v9.3: 격침 기록
                        self.strike_log.append({
                            'target': tgt.preset_name,
                            'weapon': m.name,
                            't': round(self.t, 1),
                            'sunk': sunk,
                            'hp_remaining': tgt.hp,
                        })
                    else:
                        self._log(f"[적 피격 실패] {m.name} -> {tgt.preset_name} 회피")

    # ── 7단계: 프레임 기록 ────────────────────────────────────────────────────

    def _record_frame(self):
        frame = SimFrame(self.t)
        for s in self.friendly_ships:
            # [5]=radar_factor [6]=speed_factor [7]=disabled_weapons 수 (애니메이션 피해 표시용)
            frame.friendly_ships.append((
                s.name, s.pos.x, s.pos.y, s.alive, s.hp,
                s.radar_factor, s.speed_factor, len(s.disabled_weapons)
            ))
            frame.ship_channels.append((s.name, s.channels_used, s.max_channels))
        for et in self.enemy_threats:
            frame.enemy_ships.append(
                (et.uid, et.preset_name, et.pos.x, et.pos.y, et.alive, et.hp,
                 et.altitude_m))
        for m in self.missiles:
            if m.alive:
                frame.missiles.append(
                    (m.uid, m.pos.x, m.pos.y, m.mtype, m.name,
                     self._missile_disp_alt(m)))
        frame.events = list(self._tick_events)
        self.frames.append(frame)
        self._tick_events.clear()

    def _missile_disp_alt(self, m) -> float:
        """3D 시각화용 미사일 고도(m). 탄도/HGV는 포물선 궤도 추정."""
        # v16.2: 활공 HGV는 altitude_m이 이미 실제 비행 고도(활공→종말) → 그대로 표시
        if self._hgv_glide and getattr(m, 'is_hgv', False):
            return m.altitude_m
        # v20.2a: 종말 강하 탄도도 altitude_m이 실제 비행 고도 → 그대로 표시(교전=표시 정합)
        if self._bal_descent and getattr(m, 'is_ballistic', False):
            return m.altitude_m
        if (m.is_ballistic or m.is_hgv) and getattr(m, '_init_dist', 0) > 0:
            target = m.target
            if target and hasattr(target, 'pos'):
                rem = m.pos.dist_to(target.pos)
                progress = max(0.0, min(1.0, 1.0 - rem / m._init_dist))
                peak = getattr(m, '_peak_alt_m', 50_000.0)
                return max(0.0, 4.0 * peak * progress * (1.0 - progress))
        return m.altitude_m

    # ── 종료 조건 ─────────────────────────────────────────────────────────────

    # ── v12.4: 동적 침수·복원력 모델 (enable_flooding ON 경로) ────────────────
    def _apply_flood_hit(self, m: 'MissileObj', tgt: 'FriendlyShipObj'):
        """피격 무기 종류·위치(수선 위/아래)에 따라 격실 침수 유발."""
        if not self.cfg.get('enable_flooding', False):
            return
        if not tgt.alive or tgt.is_sub_hull or tgt._surv is None:
            return   # 치명적 HP 피격으로 이미 격침된 함정은 침수 무의미
        # 무기 위력·피격위치 분류
        if getattr(m, 'is_torpedo', False):
            wclass, pcat = 'torpedo', 'torpedo'
        elif getattr(m, 'is_ballistic', False) or getattr(m, 'is_hgv', False):
            wclass, pcat = 'heavy', 'ballistic'      # 상부 수직 강하 — 수선하 드묾
        else:
            wclass = 'heavy' if m.speed_ms >= 600 else 'medium'  # 초음속=대형 탄두
            pcat   = 'missile'
        if random.random() >= FLOOD_BELOW_WL_PROB.get(pcat, 0.30):
            return   # 수선상 명중 — 침수 없음 (서브시스템 피해만)
        comp   = max(1, tgt._surv['compartments'])
        breach = FLOOD_WARHEAD_FACTOR[wclass] / math.sqrt(comp)
        tgt.add_flooding(breach)
        self._log(f"[침수] {tgt.name} 수선하 피격 — 격실 침수 (침수율 {tgt.flood:.0%})")

    def _update_flooding(self):
        """매 틱 침수 진행 + 손상통제(펌프) 경쟁 → 복원력 한계 초과 시 침몰."""
        if not self.cfg.get('enable_flooding', False):
            return
        for ship in self.friendly_ships:
            if not ship.alive or ship.is_sub_hull or ship._surv is None:
                continue
            if ship.flood <= 0.0 and ship.flood_rate <= 0.0:
                continue
            rb  = ship._surv['reserve_buoyancy']
            net = ship.flood_rate - ship._surv['dc_rating']   # 유입 vs 펌프 배수
            ship.flood += net * DT
            if ship.flood <= 0.0:
                ship.flood = 0.0; ship.flood_rate = 0.0       # 완전 배수·봉쇄 → 복구
                ship._flood_warned = False
                continue
            if ship.flood >= rb:
                ship.alive = False
                ship.sunk_by_flood = True
                self._log(f"[침몰] {ship.name} 복원력 한계 초과 — 침수 침몰 "
                          f"(침수율 {ship.flood:.0%})")
                continue
            # 침수가 손상통제를 앞서는 중 → 침몰 예상 시간 1회 경고
            if net > 0 and not getattr(ship, '_flood_warned', False):
                ship._flood_warned = True
                eta = (rb - ship.flood) / net
                self._log(f"[침수 경고] {ship.name} 손상통제 역부족 — "
                          f"침수율 {ship.flood:.0%}, 침몰 예상 {eta/60:.1f}분")

    def _is_over(self) -> bool:
        active_threats = [m for m in self.missiles if m.alive and m.mtype == 'enemy_strike']
        # 이탈 중인 항공기: 재공격 가능하면 활성 위협으로 유지, 아니면 종료로 간주
        # 무장 소진 위협은 발사 불가 → 재무장 복귀 중이므로 활성 위협에서 제외
        enemy_active   = [et for et in self.enemy_threats
                          if et.alive
                          and not (self._munition_limit and et.munition_remaining <= 0)
                          and not (et.is_aircraft and et.is_retreating
                                   and et.reattack_count >= et.max_reattacks)]

        if not active_threats and not enemy_active:
            self._log("[종료] 교전 종료 - 모든 위협 소진/격침/이탈")
            return True
        if all(not s.alive for s in self.friendly_ships):
            self._log("[종료] 아군 전멸")
            return True
        return False

    # ── v10.7: 전술 의사결정 헬퍼 ───────────────────────────────────────────────

    def _make_tactical_state(self) -> 'TacticalState':
        """현재 시뮬 상태 스냅샷 → TacticalState."""
        primary = self._primary()
        threats = [
            {
                'name':     et.preset_name,
                'type':     et.info.get('type', '?'),
                'hp':       et.hp,
                'dist_km':  round(primary.pos.dist_to(et.pos) / 1000, 1),
                'speed_ms': et.speed_ms,
            }
            for et in self.enemy_threats if et.alive
        ]
        ships = [
            {
                'name':    s.name,
                'alive':   s.alive,
                'hp':      s.hp,
                'max_hp':  s._max_hp,
                'radar':   round(s.radar_factor, 2),
                'speed':   round(s.speed_factor, 2),
            }
            for s in self.friendly_ships
        ]
        return TacticalState(
            t             = self.t,
            threats       = threats,
            ships         = ships,
            intercepted   = self.stats['intercepted_threats'],
            total_threats = self.stats['total_threats'],
            shots_fired   = self.stats['total_missiles_fired'],
        )

    def _apply_tactical_choice(self, choice: dict):
        """사용자 전술 선택 반영 — cfg 일시 재설정."""
        if not choice:
            return
        # 무기 우선순위: 다음 교전 구간 _select_defense_wpn 재정의 대신 cfg 플래그 사용
        wpn_priority = choice.get('weapon_priority', 'auto')
        if wpn_priority != 'auto':
            self.cfg['_tactical_wpn_priority'] = wpn_priority
        else:
            self.cfg.pop('_tactical_wpn_priority', None)
        # 살보 수 조정
        max_salvo = choice.get('max_salvo', None)
        if max_salvo is not None:
            self.cfg['_tactical_max_salvo'] = int(max_salvo)
        else:
            self.cfg.pop('_tactical_max_salvo', None)

    # ── 메인 루프 ─────────────────────────────────────────────────────────────

    def _simulate(self):
        """Phase 5.5c: 시뮬 본체(제너레이터). 전술 결정 지점에서 yield state →
        호출자가 send(choice). run()이 cb로 구동(동작 보존), RL은 직접 send로 구동(스레드 0)."""
        pt = {'위치갱신': 0.0, '적발사': 0.0, '대공방어': 0.0,
              '아군공격': 0.0, '대잠': 0.0, '적방어': 0.0,
              '적Anti-SAM': 0.0, '교전판정': 0.0}
        _pc = _time.perf_counter
        _step_interval = 20   # step_cb 호출 간격(초) — 단일 시뮬 UI 갱신용

        # v16.7 기뢰전: 작전 해역 진입 시 기뢰 접촉 1회 판정(틱 루프 전, OFF면 무동작)
        self._apply_mine_exposure()

        while self.t <= self._sim_time_cap:
            # v12.5: 동적 기상 전이 판정
            if self._wx_dyn_enabled and self.t >= self._wx_next_check:
                self._update_weather()
                self._wx_next_check += WEATHER_STEP_INTERVAL_S

            # NEW-A: 혼합 시나리오 파도 지연 스폰
            if self._pending_threats:
                due = [s for (spawn_t, s) in self._pending_threats if spawn_t <= self.t]
                self._pending_threats = [(spawn_t, s) for (spawn_t, s) in self._pending_threats
                                         if spawn_t > self.t]
                for spec in due:
                    self._spawn_pending_threat(spec)

            # v8.15: 항모 함재기 파도 스폰
            for _ci, et in enumerate(self.enemy_threats):
                if (et.alive and et.carrier_aircraft
                        and et.carrier_wave_interval > 0
                        and self.t > 0
                        and self.t - et._last_wave_t >= et.carrier_wave_interval):
                    # v15.09.01: 전장 모드는 항공단 규모만큼만 발진(무한 생산 차단). 단발은 종전대로 2기씩.
                    spec = {'preset': et.carrier_aircraft, 'count': 2}
                    if self._enforce_wing_cap and et.carrier_wing > 0:
                        remaining = et.carrier_wing - et.wing_launched
                        if remaining <= 0:
                            continue                 # 항공단 소진 — 더는 발진 없음(재무장 복귀 시 재발진)
                        spec['count'] = min(2, remaining)
                        et.wing_launched += spec['count']
                        spec['carrier_owner'] = _ci   # v15.09.02: 재무장 복귀용 모항 식별
                    et._last_wave_t = self.t
                    self._pending_threats.append((self.t + 10.0, spec))  # 10초 후 출격

            # v9.6: 기습 잠수함 은닉 해제 탐지 이벤트
            for et in self.enemy_threats:
                if (et.is_sub and not et.ambush_revealed
                        and et.hidden_until > 0 and self.t >= et.hidden_until):
                    et.ambush_revealed = True
                    dist_km = self._primary().pos.dist_to(et.pos) / 1000
                    self._log(
                        f"[⚠ 기습 탐지!] {et.preset_name} 잠수함 은닉 해제 "
                        f"(거리 {dist_km:.0f}km) — 즉시 교전 개시!"
                    )

            # v8.26: 생존 적 항공기 중 최대 ECM 강도 캐싱 (에어리어 재밍)
            self._active_ecm = max(
                (et.ecm_power for et in self.enemy_threats
                 if et.alive and et.is_aircraft),
                default=0.0
            )
            # v12.6: IFF p_fail 계산용 생존 위협 수 캐싱 (틱당 1회)
            self._n_alive_threats: int = sum(1 for et in self.enemy_threats if et.alive)
            _t0 = _pc(); self._update_positions(); pt['위치갱신'] += _pc() - _t0
            self._aircraft_recon()   # v16.12: 정찰 드론 OTH 탐지 확장(탐지·로깅 전 갱신)
            self._unmanned_picket_update()  # v16.12: 무인 USV·UUV 피켓 탐지 확장
            self._log_detections()   # 첫 탐지 시점 로깅 (위치 갱신 후, 적 발사 전)
            if self._adaptive_ai:    # v15.1: 적응형 전술 재평가 (발사 직전)
                self._adaptive_tactic_update()
            _t0 = _pc(); self._enemy_fire();       pt['적발사']   += _pc() - _t0
            self._arm_radar_off_check()
            _t0 = _pc(); self._friendly_defense(); pt['대공방어'] += _pc() - _t0
            self._laser_defense()   # v17.2: 지향성 에너지 무기(레이저) — OFF 시 무동작
            _t0 = _pc(); self._friendly_strike();  pt['아군공격'] += _pc() - _t0
            _t0 = _pc(); self._aircraft_asw(); self._aircraft_cap(); self._aircraft_aas(); pt['대잠'] += _pc() - _t0
            _t0 = _pc(); self._enemy_defense();    pt['적방어']   += _pc() - _t0
            _t0 = _pc(); self._enemy_anti_sam();   pt['적Anti-SAM'] += _pc() - _t0
            _t0 = _pc()
            self._check_intercepts()
            self._check_hits()
            self._update_flooding()   # v12.4: 침수 진행·침몰 판정
            pt['교전판정'] += _pc() - _t0

            # 소멸·요격 미사일의 추적 채널 반환
            for m in self.missiles:
                if not m.alive or m.intercepted:
                    self.track_radar.release(m.uid)
                    # 위협 추적 표용 — 리스트에서 빠지기 직전 적 공격 미사일 누적 (단일 시뮬만)
                    if (not self._mc_mode and m.mtype == 'enemy_strike'
                            and m.uid not in self._retired_uids):
                        self._retired_uids.add(m.uid)
                        self._retired_strikes.append(m)
            self.missiles = [m for m in self.missiles
                             if m.alive and not m.intercepted]

            if not self._mc_mode:
                self._record_frame()

            # 포팅 D: 동시 위협 수 peak 추적
            alive_count = sum(
                1 for et in self.enemy_threats
                if et.alive and not (et.is_aircraft and et.is_retreating)
            )
            if alive_count > self.stats['peak_concurrent_threats']:
                self.stats['peak_concurrent_threats'] = alive_count

            # v10.7: 전술 의사결정 모드 — 구간마다 일시정지 후 사용자 선택 반영
            # Phase 5.5c: cb 호출 대신 yield — run() 래퍼가 cb로 응답, RL은 send(choice).
            # 조건 유지(cb 없으면 yield 스킵) → 단발 bit-identical.
            if (self._tactical_pause_cb and self.t > 0
                    and int(self.t) % self._tactical_interval == 0):
                state = self._make_tactical_state()
                choice = yield state
                self._apply_tactical_choice(choice)

            # 단일 시뮬 진행 콜백 (MC 배치 워커에서는 None)
            if self._step_cb and int(self.t) % _step_interval == 0:
                vls_rem = sum(
                    getattr(s, 'vls_remaining', 0) for s in self.friendly_ships)
                last_log = self._log_entries[-1][1] if self._log_entries else ""
                self._step_cb(self.t, self._sim_time_cap, alive_count, vls_rem, last_log)

            if self._is_over():
                break

            self.t += DT

        self._phase_times = pt
        return self._compile()

    def run(self) -> dict:
        """Phase 5.5c: _simulate() 제너레이터를 cb로 구동하는 얇은 래퍼(동작 보존).
        단발(cb None)은 _simulate가 yield 안 함 → 첫 next에서 즉시 StopIteration."""
        gen = self._simulate()
        try:
            state = next(gen)
            while True:
                choice = self._tactical_pause_cb(state)
                state = gen.send(choice)
        except StopIteration as e:
            return e.value

    def _compile(self) -> dict:
        # v16.12: 무인정(USV·UUV) 손실은 인명피해 0 → friendly_ships_lost에서 분리 집계.
        # 무인정 미편성(기본 OFF) 시 is_unmanned 함정이 없어 기존과 동일(회귀 보존).
        self.stats['friendly_ships_lost']   = sum(1 for s in self.friendly_ships
                                                  if not s.alive and not s.is_unmanned)
        self.stats['unmanned_lost']         = sum(1 for s in self.friendly_ships
                                                  if not s.alive and s.is_unmanned)
        # v12.4: 침수로 침몰한 함정 수 + 생존 함정 중 침수 진행 중인 수
        self.stats['ships_sunk_by_flood']   = sum(1 for s in self.friendly_ships if getattr(s, 'sunk_by_flood', False))
        self.stats['ships_flooding']        = sum(1 for s in self.friendly_ships if s.alive and getattr(s, 'flood', 0.0) > 0.0)
        # 이탈 항공기(alive=False, is_retreating=True, intercepted=False)는 "격침" 아님
        self.stats['enemy_ships_destroyed'] = sum(
            1 for et in self.enemy_threats
            if not et.alive and (et.intercepted or not et.is_aircraft)
        )
        # 포팅 C: 항공 자산 출격 횟수 + 비용 합산
        self.stats['aircraft_sorties'] = sum(ac.sorties for ac in self.aircraft)
        self.stats['total_cost']       = (
            sum(s.total_cost for s in self.friendly_ships)
            + sum(ac.total_cost for ac in self.aircraft)
            + self._ground_cost  # v9.4: 지상 발사 자산 비용
        )

        intercept_rate = (
            min(1.0, self.stats['intercepted_threats'] / self.stats['total_threats'])
            if self.stats['total_threats'] > 0 else 1.0
        )

        # 포팅 D: 잔여 재고 합산 (REQ-07), 총 채널 수 (REQ-08)
        remaining_inv: dict = {}
        for s in self.friendly_ships:
            for wpn, cnt in s.inventory.items():
                remaining_inv[wpn] = remaining_inv.get(wpn, 0) + cnt
        total_channels = sum(s.max_channels for s in self.friendly_ships)

        ship_subsystem_damage = {
            s.name: {
                'radar_factor':    round(s.radar_factor, 3),
                'speed_factor':    round(s.speed_factor, 3),
                'channel_factor':  round(s.channel_factor, 3),
                'disabled_weapons': sorted(s.disabled_weapons),
                'alive':           s.alive,
                'hp':              s.hp,
                'max_hp':          s._max_hp,
                'hits_taken':      s.hits_taken,
            }
            for s in self.friendly_ships
        }

        # v10.6: 항모 격침/전투불능 판정
        carrier_status = {
            et.preset_name: {
                'alive': et.alive,
                'hp': et.hp,
                'max_hp': et._HP_MAP.get('항모', 5),
                'status': '격침' if not et.alive else ('전투불능' if et.hp <= 2 else '정상'),
            }
            for et in self.enemy_threats
            if et.info.get('type') == '항모'
        }

        # v14.1: 3D 전장 커버리지 돔용 — 함대 대표 1겹씩(레이더 탐지권·SAM 교전권)
        coverage = self._build_coverage()

        return {
            **self.stats,
            'intercept_rate':    intercept_rate,
            'sim_time':          self.t,
            'frames':            self.frames,
            'log':               self._log_entries,
            'friendly_ships':    self.friendly_ships,
            'enemy_ships':       self.enemy_threats,   # 하위 호환 키 유지
            'remaining_inventory': remaining_inv,
            'total_channels':      total_channels,
            'used_seed':           self.cfg.get('sim_seed', None),
            'ship_subsystem_damage': ship_subsystem_damage,
            'active_events':     [] if self._mc_mode else self._build_active_events(),
            'strike_log':        self.strike_log,              # v9.3
            'vls_depletion_t':   dict(self.vls_depletion_t),  # v9.4
            'ground_remaining':   dict(self.ground_inv),         # v9.4 / v9.11
            'ashore_sm3_fired':  self.stats['ashore_sm3_fired'],
            'thaad_fired':       self.stats['thaad_fired'],
            'lsam_fired':        self.stats['lsam_fired'],
            'chungung_fired':    self.stats['chungung_fired'],
            'patriot_fired':     self.stats['patriot_fired'],   # v20.1
            'phase_times':       dict(self._phase_times),      # v8.26: 단계별 소요시간
            'carrier_status':    carrier_status,               # v10.6
            'coverage':          coverage,                     # v14.1: 3D 커버리지 돔
            'enemy_tactic_switches': self._adaptive_switches,  # v15.01.03: 적 전술 전환 횟수
        }

    def _build_coverage(self) -> dict:
        """
        v14.1/14.3: 3D 전장 커버리지 돔 데이터.
          radar : 최대 대공 센서 함정 중심 + 함대 공유 실효 탐지거리(detect_km)
                  + 저고도(해면) 수평선 한계(low_km) + 해역(region, 지형 차폐용)
        함대 대표로 레이더 탐지권 1겹만 생성(SAM 교전권은 미표시).
        MC 모드는 frames가 없어 3D 미사용 → 빈 dict.
        """
        if self._mc_mode:
            return {}
        alive = [s for s in self.friendly_ships if s.alive]
        if not alive:
            return {}

        # 레이더 탐지권 — 최대 대공 센서 함정 중심, 함대 공유 detect_km 반경
        radar_ship = max(alive, key=lambda s: s.sensor_km.get('대공', 0))
        radar_km = self.cfg.get('detect_km') or radar_ship.sensor_km.get('대공', 0) or 0
        if radar_km <= 0:
            return {}

        # v14.3 저고도 수평선 한계 — 4/3 지구반경 모델 d≈4.12(√h_r+√h_t) km.
        # 안테나 고도 h_r≈25m(이지스 SPY급), 해면 밀착 표적 h_t≈10m → ~33km.
        horizon_km = 4.12 * (math.sqrt(_RADAR_ANT_H_M) + math.sqrt(_SEA_SKIM_H_M))
        low_km = min(float(radar_km), horizon_km)

        radar = {
            'ship':   radar_ship.name,
            'km':     float(radar_km),
            'low_km': round(low_km, 1),
            'region': self.cfg.get('fleet_region', '동해 중부'),
        }
        # v14.3 지형 차폐 — enable_terrain ON일 때만 저고도 경계를 육지 방위로 패이게
        # (엔진 탐지의 _terrain_penalty와 동일 조건). OFF면 수평선 한계 대칭 원.
        if self.cfg.get('enable_terrain', False):
            region_key = REGION_TO_ACOUSTIC_KEY.get(radar['region'], 'EAST_SEA')
            terr = _TERRAIN_SHADOW.get(region_key)
            if terr:
                radar['terrain'] = {'bearing': terr[0], 'shadow_deg': terr[1]}
        return {'radar': radar}

    def _build_active_events(self) -> list:
        """A-1: enemy_strike MissileObj → EngagementAnalysisTab 어댑터 리스트."""
        class _EvAdapter:
            __slots__ = ('label','is_active','intercepted','intercept_weapon',
                         'intercept_km','t_intercepted','gantt_bars','detect_m','enemy_info')
        evs = []
        # 요격·명중으로 빠진 위협(_retired_strikes) + 현재 비행 중인 적 공격 미사일 모두 포함
        all_strikes = self._retired_strikes + [
            m for m in self.missiles if m.mtype == 'enemy_strike']
        for m in all_strikes:
            ev = _EvAdapter()
            ev.label            = m.name
            ev.is_active        = True
            ev.intercepted      = m.intercepted
            ev.intercept_weapon = m.intercept_weapon
            ev.intercept_km     = m.intercept_km if m.intercept_km else None
            ev.t_intercepted    = m.t_intercept
            ev.detect_m         = m.detect_m
            ev.enemy_info       = m.enemy_info
            # gantt_bars: 위협 비행 전 구간 단일 바
            t_end   = m.t_intercept or self.t
            color   = '#2ecc71' if m.intercepted else '#e74c3c'
            ev.gantt_bars = [(m.name, m.t_spawn, t_end, color)]
            evs.append(ev)
        return evs


# ════════════════════════════════════════════════════════════════════════════
#  v14.1: 3D 전장 — frames → CesiumJS CZML 변환 (순수 읽기, 회귀 무영향)
# ════════════════════════════════════════════════════════════════════════════

def _low_radius_km(base_km: float, th_rad: float, radar: dict) -> float:
    """
    저고도(해면) 탐지 경계의 방위각별 반경(km).
    수평선 한계(base_km)에 지형 음영을 곱해 육지 방위 섹터를 안으로 패이게 한다.
    - 음영 없음(enable_terrain OFF·해역 미정): 대칭 원.
    - 있음: 육지 방위(terrain.bearing)에서 최대 축소, ±90° 밖은 영향 0(cos 가중).
      축소율 fmin = max(0.2, 1 - 음영각×0.1) — 설악 8.1°→0.2, 태백 3.4°→0.66, 소백 0.9°→0.91.
    th_rad: 평면 방위각(동=0, 북=π/2).
    """
    terr = radar.get('terrain')
    if not terr:
        return base_km
    # 육지 방위와의 각거리(0~π)
    d = abs(((th_rad - terr['bearing'] + math.pi) % (2.0 * math.pi)) - math.pi)
    w = max(0.0, math.cos(d))                       # ±90° 내에서만 가중
    fmin = max(0.2, 1.0 - terr['shadow_deg'] * 0.1)
    return base_km * (1.0 - (1.0 - fmin) * w)


def _spread_offset(uid: str) -> tuple:
    """동시 발사 미사일이 한 점에 겹치지 않게 황금각 나선으로 미세 분산(시각용 오프셋)."""
    digits = ''.join(ch for ch in uid if ch.isdigit())
    n = int(digits) if digits else 0
    ang = n * 2.399963229728653          # 황금각(rad) — 고르게 퍼짐
    r = 200.0 * math.sqrt(n % 40)        # 나선 반경(최대 ~1.26km, 비행거리 44km 대비 미세)
    return math.cos(ang) * r, math.sin(ang) * r


def build_czml(result: dict, epoch_iso: str = "2026-01-01T00:00:00Z") -> list:
    """
    시뮬 결과의 frames(SimFrame 리스트)를 CesiumJS CZML 패킷 리스트로 변환한다.
    - frames는 단일 시뮬에서만 존재(MC 모드는 _record_frame 생략) → MC 결과엔 document만 반환.
    - 좌표 x,y(m) → LatLon.from_xy로 위경도 변환(시뮬 직후 모듈 _ref 기준이 유효).
    - 아군 함정·적함·미사일(궤적) + 적 미사일 소멸점 마커를 패킷으로 생성.
    """
    import datetime
    frames = result.get('frames', [])
    doc = {"id": "document", "name": "battlefield", "version": "1.0"}
    if not frames:
        return [doc]

    base = datetime.datetime.fromisoformat(epoch_iso.replace("Z", "+00:00"))

    def _iso(t):
        return (base + datetime.timedelta(seconds=float(t))).strftime("%Y-%m-%dT%H:%M:%SZ")

    t0, t1 = frames[0].t, frames[-1].t
    epoch0 = _iso(0)

    # 미사일 선수집 + 표시용 ripple(같은 시각 발사를 순서대로 1.5초씩 시차 표시 — 엔진 불변, 시각 전용)
    mis: dict = {}
    for f in frames:
        for m in f.missiles:
            mis.setdefault(m[0], (m[3], m[4], []))[2].append((f.t, m[1], m[2], m[5]))
    _grp: dict = {}
    for _uid, (_mt, _mn, _sq) in mis.items():
        _grp.setdefault(_sq[0][0], []).append(_uid)
    _ripple: dict = {}
    for _spawn_t, _uids in _grp.items():
        for _i, _u in enumerate(sorted(_uids)):
            _ripple[_u] = _i * 1.5
    disp_end = t1
    for _uid, (_mt, _mn, _sq) in mis.items():
        disp_end = max(disp_end, _sq[-1][0] + _ripple.get(_uid, 0.0))

    avail = f"{_iso(t0)}/{_iso(disp_end)}"   # ripple로 미사일이 늦게 끝나는 만큼 연장
    doc["clock"] = {
        "interval": avail, "currentTime": _iso(t0),
        "multiplier": 2, "range": "LOOP_STOP", "step": "SYSTEM_CLOCK_MULTIPLIER",
    }
    packets = [doc]

    def _cart(seq, with_alt=False):
        out = []
        for it in seq:
            ll = LatLon.from_xy(it[1], it[2])
            h = float(it[3]) if (with_alt and len(it) > 3 and it[3]) else 0.0
            out += [float(it[0]), round(ll.lon, 6), round(ll.lat, 6), h]
        return out

    # ── 아군 함정 ──
    fri: dict = {}
    for f in frames:
        for s in f.friendly_ships:
            fri.setdefault(s[0], []).append((f.t, s[1], s[2]))

    # ── 레이더 탐지권 돔 (1겹, 함정 따라 이동) ──
    # ellipsoid 중심 고도 0 → 아래 반구는 지구에 가려져 위쪽 반투명 반구만 표시.
    cov = result.get('coverage') or {}
    def _dome(layer, color, line):
        info = cov.get(layer)
        if not info:
            return
        seq = fri.get(info['ship'])
        if not seq:
            return
        r = float(info['km']) * 1000.0
        packets.append({
            "id": f"coverage/{layer}", "name": f"{info['ship']} {layer}",
            "availability": avail,
            "position": {"epoch": epoch0, "cartographicDegrees": _cart(seq),
                         "forwardExtrapolationType": "HOLD"},
            "ellipsoid": {
                "radii": {"cartesian": [r, r, r]},
                "fill": True,
                "material": {"solidColor": {"color": {"rgba": color}}},
                "outline": True, "outlineColor": {"rgba": line}, "outlineWidth": 1,
                "slicePartitions": 32, "stackPartitions": 16,
            },
        })
    _dome('radar', [0, 200, 255, 28], [0, 200, 255, 90])

    # ── 저고도(해면) 탐지 경계 폴리곤 (수평선·지형 차폐 반영) ──
    # 반구는 고고도 탐지권 / 이 외곽선은 해면 밀착 표적의 탐지 한계.
    # 중심 = 레이더 함정 시작 위치 고정(교전 짧고 저속 → 이동 무시 가능).
    _radar = cov.get('radar')
    if _radar and _radar.get('low_km'):
        _seq = fri.get(_radar['ship'])
        if _seq:
            cx, cy = _seq[0][1], _seq[0][2]
            base_km = float(_radar['low_km'])
            N = 72
            pts = []
            for i in range(N):
                th = 2.0 * math.pi * i / N
                r_m = _low_radius_km(base_km, th, _radar) * 1000.0
                ll = LatLon.from_xy(cx + r_m * math.cos(th), cy + r_m * math.sin(th))
                pts += [round(ll.lon, 6), round(ll.lat, 6), 0.0]
            packets.append({
                "id": "coverage/lowalt", "name": "저고도 탐지 한계",
                "availability": avail,
                "polygon": {
                    "positions": {"cartographicDegrees": pts},
                    "height": 0,
                    "material": {"solidColor": {"color": {"rgba": [0, 210, 255, 50]}}},
                    "outline": True, "outlineColor": {"rgba": [140, 235, 255, 210]},
                },
            })

    for name, seq in fri.items():
        packets.append({
            "id": f"ship/{name}", "name": name, "availability": avail,
            "position": {"epoch": epoch0, "cartographicDegrees": _cart(seq),
                         "forwardExtrapolationType": "HOLD"},
            "point": {"pixelSize": 13, "color": {"rgba": [0, 190, 255, 255]},
                      "outlineColor": {"rgba": [255, 255, 255, 255]}, "outlineWidth": 2},
            "label": {"text": name, "font": "12px sans-serif", "scale": 0.85,
                      "pixelOffset": {"cartesian2": [0, -22]},
                      "fillColor": {"rgba": [180, 235, 255, 255]},
                      "showBackground": True, "backgroundColor": {"rgba": [0, 20, 40, 160]}},
            "path": {"width": 2, "leadTime": 0, "trailTime": 99999, "resolution": 5,
                     "material": {"solidColor": {"color": {"rgba": [0, 190, 255, 110]}}}},
        })

    # ── 적함 ──
    ene: dict = {}
    for f in frames:
        for e in f.enemy_ships:
            ene.setdefault(e[0], (e[1], []))[1].append((f.t, e[2], e[3]))
    for uid, (preset, seq) in ene.items():
        packets.append({
            "id": f"enemy/{uid}", "name": preset, "availability": avail,
            "position": {"epoch": epoch0, "cartographicDegrees": _cart(seq),
                         "forwardExtrapolationType": "HOLD"},
            "point": {"pixelSize": 12, "color": {"rgba": [255, 80, 80, 255]},
                      "outlineColor": {"rgba": [60, 0, 0, 255]}, "outlineWidth": 2},
            "label": {"text": preset, "font": "11px sans-serif", "scale": 0.8,
                      "pixelOffset": {"cartesian2": [0, -20]},
                      "fillColor": {"rgba": [255, 200, 200, 255]},
                      "showBackground": True, "backgroundColor": {"rgba": [40, 0, 0, 160]}},
        })

    # ── 미사일 궤적 + 적 미사일 소멸점 마커 (mis는 위에서 선수집) ──
    threat_times = []   # 발수 카운터용 적 미사일 [등장t, 소멸t]
    for uid, (mtype, mname, seq) in mis.items():
        ox, oy = _spread_offset(uid)
        ro = _ripple.get(uid, 0.0)
        seq = [(t + ro, x + ox, y + oy, a) for (t, x, y, a) in seq]   # 분산 + 표시용 ripple 시차
        ta, tb = seq[0][0], seq[-1][0]
        is_enemy = (mtype == 'enemy_strike')
        if is_enemy:
            threat_times.append([ta, tb])
        col = [255, 90, 60, 255] if is_enemy else [90, 230, 140, 255]
        packets.append({
            "id": f"missile/{uid}", "name": mname,
            "availability": f"{_iso(ta)}/{_iso(tb)}",
            "position": {"epoch": epoch0, "cartographicDegrees": _cart(seq, with_alt=True)},
            "point": {"pixelSize": 6, "color": {"rgba": col}},
            "path": {"width": 1.5, "leadTime": 0, "trailTime": 6, "resolution": 1,
                     "material": {"solidColor": {"color": {"rgba": col[:3] + [150]}}}},
        })
        if is_enemy:
            ll = LatLon.from_xy(seq[-1][1], seq[-1][2])
            packets.append({
                # 요격 순간 2.5초만 번쩍 — 되감기/누적 방지
                "id": f"impact/{uid}", "availability": f"{_iso(tb)}/{_iso(tb + 2.5)}",
                "position": {"cartographicDegrees": [round(ll.lon, 6), round(ll.lat, 6), 0]},
                "point": {"pixelSize": 9, "color": {"rgba": [255, 210, 0, 230]},
                          "outlineColor": {"rgba": [255, 90, 0, 255]}, "outlineWidth": 2},
            })

    doc["_threatTimes"] = threat_times   # JS 발수 카운터가 시각별로 집계
    return packets


# ════════════════════════════════════════════════════════════════════════════
#  탐지거리 자동 계산 (함대 편성 + 날씨 + 데이터링크)
# ════════════════════════════════════════════════════════════════════════════

def calculate_fleet_detect_ranges(fleet_preset_name: str, weather: str,
                                  fleet_list: list | None = None) -> dict:
    """
    함대 편성과 날씨를 기반으로 탐지거리를 자동 계산한다.
    fleet_list(임의 편성 [{name,type}]) 주어지면 프리셋 대신 사용(A1 캠페인 정밀 교전
    등 프리셋에 없는 편성용). 미지정 시 기존 동작 그대로 → 회귀 bit-identical.

    데이터링크 원칙:
      - 한국 해군 Link-16/Link-11 적용 — 편대 내 최고 성능 센서 기준 공유
      - 대공·대함 : 편대 내 max(sensor_km['대공'/'대함']) × radar_factor
      - 대잠       : 편대 내 max(sensor_km['대잠']) × sonar_factor
        (황사는 소나에 영향 없음, 풍랑·폭풍은 해상 소음으로 급감)

    반환 예시:
      {'대공': 1140, '대함': 41, '대잠': 30,
       'leading_ship': 'KDX-III', 'radar_factor': 0.95, 'sonar_factor': 0.60}
    """
    preset = fleet_list if fleet_list is not None else FLEET_PRESETS.get(fleet_preset_name, [])
    w = _make_physics_wx(weather)   # v9.13: Beaufort 물리값 적용
    rf = w.get('radar_factor', w.get('detect_range_factor', 1.0))
    sf = w.get('sonar_factor', w.get('detect_range_factor', 1.0))

    max_air = 0; max_surface = 0; max_sub = 0
    leading = '(없음)'
    for ship in preset:
        spec = SHIP_DB.get(ship['type'], {})
        s = spec.get('sensor_km', {})
        air = s.get('대공', 0)
        if air > max_air:
            max_air = air
            leading = ship.get('name', ship['type'])
        max_surface = max(max_surface, s.get('대함', 0))
        max_sub     = max(max_sub,     s.get('대잠', 0))

    return {
        '대공':         max(1, round(max_air     * rf)),
        '대함':         max(1, round(max_surface * rf)),
        '대잠':         max(1, round(max_sub     * sf)),
        'leading_ship': leading,
        'radar_factor': rf,
        'sonar_factor': sf,
    }


# ════════════════════════════════════════════════════════════════════════════
#  외부 API
# ════════════════════════════════════════════════════════════════════════════

def run_v7_simulation(cfg: dict, step_cb=None, tactical_cb=None) -> dict:
    # 탐지거리 자동 계산 (함대 + 날씨 기반, 수동 override 없을 때)
    if not cfg.get('detect_km_manual', False):
        ranges = calculate_fleet_detect_ranges(
            cfg.get('fleet_preset', '단독 작전'),
            cfg.get('weather', '맑음 (주간)'))
        cfg = dict(cfg)
        cfg['detect_km']         = ranges['대공']
        cfg['surface_detect_km'] = ranges['대함']
        cfg['sub_detect_km']     = ranges['대잠']
    else:
        cfg = dict(cfg)
    # v12.5: 동적 기상 전이가 cfg['weather']를 변경할 수 있으므로 초기값 보존
    _initial_weather = cfg.get('weather', '맑음 (주간)')
    sim = TimeStepEngine(cfg, step_cb=step_cb)
    if tactical_cb:
        sim._tactical_pause_cb = tactical_cb  # v10.7: 전술 모드 훅 주입
    result = sim.run()
    cfg['weather'] = _initial_weather  # 초기 날씨 복원 (결과·보고서 표시용)
    return result


# ════════════════════════════════════════════════════════════════════════════
#  지속 전장 엔진 (Phase 1 — 아키텍처 전환)
#  단발 살보 교전을 양측이 작전 목표를 두고 시간에 걸쳐 겨루는 지속 전장으로 전환.
#  TimeStepEngine을 상속해 물리층(요격·침수·미사일 비행 등)을 그대로 재사용하고,
#  오케스트레이션(종료조건·승패 출력·적 지속 압박)만 오버라이드한다. 부모는 무수정.
#  정본 설계: plan_battle_engine.md
# ════════════════════════════════════════════════════════════════════════════

# 표준 전장 시간 지평(초) — Phase 5.1: 1800→7200(2h). 1800 컷오프가 자연 결판 전
# 전장을 끊어 가짜 draw 생성(전면전 실측 6764s 자연종료) → 작전급으로 상향.
BATTLE_HORIZON_S = 7200

# 자원 지속성 평가에서 무제한 탄(CIWS 초기 9999 등)을 거르는 임계 — 이 값 이상이면 제외
_BATTLE_INF_AMMO = 1000

# ── 연료 소모 모델 (v15.08.01) — 실측 항속거리를 1800s 표준 작전 기준으로 정규화 ──
# 절대 항속시간(수백 시간)을 그대로 쓰면 1800s엔 거의 안 닳아 무의미 → 표준함이
# 1800s 평탄 순항 시 0.40 소모하도록 스케일하고, 함정별 실측 작전가능시간 비로 차등.
_FUEL_STD_BURN_1800 = 0.40       # 표준함이 1800s 순항 시 소모하는 연료 비율(잔여 0.60)
_FUEL_T_REF_H       = 5500 / 18  # 표준함(KDX-III) 작전가능시간[h] = 항속/순항 — 정규화 기준
_FUEL_BURN_MIN      = 0.10       # 1800s 순항 소모 하한(대형함·잠수함도 최소 소모)
_FUEL_BURN_MAX      = 0.70       # 상한(소형 고속정 1800s 내 완전 고갈·조기탈락 방지)
# 회피 기동 추가 소모: 점프 거리(300~800m/틱)는 회피 효과 표현용 추상값이라 물리 연료로
# 부적합 → 거리 대신 '회피 발생 틱당 고정량'으로 과금하고, 함정별 누적 상한을 둔다.
# (거리 비례 시 다살보 교전에서 회피가 매 틱 누적돼 연료가 0으로 붕괴 — 차등 무의미해짐)
_FUEL_MANEUVER_TICK = 0.0015     # 회피 발생 틱 1회당 추가 연료
_FUEL_MANEUVER_CAP  = 0.35       # 회피 누적 추가 소모 상한(격렬 교전이어도 차등 보존)

# ── 해상급유(RAS) 재보급 모델 (v15.09.03) — 군수지원함이 전투함 연료 보충 ──
# 화물 연료(함대 보급용)는 지원함 자체 추진 연료(_fuel_burn)와 분리. 정규화 '척분'.
_RAS_RESERVE     = {'AOE': 4.0, 'AO': 1.5}  # 지원함 화물 연료(전투함 완전급유 횟수 환산)
_RAS_RATE_PER_S  = 0.002         # 급유 스테이션당 초당 보충 연료 비율
_RAS_STATIONS    = 2             # 지원함당 동시 급유 함정 수(양현 RAS 스테이션)
_RAS_TARGET      = 0.95          # 이 잔여비까지만 보충(완전 만재 직전서 정지)

# ── v17.1: RAS 탄약 재보급 (enable_ras_rearm, 지속 전장 전용) ──
# 연료 화물(ras_reserve)과 별도로, 지원함이 주요 SAM 재장전용 탄약 화물을 보유.
# 소강기에만(위협 접근 시 중단, 연료 RAS와 조건 공유) 소진된 전투함을 재장전.
_RAS_AMMO_RESERVE  = {'AOE': 40.0, 'AO': 15.0}  # 지원함 탄약 화물(재보급 가능 주요 SAM 발수)
_RAS_REARM_PER_S   = 0.05        # 스테이션당 초당 재보급 속도(느림 — VERTREP/CONREP 반영)
_RAS_REARM_TRIGGER = 0.30        # 주요 SAM 합이 초기 재고의 30% 미만인 함정만 대상
_RAS_REARM_TARGET  = 0.80        # 초기 재고의 80%까지만 보충(완전 만재 안 함)
_RAS_MAJOR_SAM     = ('SM-3 Block IIA', 'SM-6', 'SM-2 Block IIIB')  # 재보급 대상 주요 SAM


class Objective:
    """단일 작전 목표 — {type, side, weight, progress(0..1), status}."""

    def __init__(self, otype: str, side: str, weight: float, params: dict = None):
        self.type    = otype          # 'defend_asset' | 'destroy_asset' | ...
        self.side    = side           # 'friendly' | 'enemy'
        self.weight  = weight
        self.params  = params or {}
        self.progress = 1.0 if side == 'friendly' else 0.0
        self.status  = '진행'         # '진행' | '달성' | '실패'

    def as_dict(self) -> dict:
        return {'type': self.type, 'side': self.side, 'status': self.status,
                'progress': round(self.progress, 3), 'weight': self.weight}


class BattleEngine(TimeStepEngine):
    """지속 전장 엔진. Phase 1 MVP: 자산 방어↔자산 격침 1쌍 + 목표 기반 종료·승패."""

    def __init__(self, cfg: dict, step_cb=None):
        # Phase 5.3.1: 전장 모드는 다축 스폰 기본 활성(목표지향 기동의 다축 압박 전제).
        # super().__init__ 전에 — _build_threats가 스폰 시 읽음. 명시값은 존중(시나리오가 끄면 OFF).
        cfg.setdefault('enable_multibearing', True)
        super().__init__(cfg, step_cb=step_cb)
        self.horizon_s = float(cfg.get('battle_horizon_s', BATTLE_HORIZON_S))
        # Phase 5.1: 루프 시간 상한을 horizon에 맞춰 상향(부모 기본 MAX_SIM_TIME으로는 horizon 전 끊김).
        # +60s 여유 — 마지막 틱에서 horizon 도달 종료가 정상 판정되도록.
        self._sim_time_cap = int(self.horizon_s) + 60
        self._enforce_wing_cap = True   # v15.09.01: 전장 모드는 항모 항공단 규모만큼만 함재기 발진
        # v15.09.02: 재출격 사이클 — 무장 소진 이탈한 함재기가 모항 귀환→재무장 지연 후 재발진.
        # 실제 항모 갑판 재무장 사이클 ~45분(기본 30분 전장선 거의 휴면, 작전급 긴 전장서 발현).
        self._rearm_delay  = float(cfg.get('battle_rearm_delay_s', 2700.0))
        self._rearm_queue  = []         # [(return_t, carrier_idx), ...]
        # Phase 5.2: 발사 재장전 쿨다운 — 화력을 시간축으로 펄스 분산(파상공격). 전 위협 공통.
        # 단발 모드는 부모 _reload_delay_s=0(즉시연사)으로 무영향. 기본 150s(5.0 관찰 후 튜닝).
        self._reload_delay = float(cfg.get('battle_reload_delay_s', 150.0))
        # v17.1: RAS 탄약 재보급(실험적·기본 OFF) — 지속 전장에서 소강기에 소진 SAM 재장전.
        self._ras_rearm = bool(cfg.get('enable_ras_rearm', False))
        # Phase 5.3.1: 목표지향 기동 — 수상함은 돌파선 지향(다축 분산 포진), 그 외는 기함 지향.
        self._sea_maneuver = bool(cfg.get('battle_sea_control_maneuver', True))
        # v15.08.04: 목표지향 적 AI — 항공 위협은 무장이 남는 한 계속 재접근(압박 유지).
        # 손실이 임계를 넘으면 생존 세력 전면 철수. (구 battle_air_reattacks 임시 레버 폐지)
        self._enemy_withdrawing  = False
        self._enemy_withdraw_loss = float(cfg.get('battle_enemy_withdraw_loss', 0.5))
        # Phase 5.6 self-play: ai_tactic='rl'이면 적 전술 모드(포화/분산/기만)를 RL 적 정책이
        # 결정한다. 부모 _adaptive_ai(adaptive 전용)를 rl 모드에도 켜 _enemy_fire가 _adaptive_mode를
        # 쓰게 하고, 규칙 재평가(_adaptive_tactic_update)는 skip(모드는 정책 cb가 세팅).
        # 기본(rl 아님)은 _enemy_rl=False → 적 전술 동작 불변(회귀 보존).
        self._enemy_rl = (cfg.get('ai_tactic') == 'rl')
        if self._enemy_rl:
            self._adaptive_ai = True   # 부모 속성 덮어쓰기(부모 코드 무수정)
        # 방어 자산 = 기함(primary). 시작 시점에 고정 식별(격침 판정용)
        self._asset        = self._primary()
        self._asset_max_hp = self._asset._max_hp
        w_def  = float(cfg.get('battle_w_defend',    0.5))   # Phase 2: 0.6→0.5 재배분
        w_sea  = float(cfg.get('battle_w_sea',       0.3))
        w_attr = float(cfg.get('battle_w_attrition', 0.2))   # v15.07.02
        # v15.07.01: 해역 통제 — 보호점(기함)까지 돌파선
        self._sea_line_km   = float(cfg.get('battle_sea_control_line_km', 50.0))
        self._sea_max_pen   = 0.0    # 작전 전체 최악(최대 침투) km
        self._sea_pen_now   = 0.0    # 현재 틱 침투 km
        self._frontline_tl  = []     # [(t, deepest_penetration_km), ...] 단일 시뮬만
        self._sea_tl_last_t = -1.0   # timeline 중복 t 방지
        # v15.07.02: 소모전 — 양측 전력가치(USD 조달가) 교환비. 미사일류는 제외(플랫폼만).
        self._attr_k         = float(cfg.get('battle_attrition_k', 5.0))  # 로지스틱 민감도
        fr_v0 = self._friendly_force_value()
        en_v0 = self._enemy_force_value()
        # 양측 모두 플랫폼 전력이 있을 때만 소모전 평가(순수 미사일 편대 등은 무의미 → 점수 제외)
        self._attr_active    = (fr_v0 > 0.0 and en_v0 > 0.0)
        if not self._attr_active:
            w_attr = 0.0
        self._fr_value_init  = max(1.0, fr_v0)
        self._en_value_init  = max(1.0, en_v0)
        self._fr_frac        = 1.0
        self._en_frac        = 1.0
        self._force_tl       = []     # [(t, fr_frac, en_frac), ...] 단일 시뮬만
        self._attr_tl_last_t = -1.0
        # v15.07.03: 자원 지속성(아군) — 탄약 잔여 최저비. 연료는 v15.08.01 도입 전까지 1.0 고정.
        w_sust = float(cfg.get('battle_w_sustainment', 0.2))
        self._ammo_init      = max(1.0, self._friendly_ammo_total())  # 방어+공격 인벤토리 초기 총량
        self._ammo_frac      = 1.0
        self._fuel_frac      = 1.0    # 연료 모델(v15.08.01) 도입 전까지 1.0
        self._resource_tl    = []     # [(t, ammo_frac, fuel_frac), ...] 단일 시뮬만
        self._sust_tl_last_t = -1.0
        self.objectives = [
            Objective('defend_asset',  'friendly', w_def,  {'asset': self._asset.name}),
            Objective('destroy_asset', 'enemy',    w_def,  {'asset': self._asset.name}),
            Objective('sea_control',   'friendly', w_sea,  {'line_km': self._sea_line_km}),
            Objective('sea_control',   'enemy',    w_sea,  {'line_km': self._sea_line_km}),
            Objective('attrition',     'friendly', w_attr, {}),
            Objective('attrition',     'enemy',    w_attr, {}),
            Objective('sustainment',   'friendly', w_sust, {}),
        ]

    # ── 목표 진행 갱신 ────────────────────────────────────────────────────────
    def _asset_alive(self) -> bool:
        return any(s is self._asset and s.alive for s in self.friendly_ships)

    def _sea_control_update(self):
        """해역 통제 — 보호점(기함)까지 생존 위협·적 미사일 최단거리로 침투량 갱신."""
        asset_pos = self._asset.pos
        line_km   = self._sea_line_km
        closest_m = min(
            (asset_pos.dist_to(et.pos) for et in self.enemy_threats if et.alive),
            default=float('inf'))
        closest_m = min(
            closest_m,
            min((asset_pos.dist_to(m.pos) for m in self.missiles
                 if m.alive and m.mtype == 'enemy_strike'), default=float('inf')))
        if closest_m == float('inf'):           # 위협 없음 → 침투 0
            self._sea_pen_now = 0.0
        else:
            self._sea_pen_now = max(0.0, min(line_km - closest_m / 1000.0, line_km))
        if self._sea_pen_now > self._sea_max_pen:
            self._sea_max_pen = self._sea_pen_now
        # timeline 누적 — 단일 시뮬만, 틱당 1회
        if not self._mc_mode and self.t != self._sea_tl_last_t:
            self._sea_tl_last_t = self.t
            self._frontline_tl.append((round(self.t, 1), round(self._sea_pen_now, 2)))

    # ── 소모전 전력가치 (USD 조달가 합, 미사일류 제외 = 플랫폼만) ───────────────
    def _friendly_force_value(self) -> float:
        return sum(SHIP_PROCUREMENT_USD.get(s.ship_type, 0)
                   for s in self.friendly_ships if s.alive)

    def _enemy_force_value(self) -> float:
        # 함정·잠수함·항공기만 — 미사일류는 소모품이라 전력가치에서 제외
        return sum(ENEMY_PROCUREMENT_USD.get(et.name, 0)
                   for et in self.enemy_threats
                   if et.alive and (et.is_ship or et.is_sub or et.is_aircraft))

    # ── Phase 5.2: 화력 동역학(웨이브) ─────────────────────────────────────────
    def _reload_delay_s(self, et) -> float:
        """전 위협 공통 재장전 지연 — 사거리 내 연사를 막아 화력을 파상공격으로 분산."""
        return self._reload_delay

    # ── Phase 5.3.1: 목표지향 기동(공간 확장) ──────────────────────────────────
    def _threat_move_target(self, et, primary_pos):
        """수상함=해역 통제(자기 축 상 돌파선 점) / 그 외=자산 격침(기함).
        수상함이 돌파선 원 둘레에 다축 분산 포진 → 아군 전방위 방어 강요."""
        if not self._sea_maneuver or not et.is_ship:
            return primary_pos
        dx = et.pos.x - primary_pos.x
        dy = et.pos.y - primary_pos.y
        d  = math.hypot(dx, dy)
        if d < 1.0:                      # 돌파선 안쪽 깊숙 — 0division 가드
            return primary_pos
        s = (self._sea_line_km * 1000.0) / d
        return LatLon.from_xy(primary_pos.x + dx * s, primary_pos.y + dy * s)

    # ── v15.08.04: 목표지향 적 AI ──────────────────────────────────────────────
    def _on_retreat_arrived(self, et):
        """전장 모드 후퇴 도착 처리 — 단발과 달리 max_reattacks 캡 없이 압박 유지.
        무장이 남으면 무한 재접근, 소진 시 재무장 복귀(이탈). 철수 발령 시 전원 이탈."""
        if self._enemy_withdrawing:
            et.alive = False
            if not self._mc_mode:
                self._log(f"[철수] {et.preset_name} 전장 이탈 완료")
            return
        if not (self._munition_limit and et.munition_remaining <= 0):
            et.reattack_count += 1
            et.is_retreating = False
            et.retreat_pos   = None
            if not self._mc_mode:
                self._log(f"[재접근] {et.preset_name} 압박 유지 — 재공격 개시 (무장 잔여 {et.munition_remaining})")
        else:
            et.alive = False
            # v15.09.02: 함재기는 모항 귀환 → 재무장 지연 후 재발진(풀 복귀). 모항 격침 시 회수 불가.
            owner = et.carrier_owner
            if owner is not None and self.enemy_threats[owner].alive:
                self._rearm_queue.append((self.t + self._rearm_delay, owner))
                if not self._mc_mode:
                    self._log(f"[무장 소진·귀함] {et.preset_name} 모항 복귀 — 재무장 후 재발진 예정")
            elif not self._mc_mode:
                self._log(f"[무장 소진·복귀] {et.preset_name} 재무장 위해 전장 이탈")

    def _apply_tactical_choice(self, choice: dict):
        """전장 모드 전술 선택 — 부모(무기 우선순위·살보)에 전장 전용 레버를 얹는다.
        부모 무수정: 기존 radar_off_until 상태를 BattleEngine에서만 능동 제어."""
        super()._apply_tactical_choice(choice)
        if not choice:
            return
        # v15.11.01: 능동 레이더 OFF — ARM 회피 위해 수동으로 레이더 차단.
        # 다음 결정까지(_tactical_interval) 유지. 'on'은 강제 점등 안 함(자동 ARM 회피 보존).
        if choice.get('radar') == 'off':
            p = self._primary()
            p.radar_off_until = max(p.radar_off_until, self.t + self._tactical_interval)
        # v15.11.02: 표적 우선순위 전술 — _target_sort_key 오버라이드가 읽는다.
        tp = choice.get('target_priority')
        if tp and tp != 'auto':
            self.cfg['_tactical_target_priority'] = tp
        else:
            self.cfg.pop('_tactical_target_priority', None)
        # v15.11.03: 함대 기동 자세 — _apply_ship_evasion 오버라이드가 읽는다.
        mv = choice.get('maneuver')
        if mv in ('passive', 'normal', 'aggressive'):
            self._evasion_posture = mv
        # v15.11.04: CAP 전개 자세 — _cap_posture_factors 오버라이드가 읽는다.
        cp = choice.get('cap_posture')
        if cp in ('forward', 'normal', 'defensive'):
            self._cap_posture = cp
        # v15.11.05: ECM 자세 — 부모 ECM 로직(_check_hits)이 cfg를 읽음(코드 무수정).
        # off=미사용 / normal=현행(Pk -30%) / strong=강(Pk -45%). 탄도·HGV·ARM엔 무효.
        ecm = choice.get('ecm')
        if ecm == 'off':
            self.cfg['enable_ecm'] = False
        elif ecm == 'strong':
            self.cfg['enable_ecm'] = True
            self.cfg['ecm_scale'] = 1.5
        elif ecm == 'normal':
            self.cfg['enable_ecm'] = True
            self.cfg['ecm_scale'] = 1.0
        # Phase 5.6 self-play: 적 전술 모드 주입 — RL 적 정책(또는 고정 상대)이 고른 모드를
        # _adaptive_mode에 세팅, _enemy_fire가 살보·표적집중에 반영. None이면 무영향(회귀 가드).
        em = choice.get('enemy_mode')
        if self._enemy_rl and em in ('saturation', 'dispersal', 'deception'):
            self._adaptive_mode = em

    def _adaptive_tactic_update(self):
        """Phase 5.6 self-play: 적 RL 모드면 규칙 재평가를 건너뛴다(전술 모드는 RL 정책이
        _apply_tactical_choice로 세팅). 그 외(adaptive)는 부모 규칙기반 전환 그대로."""
        if self._enemy_rl:
            return
        super()._adaptive_tactic_update()

    def _make_tactical_state(self) -> 'TacticalState':
        """전장 모드 RL 보조 관측 — 부모 스냅샷에 위협 구성·자원 잔여를 extra로 추가.
        정책이 레버를 상황별로 쓰게 한다(ECM↔탄도비율, 기동↔연료, 살보↔탄약)."""
        st = super()._make_tactical_state()
        en_msl = [m for m in self.missiles if m.alive and m.mtype == 'enemy_strike']
        n_msl = len(en_msl)
        leakers = sum(1 for m in en_msl if m.is_ballistic or m.is_hgv)  # ECM 무효
        n_thr = sum(1 for et in self.enemy_threats if et.alive)
        n_air = sum(1 for et in self.enemy_threats if et.alive and et.is_aircraft)
        st.extra = {
            'leaker_frac':   (leakers / n_msl) if n_msl else 0.0,
            'asm_inflight':  (n_msl - leakers) / 20.0,        # ECM 유효 미사일 수(정규화)
            'aircraft_frac': (n_air / n_thr) if n_thr else 0.0,
            'ammo_frac':     getattr(self, '_ammo_frac', 1.0),
            'fuel_frac':     getattr(self, '_fuel_frac', 1.0),
        }
        return st

    def _apply_ship_evasion(self):
        """전장 모드 함대 기동 자세 — `_evasion_posture`에 따라 회피 적극도 조절.
        passive=회피 안 함(연료 절약·생존 risk) / normal=현행 / aggressive=조기·대폭 회피
        (생존↑·연료↑). 자세는 _apply_tactical_choice가 RL 행동으로 세팅."""
        posture = getattr(self, '_evasion_posture', 'normal')
        if posture == 'passive':
            return
        if posture == 'aggressive':
            # 조기 회피(22km부터). 속도는 올리지 않는다 — 전속이 이미 물리 상한이다
            # (v20.5 이전엔 점프 폭을 키워 '더 적극적'을 표현했으나, 그건 함속을 넘는
            #  순간이동이었다. 적극도는 이제 '언제부터 회피하는가'로만 표현한다.)
            super()._apply_ship_evasion(evade_r_base=22_000)
        else:
            super()._apply_ship_evasion()

    def _cap_posture_factors(self):
        """전장 모드 CAP 전개 자세 — `_cap_posture`에 따라 패트롤 반경·쿨다운 조절.
        forward=공세(멀리·자주 차단, AAM 조기소진) / normal=현행 / defensive=AAM 절약(가까이·드물게)."""
        posture = getattr(self, '_cap_posture', 'normal')
        if posture == 'forward':
            return (1.4, 40.0)
        if posture == 'defensive':
            return (0.6, 90.0)
        return (1.0, 60.0)

    def _target_sort_key(self, obj, primary_pos):
        """전장 모드 표적 우선순위 전술 — cfg `_tactical_target_priority`에 따라 정렬 키 변경.
        모두 sorted(reverse=True) 내림차순 기준(클수록 먼저 교전). auto면 부모 임박도."""
        pri = self.cfg.get('_tactical_target_priority', 'auto')
        if pri == 'auto':
            return super()._target_sort_key(obj, primary_pos)
        dist = max(obj.pos.dist_to(primary_pos), 200.0)
        if pri == 'nearest':                       # 최근접 우선
            return 1.0 / dist
        if pri == 'fastest':                       # 고속 위협 우선
            return getattr(obj, 'speed_ms', 300.0)
        if pri == 'leakers':                       # 탄도·HGV(누출 위험 큰 것) 우선
            lead = 1e6 if (getattr(obj, 'is_ballistic', False)
                           or getattr(obj, 'is_hgv', False)) else 0.0
            return lead + getattr(obj, 'speed_ms', 300.0) / dist
        return super()._target_sort_key(obj, primary_pos)

    def _enemy_combat_loss_value(self) -> float:
        """전투로 격침된 적 플랫폼 전력가치(USD) 합. 자발 이탈·재무장 복귀(intercepted=False)는
        제외 — 손실로 오인하면 격침 0인데도 철수가 발동된다. (격침 표식 = intercepted)"""
        return sum(ENEMY_PROCUREMENT_USD.get(et.name, 0)
                   for et in self.enemy_threats
                   if (not et.alive) and et.intercepted
                   and (et.is_ship or et.is_sub or et.is_aircraft))

    def _process_rearm(self):
        """v15.09.02: 재무장 만기 함재기를 모항 풀에 복귀(wing_launched 감소 → 스폰 루프가 재발진).
        모항 격침 시 회수 불가(airframe 손실). 손실(격침) 기체는 큐에 없으니 영구 미복귀."""
        if not self._rearm_queue:
            return
        ready = [c for (rt, c) in self._rearm_queue if rt <= self.t]
        self._rearm_queue = [(rt, c) for (rt, c) in self._rearm_queue if rt > self.t]
        for idx in ready:
            car = self.enemy_threats[idx]
            if car.alive and car.wing_launched > 0:
                car.wing_launched -= 1   # 1기 재무장 완료 — 풀 복귀(다음 웨이브 주기에 재발진)

    def _ras_update(self):
        """v15.09.03: 해상급유(RAS) — 군수지원함이 연료 부족 전투함에 급유(화물 연료서 차감).
        위협 중(적 대함미사일 비행 중)엔 분리 기동으로 중단, 소강 시 재개. 지원함 격침 시 중단."""
        # 함대로 향하는 적 대함미사일이 떠 있으면 RAS 중단(나란히 붙어 급유 불가)
        if any(m.alive and m.mtype == 'enemy_strike' for m in self.missiles):
            return
        suppliers = [s for s in self.friendly_ships if s.alive and s.ras_reserve > 0.0]
        if not suppliers:
            return
        # 급유 대상: 비원자력 함정 중 연료 부족(< target) 순 — 가장 부족한 것 우선.
        # 지원함도 포함(화물 연료로 자함 보급 — 오일러가 병목 되지 않게). 전투함이 더 닳아 우선.
        recips = sorted(
            (s for s in self.friendly_ships
             if s.alive and not s._nuclear and s.fuel < _RAS_TARGET),
            key=lambda s: s.fuel)
        if not recips:
            return
        for sup in suppliers:
            for s in recips[:_RAS_STATIONS]:        # 양현 스테이션 — 동시 급유 함정 수
                give = min(_RAS_RATE_PER_S * DT, _RAS_TARGET - s.fuel, sup.ras_reserve)
                if give <= 0.0:
                    continue
                s.fuel = min(s.fuel_max, s.fuel + give)
                sup.ras_reserve -= give
        self._ras_rearm_update()

    def _ras_rearm_update(self):
        """v17.1: RAS 탄약 재보급 — 소강기에 소진된 주요 SAM을 재장전(화물 유한·느림).
        위협 접근·지원함 부재 시 _ras_update 진입부에서 이미 return되어 여기 도달 안 함."""
        if not self._ras_rearm:
            return
        ammo_sups = [s for s in self.friendly_ships if s.alive and s.ras_ammo_reserve >= 1.0]
        if not ammo_sups:
            return

        def _sam_frac(s):
            init = sum(s._initial_defense_stock.get(w, 0) for w in _RAS_MAJOR_SAM)
            cur  = sum(s.inventory.get(w, 0) for w in _RAS_MAJOR_SAM)
            return (cur / init) if init > 0 else 1.0

        # 히스테리시스: 30% 미만서 재보급 시작, 80% 도달 시 종료(30~80 재하락은 대기)
        for s in self.friendly_ships:
            if not s.alive or s.is_submarine:
                continue
            f = _sam_frac(s)
            if f < _RAS_REARM_TRIGGER:
                s._rearming = True
            elif f >= _RAS_REARM_TARGET:
                s._rearming = False
        needy = sorted((s for s in self.friendly_ships if s.alive and s._rearming),
                       key=_sam_frac)
        if not needy:
            return
        for sup in ammo_sups:
            for s in needy[:_RAS_STATIONS]:
                if sup.ras_ammo_reserve < 1.0:
                    break
                s._rearm_acc += _RAS_REARM_PER_S * DT   # float 누적 → 1발 단위 지급
                if s._rearm_acc < 1.0:
                    continue
                # 가장 소진된(초기 대비 잔여율 낮은) 주요 SAM부터 초기 80%까지 1발 재장전
                target_wpn, worst = None, 1.0
                for w in _RAS_MAJOR_SAM:
                    init_w = s._initial_defense_stock.get(w, 0)
                    if init_w <= 0:
                        continue
                    if s.inventory.get(w, 0) >= int(init_w * _RAS_REARM_TARGET):
                        continue
                    frac_w = s.inventory.get(w, 0) / init_w
                    if frac_w < worst:
                        worst, target_wpn = frac_w, w
                if target_wpn is None:
                    # 더 채울 SAM 없음(전 무기 상한 도달) — 재보급 완료로 종료.
                    # _rearming을 끄지 않으면 상한이 int() 내림으로 TARGET에 못 미쳐도
                    # 매 틱 needy에 남아 누적·리셋을 무한 반복하므로 여기서 정상 종료.
                    s._rearm_acc = 0.0
                    s._rearming  = False
                    continue
                s.inventory[target_wpn] = s.inventory.get(target_wpn, 0) + 1
                sup.ras_ammo_reserve -= 1.0
                s._rearm_acc -= 1.0
                self.stats['ras_missiles_resupplied'] += 1

    def _enemy_withdraw_check(self):
        """적 전투 손실이 임계 초과 시 생존 세력 전면 철수(임무 불가 — 자원 보존).
        교리상 전투력 50% 상실 ≈ 임무 수행 불가. 격침된 플랫폼 전력가치 기준(이탈·재무장 제외)."""
        if self._enemy_withdrawing or not self._attr_active:
            return
        en_loss = self._enemy_combat_loss_value() / self._en_value_init
        if en_loss < self._enemy_withdraw_loss:
            return
        self._enemy_withdrawing = True
        primary_pos = self._primary().pos
        for et in self.enemy_threats:
            if not et.alive or et.is_retreating:
                continue
            ang = et.pos.bearing_to(primary_pos) + math.pi   # 함대 반대 방향 이탈
            et.retreat_pos = LatLon.from_xy(
                et.pos.x + math.cos(ang) * 300_000,
                et.pos.y + math.sin(ang) * 300_000)
            et.is_retreating = True
        if not self._mc_mode:
            self._log(f"[적 전면 철수] 전투 손실 {en_loss*100:.0f}% — 생존 세력 임무 포기·이탈")

    def _attrition_update(self):
        """소모전 — 양측 잔존 전력가치 비율 갱신(현재값/초기값, 0..1)."""
        self._fr_frac = min(1.0, self._friendly_force_value() / self._fr_value_init)
        self._en_frac = min(1.0, self._enemy_force_value()    / self._en_value_init)
        if not self._mc_mode and self.t != self._attr_tl_last_t:
            self._attr_tl_last_t = self.t
            self._force_tl.append((round(self.t, 1),
                                   round(self._fr_frac, 3), round(self._en_frac, 3)))

    # ── 자원 지속성 (탄약 잔여비 = 방어+공격 인벤토리 합 / 초기 합) ──────────────
    def _friendly_ammo_total(self) -> float:
        # CIWS(초기 9999)처럼 사실상 무제한인 탄은 제외 — 안 그러면 분모를 지배해
        # 제약 탄약(SAM·어뢰·대함)을 전소해도 잔여비가 ~0.98로 고정돼 지표가 무의미해진다.
        # 격침 함정의 잔여 탄약은 가용 자원이 아니므로 생존 함정만 합산(_friendly_force_value와 일관).
        return sum(v for s in self.friendly_ships if s.alive
                   for inv in (s.inventory, s.strike_inventory)
                   for v in inv.values() if v < _BATTLE_INF_AMMO)

    def _fuel_update(self):
        """매 틱 연료 소모 — 순항 기본분 + 회피 기동 추가분(틱당 고정·누적 상한). 원자력·격침 제외."""
        # Phase 5.1: 순항 소모를 horizon에 연동. _fuel_burn_per_s는 1800s 앵커라 긴 전장
        # (7200s)에선 종료 전 0 포화 → 지속성 목표가 상수화됨. 1800/horizon으로 스케일해
        # 전장 길이와 무관하게 "종료 시 잔여 0.60(표준함)" 설계 의도를 유지(회귀 시 1.0).
        # 회피 추가분은 누적 상한(_FUEL_MANEUVER_CAP)이 있어 horizon에 이미 robust — 미스케일.
        cruise_scale = 1800.0 / self.horizon_s if self.horizon_s > 0 else 1.0
        for s in self.friendly_ships:
            if not s.alive or s._nuclear:
                continue
            burn = s._fuel_burn_per_s * DT * cruise_scale
            # 아군 함정은 평상시 고정 진형 — pos가 1m 넘게 변하면 이번 틱 회피 기동 발생.
            # 회피 점프 거리는 추상값이라 거리 비례 대신 틱당 고정량으로 과금(누적 상한 적용).
            if s._prev_fuel_pos is not None and s.pos.dist_to(s._prev_fuel_pos) > 1.0:
                add = min(_FUEL_MANEUVER_TICK, _FUEL_MANEUVER_CAP - s._maneuver_fuel_used)
                if add > 0.0:
                    s._maneuver_fuel_used += add
                    burn += add
            s.fuel = max(0.0, s.fuel - burn)
            s._prev_fuel_pos = s.pos.copy()

    def _sustainment_update(self):
        """자원 지속성 — 탄약·연료 잔여비 갱신. 둘 중 낮은 쪽이 progress."""
        self._ammo_frac = min(1.0, self._friendly_ammo_total() / self._ammo_init)
        # 연료: 생존·비원자력 함정 중 최저 잔여비(원자력·격침 함정 제외). 없으면 1.0.
        fuels = [s.fuel / s.fuel_max for s in self.friendly_ships
                 if s.alive and not s._nuclear and s.fuel_max > 0]
        self._fuel_frac = min(fuels) if fuels else 1.0
        if not self._mc_mode and self.t != self._sust_tl_last_t:
            self._sust_tl_last_t = self.t
            self._resource_tl.append((round(self.t, 1),
                                      round(self._ammo_frac, 3), round(self._fuel_frac, 3)))

    def _update_objectives(self):
        if self._asset_alive() and self._asset_max_hp:
            frac = max(0.0, self._asset.hp / self._asset_max_hp)
        elif self._asset_alive():
            frac = 1.0
        else:
            frac = 0.0
        self._sea_control_update()
        line_km   = self._sea_line_km
        breached  = self._sea_max_pen >= line_km                 # 자산 도달
        held      = self._sea_max_pen <= 0.0                     # 완전 저지
        sea_prog_f = 1.0 - (self._sea_max_pen / line_km if line_km else 0.0)
        # v15.07.02: 소모전 교환비 — 아군이 적을 더 깎을수록 progress↑ (로지스틱)
        if self._attr_active:
            self._attrition_update()
            x = (1.0 - self._en_frac) - (1.0 - self._fr_frac)    # 적손실율 - 아군손실율
            # exp 인자 clamp: attr_k 극단값(경계 fuzzing)에서 math.exp 오버플로 방지.
            # 정상값(attr_k≈5, x∈[-1,1])에선 무영향 — z∈[-5,5]가 [-700,700] 안.
            _z = max(-700.0, min(700.0, self._attr_k * x))
            attr_prog_f = 1.0 / (1.0 + math.exp(-_z))
            en_wiped = self._en_frac <= 0.0
            fr_wiped = self._fr_frac <= 0.0
        # v15.07.03: 자원 지속성 — 탄약·연료 최저 잔여비
        self._sustainment_update()
        sust_prog_f = min(self._ammo_frac, self._fuel_frac)
        for ob in self.objectives:
            if ob.type == 'defend_asset':
                ob.progress = frac
                ob.status   = '실패' if frac <= 0.0 else '진행'
            elif ob.type == 'destroy_asset':
                ob.progress = 1.0 - frac
                ob.status   = '달성' if frac <= 0.0 else '진행'
            elif ob.type == 'sea_control' and ob.side == 'friendly':
                ob.progress = sea_prog_f
                ob.status   = '실패' if breached else ('달성' if held else '진행')
            elif ob.type == 'sea_control' and ob.side == 'enemy':
                ob.progress = 1.0 - sea_prog_f
                ob.status   = '달성' if breached else '진행'
            elif ob.type == 'attrition' and ob.side == 'friendly':
                if not self._attr_active:
                    ob.progress = 0.5; ob.status = '해당없음'
                else:
                    ob.progress = attr_prog_f
                    ob.status   = '달성' if en_wiped else ('실패' if fr_wiped else '진행')
            elif ob.type == 'attrition' and ob.side == 'enemy':
                if not self._attr_active:
                    ob.progress = 0.5; ob.status = '해당없음'
                else:
                    ob.progress = 1.0 - attr_prog_f
                    ob.status   = '달성' if fr_wiped else '진행'
            elif ob.type == 'sustainment' and ob.side == 'friendly':
                ob.progress = sust_prog_f
                ob.status   = '실패' if sust_prog_f <= 0.0 else '진행'

    # ── 종료조건 (목표 기반 — 부모의 위협소진 종료를 대체) ─────────────────────
    def _is_over(self) -> bool:
        self._fuel_update()        # 매 틱 1회 연료 소모 (루프에서 _is_over만 호출 — 이중차감 없음)
        self._ras_update()             # v15.09.03: 소강 시 해상급유(연료 소모 후, 지속성 평가 전)
        self._update_objectives()
        self._process_rearm()          # v15.09.02: 재무장 만기 함재기 모항 풀 복귀
        self._enemy_withdraw_check()   # v15.08.04: 손실 임계 초과 시 적 전면 철수
        if not self._asset_alive():
            if not self._mc_mode:
                self._log(f"[종료] 방어 자산 {self._asset.name} 격침 — 적 승")
            return True
        if all(not s.alive for s in self.friendly_ships):
            if not self._mc_mode:
                self._log("[종료] 아군 전멸 — 적 승")
            return True
        if self.t >= self.horizon_s:
            if not self._mc_mode:
                self._log(f"[종료] 작전 시간({int(self.horizon_s)}s) 종료 — 목표 기반 판정")
            return True
        # 적 위협 소멸 판정 → 아군 승. 무장 소진(복귀 중)·전면 철수 발령 위협은 제외.
        # (v15.08.04: 압박 유지로 재접근하는 위협은 무장 남는 한 활성 — max_reattacks 캡 폐지)
        enemy_active = [et for et in self.enemy_threats
                        if et.alive
                        and not (self._munition_limit and et.munition_remaining <= 0)
                        and not self._enemy_withdrawing]
        active_missiles = [m for m in self.missiles
                           if m.alive and m.mtype == 'enemy_strike']
        if not enemy_active and not active_missiles:
            if not self._mc_mode:
                self._log("[종료] 적 위협 격멸 — 자산 방어 성공")
            return True
        return False

    # ── 승패 점수 (요격률 대체 핵심 지표) ──────────────────────────────────────
    def _score_outcome(self):
        self._update_objectives()
        fw = sum(ob.weight for ob in self.objectives if ob.side == 'friendly') or 1.0
        ew = sum(ob.weight for ob in self.objectives if ob.side == 'enemy') or 1.0
        f  = sum(ob.weight * ob.progress for ob in self.objectives if ob.side == 'friendly') / fw
        e  = sum(ob.weight * ob.progress for ob in self.objectives if ob.side == 'enemy') / ew
        # 점수 [0,1] clamp: weight 극단값(음수 등, 경계 fuzzing)에서 점수 범위 이탈 방지.
        # 정상값(양수 weight, progress∈[0,1])에선 이미 [0,1]이라 무영향.
        f = min(1.0, max(0.0, f)); e = min(1.0, max(0.0, e))
        margin = float(self.cfg.get('battle_draw_margin', 0.1))
        if   f - e >  margin: outcome = 'win'
        elif e - f >  margin: outcome = 'loss'
        else:                 outcome = 'draw'
        return outcome, f, e

    # ── 출력 (요격률 → outcome 중심으로 확장) ──────────────────────────────────
    def _compile(self) -> dict:
        result = super()._compile()       # 물리 통계·intercept_rate(보조 지표) 유지
        outcome, fscore, escore = self._score_outcome()
        result['outcome']          = outcome
        result['friendly_score']   = round(fscore, 3)
        result['enemy_score']      = round(escore, 3)
        result['objectives']       = [ob.as_dict() for ob in self.objectives]
        result['battle_horizon_s'] = self.horizon_s
        if not self._mc_mode:      # timeline은 단일 시뮬 전용 (frames와 동일 패턴)
            result['timeline'] = {'frontline_km': self._frontline_tl,
                                  'force_ratio':  self._force_tl,
                                  'resource_min': self._resource_tl}
        return result


def run_battle_simulation(cfg: dict, step_cb=None, tactical_cb=None) -> dict:
    """지속 전장 1회 실행. run_v7_simulation의 전장 버전 — 부모 함수는 무수정."""
    if not cfg.get('detect_km_manual', False):
        ranges = calculate_fleet_detect_ranges(
            cfg.get('fleet_preset', '단독 작전'),
            cfg.get('weather', '맑음 (주간)'))
        cfg = dict(cfg)
        cfg['detect_km']         = ranges['대공']
        cfg['surface_detect_km'] = ranges['대함']
        cfg['sub_detect_km']     = ranges['대잠']
    else:
        cfg = dict(cfg)
    _initial_weather = cfg.get('weather', '맑음 (주간)')
    sim = BattleEngine(cfg, step_cb=None)
    if step_cb:   # 진행바 총량을 MAX_SIM_TIME(3600)이 아닌 전장 시간 지평으로 보고
        _horizon = sim.horizon_s
        sim._step_cb = lambda t, _tmax, alive, vls, last: step_cb(t, _horizon, alive, vls, last)
    if tactical_cb:
        sim._tactical_pause_cb = tactical_cb
    result = sim.run()
    cfg['weather'] = _initial_weather
    return result


def _mc_run_one(cfg: dict) -> dict:
    """MC 1회 — 전장 모드면 지속 전장 엔진, 아니면 단발 교전."""
    if cfg.get('enable_battle_mode', False):
        return run_battle_simulation(cfg)
    return run_v7_simulation(cfg)


def _battle_agg(outcomes: list, fscores: list, n: int) -> dict:
    """전장 모드 승률 집계 — outcomes 비면 빈 dict 반환(단발 모드 무영향)."""
    if not outcomes:
        return {}
    _n = max(n, len(outcomes), 1)
    return {
        'win_rate':            outcomes.count('win')  / _n,
        'loss_rate':           outcomes.count('loss') / _n,
        'draw_rate':           outcomes.count('draw') / _n,
        'mean_friendly_score': float(np.mean(fscores)) if fscores else 0.0,
    }


# ════════════════════════════════════════════════════════════════════════════
#  몬테카를로 분석
# ════════════════════════════════════════════════════════════════════════════

def monte_carlo_v7(cfg: dict, n: int = 200, desc: str = '',
                   progress_cb=None) -> dict:
    """
    run_v7_simulation을 n회 반복해 통계를 집계한다.

    반환 dict 키:
      intercept_rates   : list[float]   — 회차별 요격률
      friendly_hits     : list[int]
      enemy_destroyed   : list[int]
      friendly_lost     : list[int]
      total_costs       : list[float]
      weapon_usage      : dict[str, list[int]]  — 무기별 회차별 소모량
      ship_hits         : dict[str, list[int]]  — 함정별 회차별 피격 횟수
      mean_intercept    : float
      std_intercept     : float
      full_pass_rate    : float          — 요격률 1.0 비율
    """
    cfg = dict(cfg); cfg['mc_mode'] = True
    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}   # {무기명: [회차별 소모량]}
    ship_hits_mc: dict = {}   # {함정명: [회차별 피격]}
    weapon_zero:  dict = {}   # {무기명: 소진(잔여=0) 횟수}
    # v9.4: VLS 고갈 통계
    vls_dep_count: int = 0
    vls_dep_times: list = []
    # v12.4·v12.6: 침수·IFF 통계
    flood_sunk: list = []; flood_on: list = []; iff_fail: list = []; iff_frat: list = []
    mine_struck: list = []; mine_lost: list = []
    recon_loss:  list = []
    unmanned_lost: list = []
    ras_resupplied: list = []
    laser_kills: list = []
    lsam_fired: list = []; chungung_fired: list = []; patriot_fired: list = []
    outcomes: list = []; fscores: list = []   # 전장 모드 승률 집계

    step = max(1, n // 5)
    if desc:
        print(f'  [{desc}] {n}회 MC 시작... ', end='', flush=True)

    base_seed = cfg.get('sim_seed', None)
    for i in range(n):
        # 회차마다 다른 시드 (기반 시드 + 회차번호)
        run_cfg = dict(cfg)
        if base_seed:
            run_cfg['sim_seed'] = int(base_seed) + i
        r = _mc_run_one(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])
        flood_sunk.append(r.get('ships_sunk_by_flood', 0))
        flood_on.append(r.get('ships_flooding', 0))
        iff_fail.append(r.get('iff_failures', 0))
        iff_frat.append(r.get('iff_fratricide', 0))
        mine_struck.append(r.get('mines_struck', 0))
        mine_lost.append(r.get('ships_lost_to_mine', 0))
        recon_loss.append(r.get('recon_losses', 0))
        unmanned_lost.append(r.get('unmanned_lost', 0))
        ras_resupplied.append(r.get('ras_missiles_resupplied', 0))
        laser_kills.append(r.get('laser_kills', 0))
        lsam_fired.append(r.get('lsam_fired', 0))
        chungung_fired.append(r.get('chungung_fired', 0))
        patriot_fired.append(r.get('patriot_fired', 0))
        _oc = r.get('outcome')
        if _oc:
            outcomes.append(_oc); fscores.append(r.get('friendly_score', 0.0))

        # 무기별 소모량 (초기 재고 - 잔여 재고) + 소진 횟수 집계
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            if wpn not in weapon_usage:
                weapon_usage[wpn] = []
            weapon_usage[wpn].append(remaining)
            if remaining == 0:
                weapon_zero[wpn] = weapon_zero.get(wpn, 0) + 1

        # 함정별 피격 횟수
        for ship in r.get('friendly_ships', []):
            sname = ship.name
            hits = getattr(ship, 'hits_taken', 0)
            if sname not in ship_hits_mc:
                ship_hits_mc[sname] = []
            ship_hits_mc[sname].append(hits)

        # v9.4: VLS 고갈 시각 수집
        dep_t = r.get('vls_depletion_t', {})
        if dep_t:
            vls_dep_count += 1
            vls_dep_times.extend(dep_t.values())

        if desc and (i + 1) % step == 0:
            print(f'{(i + 1) * 100 // n}%', end=' ', flush=True)
        if progress_cb:
            progress_cb(i + 1, n)

    if desc:
        print('완료')

    arr = np.array(rates)
    # 무기별 평균 잔여 재고 (소모량 = 초기 - 평균 잔여)
    weapon_avg_remaining = {k: float(np.mean(v)) for k, v in weapon_usage.items()}
    ship_avg_hits = {k: float(np.mean(v)) for k, v in ship_hits_mc.items()}
    dest_arr = np.array(e_dest, dtype=float)
    return {
        'intercept_rates':         rates,
        'friendly_hits':           f_hits,
        'enemy_destroyed':         e_dest,
        'friendly_lost':           f_lost,
        'total_costs':             costs,
        'weapon_avg_remaining':    weapon_avg_remaining,
        'weapon_exhaustion_rates': {k: v / n for k, v in weapon_zero.items()},
        'ship_avg_hits':           ship_avg_hits,
        'mean_intercept':          float(arr.mean()),
        'std_intercept':           float(arr.std()),
        'full_pass_rate':          float((arr == 1.0).mean()),
        'n':                       n,
        # v9.3: 공격 임무 격침 통계
        'mean_enemy_destroyed':    float(dest_arr.mean()),
        'max_enemy_destroyed':     int(dest_arr.max()) if len(dest_arr) else 0,
        # v9.4: VLS 고갈 통계
        'vls_depletion_rate':      vls_dep_count / n,
        'vls_depletion_t_mean':    float(np.mean(vls_dep_times)) if vls_dep_times else None,
        # v12.4·v12.6: 침수·IFF 회당 평균
        'mean_ships_sunk_by_flood': float(np.mean(flood_sunk)),
        'mean_ships_flooding':      float(np.mean(flood_on)),
        'mean_iff_failures':        float(np.mean(iff_fail)),
        'mean_iff_fratricide':      float(np.mean(iff_frat)),
        'mean_mines_struck':        float(np.mean(mine_struck)),
        'mean_ships_lost_to_mine':  float(np.mean(mine_lost)),
        'mean_recon_losses':        float(np.mean(recon_loss)) if recon_loss else 0.0,
        'mean_unmanned_lost':       float(np.mean(unmanned_lost)) if unmanned_lost else 0.0,
        'mean_ras_resupplied':      float(np.mean(ras_resupplied)) if ras_resupplied else 0.0,
        'mean_laser_kills':         float(np.mean(laser_kills)) if laser_kills else 0.0,
        'mean_lsam_fired':          float(np.mean(lsam_fired)) if lsam_fired else 0.0,
        'mean_chungung_fired':      float(np.mean(chungung_fired)) if chungung_fired else 0.0,
        'mean_patriot_fired':       float(np.mean(patriot_fired)) if patriot_fired else 0.0,
        **_battle_agg(outcomes, fscores, n),
    }


def _mc_batch_worker(args: tuple) -> tuple:
    """ProcessPoolExecutor 배치 워커 — PyQt6 의존성 없는 순수 엔진 함수."""
    cfg, n, seed_offset = args
    cfg = dict(cfg); cfg['mc_mode'] = True
    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}
    weapon_zero:  dict = {}
    ship_hits_mc: dict = {}
    phase_times_acc: dict = {}   # v8.26: 배치 내 단계별 시간 누적
    flood_sunk: list = []; flood_on: list = []; iff_fail: list = []; iff_frat: list = []
    mine_struck: list = []; mine_lost: list = []
    recon_loss:  list = []
    unmanned_lost: list = []
    ras_resupplied: list = []
    laser_kills: list = []
    lsam_fired: list = []; chungung_fired: list = []; patriot_fired: list = []
    outcomes: list = []; fscores: list = []
    base_seed = cfg.get('sim_seed', None)
    for i in range(n):
        run_cfg = dict(cfg)
        if base_seed:
            run_cfg['sim_seed'] = int(base_seed) + seed_offset + i
        r = _mc_run_one(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])
        flood_sunk.append(r.get('ships_sunk_by_flood', 0))
        flood_on.append(r.get('ships_flooding', 0))
        iff_fail.append(r.get('iff_failures', 0))
        iff_frat.append(r.get('iff_fratricide', 0))
        mine_struck.append(r.get('mines_struck', 0))
        mine_lost.append(r.get('ships_lost_to_mine', 0))
        recon_loss.append(r.get('recon_losses', 0))
        unmanned_lost.append(r.get('unmanned_lost', 0))
        ras_resupplied.append(r.get('ras_missiles_resupplied', 0))
        laser_kills.append(r.get('laser_kills', 0))
        lsam_fired.append(r.get('lsam_fired', 0))
        chungung_fired.append(r.get('chungung_fired', 0))
        patriot_fired.append(r.get('patriot_fired', 0))
        _oc = r.get('outcome')
        if _oc:
            outcomes.append(_oc); fscores.append(r.get('friendly_score', 0.0))
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            weapon_usage.setdefault(wpn, []).append(remaining)
            if remaining == 0:
                weapon_zero[wpn] = weapon_zero.get(wpn, 0) + 1
        for ship in r.get('friendly_ships', []):
            ship_hits_mc.setdefault(ship.name, []).append(
                getattr(ship, 'hits_taken', 0))
        for k, v in r.get('phase_times', {}).items():
            phase_times_acc[k] = phase_times_acc.get(k, 0.0) + v
    # 배치 평균으로 변환
    if n > 0:
        phase_times_avg = {k: v / n for k, v in phase_times_acc.items()}
    else:
        phase_times_avg = {}
    extra_stats = {'ships_sunk_by_flood': flood_sunk, 'ships_flooding': flood_on,
                   'iff_failures': iff_fail, 'iff_fratricide': iff_frat,
                   'mines_struck': mine_struck, 'ships_lost_to_mine': mine_lost,
                   'recon_losses': recon_loss,
                   'unmanned_lost': unmanned_lost,
                   'ras_missiles_resupplied': ras_resupplied,
                   'laser_kills': laser_kills,
                   'lsam_fired': lsam_fired, 'chungung_fired': chungung_fired,
                   'patriot_fired': patriot_fired,
                   'outcome': outcomes, 'friendly_score': fscores}
    return rates, f_hits, e_dest, f_lost, costs, weapon_usage, ship_hits_mc, weapon_zero, phase_times_avg, extra_stats


def _mc_lhs_batch_worker(args: tuple) -> tuple:
    """LHS MC 배치 워커 — 미리 계산된 샘플 슬라이스와 파라미터 정의를 받아 실행."""
    cfg_base, samples, param_defs = args
    cfg_base = dict(cfg_base); cfg_base['mc_mode'] = True
    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}
    ship_hits_mc: dict = {}
    flood_sunk: list = []; flood_on: list = []; iff_fail: list = []; iff_frat: list = []
    mine_struck: list = []; mine_lost: list = []
    recon_loss:  list = []
    unmanned_lost: list = []
    ras_resupplied: list = []
    laser_kills: list = []
    lsam_fired: list = []; chungung_fired: list = []; patriot_fired: list = []
    outcomes: list = []; fscores: list = []
    for sample in samples:
        run_cfg = dict(cfg_base)
        for j, (key, lo, hi, _) in enumerate(param_defs):
            run_cfg[key] = float(lo + sample[j] * (hi - lo))
        r = _mc_run_one(run_cfg)
        rates.append(r['intercept_rate'])
        f_hits.append(r['friendly_hits'])
        e_dest.append(r['enemy_ships_destroyed'])
        f_lost.append(r['friendly_ships_lost'])
        costs.append(r['total_cost'])
        flood_sunk.append(r.get('ships_sunk_by_flood', 0))
        flood_on.append(r.get('ships_flooding', 0))
        iff_fail.append(r.get('iff_failures', 0))
        iff_frat.append(r.get('iff_fratricide', 0))
        mine_struck.append(r.get('mines_struck', 0))
        mine_lost.append(r.get('ships_lost_to_mine', 0))
        recon_loss.append(r.get('recon_losses', 0))
        unmanned_lost.append(r.get('unmanned_lost', 0))
        ras_resupplied.append(r.get('ras_missiles_resupplied', 0))
        laser_kills.append(r.get('laser_kills', 0))
        lsam_fired.append(r.get('lsam_fired', 0))
        chungung_fired.append(r.get('chungung_fired', 0))
        patriot_fired.append(r.get('patriot_fired', 0))
        _oc = r.get('outcome')
        if _oc:
            outcomes.append(_oc); fscores.append(r.get('friendly_score', 0.0))
        for wpn, remaining in r.get('remaining_inventory', {}).items():
            weapon_usage.setdefault(wpn, []).append(remaining)
        for ship in r.get('friendly_ships', []):
            ship_hits_mc.setdefault(ship.name, []).append(getattr(ship, 'hits_taken', 0))
    extra_stats = {'ships_sunk_by_flood': flood_sunk, 'ships_flooding': flood_on,
                   'iff_failures': iff_fail, 'iff_fratricide': iff_frat,
                   'mines_struck': mine_struck, 'ships_lost_to_mine': mine_lost,
                   'recon_losses': recon_loss,
                   'unmanned_lost': unmanned_lost,
                   'ras_missiles_resupplied': ras_resupplied,
                   'laser_kills': laser_kills,
                   'lsam_fired': lsam_fired, 'chungung_fired': chungung_fired,
                   'patriot_fired': patriot_fired,
                   'outcome': outcomes, 'friendly_score': fscores}
    return rates, f_hits, e_dest, f_lost, costs, weapon_usage, ship_hits_mc, extra_stats


# ════════════════════════════════════════════════════════════════════════════
#  분석 고도화: LHS / CVaR / Stress Test / Sobol 민감도
# ════════════════════════════════════════════════════════════════════════════

# LHS 샘플링 대상 불확실 파라미터: (cfg_key, 하한, 상한, 표시명)
_LHS_PARAM_DEFS = [
    ('pk_scale',         0.70, 1.30, 'SAM Pk 배율'),
    ('detect_scale',     0.70, 1.30, '탐지거리 배율'),
    ('cd_scale',         0.80, 1.50, 'C&D 시간 배율'),
    ('ecm_scale',        0.50, 1.50, 'ECM 효과 배율'),
    ('threat_spd_scale', 0.80, 1.30, '위협 속도 배율'),
    ('decoy_stock',      0.0,  4.0,  '기만기 재고'),
]

# 스트레스 테스트 2D 그리드 정의
STRESS_DIMS = {
    'channel_degrade': {
        'label':  '유도 채널 감소 (%)',
        'values': [0, 25, 50, 75],
    },
    'radar_degrade': {
        'label':  '레이더 성능 감소 (%)',
        'values': [0, 25, 50],
    },
}


def compute_cvar(rates: list, alpha: float = 0.05) -> float:
    """하위 alpha% 요격률의 평균 (Conditional Value at Risk — 최악 시나리오 평균)."""
    if not rates:
        return 0.0
    sorted_r = sorted(rates)
    n_tail   = max(1, int(len(sorted_r) * alpha))
    return float(np.mean(sorted_r[:n_tail]))


def monte_carlo_lhs(cfg: dict, n: int = 10_000,
                    progress_cb=None) -> dict:
    """
    Latin Hypercube Sampling 기반 MC (멀티프로세싱 + mc_mode 최적화).
    불확실 파라미터 6종을 LHS로 공간 균등 샘플링하여 순수 MC 대비 3~5× 빠른 수렴.

    반환: monte_carlo_v7 형식 dict + 'cvar', 'method' 키
    """
    from concurrent.futures import ProcessPoolExecutor, as_completed
    import os

    try:
        from scipy.stats.qmc import LatinHypercube
        d = len(_LHS_PARAM_DEFS)
        seed_val = cfg.get('sim_seed', None)
        sampler  = LatinHypercube(d=d, seed=int(seed_val) if seed_val else None)
        samples  = sampler.random(n=n)   # (n, d) in [0,1]
    except ImportError:
        samples = np.random.rand(n, len(_LHS_PARAM_DEFS))

    rates, f_hits, e_dest, f_lost, costs = [], [], [], [], []
    weapon_usage: dict = {}
    ship_hits_mc: dict = {}
    flood_sunk: list = []; flood_on: list = []; iff_fail: list = []; iff_frat: list = []
    mine_struck: list = []; mine_lost: list = []
    recon_loss:  list = []
    unmanned_lost: list = []
    ras_resupplied: list = []
    laser_kills: list = []
    lsam_fired: list = []; chungung_fired: list = []; patriot_fired: list = []
    outcomes: list = []; fscores: list = []

    n_workers = min(os.cpu_count() or 4, 8)
    batch_size = max(1, n // n_workers)
    batches = [samples[i:i + batch_size] for i in range(0, n, batch_size)]
    args_list = [(cfg, b.tolist(), _LHS_PARAM_DEFS) for b in batches]

    done = 0
    try:
        with ProcessPoolExecutor(max_workers=n_workers) as pool:
            futs = {pool.submit(_mc_lhs_batch_worker, a): len(a[1]) for a in args_list}
            for fut in as_completed(futs):
                br, bh, bd, bl, bc, bwu, bsh, bxs = fut.result()
                rates.extend(br); f_hits.extend(bh); e_dest.extend(bd)
                f_lost.extend(bl); costs.extend(bc)
                flood_sunk.extend(bxs['ships_sunk_by_flood'])
                flood_on.extend(bxs['ships_flooding'])
                iff_fail.extend(bxs['iff_failures'])
                iff_frat.extend(bxs['iff_fratricide'])
                mine_struck.extend(bxs['mines_struck'])
                mine_lost.extend(bxs['ships_lost_to_mine'])
                recon_loss.extend(bxs.get('recon_losses', []))
                unmanned_lost.extend(bxs.get('unmanned_lost', []))
                ras_resupplied.extend(bxs.get('ras_missiles_resupplied', []))
                laser_kills.extend(bxs.get('laser_kills', []))
                lsam_fired.extend(bxs.get('lsam_fired', []))
                chungung_fired.extend(bxs.get('chungung_fired', []))
                patriot_fired.extend(bxs.get('patriot_fired', []))
                outcomes.extend(bxs.get('outcome', []))
                fscores.extend(bxs.get('friendly_score', []))
                for k, v in bwu.items(): weapon_usage.setdefault(k, []).extend(v)
                for k, v in bsh.items(): ship_hits_mc.setdefault(k, []).extend(v)
                done += futs[fut]
                if progress_cb:
                    progress_cb(done, n)
    except Exception:
        # 멀티프로세싱 실패 시 직렬 폴백
        cfg2 = dict(cfg); cfg2['mc_mode'] = True
        for i, sample in enumerate(samples):
            run_cfg = dict(cfg2)
            for j, (key, lo, hi, _) in enumerate(_LHS_PARAM_DEFS):
                run_cfg[key] = float(lo + sample[j] * (hi - lo))
            r = _mc_run_one(run_cfg)
            rates.append(r['intercept_rate']); f_hits.append(r['friendly_hits'])
            e_dest.append(r['enemy_ships_destroyed']); f_lost.append(r['friendly_ships_lost'])
            costs.append(r['total_cost'])
            flood_sunk.append(r.get('ships_sunk_by_flood', 0))
            flood_on.append(r.get('ships_flooding', 0))
            iff_fail.append(r.get('iff_failures', 0))
            iff_frat.append(r.get('iff_fratricide', 0))
            mine_struck.append(r.get('mines_struck', 0))
            mine_lost.append(r.get('ships_lost_to_mine', 0))
            recon_loss.append(r.get('recon_losses', 0))
            unmanned_lost.append(r.get('unmanned_lost', 0))
            ras_resupplied.append(r.get('ras_missiles_resupplied', 0))
            laser_kills.append(r.get('laser_kills', 0))
            lsam_fired.append(r.get('lsam_fired', 0))
            chungung_fired.append(r.get('chungung_fired', 0))
            patriot_fired.append(r.get('patriot_fired', 0))
            _oc = r.get('outcome')
            if _oc:
                outcomes.append(_oc); fscores.append(r.get('friendly_score', 0.0))
            for wpn, rem in r.get('remaining_inventory', {}).items():
                weapon_usage.setdefault(wpn, []).append(rem)
            for ship in r.get('friendly_ships', []):
                ship_hits_mc.setdefault(ship.name, []).append(getattr(ship, 'hits_taken', 0))
            if progress_cb:
                progress_cb(i + 1, n)

    arr = np.array(rates)
    return {
        'intercept_rates':          rates,
        'friendly_hits':            f_hits,
        'enemy_destroyed':          e_dest,
        'friendly_lost':            f_lost,
        'total_costs':              costs,
        'weapon_avg_remaining':     {k: float(np.mean(v)) for k, v in weapon_usage.items()},
        # MC 3경로 정합: monte_carlo_v7·_mc_batch_worker가 채우는 소진율을 LHS도 제공(잔여≤0 비율)
        'weapon_exhaustion_rates':  {k: sum(1 for x in v if x <= 0) / len(v)
                                     for k, v in weapon_usage.items() if v},
        'ship_avg_hits':            {k: float(np.mean(v)) for k, v in ship_hits_mc.items()},
        'mean_intercept':           float(arr.mean()),
        'std_intercept':            float(arr.std()),
        'full_pass_rate':           float((arr == 1.0).mean()),
        'cvar':                     compute_cvar(rates),
        'n':                        n,
        'method':                   'LHS',
        'mean_enemy_destroyed':     float(np.mean(e_dest)) if e_dest else 0.0,
        'max_enemy_destroyed':      int(max(e_dest)) if e_dest else 0,
        'mean_ships_sunk_by_flood': float(np.mean(flood_sunk)) if flood_sunk else 0.0,
        'mean_ships_flooding':      float(np.mean(flood_on)) if flood_on else 0.0,
        'mean_iff_failures':        float(np.mean(iff_fail)) if iff_fail else 0.0,
        'mean_iff_fratricide':      float(np.mean(iff_frat)) if iff_frat else 0.0,
        'mean_mines_struck':        float(np.mean(mine_struck)) if mine_struck else 0.0,
        'mean_ships_lost_to_mine':  float(np.mean(mine_lost)) if mine_lost else 0.0,
        'mean_recon_losses':        float(np.mean(recon_loss)) if recon_loss else 0.0,
        'mean_unmanned_lost':       float(np.mean(unmanned_lost)) if unmanned_lost else 0.0,
        'mean_ras_resupplied':      float(np.mean(ras_resupplied)) if ras_resupplied else 0.0,
        'mean_laser_kills':         float(np.mean(laser_kills)) if laser_kills else 0.0,
        'mean_lsam_fired':          float(np.mean(lsam_fired)) if lsam_fired else 0.0,
        'mean_chungung_fired':      float(np.mean(chungung_fired)) if chungung_fired else 0.0,
        'mean_patriot_fired':       float(np.mean(patriot_fired)) if patriot_fired else 0.0,
        **_battle_agg(outcomes, fscores, n),
    }


def _sim_rate_worker(cfg: dict) -> float:
    """단일 시뮬을 돌려 요격률만 반환 — 프로세스 풀 병렬화용 top-level 워커(피클 가능)."""
    return run_v7_simulation(cfg)['intercept_rate']


def stress_test_grid(cfg: dict, n_per_cell: int = 500,
                     progress_cb=None, map_fn=map) -> dict:
    """
    2D 스트레스 테스트: 채널 감소(%) × 레이더 성능 감소(%) 그리드 요격률 매트릭스.

    n_per_cell: 셀당 시뮬 횟수 (빠름=300, 표준=500, 정밀=3000)
    map_fn: (fn, iterable) → results. 기본 직렬(map). app_main가 풀 기반 병렬 map 주입 시
            8코어 병렬 실행. 각 시뮬은 seed가 고정된 독립 작업이라 결과는 직렬과 동일.
    """
    ch_vals  = STRESS_DIMS['channel_degrade']['values']   # [0, 25, 50, 75]
    rad_vals = STRESS_DIMS['radar_degrade']['values']     # [0, 25, 50]
    grid       = np.zeros((len(ch_vals), len(rad_vals)))
    cvar_grid  = np.zeros_like(grid)
    total_cells = len(ch_vals) * len(rad_vals)
    base_seed   = cfg.get('sim_seed', None)

    # 전 셀 × 전 반복의 cfg를 평탄 리스트로 생성 (seed = base + 전역 인덱스, 직렬과 동일)
    cells: list = [(i, j) for i in range(len(ch_vals)) for j in range(len(rad_vals))]
    flat_cfgs: list = []
    for done, (i, j) in enumerate(cells):
        cell_cfg = dict(cfg)
        cell_cfg['detect_scale'] = 1.0 - rad_vals[j] / 100.0
        cell_cfg['pk_scale']     = max(0.1, 1.0 - ch_vals[i] / 200.0)
        for k in range(n_per_cell):
            run_cfg = dict(cell_cfg)
            if base_seed:
                run_cfg['sim_seed'] = int(base_seed) + done * n_per_cell + k
            flat_cfgs.append(run_cfg)

    # 병렬(또는 직렬) 실행 — map_fn은 순서를 보존하므로 셀별 슬라이스 재조립이 안전
    all_rates: list = []
    for idx, rate in enumerate(map_fn(_sim_rate_worker, flat_cfgs)):
        all_rates.append(rate)
        if progress_cb and (idx + 1) % n_per_cell == 0:
            progress_cb((idx + 1) // n_per_cell, total_cells)

    for done, (i, j) in enumerate(cells):
        cell_rates = all_rates[done * n_per_cell:(done + 1) * n_per_cell]
        grid[i, j]      = float(np.mean(cell_rates))
        cvar_grid[i, j] = compute_cvar(cell_rates)

    return {
        'grid':      grid.tolist(),
        'cvar_grid': cvar_grid.tolist(),
        'ch_vals':   ch_vals,
        'rad_vals':  rad_vals,
        'ch_label':  STRESS_DIMS['channel_degrade']['label'],
        'rad_label': STRESS_DIMS['radar_degrade']['label'],
        'n_per_cell': n_per_cell,
    }


def sobol_analysis(cfg: dict, n_sobol: int = 4096, n_per_point: int = 1,
                   progress_cb=None, map_fn=map) -> dict:
    """
    Sobol 1차/전체 민감도 지수 — 정밀 모드 전용.

    n_per_point: 각 파라미터 샘플 포인트당 시뮬레이션 반복 수.
      - n_per_point=1 (기본): 총 N×(D+2) ≈ 32,768회. 빠르지만 확률 노이즈 있음.
      - n_per_point=3: 총 ≈ 98,304회. 표준편차 √3 ≈ 1.7× 감소.
      - n_per_point=5: 총 ≈ 163,840회. 표준편차 √5 ≈ 2.2× 감소.
    확률적 시뮬레이션에서 n_per_point≥3 권장.
    """
    try:
        from SALib.sample import saltelli
        from SALib.analyze import sobol as sobol_analyze
    except ImportError:
        return {'error': 'SALib 미설치 — pip install SALib'}

    param_names = [p[0] for p in _LHS_PARAM_DEFS]
    problem = {
        'num_vars': len(_LHS_PARAM_DEFS),
        'names':    param_names,
        'bounds':   [[p[1], p[2]] for p in _LHS_PARAM_DEFS],
    }

    param_values  = saltelli.sample(problem, n_sobol, calc_second_order=False)
    n_sobol_pts   = len(param_values)
    total_runs    = n_sobol_pts * n_per_point
    Y = np.zeros(n_sobol_pts)
    base_seed = cfg.get('sim_seed', None)

    # 전 포인트 × n_per_point의 cfg를 평탄 리스트로 생성 (seed = base + i*npp + k, 직렬과 동일)
    flat_cfgs: list = []
    for i, pv in enumerate(param_values):
        run_cfg = dict(cfg)
        for j, key in enumerate(param_names):
            run_cfg[key] = float(pv[j])
        for k in range(n_per_point):
            rc = dict(run_cfg)
            if n_per_point > 1 and base_seed:
                rc['sim_seed'] = int(base_seed) + i * n_per_point + k
            flat_cfgs.append(rc)

    # 병렬(또는 직렬) 실행 — 순서 보존이라 포인트별 평균 재조립이 안전
    all_rates: list = []
    for idx, rate in enumerate(map_fn(_sim_rate_worker, flat_cfgs)):
        all_rates.append(rate)
        if progress_cb and (idx + 1) % n_per_point == 0:
            progress_cb((idx + 1) // n_per_point, n_sobol_pts)
    for i in range(n_sobol_pts):
        Y[i] = float(np.mean(all_rates[i * n_per_point:(i + 1) * n_per_point]))

    Si = sobol_analyze.analyze(
        problem, Y, calc_second_order=False, print_to_console=False)
    return {
        'S1':          Si['S1'].tolist(),
        'ST':          Si['ST'].tolist(),
        'S1_conf':     Si['S1_conf'].tolist(),
        'ST_conf':     Si['ST_conf'].tolist(),
        'names':       [p[3] for p in _LHS_PARAM_DEFS],
        'n_runs':      total_runs,
        'n_per_point': n_per_point,
    }


# ════════════════════════════════════════════════════════════════════════════
#  최적 무기 조합 추천 — 그리드 서치 + 정밀 검증
# ════════════════════════════════════════════════════════════════════════════

_OPTIMIZE_WEAPONS = [
    ('SM-3 Block IIA',  'sm3_stock'),
    ('SM-6',            'sm6_stock'),
    ('SM-2 Block IIIB', 'sm2_stock'),
    ('RIM-116 RAM',     'ram_stock'),
]


def _optimize_coarse_worker(args):
    """조합 1개의 조악한 MC 평가 — 풀 병렬화용 top-level 워커. (요격률, 조합) 반환."""
    cfg, combo, n = args
    run_cfg = {**cfg}
    for (_, key), val in zip(_OPTIMIZE_WEAPONS, combo):
        run_cfg[key] = val
    return (monte_carlo_v7(run_cfg, n=n)['mean_intercept'], combo)


def _sim_metrics_worker(cfg):
    """단일 시뮬의 (요격률, 비용, 적 격침수)만 반환 — 정밀 검증 시뮬 단위 병렬화용 워커.
    결과 dict의 함정 객체 등 비피클 항목을 피해 스칼라 3개만 전달."""
    r = run_v7_simulation(cfg)
    return (r['intercept_rate'], r['total_cost'], r['enemy_ships_destroyed'])


def optimize_weapon_loadout_v7(cfg: dict,
                                budget:     int = 64,
                                step:       int = 8,
                                max_per:    int = 32,
                                coarse_n:   int = 20,
                                fine_n:     int = 200,
                                top_k:      int = 5,
                                progress_cb = None,
                                map_fn = map) -> list:
    """
    VLS 예산 안에서 최적 무기 조합 탐색 (그리드 서치 + 정밀 검증).

    1단계: 모든 조합을 coarse_n 회 MC로 빠르게 평가 (조합 단위 병렬 — 조합 수가 코어보다 많음)
    2단계: 상위 top_k 조합을 fine_n 회 MC로 정밀 검증 (시뮬 단위 병렬 — top_k가 코어보다 적어
           조합 단위로는 코어가 남으므로, 전 시뮬을 펼쳐 분산)
    map_fn: 평가를 분산할 map (기본 직렬). app_main가 풀 기반 병렬 map 주입 시 8코어 병렬.
            각 평가는 독립이고 seed가 고정이라 결과(요격률·순위)는 직렬과 동일.
    반환: [{'combo': dict, 'rate': float, 'std': float, 'total': int}, ...]
    """
    import itertools as _it

    vals   = list(range(0, max_per + 1, step))          # [0, 8, 16, 24, 32]
    combos = [
        c for c in _it.product(vals, repeat=len(_OPTIMIZE_WEAPONS))
        if sum(c) <= budget and sum(c) > 0              # 0발 조합 제외
    ]

    # 1단계: 조악한 MC로 빠른 탐색 (조합 단위 병렬)
    total = len(combos)
    coarse_args = [(cfg, combo, coarse_n) for combo in combos]
    coarse_results: list = []
    for idx, res in enumerate(map_fn(_optimize_coarse_worker, coarse_args)):
        coarse_results.append(res)
        if progress_cb:
            progress_cb(idx + 1, total, 'coarse')

    coarse_results.sort(reverse=True)

    # 2단계: 상위 top_k 정밀 검증 — 전 시뮬(top_k × fine_n)을 펼쳐 시뮬 단위로 병렬화.
    # seed = base_seed + i (monte_carlo_v7와 동일)이라 요격률·순위는 직렬 검증과 일치.
    top_combos = [combo for _, combo in coarse_results[:top_k]]
    base_seed  = cfg.get('sim_seed', None)
    fine_cfgs: list = []
    for combo in top_combos:
        run_cfg = {**cfg, 'mc_mode': True}
        for (_, key), val in zip(_OPTIMIZE_WEAPONS, combo):
            run_cfg[key] = val
        for i in range(fine_n):
            rc = dict(run_cfg)
            if base_seed:
                rc['sim_seed'] = int(base_seed) + i
            fine_cfgs.append(rc)

    metrics: list = []
    for idx, m in enumerate(map_fn(_sim_metrics_worker, fine_cfgs)):
        metrics.append(m)
        if progress_cb and (idx + 1) % fine_n == 0:
            progress_cb(total + (idx + 1) // fine_n, total + top_k, 'fine')

    final: list = []
    for ci, combo in enumerate(top_combos):
        seg   = metrics[ci * fine_n:(ci + 1) * fine_n]
        arr   = np.array([m[0] for m in seg])
        costs = [m[1] for m in seg]
        edest = [m[2] for m in seg]
        combo_dict = {wpn: val for (wpn, _), val in zip(_OPTIMIZE_WEAPONS, combo)}
        combo_cost = sum(cnt * FRIENDLY_DB.get(wpn, {}).get('cost_usd', 0)
                         for wpn, cnt in combo_dict.items())
        mean_cost   = float(np.mean(costs)) if costs else 0.0
        mean_e_dest = float(np.mean(edest)) if edest else 0.0
        final.append({
            'combo':      combo_dict,
            'rate':       float(arr.mean()),
            'std':        float(arr.std()),
            'total':      sum(combo),
            'combo_cost': combo_cost,
            'mean_cost':  mean_cost,
            'cost_per_kill': (mean_cost / mean_e_dest if mean_e_dest > 0 else float('inf')),
        })

    return sorted(final, key=lambda x: -x['rate'])


# ════════════════════════════════════════════════════════════════════════════
#  v15.1: 적정 편대 추천 — 후보 편대 프리셋을 MC 평가해 성능·비용효과 순위 산출
# ════════════════════════════════════════════════════════════════════════════

# 후보 편대 — 한국 단독 / 한미 연합 두 그룹으로 분리 (FLEET_PRESETS 키)
_FLEET_CANDIDATES_KR = [
    '단독 작전', '기동전단 기본', 'BMD 중점', '대잠 중점', '대잠전단',
    '최대 편대', '이지스 기동전단', '이지스 기동전단 (강화)', '전 이지스 기동전단',
    '독도함 상륙전단', '동해 해역방어 (1함대)', '서해 해역방어 (2함대)',
]
_FLEET_CANDIDATES_COMBINED = [
    '한미 기동전단 기본', '한미 기동전단 강화', '한미 항모전단 지원',
]


def fleet_procurement_cost(preset_name: str) -> float:
    """편대 조달비용(USD) = Σ함정 조달가 + Σ탑재 무기 재고비(CIWS 무한재고 제외)."""
    total = 0.0
    for ship in FLEET_PRESETS.get(preset_name, []):
        stype = ship['type']
        total += SHIP_PROCUREMENT_USD.get(stype, 0)
        spec = SHIP_DB.get(stype, {})
        for wpn, cnt in spec.get('default_inventory', {}).items():
            if cnt >= 9999:        # CIWS 등 무한재고 마커 제외
                continue
            total += cnt * FRIENDLY_DB.get(wpn, {}).get('cost_usd', 0)
        for wpn, cnt in spec.get('default_strike_inventory', {}).items():
            total += cnt * FRIENDLY_DB.get(wpn, {}).get('cost_usd', 0)
    return total


def _fleet_reason(preset_name: str) -> str:
    """편대 구성에서 강점을 추출해 '왜 이 편대인지' 자연어로 생성."""
    ships = FLEET_PRESETS.get(preset_name, [])
    aegis_types = ('KDX-III-B2', 'KDX-III-B1', 'DDG-51', 'CG-47')
    n_aegis = sum(1 for s in ships if s['type'] in aegis_types)
    n_bmd   = sum(1 for s in ships if 'BMD' in SHIP_DB.get(s['type'], {}).get('role', []))
    n_sub   = sum(1 for s in ships if SHIP_DB.get(s['type'], {}).get('is_submarine'))
    n_heli  = sum(1 for s in ships if s['type'] in ('LPH', 'CVN', 'LPD'))  # 대잠 헬기 모함급
    channels = sum(SHIP_DB.get(s['type'], {}).get('max_channels', 0) for s in ships)
    n_total = len(ships)

    parts = []
    if n_aegis >= 2:
        parts.append(f"이지스 {n_aegis}척·동시교전 {channels}채널 → 포화 공격 분산 대응 우수")
    elif n_aegis == 1:
        parts.append(f"이지스 1척 중심({channels}채널) → 표준 위협 대응")
    else:
        parts.append(f"방공 채널 {channels}개 → 제한적 대공 방어")
    if n_bmd >= 1:
        parts.append(f"BMD 자산 {n_bmd}척 → 탄도·극초음속 요격 가능")
    if n_sub >= 1 or n_heli >= 1:
        parts.append("대잠 헬기·잠수함 보유 → 수중 위협 대응 우위")
    if n_total >= 6:
        parts.append("대규모 편성 → 전면전·장기 교전 지속력")
    elif n_total <= 1:
        parts.append("최소 편성 → 저비용·제한 위협용")
    return ' / '.join(parts)


def _fleet_metrics_worker(args):
    """편대 1개의 단일 시뮬 평가 — (preset, 1차지표, 생존율, 승리flag) 반환(피클 안전).
    단발: 1차지표=요격률, 승리flag=None / 전장: 1차지표=임무점수, 승리flag=1.0(승)/0.0."""
    preset_name, cfg, seed = args
    run_cfg = {**cfg, 'fleet_preset': preset_name, 'mc_mode': True}
    if seed is not None:
        run_cfg['sim_seed'] = seed
    if run_cfg.get('enable_battle_mode'):
        r = run_battle_simulation(run_cfg)
        ships   = r.get('friendly_ships', [])
        n_ships = len(ships) if ships else 1
        survival = max(0.0, 1.0 - r.get('friendly_ships_lost', 0) / n_ships)
        win = 1.0 if r.get('outcome') == 'win' else 0.0
        return (preset_name, r.get('friendly_score', 0.0), survival, win)
    r = run_v7_simulation(run_cfg)
    ships   = r.get('friendly_ships', [])
    n_ships = len(ships) if ships else 1
    survival = max(0.0, 1.0 - r.get('friendly_ships_lost', 0) / n_ships)
    return (preset_name, r['intercept_rate'], survival, None)


def recommend_fleet_v7(cfg: dict,
                       candidates,
                       n:           int = 150,
                       progress_cb = None,
                       map_fn = map) -> list:
    """
    후보 편대들을 동일 위협(현재 cfg)에 대해 MC 평가 → 성능·비용효과 순위.

    각 후보 × n 시뮬을 펼쳐 시뮬 단위로 병렬화(seed = base_seed + i 고정 → 결정론).
    단발 성능 = 요격률 0.6 + 생존율 0.4 / 전장 성능 = 승률 0.6 + 임무점수 0.4.
    비용효과 = 성능 / 정규화 조달비용.
    반환(성능순 정렬): 단발 [{preset, rate, std, survival, fleet_cost, perf_score,
                         cost_eff, reason}, ...] / 전장은 추가로 battle·win_rate·mission_score.
    """
    candidates = list(candidates)
    battle = bool(cfg.get('enable_battle_mode'))
    if battle:
        # 전장 시뮬은 단발 대비 ~45배 무거움 → 후보 랭킹용 표본 축소(과도한 대기 방지)
        n = min(n, 40)
    base_seed  = cfg.get('sim_seed', None)
    args = []
    for preset in candidates:
        for i in range(n):
            seed = (int(base_seed) + i) if base_seed else None
            args.append((preset, dict(cfg), seed))

    flat = []
    for idx, m in enumerate(map_fn(_fleet_metrics_worker, args)):
        flat.append(m)
        if progress_cb and (idx + 1) % n == 0:
            progress_cb((idx + 1) // n, len(candidates), 'eval')

    out = []
    for ci, preset in enumerate(candidates):
        seg   = flat[ci * n:(ci + 1) * n]
        prim  = np.array([s[1] for s in seg])   # 단발=요격률 / 전장=임무점수
        survs = np.array([s[2] for s in seg])
        surv  = float(survs.mean())
        cost  = fleet_procurement_cost(preset)
        if battle:
            wins     = np.array([s[3] for s in seg], dtype=float)
            win_rate = float(wins.mean())
            mission  = float(prim.mean())
            out.append({
                'preset':        preset,
                'battle':        True,
                'win_rate':      win_rate,
                'mission_score': mission,
                'survival':      surv,
                'std':           float(prim.std()),
                'fleet_cost':    cost,
                'perf_score':    win_rate * 0.6 + mission * 0.4,
                'reason':        _fleet_reason(preset),
            })
        else:
            rate = float(prim.mean())
            out.append({
                'preset':     preset,
                'rate':       rate,
                'std':        float(prim.std()),
                'survival':   surv,
                'fleet_cost': cost,
                'perf_score': rate * 0.6 + surv * 0.4,
                'reason':     _fleet_reason(preset),
            })

    # 비용효과: 그룹 내 최소 조달비용 기준 정규화 (같은 성능이면 싼 편대 우위)
    min_cost = min((o['fleet_cost'] for o in out if o['fleet_cost'] > 0), default=1.0)
    for o in out:
        c = o['fleet_cost'] if o['fleet_cost'] > 0 else min_cost
        o['cost_eff'] = o['perf_score'] * (min_cost / c)

    return sorted(out, key=lambda x: -x['perf_score'])


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: REQ 요구조건 판정
# ════════════════════════════════════════════════════════════════════════════

REQ_ITEMS_V7 = [
    {'id': 'REQ-01', 'name': 'MC 평균 요격률 ≥ 95%',  'desc': 'MC 평균 요격률 ≥ 95% (운 배제, 실력 기준)'},
    {'id': 'REQ-02', 'name': '응답시간 충족',          'desc': f'첫 SAM 발사 ≤ {MAX_RESPONSE_TIME_S}s'},
    {'id': 'REQ-04', 'name': '완전 요격 달성률 ≥ 90%', 'desc': 'MC 완전 요격(100%) 달성 비율 ≥ 90%'},
    {'id': 'REQ-05', 'name': '아군 무피격 (MC)',        'desc': 'MC 시뮬의 85% 이상에서 아군 피격 0회'},
    {'id': 'REQ-06', 'name': '다층 방어 확인',          'desc': '발사 미사일 수 ≥ 위협 수 (재교전 여력)'},
    {'id': 'REQ-07', 'name': '주요 SAM 잔여 ≥ 20%',    'desc': 'SM-3·SM-6·SM-2 합산 잔여 ≥ 초기 재고 20%'},
    {'id': 'REQ-08', 'name': '채널 한계 미초과',        'desc': '최대 동시 위협 ≤ 편대 총 채널'},
]


_SAM_STOCK_KEYS = {
    'SM-3 Block IIA':  'sm3_stock',
    'SM-6':            'sm6_stock',
    'SM-2 Block IIIB': 'sm2_stock',
}


def evaluate_req_v7(result: dict, mc: dict, cfg: dict = None) -> tuple:
    """REQ_ITEMS_V7 7항목 판정. (verdicts: list[bool], details: list[str]) 반환."""
    tfirst   = result.get('t_first_fire', -1.0)
    fired    = result.get('total_missiles_fired', 0)
    threats  = result['total_threats']
    peak_et  = result.get('peak_concurrent_threats', 0)
    tot_ch   = result.get('total_channels', 16)
    rem_inv  = result.get('remaining_inventory', {})

    # REQ-01: MC 평균 요격률 ≥ 95% (단일 시뮬 운 배제)
    req1 = mc['mean_intercept'] >= 0.95

    req2 = 0 <= tfirst <= MAX_RESPONSE_TIME_S

    req4 = mc['full_pass_rate'] >= 0.90

    # REQ-05: MC 시뮬의 85% 이상에서 아군 피격 0회
    hits_list = mc.get('friendly_hits', [])
    zero_hit_rate = (sum(1 for h in hits_list if h == 0) / len(hits_list)) if hits_list else 1.0
    req5 = zero_hit_rate >= 0.85

    req6 = (fired >= threats) if threats > 0 else True

    # REQ-07: 주요 SAM(SM-3·SM-6·SM-2) 잔여 ≥ 초기 재고 20%
    if cfg:
        init_sam = sum(cfg.get(v, 0) for v in _SAM_STOCK_KEYS.values())
        rem_sam  = sum(rem_inv.get(k, 0) for k in _SAM_STOCK_KEYS)
        req7 = (rem_sam / init_sam >= 0.20) if init_sam > 0 else True
        req7_detail = f"주요 SAM 잔여 {rem_sam}발 / 초기 {init_sam}발 ({rem_sam/init_sam:.0%})" if init_sam > 0 else "재고 없음"
    else:
        req7 = any(v > 0 for v in rem_inv.values())
        req7_detail = f"잔여 {'확보됨' if req7 else '전량 소진!'} ({sum(rem_inv.values())}발)"

    req8 = peak_et <= tot_ch

    verdicts = [req1, req2, req4, req5, req6, req7, req8]
    details  = [
        f"MC 평균 요격률 {mc['mean_intercept']:.1%} {'≥' if req1 else '<'} 95%",
        f"첫 발사 {tfirst:.0f}s ≤ {MAX_RESPONSE_TIME_S}s" if tfirst >= 0 else "발사 없음",
        f"MC 완전 성공률 {mc['full_pass_rate']:.1%} {'≥' if req4 else '<'} 90%",
        f"MC 무피격 비율 {zero_hit_rate:.1%} {'≥' if req5 else '<'} 85%",
        f"발사 {fired}발 / 위협 {threats}개",
        req7_detail,
        f"최대 동시 위협 {peak_et} ≤ 채널 {tot_ch}",
    ]
    return verdicts, details


# 지속 전장 모드 목표 → 표시 명칭 (app_main _otype과 일치)
_BATTLE_OBJ_NAME = {
    'defend_asset': '자산 방어', 'destroy_asset': '자산 격침',
    'sea_control': '해역 통제', 'attrition': '소모전', 'sustainment': '작전 지속성',
}
# 작전 목표 달성 판정 임계 — progress 0~1 정규화 기준 '우세'
_BATTLE_REQ_PROGRESS = 0.50
_BATTLE_REQ_WINRATE  = 0.50


def evaluate_req_battle_v7(result: dict, mc: dict, cfg: dict = None) -> tuple:
    """지속 전장 모드 REQ 판정 — 단발 요격률 기준 대신 엔진 작전 목표(progress)·
    승패를 그대로 판정으로 승격. 시나리오에 편성된 목표가 곧 REQ(동적).
    반환: (items, verdicts, details) — 3-tuple (단발 evaluate_req_v7는 2-tuple, 별도).
      ▸종합: 작전 승리(MC 승률, 없으면 단일 outcome)
      ▸목표별: result['objectives']의 아군 목표 progress ≥ 0.5 (편성된 만큼 동적)
      ▸공통 능력: 응답시간·채널 한계 (작전 성패 무관 시스템 요구성능)
    """
    items, verdicts, details = [], [], []
    bidx = 1

    # ── 종합: 작전 승리 ──────────────────────────────────────────────
    wr = mc.get('win_rate')
    if wr is not None:
        v = wr >= _BATTLE_REQ_WINRATE
        items.append({'id': f'REQ-B{bidx}', 'name': f'작전 승리 (MC 승률 ≥ {_BATTLE_REQ_WINRATE:.0%})',
                      'desc': f'MC 작전 승률 ≥ {_BATTLE_REQ_WINRATE:.0%}'})
        details.append(f"MC 승률 {wr:.0%} {'≥' if v else '<'} {_BATTLE_REQ_WINRATE:.0%}")
    else:
        oc = result.get('outcome')
        v = (oc == 'win')
        items.append({'id': f'REQ-B{bidx}', 'name': '작전 승리',
                      'desc': '단일 작전에서 승리(임무 달성)'})
        _ocname = {'win': '승리', 'loss': '패배', 'draw': '무승부'}.get(oc, oc)
        details.append(f"작전 결과 {_ocname}")
    verdicts.append(v); bidx += 1

    # ── 목표별: 편성된 아군 목표 progress 그대로 (동적) ──────────────
    for ob in result.get('objectives', []):
        if ob.get('side') != 'friendly':
            continue
        prog = ob.get('progress', 0.0)
        v = prog >= _BATTLE_REQ_PROGRESS
        nm = _BATTLE_OBJ_NAME.get(ob.get('type'), ob.get('type'))
        items.append({'id': f'REQ-B{bidx}', 'name': f'{nm} 달성',
                      'desc': f'{nm} 달성도 ≥ {_BATTLE_REQ_PROGRESS:.0%}'})
        details.append(f"{nm} 달성도 {prog:.0%} {'≥' if v else '<'} {_BATTLE_REQ_PROGRESS:.0%}")
        verdicts.append(v); bidx += 1

    # ── 공통 능력: 응답시간 (데이터 있을 때만 — 전장 _compile 미채움 시 제외) ─
    tfirst = result.get('t_first_fire', -1.0)
    if tfirst >= 0:
        v = tfirst <= MAX_RESPONSE_TIME_S
        items.append({'id': f'REQ-B{bidx}', 'name': '응답시간 충족',
                      'desc': f'첫 SAM 발사 ≤ {MAX_RESPONSE_TIME_S}s'})
        details.append(f"첫 발사 {tfirst:.0f}s ≤ {MAX_RESPONSE_TIME_S}s")
        verdicts.append(v); bidx += 1

    # ── 공통 능력: 채널 한계 (관측된 동시 위협이 있을 때만) ──────────
    peak_et = result.get('peak_concurrent_threats', 0)
    if peak_et > 0:
        tot_ch = result.get('total_channels', 16)
        v = peak_et <= tot_ch
        items.append({'id': f'REQ-B{bidx}', 'name': '채널 한계 미초과',
                      'desc': '최대 동시 위협 ≤ 편대 총 채널'})
        details.append(f"최대 동시 위협 {peak_et} ≤ 채널 {tot_ch}")
        verdicts.append(v); bidx += 1

    return items, verdicts, details


# ════════════════════════════════════════════════════════════════════════════
#  REQ 달성 최소 재고 역산
# ════════════════════════════════════════════════════════════════════════════

# cfg 키 ↔ 무기명 매핑 (포팅 A _def_map과 동일 구조)
_STOCK_CFG_KEY: dict = {
    'SM-3 Block IIA':   'sm3_stock',
    'SM-6':             'sm6_stock',
    'SM-2 Block IIIB':  'sm2_stock',
    'RIM-116 RAM':      'ram_stock',
    '홍상어 (대잠)':    'hongsango_stock',
    '청상어 (경어뢰)':  'cheongsango_stock',
}


def find_min_stock_v7(
    cfg: dict,
    weapon_name: str,
    target_rate: float = 0.90,
    mc_n: int = 40,
) -> int:
    """
    이진 탐색으로 REQ-04(MC 완전 요격 성공률 ≥ target_rate) 달성에 필요한
    weapon_name 의 최소 함정당 재고를 반환.
      ≥ 0 : 달성 가능한 최소 재고
      -1  : 최대값(SHIP_DB 기본)에서도 달성 불가
    """
    stock_key = _STOCK_CFG_KEY.get(weapon_name)
    if stock_key is None:
        return -1

    max_val = FRIENDLY_DB.get(weapon_name, {}).get('stock', 48)

    # 상한에서도 미달성이면 불가
    if monte_carlo_v7({**cfg, stock_key: max_val}, mc_n)['full_pass_rate'] < target_rate:
        return -1

    lo, hi = 0, max_val
    while lo < hi:
        mid = (lo + hi) // 2
        rate = monte_carlo_v7({**cfg, stock_key: mid}, mc_n)['full_pass_rate']
        if rate >= target_rate:
            hi = mid
        else:
            lo = mid + 1
    return lo


def _min_stock_worker(args):
    """무기 1종의 최소 재고 이진 탐색 — 풀 병렬화용 top-level 워커.
    (무기명, {min_stock, current_stock, achievable}) 반환."""
    cfg, wpn, target_rate, mc_n = args
    key     = _STOCK_CFG_KEY[wpn]
    current = cfg.get(key, FRIENDLY_DB.get(wpn, {}).get('stock', 0))
    min_s   = find_min_stock_v7(cfg, wpn, target_rate, mc_n)
    return (wpn, {
        'min_stock':     min_s,
        'current_stock': current,
        'achievable':    min_s >= 0,
    })


def find_all_min_stocks_v7(
    cfg: dict,
    target_rate: float = 0.90,
    mc_n: int = 40,
    progress_cb=None,
    map_fn=map,
) -> dict:
    """
    주요 무기 6종의 최소 함정당 재고를 탐색 (무기 단위 병렬).
    map_fn: 무기별 탐색을 분산할 map (기본 직렬). app_main가 풀 기반 병렬 map 주입 시
            무기 6종을 동시 탐색. 무기끼리 독립이라 결과는 직렬과 동일.
    반환: {weapon_name: {'min_stock': int, 'current_stock': int, 'achievable': bool}}
    """
    weapons = list(_STOCK_CFG_KEY.keys())
    args = [(cfg, wpn, target_rate, mc_n) for wpn in weapons]
    results = {}
    for idx, (wpn, rec) in enumerate(map_fn(_min_stock_worker, args)):
        results[wpn] = rec
        if progress_cb:
            progress_cb(idx + 1, len(weapons), wpn)
    return results


# ════════════════════════════════════════════════════════════════════════════
#  자동 취약점 진단
# ════════════════════════════════════════════════════════════════════════════

def diagnose_vulnerabilities_v7(result: dict, mc: dict, cfg: dict) -> list:
    """
    MC 결과를 자동 분석하여 취약점 진단 카드 목록을 반환.
    각 카드: {'severity': 'HIGH'|'MED'|'LOW'|'OK', 'title', 'detail', 'suggestion'}
    """
    cards = []

    mean_ir   = mc['mean_intercept']
    full_pass = mc['full_pass_rate']
    std_ir    = mc['std_intercept']
    mean_hits = float(np.mean(mc['friendly_hits'])) if mc['friendly_hits'] else 0.0
    peak_et   = result.get('peak_concurrent_threats', 0)
    tot_ch    = result.get('total_channels', 16)
    rem_inv   = result.get('remaining_inventory', {})
    t_first   = result.get('t_first_fire', -1.0)
    w_avg_rem = mc.get('weapon_avg_remaining', {})

    # ── 1. 완전 요격 성공률 미달 ──────────────────────────────────────────
    if full_pass < 0.90:
        sev = 'HIGH' if full_pass < 0.70 else 'MED'
        # 가장 많이 소진된 무기 파악 → 구체적 개선 제안
        most_depleted = min(w_avg_rem, key=lambda k: w_avg_rem[k], default=None)
        if most_depleted and most_depleted in _STOCK_CFG_KEY:
            key     = _STOCK_CFG_KEY[most_depleted]
            cur_stk = cfg.get(key, FRIENDLY_DB.get(most_depleted, {}).get('stock', 0))
            sugg    = f'• {most_depleted} 재고를 {cur_stk}→{cur_stk + 12}발로 증가 검토\n• MC 횟수 증가로 정밀도 향상'
        else:
            sugg = '• 주요 SAM 재고 증가\n• CEC 활성화 또는 함정 증원 검토'
        cards.append({
            'severity':   sev,
            'title':      f'완전 요격 성공률 미달  ({full_pass:.0%} < REQ 90%)',
            'detail':     (f'MC {mc["n"]}회 중 {full_pass:.0%}만 모든 위협 요격. '
                           f'평균 요격률 {mean_ir:.1%} (편차 ±{std_ir:.1%}).'),
            'suggestion': sugg,
        })

    # ── 2. 아군 피격 빈발 ─────────────────────────────────────────────────
    if mean_hits > 0.3:
        sev = 'HIGH' if mean_hits >= 1.5 else 'MED'
        most_hit_ship = ''
        ship_avg = mc.get('ship_avg_hits', {})
        if ship_avg:
            sh = max(ship_avg, key=ship_avg.get)
            most_hit_ship = f'  가장 많이 피격: {sh} (평균 {ship_avg[sh]:.1f}회)'
        cards.append({
            'severity':   sev,
            'title':      f'아군 함정 피격 빈발  (MC 평균 {mean_hits:.1f}회)',
            'detail':     f'종말 방어 단계(RAM/CIWS) 취약 또는 ECM·회피 기동 효과 부족.{most_hit_ship}',
            'suggestion': '• RIM-116 RAM 재고 증가\n• 함정 회피 기동 활성화\n• ECM(AN/SLQ-32) 옵션 활성화',
        })

    # ── 3. 채널 포화 ──────────────────────────────────────────────────────
    if peak_et > 0 and tot_ch > 0:
        ratio = peak_et / tot_ch
        if ratio >= 1.0:
            cards.append({
                'severity':   'HIGH',
                'title':      f'채널 포화 발생  (동시 위협 {peak_et} > 채널 {tot_ch})',
                'detail':     f'최대 {peak_et}개 위협이 동시 접근 — 교전 채널 {tot_ch}개 초과. 일부 위협 무대응.',
                'suggestion': '• CEC(협동교전능력) 활성화로 채널 공유\n• 추가 함정 편입\n• 발사 간격(launch_interval_s) 단축',
            })
        elif ratio >= 0.80:
            cards.append({
                'severity':   'MED',
                'title':      f'채널 포화 근접  ({peak_et}/{tot_ch} = {ratio:.0%})',
                'detail':     f'채널 사용률 {ratio:.0%} — 위협 추가 시 포화 임박.',
                'suggestion': '• CEC 활성화 또는 함정 증원 검토',
            })

    # ── 4. 주요 무기 소진 ─────────────────────────────────────────────────
    key_weapons = list(_STOCK_CFG_KEY.keys())
    for wpn in key_weapons:
        avg_rem = w_avg_rem.get(wpn, -1.0)
        if avg_rem < 0:
            continue  # 해당 무기 미사용 시나리오
        if avg_rem < 2.0:
            sev     = 'HIGH' if avg_rem < 0.5 else 'MED'
            cfg_key = _STOCK_CFG_KEY[wpn]
            cur_stk = cfg.get(cfg_key, FRIENDLY_DB.get(wpn, {}).get('stock', 0))
            cards.append({
                'severity':   sev,
                'title':      f'{wpn} 재고 고갈 위험  (MC 평균 잔여 {avg_rem:.1f}발)',
                'detail':     f'평균적으로 {wpn}이 거의 소진됨 (현재 재고: 함정당 {cur_stk}발).',
                'suggestion': f'• {wpn} 재고를 {cur_stk}→{cur_stk + 12}발로 증가 검토',
            })

    # ── 5. 응답시간 초과 ──────────────────────────────────────────────────
    if 0 <= t_first > MAX_RESPONSE_TIME_S:
        cards.append({
            'severity':   'MED',
            'title':      f'응답시간 초과  (첫 SAM 발사 {t_first:.0f}s > REQ {MAX_RESPONSE_TIME_S}s)',
            'detail':     'C&D + 확인 절차 후 첫 발사까지 지나치게 오래 소요. REQ-02 불충족.',
            'suggestion': '• C&D 시간(cd_time_s) 단축\n• 탐지거리 확대로 사전 추적 가능',
        })

    # ── 6. 높은 변동성 ────────────────────────────────────────────────────
    if std_ir > 0.15 and mean_ir < 0.98:
        cards.append({
            'severity':   'LOW',
            'title':      f'요격률 불안정  (표준편차 {std_ir:.1%})',
            'detail':     f'MC 결과 편차 큼 — {mean_ir:.0%}±{std_ir:.0%}. 특정 조건에서 방어 붕괴 가능.',
            'suggestion': '• MC 횟수 증가(200회 이상)로 신뢰도 향상\n• CEC 활성화로 일관성 확보',
        })

    # ── 7. 지상 BMD 자산 소진 ────────────────────────────────────────────
    if cfg:
        ground_rem = result.get('ground_remaining', {})
        for asset, cfg_key, label in [
            ('SM-3 (어쇼어)', 'ashore_sm3_stock', '이지스 어쇼어'),
            ('THAAD 요격탄',  'thaad_stock',       'THAAD'),
            ('L-SAM',         'lsam_stock',        'L-SAM'),
            ('천궁-II',       'chungung_stock',    '천궁-II'),
            ('PAC-3 MSE',     'patriot_stock',     '패트리엇 PAC-3'),
        ]:
            init_stock = cfg.get(cfg_key, 0)
            if init_stock > 0:
                rem   = ground_rem.get(asset, 0)
                ratio = rem / init_stock
                if ratio < 0.20:
                    sev = 'HIGH' if ratio < 0.05 else 'MED'
                    cards.append({
                        'severity':   sev,
                        'title':      f'{label} 탄약 고갈  (잔여 {rem}발 / 초기 {init_stock}발, {ratio:.0%})',
                        'detail':     '지상 BMD 1차 방어망 약화. 탄도·HGV 위협이 함정까지 도달 가능.',
                        'suggestion': f'• {label} 탄약 재고 {init_stock + 12}발 이상으로 증가\n• 함정 SM-3 재고 보충으로 백업 강화',
                    })

    # ── 이상 없음 ─────────────────────────────────────────────────────────
    if not cards:
        cards.append({
            'severity':   'OK',
            'title':      '취약점 없음 — 방어 태세 양호',
            'detail':     (f'완전 요격 성공률 {full_pass:.0%} · '
                           f'MC 평균 요격률 {mean_ir:.1%} · '
                           f'아군 평균 피격 {mean_hits:.1f}회'),
            'suggestion': '더 어려운 시나리오(전방위 포화·혼합 공격)로 한계 탐색 권장.',
        })

    return cards


# ════════════════════════════════════════════════════════════════════════════
#  교전 후 브리핑 자동 생성
# ════════════════════════════════════════════════════════════════════════════

def generate_briefing(result: dict, mc: dict, cfg: dict) -> str:
    """시뮬 결과를 군사 보고서 형식의 서술형 텍스트로 자동 생성."""
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    mean_ir   = mc['mean_intercept']
    full_pass = mc['full_pass_rate']
    cvar      = mc.get('cvar')
    mean_hits = float(np.mean(mc['friendly_hits'])) if mc['friendly_hits'] else 0.0
    n         = mc['n']

    fleet_preset = cfg.get('fleet_preset', '—')
    weather      = cfg.get('weather', '—')
    enemy_label  = (cfg.get('mixed_scenario')
                    or cfg.get('enemy_fleet_preset')
                    or cfg.get('enemy_fleet_mode', '—'))

    total_threats = result.get('total_threats', 0)
    enemy_dest    = result.get('enemy_ships_destroyed', 0)
    friendly_lost = result.get('friendly_ships_lost', 0)
    mean_cost     = mc.get('mean_cost', result.get('total_cost', 0))
    aircraft_sort = result.get('aircraft_sorties', 0)

    mission_ok  = mean_ir >= 0.95 and full_pass >= 0.90
    mission_str = '✅ 임무 성공' if mission_ok else '❌ 임무 실패'

    verdicts, details = evaluate_req_v7(result, mc, cfg)
    cards = diagnose_vulnerabilities_v7(result, mc, cfg)

    SEP  = '━' * 54
    SEP2 = '─' * 54
    _SEV = {'HIGH': '🔴 위험', 'MED': '🟡 경고', 'LOW': '🔵 주의', 'OK': '🟢 양호'}

    lines = [
        SEP,
        '  이지스 기동전단 교전 후 브리핑 (자동 생성)',
        f'  작성 일시: {now}',
        SEP, '',
        '【1. 작전 개요】',
        f'  편대 구성  : {fleet_preset}',
        f'  위협 환경  : {weather} / {enemy_label}',
        f'  총 위협 수 : {total_threats}개',
        f'  MC 시뮬    : {n}회',
        '',
        '【2. 교전 결과 요약】',
        f'  요격률 (MC 평균)     : {mean_ir:.1%}  →  {mission_str}',
        f'  완전 요격 달성률     : {full_pass:.1%}',
    ]
    if cvar is not None:
        lines.append(f'  최악 CVaR (하위 5%)  : {cvar:.1%}')
    lines += [
        f'  아군 피격 (MC 평균)  : {mean_hits:.1f}회',
        f'  아군 함정 손실       : {friendly_lost}척',
        f'  적 함정 격침         : {enemy_dest}척',
        f'  총 교전 비용 (평균)  : ${mean_cost:,.0f}',
    ]
    if aircraft_sort > 0:
        lines.append(f'  항공 출격 횟수       : {aircraft_sort}회')
    lines += ['', '【3. REQ 판정 결과】']
    for req, v, d in zip(REQ_ITEMS_V7, verdicts, details):
        mark = '✅' if v else '❌'
        lines.append(f'  {mark} {req["id"]}  {req["name"]}  →  {"PASS" if v else "FAIL"}')
        lines.append(f'       {d}')

    lines += ['', '【4. 주요 취약점】']
    for card in cards:
        lines.append(f'  {_SEV.get(card["severity"], "")}  {card["title"]}')
        lines.append(f'       {card["detail"]}')

    lines += ['', '【5. 권고사항】']
    suggestions = []
    for card in cards:
        if card['severity'] in ('HIGH', 'MED') and card.get('suggestion'):
            for s in card['suggestion'].split('\n'):
                s = s.strip().lstrip('•').strip()
                if s and s not in suggestions:
                    suggestions.append(s)
    if suggestions:
        for i, s in enumerate(suggestions, 1):
            lines.append(f'  {i}. {s}')
    else:
        lines.append('  현재 주요 권고사항 없음. 방어 태세 양호.')

    lines += [
        '',
        SEP,
        '  본 브리핑은 시뮬레이션 데이터 기반 자동 생성 문서입니다.',
        '  Pk 수치는 공개 자료 기반 추정값 (±15~20%) — 실측 데이터 아님.',
        SEP,
    ]
    return '\n'.join(lines)


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: 날씨별 시나리오 비교
# ════════════════════════════════════════════════════════════════════════════

_SCENARIO_WEATHERS = [
    ('최선 (맑음)',  '맑음 (주간)'),
    ('평균 (흐림)',  '흐림 (박무)'),
    ('최악 (폭풍)',  '폭풍 (해상 악화)'),
]


def scenario_comparison_v7(cfg: dict, n: int = 200) -> dict:
    """날씨 3종 MC 비교. {label: mc_dict + mean_cost} 반환."""
    results = {}
    for label, weather in _SCENARIO_WEATHERS:
        c = dict(cfg)
        c['weather'] = weather
        mc = monte_carlo_v7(c, n=n, desc=f'시나리오: {label}')
        results[label] = {
            **mc,
            'mean_cost': float(np.mean(mc['total_costs'])),
        }
    return results


# ════════════════════════════════════════════════════════════════════════════
#  CEC 효과 비교
# ════════════════════════════════════════════════════════════════════════════

def cec_comparison_v7(cfg: dict, n: int = 200) -> dict:
    """
    CEC ON/OFF/두절 3종 MC 비교.
    반환: {'CEC ON': mc_dict, 'CEC OFF (독립교전)': mc_dict, 'CEC 두절 (재밍)': mc_dict}
    """
    results = {}
    scenarios = [
        ('CEC ON',            {'enable_cec_preassign': True,  'enable_cec_jammed': False}),
        ('CEC OFF (독립교전)', {'enable_cec_preassign': False, 'enable_cec_jammed': False}),
        ('CEC 두절 (재밍)',   {'enable_cec_preassign': False, 'enable_cec_jammed': True}),
    ]
    for label, overrides in scenarios:
        c = {**cfg, **overrides}
        mc = monte_carlo_v7(c, n=n, desc=f'CEC: {label}')
        results[label] = {
            **mc,
            'mean_cost': float(np.mean(mc['total_costs'])) if mc.get('total_costs') else 0.0,
        }
    return results


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: A vs B 시나리오 비교
# ════════════════════════════════════════════════════════════════════════════

def _ab_mc_worker(args):
    """A/B 비교용 단일 MC 실행 워커 — 풀 병렬화용 top-level(피클 가능). mc dict 반환."""
    cfg, n = args
    return monte_carlo_v7(cfg, n=n)


def _heatmap_cell_worker(args):
    """히트맵 셀 1개 MC 실행 워커 — 풀 병렬화용 top-level. (r, c, 평균요격률) 반환."""
    cfg, r, c, n = args
    return (r, c, float(monte_carlo_lhs(cfg, n=n).get('mean_intercept', 0.0)))


def compare_ab_v7(cfg_a: dict, cfg_b: dict, n: int = 200, map_fn=map) -> dict:
    """
    두 cfg로 MC를 각각 실행해 비교 dict를 반환.
    map_fn: 두 MC를 분산 (기본 직렬). app_main가 풀 병렬 map 주입 시 A·B 동시 실행.
    반환: {'a': mc_dict, 'b': mc_dict, 'delta_intercept': float, 'delta_cost': float}
    """
    mc_a, mc_b = list(map_fn(_ab_mc_worker, [(cfg_a, n), (cfg_b, n)]))
    return {
        'a':               mc_a,
        'b':               mc_b,
        'delta_intercept': mc_b['mean_intercept'] - mc_a['mean_intercept'],
        'delta_cost':      float(np.mean(mc_b['total_costs'])) - float(np.mean(mc_a['total_costs'])),
    }


# ════════════════════════════════════════════════════════════════════════════
#  포팅 D: 시나리오 저장 / 불러오기
# ════════════════════════════════════════════════════════════════════════════

import json as _json


def save_scenario_v7(cfg: dict, path: str):
    """cfg를 JSON으로 저장. 직렬화 불가능한 값은 제외."""
    serializable = {}
    for k, v in cfg.items():
        try:
            _json.dumps(v)
            serializable[k] = v
        except (TypeError, ValueError):
            pass
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(serializable, f, ensure_ascii=False, indent=2)


def load_scenario_v7(path: str) -> dict:
    """JSON 파일에서 cfg를 불러온다."""
    with open(path, 'r', encoding='utf-8') as f:
        return _json.load(f)


# ════════════════════════════════════════════════════════════════════════════
#  PNG 차트 생성
# ════════════════════════════════════════════════════════════════════════════

_BG   = '#0a0e1a'
_GRID = '#1e2a3a'
_ACC  = '#3498db'

def _ax_style(ax, title: str):
    ax.set_facecolor(_BG)
    ax.tick_params(colors='#aab', labelsize=11)
    for sp in ax.spines.values():
        sp.set_color(_GRID)
    ax.set_title(title, color='#dde', fontsize=13, fontweight='bold', pad=6)


def plot_v7(result: dict, mc: dict, cfg: dict,
            img_path: str = '이지스_기동전단_v7_분석.png') -> str:
    """
    단일 시뮬 결과(result) + MC 통계(mc)를 6개 서브플롯으로 시각화.
    img_path에 저장 후 경로 반환.
    """
    fig = _MplFigure(figsize=(16, 10), facecolor=_BG)
    _FigureCanvasAgg(fig)
    fig.suptitle(
        f"이지스 기동전단 통합 방어 시뮬레이터 v7.0\n"
        f"시나리오: {cfg.get('fleet_preset','?')} | "
        f"날씨: {cfg.get('weather','?')} | "
        f"MC {mc['n']}회",
        color='white', fontsize=16, fontweight='bold', y=0.98,
    )

    gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.45, wspace=0.35,
                           left=0.07, right=0.97, top=0.90, bottom=0.08)

    # ── (0,0) 요격률 히스토그램 ──────────────────────────────────────────────
    ax0 = fig.add_subplot(gs[0, 0])
    _ax_style(ax0, '요격률 분포 (MC)')
    ax0.hist(mc['intercept_rates'], bins=20, color=_ACC, edgecolor='#0a0e1a', alpha=0.85)
    ax0.axvline(mc['mean_intercept'], color='#e74c3c', lw=1.5, ls='--',
                label=f"평균 {mc['mean_intercept']:.1%}")
    ax0.set_xlabel('요격률', color='#aab', fontsize=11)
    ax0.set_ylabel('빈도', color='#aab', fontsize=11)
    ax0.legend(fontsize=10, facecolor=_BG, labelcolor='white', edgecolor=_GRID)
    ax0.xaxis.set_major_formatter(_FuncFormatter(lambda v, _: f'{v:.0%}'))

    # ── (0,1) 아군 피격 분포 ─────────────────────────────────────────────────
    ax1 = fig.add_subplot(gs[0, 1])
    _ax_style(ax1, '아군 피격 횟수 분포 (MC)')
    ax1.hist(mc['friendly_hits'], bins=range(0, max(mc['friendly_hits']) + 2),
             color='#e74c3c', edgecolor='#0a0e1a', alpha=0.85, align='left')
    ax1.set_xlabel('피격 횟수', color='#aab', fontsize=11)
    ax1.set_ylabel('빈도', color='#aab', fontsize=11)

    # ── (0,2) 적 격침 분포 ───────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[0, 2])
    _ax_style(ax2, '적 플랫폼 격침 수 분포 (MC)')
    max_dest = max(mc['enemy_destroyed']) if mc['enemy_destroyed'] else 1
    ax2.hist(mc['enemy_destroyed'], bins=range(0, max_dest + 2),
             color='#2ecc71', edgecolor='#0a0e1a', alpha=0.85, align='left')
    ax2.set_xlabel('격침 수', color='#aab', fontsize=11)
    ax2.set_ylabel('빈도', color='#aab', fontsize=11)

    # ── (1,0) 무기 소모 현황 (단일 시뮬) ─────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 0])
    _ax_style(ax3, '무기 소모 현황 (단일 시뮬)')
    ships = result.get('friendly_ships', [])
    wpn_used: dict = {}
    for s in ships:
        spec = SHIP_DB.get(s.ship_type, {})
        default_inv = spec.get('default_inventory', {})
        for wpn, orig in default_inv.items():
            used = orig - s.inventory.get(wpn, orig)
            if used > 0:
                wpn_used[wpn] = wpn_used.get(wpn, 0) + used
    if wpn_used:
        labels = list(wpn_used.keys())
        values = [wpn_used[k] for k in labels]
        colors = [_ACC if i % 2 == 0 else '#5dade2' for i in range(len(labels))]
        bars = ax3.barh(labels, values, color=colors, edgecolor='#0a0e1a')
        ax3.bar_label(bars, padding=3, color='white', fontsize=10)
        ax3.set_xlabel('발사 수', color='#aab', fontsize=11)
        ax3.tick_params(axis='y', labelsize=10)
    else:
        ax3.text(0.5, 0.5, '발사 없음', color='#aab', ha='center', va='center',
                 transform=ax3.transAxes)

    # ── (1,1) 비용 분포 ──────────────────────────────────────────────────────
    ax4 = fig.add_subplot(gs[1, 1])
    _ax_style(ax4, '총 교전 비용 분포 (MC)')
    costs_m = [c / 1_000_000 for c in mc['total_costs']]
    ax4.hist(costs_m, bins=20, color='#f39c12', edgecolor='#0a0e1a', alpha=0.85)
    mean_m = np.mean(costs_m)
    ax4.axvline(mean_m, color='#e74c3c', lw=1.5, ls='--',
                label=f'평균 ${mean_m:.1f}M')
    ax4.set_xlabel('비용 (백만 USD)', color='#aab', fontsize=11)
    ax4.set_ylabel('빈도', color='#aab', fontsize=11)
    ax4.legend(fontsize=10, facecolor=_BG, labelcolor='white', edgecolor=_GRID)

    # ── (1,2) 핵심 수치 요약 ────────────────────────────────────────────────
    ax5 = fig.add_subplot(gs[1, 2])
    ax5.set_facecolor(_BG)
    ax5.axis('off')
    _ax_style(ax5, '핵심 수치 요약')

    enemy_count = sum(s.get('count', 1) for s in cfg.get('enemy_fleet', []))
    summary_lines = [
        ('적 편대 규모',     f"{enemy_count}기/척"),
        ('교전 지속(모사)',  f"{result['sim_time']:.0f}s"),
        ('총 위협 수',       f"{result['total_threats']}발/기"),
        ('요격 성공 (단일)', f"{result['intercepted_threats']}발/기"),
        ('', ''),
        ('MC 평균 요격률',   f"{mc['mean_intercept']:.1%}"),
        ('MC 표준편차',      f"±{mc['std_intercept']:.1%}"),
        ('완전요격 비율',    f"{mc['full_pass_rate']:.1%}"),
        ('', ''),
        ('아군 피격 (단일)', f"{result['friendly_hits']}회"),
        ('적 격침 (단일)',   f"{result['enemy_ships_destroyed']}기/척"),
        ('아군 손실 (단일)', f"{result['friendly_ships_lost']}척"),
        ('총 비용 (단일)',   f"${result['total_cost']/1e6:.1f}M"),
    ]
    y = 0.97
    for label, val in summary_lines:
        if not label:
            y -= 0.04
            continue
        ax5.text(0.04, y, label, color='#7fb3d3', fontsize=11, transform=ax5.transAxes, va='top')
        ax5.text(0.96, y, val,   color='white',   fontsize=11, transform=ax5.transAxes, va='top', ha='right', fontweight='bold')
        y -= 0.075

    fig.savefig(img_path, dpi=150, bbox_inches='tight', facecolor=_BG)
    fig.clf()
    print(f"  그래프 저장: '{img_path}'")
    return img_path


# ════════════════════════════════════════════════════════════════════════════
#  Excel 보고서 생성
# ════════════════════════════════════════════════════════════════════════════

def save_excel_report_v7(result: dict, mc: dict, cfg: dict,
                          img_path: str = '',
                          xlsx_path: str = '이지스_기동전단_v7_보고서.xlsx'):
    """
    Sheet1: MC 통계 요약
    Sheet2: 무기 소모 현황
    Sheet3: 교전 로그
    Sheet4: PNG 차트 삽입 (이미지 있을 때)
    """
    wb = Workbook()
    tb = Border(**{s: Side(style='thin', color='CCCCCC')
                   for s in ('left', 'right', 'top', 'bottom')})

    def cs(ws, r, c, v, bold=False, bg=None, center=True, color='000000'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, size=10, name='Malgun Gothic', color=color)
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center', wrap_text=True,
        )
        cell.border = tb
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    def title_row(ws, r, text, cols='A:F', bg='1A252F'):
        end_col = cols.split(':')[1]
        ws.merge_cells(f'A{r}:{end_col}{r}')
        cell = ws.cell(row=r, column=1, value=text)
        cell.font      = Font(bold=True, size=13, color='FFFFFF', name='Malgun Gothic')
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 28

    def hdr(ws, r, headers, bg='2C3E50'):
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.font      = Font(bold=True, size=10, color='FFFFFF', name='Malgun Gothic')
            cell.fill      = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = tb

    # ── Sheet1: MC 통계 요약 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'MC 통계 요약'
    ws1.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [20, 20, 18, 18, 18, 18]):
        ws1.column_dimensions[col].width = w

    title_row(ws1, 1, f'이지스 기동전단 v7.0 — MC {mc["n"]}회 통계 요약')
    hdr(ws1, 2, ['항목', '단일 시뮬', 'MC 평균', 'MC 표준편차', 'MC 최솟값', 'MC 최댓값'])

    rows = [
        ('요격률',
         f"{result['intercept_rate']:.1%}",
         f"{mc['mean_intercept']:.1%}",
         f"±{mc['std_intercept']:.1%}",
         f"{min(mc['intercept_rates']):.1%}",
         f"{max(mc['intercept_rates']):.1%}"),
        ('아군 피격',
         result['friendly_hits'],
         f"{np.mean(mc['friendly_hits']):.1f}",
         f"±{np.std(mc['friendly_hits']):.1f}",
         min(mc['friendly_hits']),
         max(mc['friendly_hits'])),
        ('적 격침',
         result['enemy_ships_destroyed'],
         f"{np.mean(mc['enemy_destroyed']):.1f}",
         f"±{np.std(mc['enemy_destroyed']):.1f}",
         min(mc['enemy_destroyed']),
         max(mc['enemy_destroyed'])),
        ('아군 함정 손실',
         result['friendly_ships_lost'],
         f"{np.mean(mc['friendly_lost']):.1f}",
         f"±{np.std(mc['friendly_lost']):.1f}",
         min(mc['friendly_lost']),
         max(mc['friendly_lost'])),
        ('총 비용 (USD)',
         f"${result['total_cost']:,.0f}",
         f"${np.mean(mc['total_costs']):,.0f}",
         f"±${np.std(mc['total_costs']):,.0f}",
         f"${min(mc['total_costs']):,.0f}",
         f"${max(mc['total_costs']):,.0f}"),
        ('완전 요격 비율',
         '—',
         f"{mc['full_pass_rate']:.1%}",
         '—', '—', '—'),
    ]
    for i, row in enumerate(rows):
        bg = 'D5F5E3' if i % 2 == 0 else 'EBF5FB'
        for j, val in enumerate(row, 1):
            cs(ws1, i + 3, j, val, bg=bg, center=(j > 1))

    # 시나리오 파라미터 블록
    ws1.cell(row=len(rows) + 5, column=1, value='【시나리오 파라미터】').font = Font(bold=True, size=11, name='Malgun Gothic')
    params = [
        ('편대 프리셋',   cfg.get('fleet_preset', '?')),
        ('날씨',          cfg.get('weather', '?')),
        ('탐지 거리',     f"{cfg.get('detect_km', '?')} km"),
        ('해성-II 재고',  cfg.get('haesong2_stock', 0)),
        ('해성-I  재고',  cfg.get('haesong1_stock', 0)),
        ('하푼 재고',     cfg.get('harpoon_stock', 0)),
    ]
    for k, (label, val) in enumerate(params):
        cs(ws1, len(rows) + 6 + k, 1, label, center=False)
        cs(ws1, len(rows) + 6 + k, 2, val,   center=False)

    # ── Sheet2: 무기 소모 현황 ───────────────────────────────────────────────
    ws2 = wb.create_sheet('무기 소모 현황')
    ws2.sheet_view.showGridLines = False
    for col, w in zip('ABCDE', [24, 16, 16, 16, 16]):
        ws2.column_dimensions[col].width = w

    title_row(ws2, 1, '함정별 무기 소모 현황 (단일 시뮬)')
    hdr(ws2, 2, ['함정명', '무기', '초기 재고', '소모량', '잔여량'])

    row_idx = 3
    for s in result.get('friendly_ships', []):
        spec        = SHIP_DB.get(s.ship_type, {})
        default_inv = spec.get('default_inventory', {})
        for wpn, orig in default_inv.items():
            remaining = s.inventory.get(wpn, orig)
            used      = orig - remaining
            bg = 'FADBD8' if used > 0 else 'F2F3F4'
            cs(ws2, row_idx, 1, s.name,      bg=bg, center=False)
            cs(ws2, row_idx, 2, wpn,         bg=bg, center=False)
            cs(ws2, row_idx, 3, orig,        bg=bg)
            cs(ws2, row_idx, 4, used,        bg='E74C3C' if used > 0 else bg,
               color='FFFFFF' if used > 0 else '000000', bold=(used > 0))
            cs(ws2, row_idx, 5, remaining,   bg=bg)
            row_idx += 1

    # ── Sheet3: 교전 로그 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('교전 로그')
    ws3.sheet_view.showGridLines = False
    for col, w in zip('AB', [12, 80]):
        ws3.column_dimensions[col].width = w

    title_row(ws3, 1, '교전 로그 (단일 시뮬)', cols='A:B')
    hdr(ws3, 2, ['시각 (s)', '이벤트'])

    for i, (t, msg) in enumerate(result.get('log', [])):
        bg = 'EBF5FB' if i % 2 == 0 else 'FDFEFE'
        cs(ws3, i + 3, 1, f'{t:.0f}', bg=bg)
        cs(ws3, i + 3, 2, msg,        bg=bg, center=False)

    # ── Sheet4: 교전 후 브리핑 ──────────────────────────────────────────────
    ws4 = wb.create_sheet('교전 후 브리핑')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 70
    title_row(ws4, 1, '교전 후 브리핑 (자동 생성)', cols='A:A')
    briefing_text = generate_briefing(result, mc, cfg)
    for i, line in enumerate(briefing_text.split('\n'), 2):
        cell = ws4.cell(row=i, column=1, value=line)
        cell.font      = Font(name='Malgun Gothic', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws4.row_dimensions[i].height = 16

    # ── Sheet5: PNG 차트 ─────────────────────────────────────────────────────
    if img_path and _CAN_IMG and os.path.exists(img_path):
        ws5 = wb.create_sheet('분석 차트')
        ws5.sheet_view.showGridLines = False
        title_row(ws5, 1, 'MC 분석 차트', cols='A:L')
        img_obj = XLImage(img_path)
        img_obj.anchor = 'A2'
        ws5.add_image(img_obj)

    wb.save(xlsx_path)
    print(f"  엑셀 보고서 저장: '{xlsx_path}'")
    return xlsx_path


# ════════════════════════════════════════════════════════════════════════════
#  결과 JSON 내보내기
# ════════════════════════════════════════════════════════════════════════════

def save_json_report_v7(result: dict, mc: dict, path: str):
    """
    단일 시뮬 결과 + MC 통계를 JSON으로 저장.
    FriendlyShipObj 등 직렬화 불가 객체는 요약 딕셔너리로 변환.
    """
    def _safe(v):
        if isinstance(v, (int, float, str, bool, type(None))):
            return v
        if isinstance(v, (list, tuple)):
            return [_safe(x) for x in v]
        if isinstance(v, dict):
            return {k2: _safe(v2) for k2, v2 in v.items()}
        # 객체: 기본 속성만 추출
        return str(v)

    summary = {
        'result': {
            'intercept_rate':      result.get('intercept_rate'),
            'total_threats':       result.get('total_threats'),
            'intercepted_threats': result.get('intercepted_threats'),
            'friendly_hits':       result.get('friendly_hits'),
            'enemy_ships_destroyed': result.get('enemy_ships_destroyed'),
            'friendly_ships_lost': result.get('friendly_ships_lost'),
            'total_cost':          result.get('total_cost'),
            'sim_time':            result.get('sim_time'),
            't_first_fire':        result.get('t_first_fire'),
            'total_missiles_fired': result.get('total_missiles_fired'),
            'remaining_inventory': result.get('remaining_inventory', {}),
            'ships': [
                {
                    'name':         s.name,
                    'type':         s.ship_type,
                    'alive':        s.alive,
                    'hits_taken':   getattr(s, 'hits_taken', 0),
                    'total_cost':   s.total_cost,
                    'inventory':    dict(s.inventory),
                }
                for s in result.get('friendly_ships', [])
            ],
        },
        'mc': {
            'n':               mc.get('n'),
            'mean_intercept':  mc.get('mean_intercept'),
            'std_intercept':   mc.get('std_intercept'),
            'full_pass_rate':  mc.get('full_pass_rate'),
            'mean_cost':       float(np.mean(mc['total_costs'])) if mc.get('total_costs') else 0,
            'weapon_avg_remaining': mc.get('weapon_avg_remaining', {}),
            'ship_avg_hits':   mc.get('ship_avg_hits', {}),
        },
    }
    with open(path, 'w', encoding='utf-8') as f:
        _json.dump(summary, f, ensure_ascii=False, indent=2, default=_safe)
    print(f"  JSON 보고서 저장: '{path}'")
    return path


# ════════════════════════════════════════════════════════════════════════════
#  단독 실행 테스트 — 32종 혼합 시나리오
# ════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import sys

    scenario = sys.argv[1] if len(sys.argv) > 1 else 'mixed'

    SCENARIOS = {
        # 수상함 대결
        'surface': {
            'fleet_preset': '기동전단 기본',
            'weather':      '맑음 (주간)',
            'detect_km':    200,
            'haesong2_stock': 8,
            'haesong1_stock': 0,
            'harpoon_stock':  4,
            'enemy_fleet': [
                {'preset': '055형 대형 구축함', 'count': 1},
                {'preset': '052D형 구축함',     'count': 2},
            ],
        },
        # 항공 위협
        'air': {
            'fleet_preset': 'BMD 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    300,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': 'J-20 (위룡)',    'count': 2},
                {'preset': 'J-16 (플랭커-D)', 'count': 2},
                {'preset': 'H-6 (폭격기)',    'count': 1},
            ],
        },
        # 탄도/순항 미사일
        'missile': {
            'fleet_preset': 'BMD 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    400,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': 'DF-21D (대함 탄도)',        'count': 2},
                {'preset': 'DF-17 (극초음속 활공)',     'count': 1},
                {'preset': 'KN-23 (북한 이스칸데르)',   'count': 2},
                {'preset': 'YJ-12 (초음속 순항)',       'count': 3},
            ],
        },
        # 잠수함
        'sub': {
            'fleet_preset': '대잠 중점',
            'weather':      '맑음 (주간)',
            'detect_km':    200,
            'sub_detect_km': 50,
            'haesong2_stock': 0,
            'haesong1_stock': 0,
            'harpoon_stock':  0,
            'enemy_fleet': [
                {'preset': '093형 잠수함 (위안급)', 'count': 2},
                {'preset': '039형 잠수함 (송급)',   'count': 1},
            ],
        },
        # 32종 혼합
        'mixed': {
            'fleet_preset': '최대 편대',
            'weather':      '맑음 (주간)',
            'detect_km':    300,
            'sub_detect_km': 50,
            'haesong2_stock': 12,
            'haesong1_stock': 4,
            'harpoon_stock':  8,
            'enemy_fleet': [
                {'preset': '055형 대형 구축함',         'count': 1},
                {'preset': '052D형 구축함',             'count': 1},
                {'preset': 'J-20 (위룡)',               'count': 1},
                {'preset': 'H-6 (폭격기)',              'count': 1},
                {'preset': 'DF-21D (대함 탄도)',        'count': 1},
                {'preset': 'DF-17 (극초음속 활공)',     'count': 1},
                {'preset': 'KN-23 (북한 이스칸데르)',   'count': 1},
                {'preset': 'YJ-12 (초음속 순항)',       'count': 2},
                {'preset': '039형 잠수함 (송급)', 'count': 1},
            ],
        },
    }

    cfg = SCENARIOS.get(scenario, SCENARIOS['mixed'])

    MC_N = int(sys.argv[2]) if len(sys.argv) > 2 else 200

    print("=" * 66)
    print(f"  이지스 기동전단 통합 방어 시뮬레이터 v7.0  [시나리오: {scenario}]")
    print("=" * 66)

    result = run_v7_simulation(cfg)

    print(f"  시뮬 종료 시각  : {result['sim_time']:.0f}s")
    print(f"  총 위협 수      : {result['total_threats']}발/기")
    print(f"  요격 성공       : {result['intercepted_threats']}발/기")
    print(f"  요격률          : {result['intercept_rate']:.1%}")
    print(f"  아군 피격       : {result['friendly_hits']}회")
    print(f"  적 피격         : {result['enemy_hits']}회")
    print(f"  적 위협 격침    : {result['enemy_ships_destroyed']}기/척")
    print(f"  아군 함정 손실  : {result['friendly_ships_lost']}척")
    print(f"  총 비용         : ${result['total_cost']:,.0f}")
    print("-" * 66)
    print("  교전 로그:")
    for t, msg in result['log']:
        print(f"  [{t:5.0f}s] {msg}")
    print("=" * 66)

    # ── MC 분석 + 보고서 생성 ────────────────────────────────────────────────
    print(f"\n  몬테카를로 분석 ({MC_N}회) 시작...")
    mc = monte_carlo_v7(cfg, n=MC_N, desc=scenario)

    print(f"  MC 평균 요격률 : {mc['mean_intercept']:.1%}  "
          f"(±{mc['std_intercept']:.1%})")
    print(f"  완전 요격 비율 : {mc['full_pass_rate']:.1%}")

    img_path  = f'이지스_기동전단_v7_{scenario}_분석.png'
    xlsx_path = f'이지스_기동전단_v7_{scenario}_보고서.xlsx'

    plot_v7(result, mc, cfg, img_path=img_path)
    save_excel_report_v7(result, mc, cfg,
                         img_path=img_path, xlsx_path=xlsx_path)

    print("\n  ※ 실행 방법:")
    print(f"     python engine_combat.py [시나리오] [MC횟수]")
    print(f"     예) python engine_combat.py mixed 500")
    print("=" * 66)

